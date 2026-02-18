import discord
from discord.ext import commands
import subprocess
import sys
import openpyxl
from openpyxl import load_workbook
import os

# Database imports
from db_helper import (
    init_database,
    get_all_usernames,
    get_database_stats,
    get_user_stats,
    get_user_stats_with_deltas,
    get_user_meta,
    get_all_user_meta,
    update_user_meta,
    user_exists,
    get_discord_id,
    set_discord_link,
    get_all_user_links,
    get_default_username,
    set_default_username,
    get_all_default_users,
    get_tracked_streaks,
    update_tracked_streaks,
    get_all_tracked_streaks,
    add_tracked_user,
    remove_tracked_user,
    is_tracked_user,
    get_tracked_users,
    set_tracked_users,
    get_db_connection,
    delete_user,
    register_user,
    is_registered_user,
    store_uuid,
    get_uuid_for_username,
    resolve_username_to_uuid,
    update_username_for_uuid,
    get_hotbar_layouts,
    # Guild functions
    add_tracked_guild,
    remove_tracked_guild,
    is_tracked_guild,
    is_registered_guild,
    get_tracked_guilds,
    get_guilds_for_periodic_updates,
    guild_has_tracked_members,
    update_guild_exp,
    get_guild_exp,
    get_all_guilds,
    guild_exists,
    log_guild_historical_snapshot,
    get_guild_historical_data
)
import re
import shutil
import sqlite3
from zoneinfo import ZoneInfo
import json
from pathlib import Path
import io
import math
import requests
import asyncio
import time
from typing import Optional, Union
from api_get import experience_delta_to_level_delta
try:
    from PIL import Image, ImageDraw, ImageFont
except Exception:
    Image = None

# Get the directory where discord.py is located
BOT_DIR = Path(__file__).parent.absolute()

# tracked/users + creator info
# TRACKED_FILE - now in database (tracked_users table)
# JSON files are now replaced with database tables
# USER_LINKS_FILE, USER_COLORS_FILE, DEFAULT_USERS_FILE, TRACKED_STREAKS_FILE - now in database
CREATOR_NAME = "chuckegg"
# Optionally set a numeric Discord user ID for direct DM (recommended for reliability)
CREATOR_ID = "542467909549555734"
ADMIN_IDS = ["542467909549555734", "1040340714824937554"]
ADMIN_NAMES = ["chuckegg", "felix.6554"]
CREATOR_TZ = ZoneInfo("America/New_York")

START_TIME = time.time()

# API Request Tracking System
class APIRequestTracker:
    """Track API requests in 5-minute windows."""
    def __init__(self):
        self.lock = asyncio.Lock()
        self.current_window_start = None
        self.requests = {
            'player': 0,
            'guild': 0,
            'status': 0,
            'other': 0
        }
    
    def _get_current_window(self):
        """Get the start time of the current 5-minute window."""
        now = time.time()
        # Round down to nearest 5-minute mark
        return int(now // 300) * 300
    
    async def log_request(self, request_type: str):
        """Log an API request of the given type."""
        async with self.lock:
            current_window = self._get_current_window()
            
            # Reset if we're in a new window
            if self.current_window_start != current_window:
                self.current_window_start = current_window
                self.requests = {
                    'player': 0,
                    'guild': 0,
                    'status': 0,
                    'other': 0
                }
            
            # Increment the appropriate counter
            if request_type in self.requests:
                self.requests[request_type] += 1
            else:
                self.requests['other'] += 1
    
    async def get_stats(self):
        """Get current API request statistics."""
        async with self.lock:
            current_window = self._get_current_window()
            
            # Reset if we're in a new window
            if self.current_window_start != current_window:
                self.current_window_start = current_window
                self.requests = {
                    'player': 0,
                    'guild': 0,
                    'status': 0,
                    'other': 0
                }
            
            total = sum(self.requests.values())
            return {
                'total': total,
                'breakdown': dict(self.requests),
                'window_start': self.current_window_start
            }

API_TRACKER = APIRequestTracker()

# Database Repair Functions
DB_PATH = str(BOT_DIR / "stats.db")
BACKUP_DIR = str(BOT_DIR / "backups")

def check_database_integrity(db_path):
    """
    Check if database is corrupted.
    Returns: (is_valid, error_message)
    """
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Run integrity check
        cursor.execute("PRAGMA integrity_check;")
        result = cursor.fetchone()
        
        conn.close()
        
        if result[0] == "ok":
            return True, None
        else:
            return False, result[0]
            
    except sqlite3.DatabaseError as e:
        return False, str(e)
    except Exception as e:
        return False, f"Unexpected error: {str(e)}"

def find_latest_valid_backup():
    """
    Find the most recent non-corrupted backup file.
    Returns: (backup_path, timestamp_str) or (None, None)
    """
    if not os.path.exists(BACKUP_DIR):
        return None, None
        
    backups = [f for f in os.listdir(BACKUP_DIR) if f.startswith("stats_") and f.endswith(".db")]
    if not backups:
        return None, None
        
    # Sort by filename (which includes timestamp) - newest first
    backups.sort(reverse=True)
    
    # Check each backup until we find a valid one
    for backup_file in backups:
        backup_path = os.path.join(BACKUP_DIR, backup_file)
        is_valid, _ = check_database_integrity(backup_path)
        if is_valid:
            # Extract timestamp from filename (stats_2026-02-08_23-00-00.db)
            timestamp_str = backup_file.replace("stats_", "").replace(".db", "")
            return backup_path, timestamp_str
    
    return None, None

def try_repair_database(db_path):
    """
    Attempt to repair database using SQLite's dump and restore method.
    Returns: (success, message)
    """
    try:
        # Create temporary file for dump
        dump_file = f"{db_path}.dump.sql"
        repaired_file = f"{db_path}.repaired"
        
        # Dump the database (this reads all accessible data)
        conn = sqlite3.connect(db_path)
        with open(dump_file, 'w') as f:
            for line in conn.iterdump():
                f.write(f'{line}\n')
        conn.close()
        
        # Create new database from dump
        if os.path.exists(repaired_file):
            os.remove(repaired_file)
            
        conn_new = sqlite3.connect(repaired_file)
        cursor = conn_new.cursor()
        with open(dump_file, 'r') as f:
            cursor.executescript(f.read())
        conn_new.commit()
        conn_new.close()
        
        # Check integrity of repaired database
        is_valid, error = check_database_integrity(repaired_file)
        if not is_valid:
            os.remove(dump_file)
            os.remove(repaired_file)
            return False, f"Repaired database still has issues: {error}"
            
        # Backup corrupted database
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        corrupted_backup = f"stats.db.corrupted.{timestamp}"
        shutil.copy2(db_path, corrupted_backup)
        
        # Replace with repaired database
        shutil.move(repaired_file, db_path)
        
        # Clean up dump file
        os.remove(dump_file)
        
        return True, f"Database repaired successfully! Corrupted database saved to: {corrupted_backup}"
        
    except Exception as e:
        # Clean up temporary files
        if os.path.exists(dump_file):
            os.remove(dump_file)
        if os.path.exists(repaired_file):
            os.remove(repaired_file)
        return False, f"Repair failed: {str(e)}"

def restore_from_backup(backup_path, db_path):
    """
    Restore database from backup.
    Returns: (success, message)
    """
    if not os.path.exists(backup_path):
        return False, f"Backup file not found: {backup_path}"
        
    try:
        # First, validate the backup (should already be validated, but double-check)
        is_valid, error = check_database_integrity(backup_path)
        if not is_valid:
            return False, f"Backup file is corrupted: {error}"
            
        # Create a backup of the corrupted database
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        corrupted_backup = f"stats.db.corrupted.{timestamp}"
        if os.path.exists(db_path):
            shutil.copy2(db_path, corrupted_backup)
        
        # Restore from backup
        shutil.copy2(backup_path, db_path)
        
        # Verify the restored database
        is_valid, error = check_database_integrity(db_path)
        if is_valid:
            return True, f"Database restored from backup. Corrupted database saved to: {corrupted_backup}"
        else:
            return False, f"Restored database failed integrity check: {error}"
            
    except Exception as e:
        return False, f"Error during restore: {str(e)}"

async def perform_database_repair():
    """
    Perform database repair with the following strategy:
    1. Try to repair the database using dump/restore
    2. ONLY IF REPAIR FAILS: Find latest valid backup and restore from it
    
    Returns: (success, message, backup_used)
    """
    # Check if database is corrupted
    is_valid, error = check_database_integrity(DB_PATH)
    if is_valid:
        return True, "Database is healthy - no repair needed", None
    
    # Try repair first
    success, message = try_repair_database(DB_PATH)
    if success:
        return True, message, None
    
    # Repair failed - find latest valid backup
    backup_path, timestamp_str = find_latest_valid_backup()
    if not backup_path:
        return False, "Repair failed and no valid backups found", None
    
    # Restore from backup
    success, restore_message = restore_from_backup(backup_path, DB_PATH)
    if success:
        return True, f"Repair failed. {restore_message}\nRestored from backup: {timestamp_str}", timestamp_str
    else:
        return False, f"Repair failed and restore also failed: {restore_message}", None

# Font cache to avoid repeatedly searching for fonts
_FONT_CACHE = {}

def _get_font_path(font_name: str) -> str:
    """Find the full path to a TrueType font file.
    
    Searches local fonts directory first, then common system font directories 
    on Windows, Linux, and macOS.
    
    Args:
        font_name: Name of the font file (e.g., 'DejaVuSans.ttf')
    
    Returns:
        Full path to the font file if found, otherwise returns the font_name as-is
        (will fall back to default font if not found)
    """
    if font_name in _FONT_CACHE:
        return _FONT_CACHE[font_name]
    
    # Check local fonts directory first (bundled with bot)
    local_fonts_dir = os.path.join(BOT_DIR, 'fonts')
    local_font_path = os.path.join(local_fonts_dir, font_name)
    if os.path.exists(local_font_path):
        _FONT_CACHE[font_name] = local_font_path
        return local_font_path
    
    # Common font directories by OS
    font_dirs = []
    
    if sys.platform == 'win32':
        # Windows font directories
        font_dirs = [
            os.path.expandvars(r'%WINDIR%\Fonts'),
            os.path.expandvars(r'%SystemRoot%\Fonts'),
        ]
    elif sys.platform == 'darwin':
        # macOS font directories
        font_dirs = [
            os.path.expanduser('~/Library/Fonts'),
            '/Library/Fonts',
            '/System/Library/Fonts',
        ]
    else:
        # Linux and other Unix-like systems
        font_dirs = [
            os.path.expanduser('~/.fonts'),
            '/usr/share/fonts',
            '/usr/local/share/fonts',
            '/usr/share/fonts/truetype',
            '/usr/share/fonts/truetype/dejavu',
        ]
    
    # Search for the font
    for directory in font_dirs:
        font_path = os.path.join(directory, font_name)
        if os.path.exists(font_path):
            _FONT_CACHE[font_name] = font_path
            return font_path
    
    # If not found, return the original name and let Pillow handle it
    _FONT_CACHE[font_name] = font_name
    return font_name

def _load_font(font_name: str, font_size: int):
    """Load a font from the fonts directory."""
    path = os.path.join(BOT_DIR, 'fonts', font_name)
    try:
        return ImageFont.truetype(path, font_size)
    except:
        return ImageFont.load_default()

def _load_font_with_fallback(font_name: str, font_size: int):
    """Load primary font and fallback font for emoji/symbol support."""
    primary = _load_font(font_name, font_size)
    
    # Try to load a font with better symbol support as fallback
    fallback = None
    try:
        # Try Noto Sans which has good symbol coverage
        noto_path = '/usr/share/fonts/truetype/noto/NotoSans-Regular.ttf'
        if os.path.exists(noto_path):
            fallback = ImageFont.truetype(noto_path, font_size)
    except:
        pass
    
    return primary, fallback

def _draw_text_with_fallback(draw, xy, text, font, fallback_font, fill):
    """Draw text trying primary font, falling back to secondary for unsupported chars."""
    if fallback_font is None:
        draw.text(xy, text, font=font, fill=fill)
        return
    
    # Check each character and use fallback for symbols that don't render well
    x, y = xy
    for char in text:
        # These are characters known to have issues with DejaVu but work in Noto
        needs_fallback = ord(char) > 0x2000 and ord(char) not in range(0x20, 0x7F)
        
        current_font = fallback_font if needs_fallback else font
        draw.text((x, y), char, font=current_font, fill=fill)
        
        # Get bbox to advance x position
        bbox = draw.textbbox((x, y), char, font=current_font)
        x = bbox[2]

LOCK_FILE = str(BOT_DIR / "stats.xlsx.lock")  # Kept for backward compatibility
DB_FILE = BOT_DIR / "stats.db"

class FileLock:
    """Simple file-based lock to prevent concurrent Excel writes."""
    def __init__(self, lock_file, timeout=20, delay=0.1):
        self.lock_file = lock_file
        self.timeout = timeout
        self.delay = delay
        self._fd = None

    def __enter__(self):
        start_time = time.time()
        while True:
            try:
                # Exclusive creation of lock file
                self._fd = os.open(self.lock_file, os.O_CREAT | os.O_EXCL | os.O_RDWR)
                break
            except FileExistsError:
                # Check for stale lock (older than 60 seconds)
                try:
                    if os.path.exists(self.lock_file) and time.time() - os.stat(self.lock_file).st_mtime > 300:
                        try:
                            os.remove(self.lock_file)
                        except OSError:
                            pass
                        continue
                except OSError:
                    pass
                if time.time() - start_time >= self.timeout:
                    raise TimeoutError(f"Could not acquire lock on {self.lock_file}")
                time.sleep(self.delay)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._fd is not None:
            os.close(self._fd)
            try:
                os.remove(self.lock_file)
            except OSError:
                pass

# Global Cache System
class StatsCache:
    def __init__(self):
        self.data = {}
        self.last_mtime = 0
        self.lock = asyncio.Lock()
        self.db_path = DB_FILE

    async def get_data(self):
        """Get cached data, reloading from database if it has changed."""
        if not self.db_path.exists():
            return {}

        try:
            current_mtime = self.db_path.stat().st_mtime
            
            # Double-check locking to prevent multiple reloads
            if current_mtime > self.last_mtime:
                async with self.lock:
                    # Check again inside lock
                    if current_mtime > self.last_mtime:
                        print(f"[CACHE] Database changed. Reloading stats cache...")
                        self.data = await asyncio.to_thread(self._load_from_database)
                        self.last_mtime = current_mtime
                        print(f"[CACHE] Reload complete. Cached {len(self.data)} users.")
            
            return self.data
        except Exception as e:
            print(f"[CACHE] Error accessing cache: {e}")
            return self.data

    def _load_from_database(self):
        """Load all user data from SQLite database."""
        cache = {}

        try:
            # Get all usernames
            usernames = get_all_usernames()
            
            for username in usernames:
                try:
                    # Get stats with deltas
                    stats = get_user_stats_with_deltas(username)
                    
                    # Get metadata
                    meta_db = get_user_meta(username)
                    
                    if not stats:
                        continue
                    
                    user_cache = {"stats": stats, "meta": {}}
                    
                    # Calculate level from stats if available
                    level = int(stats.get('level', {}).get('lifetime', 0))
                    if level == 0 and 'experience' in stats:
                        level = int(stats['experience']['lifetime'] / 5000)
                    
                    # Get user color info from database
                    ign_color = meta_db.get('ign_color') if meta_db else None
                    if not ign_color:
                        ign_color = get_rank_color_hex(meta_db.get('rank') if meta_db else None)
                    
                    g_tag = meta_db.get('guild_tag') if meta_db else None
                    g_hex = meta_db.get('guild_hex', '#AAAAAA') if meta_db else '#AAAAAA'
                    
                    user_cache["meta"] = {
                        "level": level,
                        "icon": get_prestige_icon(level),
                        "ign_color": ign_color,
                        "guild_tag": g_tag,
                        "guild_hex": g_hex,
                        "username": username
                    }
                    
                    cache[username] = user_cache
                    
                except Exception as e:
                    print(f"[CACHE] Error loading user {username}: {e}")
                    continue
            
            return cache
            
        except Exception as e:
            print(f"[CACHE] Error loading from database: {e}")
            return {}

    async def update_cache_entry(self, username: str, processed_stats: dict):
        """Update a single user's cache entry from api_get.py output without reloading database."""
        async with self.lock:
            # Ensure data dict exists
            if not self.data:
                self.data = {}
            
            # Get meta from database to ensure colors/guilds are preserved
            meta_db = get_user_meta(username)
            
            # Calculate meta
            level = int(processed_stats.get('level', {}).get('lifetime', 0))
            if level == 0 and 'experience' in processed_stats:
                level = int(processed_stats['experience']['lifetime'] / 5000)
            
            ign_color = meta_db.get('ign_color') if meta_db else None
            if not ign_color:
                ign_color = get_rank_color_hex(meta_db.get('rank') if meta_db else None)
            
            g_tag = meta_db.get('guild_tag') if meta_db else None
            g_hex = meta_db.get('guild_hex', '#AAAAAA') if meta_db else '#AAAAAA'

            # Update cache
            user_cache = {"stats": processed_stats, "meta": {"level": level, "icon": get_prestige_icon(level), "ign_color": ign_color, "guild_tag": g_tag, "guild_hex": g_hex, "username": username}}
            self.data[username] = user_cache

            # Update streaks if applicable
            try:
                update_streaks_from_stats(username, processed_stats)
            except Exception as e:
                print(f"[STREAK] Failed to update streaks for {username}: {e}")
            
            # Update mtime to prevent the next get_data call from reloading
            if self.db_path.exists():
                self.last_mtime = self.db_path.stat().st_mtime
            
            return user_cache

    async def refresh(self):
        """Force reload of cache from database."""
        async with self.lock:
            print("[CACHE] Forcing cache refresh...")
            self.data = await asyncio.to_thread(self._load_from_database)
            if self.db_path.exists():
                self.last_mtime = self.db_path.stat().st_mtime
            print(f"[CACHE] Refresh complete. Cached {len(self.data)} users.")
    
    async def invalidate(self):
        """Invalidate cache to force reload on next access."""
        async with self.lock:
            self.last_mtime = 0
            print("[CACHE] Cache invalidated. Will reload on next access.")

STATS_CACHE = StatsCache()

def safe_save_workbook(wb, filepath: str) -> bool:
    """Safely save a workbook using atomic write to prevent corruption.
    
    Writes to a temp file first, then atomically replaces the target file.
    
    Args:
        wb: The openpyxl Workbook object to save
        filepath: Path to the Excel file
        
    Returns:
        bool: True if save succeeded, False otherwise
    """
    temp_path = str(filepath) + ".tmp"
    backup_path = str(filepath) + ".backup"
    
    try:
        # 1. Save to temporary file first
        wb.save(temp_path)
        
        # 2. Create backup of existing file
        if os.path.exists(str(filepath)):
            try:
                shutil.copy2(str(filepath), backup_path)
            except Exception as backup_err:
                print(f"[WARNING] Failed to create backup: {backup_err}")
        
        # 3. Atomic replace
        os.replace(temp_path, str(filepath))
        print(f"[SAVE] Successfully saved: {filepath}")
        
        # 4. Cleanup backup
        if os.path.exists(backup_path):
            try:
                os.remove(backup_path)
            except Exception:
                pass  # Not critical if backup removal fails
        
        return True
        
    except Exception as save_err:
        print(f"[ERROR] Failed to save workbook: {save_err}")
        # Clean up temp file
        if os.path.exists(temp_path):
            try:
                os.remove(temp_path)
            except:
                pass
        
        return False
        
    finally:
        # Always try to close the workbook
        try:
            wb.close()
            print(f"[CLEANUP] Workbook closed")
        except Exception as close_err:
            print(f"[WARNING] Error closing workbook: {close_err}")

# sanitize output for Discord (remove problematic unicode/control chars)
def sanitize_output(text: str) -> str:
    if text is None:
        return ""
    # Replace a few common emoji with ASCII labels
    replacements = {
        '✅': '[OK]',
        '❌': '[ERROR]',
        '⚠️': '[WARNING]',
        '📊': '[DATA]',
        '📋': '[INFO]',
        '⏭️': '[SKIP]',
    }
    for k, v in replacements.items():
        text = text.replace(k, v)

    # Remove C0 control chars except newline and tab
    text = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', str(text))
    # Collapse very long whitespace
    text = re.sub(r"\s{3,}", ' ', text)
    return text


def validate_and_normalize_ign(ign: str):
    s = str(ign or '').strip()
    if not re.fullmatch(r'^[A-Za-z0-9_]{3,16}$', s):
        return False, None
    try:
        r = requests.get(f'https://api.mojang.com/users/profiles/minecraft/{s}', timeout=5)
        if r.status_code == 200:
            data = r.json()
            name = data.get('name')
            if isinstance(name, str) and re.fullmatch(r'^[A-Za-z0-9_]{3,16}$', name):
                return True, name
            return True, s
        if r.status_code in (204, 404):
            return False, None
        return True, s
    except Exception:
        return True, s


def _to_number(val):
    if val is None: return 0
    if isinstance(val, (int, float)): return val
    s = str(val).replace(".", "").replace(",", "").strip()
    try: return float(s)
    except: return 0

# Helper function to run scripts with proper working directory
def run_script(script_name, args, timeout=30):
    """Run a Python script in the bot directory with proper working directory"""
    return subprocess.run(
        [sys.executable, script_name, *args],
        cwd=str(BOT_DIR),
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        timeout=timeout
    )

def run_script_batch(script_name, args):
    """Run a batch script with extended timeout (5 minutes for large user lists)"""
    return subprocess.run(
        [sys.executable, script_name, *args],
        cwd=str(BOT_DIR),
        capture_output=True,
        text=True,
        encoding='utf-8',
        errors='replace',
        timeout=300  # 5 minutes for batch operations
    )

async def ensure_user_cached(ign: str, timeout: int = 60) -> tuple[bool, str]:
    """
    Ensure a user's data is cached. If not in cache, fetch it.
    Returns (success, actual_ign) tuple.
    """
    cache_data = await STATS_CACHE.get_data()
    key = ign.casefold()
    
    # Check if already cached
    for name, data in cache_data.items():
        if name.casefold() == key:
            return True, name
    
    # Not cached, fetch it
    print(f"[CACHE] User {ign} not in cache, fetching now...")
    try:
        result = await asyncio.to_thread(run_script, "api_get.py", ["-ign", ign], timeout=timeout)
        if result.returncode != 0:
            print(f"[CACHE] Failed to fetch {ign}: {result.stderr}")
            return False, ign
        
        # Verify it's now cached
        cache_data = await STATS_CACHE.get_data()
        for name, data in cache_data.items():
            if name.casefold() == key:
                return True, name
        
        print(f"[CACHE] Fetched {ign} but still not in cache")
        return False, ign
    except Exception as e:
        print(f"[CACHE] Exception fetching {ign}: {e}")
        return False, ign

# additional imports for background tasks
import asyncio
import datetime
import random

def format_playtime(seconds: int) -> str:
    if not isinstance(seconds, (int, float)) or seconds <= 0:
        return "0s"
    seconds = int(seconds)
    days = seconds // 86400
    rem = seconds % 86400
    hours = rem // 3600
    minutes = (rem % 3600) // 60
    if days > 0: return f"{days}d {hours}h"
    if hours > 0: return f"{hours}h {minutes}m"
    return f"{minutes}m"


# Prestige icons per 100 levels (index 0 = levels 0-99)
PRESTIGE_ICONS = [
    "❤", "✙", "✫", "✈", "✠", "♙", "⚡", "☢", "✏", "☯",
    "☃️", "۞", "✤", "♫", "♚", "❉", "Σ", "￡", "✖", "❁",
    "✚", "✯", "✆", "❥", "☾⋆⁺", "⚜", "✦", "⚝", "✉", "ツ",
    "❣", "✮", "✿", "✲", "❂", "ƒ", "$", "⋚⋚", "Φ", "✌",
]

# Prestige colors (RGB tuples for Discord embed colors)
# Levels: 0, 100, 200, 300, 400, 500, 600, 700, 800, 900, 1000+
PRESTIGE_COLORS = {
    0: (119, 119, 119),      # GRAY (§7)
    100: (255, 255, 255),    # WHITE (§f)
    200: (255, 85, 85),      # RED (§c)
    300: (255, 170, 0),      # GOLD (§6)
    400: (255, 255, 85),     # YELLOW (§e)
    500: (85, 255, 85),      # LIGHT_GREEN (§a)
    600: (0, 170, 170),      # DARK_AQUA (§3)
    700: (170, 0, 170),      # DARK_PURPLE (§5)
    800: (255, 85, 255),     # LIGHT_PURPLE (§d)
    900: None,               # Rainbow
    1000: (255, 255, 255),   # WHITE (§f)
    1100: (170, 170, 170),   # &7 -> GRAY
    1200: (255, 85, 85),     # &c -> RED
    1300: (255, 170, 0),     # &6 -> GOLD
    1400: (255, 255, 85),    # &e -> YELLOW
    1500: (85, 255, 85),     # &a -> GREEN
    1600: (0, 170, 170),     # &3 -> DARK_AQUA
    1700: (255, 85, 255),    # &d -> LIGHT_PURPLE
    1800: (170, 0, 170),     # &5 -> DARK_PURPLE
    1900: None,              # Rainbow
    2000: None,              # Rainbow (multi-color pattern)
    2100: None,              # Rainbow (multi-color pattern)
    2200: None,              # Rainbow (multi-color pattern)
    2300: None,              # Rainbow (multi-color pattern)
    2400: None,              # Rainbow (multi-color pattern)
    2500: None,              # Rainbow (multi-color pattern)
    2600: None,              # Rainbow (multi-color pattern)
    2700: None,              # Rainbow (multi-color pattern)
    2800: None,              # Rainbow (multi-color pattern)
    2900: None,              # Rainbow
    3000: None,              # Rainbow (multi-color pattern)
    3100: None,              # Rainbow (multi-color pattern)
    3200: None,              # Rainbow (multi-color pattern)
    3300: None,              # Rainbow (multi-color pattern)
    3400: None,              # Rainbow (multi-color pattern)
    3500: None,              # Rainbow (multi-color pattern)
    3600: None,              # Rainbow (multi-color pattern)
    3700: None,              # Rainbow (multi-color pattern)
    3800: None,              # Rainbow (multi-color pattern)
    3900: None,              # Rainbow
    4000: None,              # Rainbow (multi-color pattern)
    4100: (255, 255, 255),   # &f -> WHITE (single color)
    4200: (255, 85, 85),     # &c -> RED (single color)
    4300: (255, 170, 0),     # &6 -> GOLD (single color)
    4400: (255, 255, 85),    # &e -> YELLOW (single color)
    4500: (85, 255, 85),     # &a -> GREEN (single color)
    4600: (0, 170, 170),     # &3 -> DARK_AQUA (single color)
    4700: (255, 85, 255),    # &d -> LIGHT_PURPLE (single color)
    4800: (170, 0, 170),     # &5 -> DARK_PURPLE (single color)
    4900: None,              # Rainbow
    5000: None,              # Rainbow (High level default)
}


def get_prestige_icon(level: int) -> str:
    try:
        lvl = int(level)
    except Exception:
        lvl = 0
    base = (lvl // 100) * 100
    # If a raw pattern exists and contains an icon, extract it (strip color codes)
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    if raw:
        stripped = re.sub(r'&[0-9a-fA-F]', '', raw)
        # Look for content inside brackets
        m = re.search(r"\[(.*?)\]", stripped)
        if m:
            inner = m.group(1)
            # remove leading digits (the level number) to get icon
            icon = re.sub(r'^[0-9]+', '', inner).strip()
            if icon:
                return icon

    # Fallback to PRESTIGE_ICONS list
    idx = max(0, lvl // 100)
    if idx >= len(PRESTIGE_ICONS):
        idx = len(PRESTIGE_ICONS) - 1
    return PRESTIGE_ICONS[idx]

def get_prestige_color(level: int) -> tuple:
    """Get RGB color tuple for a given prestige level.
    Supports levels 0-1000. Returns default dark gray for levels outside this range.
    """
    try:
        lvl = int(level)
    except Exception:
        lvl = 0

    base = (lvl // 100) * 100

    # If a raw pattern exists for this prestige base, prefer its first color code
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    if raw:
        m = re.search(r'&([0-9a-fA-F])', raw)
        if m:
            code = m.group(1).lower()
            hexcol = MINECRAFT_CODE_TO_HEX.get(code)
            if hexcol:
                return hex_to_rgb(hexcol)

    # Otherwise fall back to explicit PRESTIGE_COLORS mapping
    for prestige_level in sorted(PRESTIGE_COLORS.keys(), reverse=True):
        if lvl >= prestige_level:
            color = PRESTIGE_COLORS[prestige_level]
            # Handle Rainbow (None) by returning a default color or cycling
            if color is None:
                return (255, 100, 200)
            return color

    # Fallback to gray if below 0
    return (119, 119, 119)

def level_to_experience_required(level: int) -> int:
    """Calculate the total XP required to reach a specific level from level 1.
    
    Uses the Wool Games prestige system:
    - Each prestige is 100 levels and requires 490,000 XP total
    - Within each prestige:
      - Level X+0 to X+1: 1,000 XP
      - Level X+1 to X+2: 2,000 XP
      - Level X+2 to X+3: 3,000 XP
      - Level X+3 to X+4: 4,000 XP
      - Level X+4 to X+5: 5,000 XP
      - Level X+5 to X+100: 5,000 XP each (95 levels)
    
    Args:
        level: Target level (1-based, as displayed in-game)
    
    Returns:
        Total XP required to reach this level
    """
    if level <= 1:
        return 0
    
    # Convert to 0-based for calculation
    level_zero_based = level - 1
    
    XP_PER_PRESTIGE = 490000
    prestige_count = level_zero_based // 100
    level_in_prestige = level_zero_based % 100
    
    # XP from completed prestiges
    total_xp = prestige_count * XP_PER_PRESTIGE
    
    # XP within current prestige
    if level_in_prestige > 0:
        # Levels 0-1: 1000 XP
        total_xp += 1000
    if level_in_prestige > 1:
        # Levels 1-2: 2000 XP
        total_xp += 2000
    if level_in_prestige > 2:
        # Levels 2-3: 3000 XP
        total_xp += 3000
    if level_in_prestige > 3:
        # Levels 3-4: 4000 XP
        total_xp += 4000
    if level_in_prestige > 4:
        # Levels 4-5: 5000 XP
        total_xp += 5000
    if level_in_prestige > 5:
        # Levels 5-100: 5000 XP each
        total_xp += (level_in_prestige - 5) * 5000
    
    return total_xp

def get_xp_for_next_level(current_level: int) -> int:
    """Get the XP required to go from current_level to current_level+1.
    
    Args:
        current_level: Current level (1-based)
    
    Returns:
        XP required for the next level
    """
    level_in_prestige = (current_level - 1) % 100
    
    if level_in_prestige == 0:
        return 1000
    elif level_in_prestige == 1:
        return 2000
    elif level_in_prestige == 2:
        return 3000
    elif level_in_prestige == 3:
        return 4000
    else:
        return 5000

def get_ansi_color_code(level: int) -> str:
    """Get ANSI color code for a given prestige level."""
    color = get_prestige_color(level)
    
    # Map RGB to closest basic ANSI color for Discord compatibility
    r, g, b = color
    
    # Determine which basic ANSI color is closest
    if r > 200 and g > 200 and b > 200:
        return "\u001b[0;37m"  # White
    elif r < 100 and g < 100 and b < 100:
        return "\u001b[0;30m"  # Gray
    elif r > 200 and g < 100 and b < 100:
        return "\u001b[0;31m"  # Red
    elif r > 200 and g > 150 and b < 100:
        return "\u001b[0;33m"  # Yellow/Gold
    elif r < 100 and g > 200 and b < 100:
        return "\u001b[0;32m"  # Green
    elif r < 100 and g > 150 and b > 150:
        return "\u001b[0;36m"  # Cyan
    elif r > 150 and g < 100 and b > 150:
        return "\u001b[0;35m"  # Magenta/Pink
    elif r > 200 and g > 200 and b < 100:
        return "\u001b[0;33m"  # Yellow
    else:
        return "\u001b[0;37m"  # Default White

def make_bold_ansi(code: str) -> str:
    """Convert a basic ANSI color code to bold variant.
    Expects codes like "\u001b[0;33m" and returns "\u001b[1;33m".
    """
    if not code:
        return code
    # If already contains bold flag, return as-is
    if "1;" in code or "\u001b[1m" in code:
        return code
    # If code already contains bold or is empty, return it
    if not code:
        return code
    if "1;" in code or "\u001b[1m" in code:
        return code
    # For any CSI like '\x1b[...m', insert '1;' after the '[' if not present
    m = re.match(r"^(\x1b\[)(?!1;)(.*)m$", code)
    if m:
        return f"{m.group(1)}1;{m.group(2)}m"
    return code


# Mapping of Minecraft color codes (§) to approximate ANSI codes for inline coloring
# Official Minecraft-ish main hex colors for § codes (main hex values)
MINECRAFT_CODE_TO_HEX = {
    '0': '#000000',
    '1': '#0000AA',
    '2': '#00AA00',
    '3': '#00AAAA',
    '4': '#AA0000',
    '5': '#AA00AA',
    '6': '#FFAA00',
    '7': '#AAAAAA',
    '8': '#555555',
    '9': '#5555FF',
    'a': '#55FF55',
    'b': '#55FFFF',
    'c': '#FF5555',
    'd': '#FFD3F5',
    'e': '#FFFF55',
    'f': '#FFFFFF',
}

# Minecraft color name to hex (from Hypixel API)
MINECRAFT_NAME_TO_HEX = {
    'BLACK': '#000000',
    'DARK_BLUE': '#0000AA',
    'DARK_GREEN': '#00AA00',
    'DARK_AQUA': '#00AAAA',
    'DARK_RED': '#AA0000',
    'DARK_PURPLE': '#AA00AA',
    'GOLD': '#FFAA00',
    'GRAY': '#AAAAAA',
    'DARK_GRAY': '#555555',
    'BLUE': '#5555FF',
    'GREEN': '#55FF55',
    'AQUA': '#55FFFF',
    'RED': '#FF5555',
    'LIGHT_PURPLE': '#FF55FF',
    'YELLOW': '#FFFF55',
    'WHITE': '#FFFFFF',
}

def hex_to_rgb(h: str) -> tuple:
    h = h.lstrip('#')
    return tuple(int(h[i:i+2], 16) for i in (0, 2, 4))

def hex_to_ansi(h: str, background: bool = False) -> str:
    r, g, b = hex_to_rgb(h)
    if background:
        return f"\u001b[48;2;{r};{g};{b}m"
    return f"\u001b[38;2;{r};{g};{b}m"

def rgb_to_ansi256_index(r: int, g: int, b: int) -> int:
    """Convert RGB 0-255 to xterm-256 color index."""
    # Grayscale approximation
    if r == g == b:
        if r < 8:
            return 16
        if r > 248:
            return 231
        return 232 + int((r - 8) / 247 * 24)

    # 6x6x6 color cube
    ri = int(round((r / 255) * 5))
    gi = int(round((g / 255) * 5))
    bi = int(round((b / 255) * 5))
    return 16 + (36 * ri) + (6 * gi) + bi

def rgb_to_ansi256_escape(r: int, g: int, b: int, background: bool = False) -> str:
    idx = rgb_to_ansi256_index(r, g, b)
    if background:
        return f"\u001b[48;5;{idx}m"
    return f"\u001b[38;5;{idx}m"

def hex_to_ansi256(h: str, background: bool = False) -> str:
    r, g, b = hex_to_rgb(h)
    return rgb_to_ansi256_escape(r, g, b, background=background)

# Map Minecraft color codes to chosen xterm-256 indices for clearer, distinct colors
# These indices were selected to improve visual separation between gold/yellow/green
_MINECRAFT_256_INDEX = {
    '0': 16,   # black
    '1': 19,   # dark_blue
    '2': 28,   # dark_green
    '3': 37,   # dark_aqua
    '4': 124,  # dark_red
    '5': 127,  # dark_purple
    '6': 214,  # gold/orange
    '7': 248,  # gray
    '8': 239,  # dark_gray
    '9': 75,   # blue
    'a': 46,   # bright green
    'b': 51,   # aqua
    'c': 203,  # red
    'd': 201,  # pink
    'e': 227,  # yellow
    'f': 15,   # white
}

def get_rank_color_hex(rank: Optional[str]) -> str:
    """Get the default hex color for a rank."""
    if not rank:
        return "#AAAAAA"  # Gray for no rank/default
    
    rank_upper = rank.upper()
    rank_colors = {
        "ADMIN": "#FF5555",           # Red
        "SUPERSTAR": "#FFAA00",       # Gold (MVP++)
        "MVP_PLUS": "#55FFFF",        # Aqua
        "MVP_PLUS_PLUS": "#FFAA00",   # Gold
        "MVP": "#55FFFF",             # Aqua
        "VIP_PLUS": "#55FF55",        # Green
        "VIP": "#55FF55",             # Green
    }
    return rank_colors.get(rank_upper, "#AAAAAA") # Default to Gray for unlisted ranks

MINECRAFT_CODE_TO_ANSI_SGR = {k: f"\u001b[38;5;{idx}m" for k, idx in _MINECRAFT_256_INDEX.items()}

# Keep the 24-bit hex map for embed accent colors
MINECRAFT_CODE_TO_ANSI = {k: hex_to_ansi(v) for k, v in MINECRAFT_CODE_TO_HEX.items()}

# Patterns for multi-colored prestige prefixes. Key = prestige base (e.g. 1900),
# For flexibility we store raw Minecraft-style color sequences per prestige.
# Each string uses '&' followed by hex code, e.g. '&c[&61&e9&a0&30&5✖&d]'.
# The runtime parser below converts those into (code, text) pieces.
PRESTIGE_RAW_PATTERNS = {
    0: "&7[0❤]",
    100: "&f[100✙]",
    200: "&c[200✫]",
    300: "&6[300✈]",
    400: "&e[400✠]",
    500: "&a[500♙]",
    600: "&3[600⚡]",
    700: "&5[700✠]",
    800: "&d[800✏]",
    900: "&c[&69&e0&a0&b✏&d]",
    1000: "&0[&f1000☯&0]",
    1100: "&0[&81100☃️&0]",
    1200: "&0[&c1200✤&0]",
    1300: "&0[&61300✤&0]",
    1400: "&0[&e1400♫&0]",
    1500: "&0[&a1500♚&0]",
    1600: "&0[&31600❉&0]",
    1700: "&0[&51700Σ&0]",
    1800: "&0[&d1800✖&0]",
    1900: "&c[&61&e9&a0&30&5✖&d]",
    2000: "&0[2&80&700&f❁]",
    2100: "&f[2&710&80&0✚]",
    2200: "&f[2&e20&60&c✯]",
    2300: "&c[2&630&e0&a✆]",
    2400: "&b[2&340&50&d❥]",
    2500: "&f[2&a500&2☾⋆⁺]",
    2600: "&f[2&b60&30&9⚜&1]",
    2700: "&f[2&d700&5✦]",
    2800: "&c[2&480&50&d✉]",
    2900: "&d[&52&39&a0&e0&6✉&c]",
    3000: "&f[&03&80&00&80&0❣&f]",
    3100: "&0[&f3&71&f0&70&f❣&0]",
    3200: "&0[&c3&42&c0&40&c✮&0]",
    3300: "&0[&63&c3&60&c0&6✿&0]",
    3400: "&0[&e3&64&e0&60&e✲&0]",
    3500: "&0[&a3&25&a0&20&a❂&0]",
    3600: "&0[&33&16&30&10&3ƒ&0]",
    3700: "&0[&d3&57&d0&50&d$&0]",
    3800: "&0[&53&48&50&40&5⋚⋚&0]",
    3900: "&4[&63&e9&20&10&5Φ&d]",
    4000: "&0[4&80&70&80&0✌]",
    4900: "&4[&64&e9&a0&30&5✖&d]",
}

def _parse_raw_pattern(raw: str) -> list:
    """Parse a raw pattern into list of (code, text) pieces."""
    parts = []
    cur_code = None
    buf = ''
    i = 0
    while i < len(raw):
        ch = raw[i]
        if ch == '&' and i + 1 < len(raw):
            if buf:
                parts.append((cur_code or 'f', buf))
                buf = ''
            cur_code = raw[i+1].lower()
            i += 2
            continue
        else:
            buf += ch
            i += 1
    if buf:
        parts.append((cur_code or 'f', buf))
    return parts

def get_prestige_segments(level: int, icon: str) -> list:
    """Generate colored text segments for a prestige level."""
    base = (level // 100) * 100
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    segments = []

    # Check if this should be rainbow (explicitly defined as None in colors, or ends in 900)
    # Also treat anything >= 5000 as rainbow based on user feedback
    is_rainbow = (base in PRESTIGE_COLORS and PRESTIGE_COLORS[base] is None) or (base % 1000 == 900) or (base >= 5000)
    
    if raw:
        parts = _parse_raw_pattern(raw)
        concat = ''.join(t for (_, t) in parts)
        m = re.search(r"\d+", concat)
        
        if m:
            num_start, num_end = m.start(), m.end()
            pos = 0
            replaced = False
            
            for code, text in parts:
                part_start = pos
                part_end = pos + len(text)
                pos = part_end
                hexcol = MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF')
                
                if part_end <= num_start or part_start >= num_end:
                    segments.append((hexcol, text))
                    continue
                
                # Prefix before number
                prefix_len = max(0, num_start - part_start)
                if prefix_len > 0:
                    segments.append((hexcol, text[:prefix_len]))
                
                # Replace with actual level
                if not replaced:
                    if is_rainbow:
                        colors_in_span = []
                        pos2 = 0
                        for c_code, c_text in parts:
                            part_s = pos2
                            part_e = pos2 + len(c_text)
                            pos2 = part_e
                            overlap_s = max(part_s, num_start)
                            overlap_e = min(part_e, num_end)
                            if overlap_e > overlap_s:
                                hexcol_span = MINECRAFT_CODE_TO_HEX.get(c_code.lower(), '#FFFFFF')
                                for _ in range(overlap_e - overlap_s):
                                    colors_in_span.append(hexcol_span)
                        
                        if not colors_in_span:
                            RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
                            colors_in_span = [MINECRAFT_CODE_TO_HEX.get(c, '#FFFFFF') for c in RAINBOW_CODES]
                        
                        for i, ch in enumerate(str(level)):
                            col = colors_in_span[i % len(colors_in_span)]
                            segments.append((col, ch))
                    else:
                        segments.append((hexcol, str(level)))
                    replaced = True
                
                # Suffix after number
                suffix_start_in_part = max(0, num_end - part_start)
                if suffix_start_in_part < len(text):
                    segments.append((hexcol, text[suffix_start_in_part:]))
        else:
            segments = [(MINECRAFT_CODE_TO_HEX.get(code, '#FFFFFF'), text) for code, text in parts]
    else:
        if is_rainbow:
            # Default rainbow behavior for undefined high levels (e.g. 4900)
            bracket_col = MINECRAFT_CODE_TO_HEX.get('8', '#555555') # Dark Gray
            segments.append((bracket_col, "["))
            
            RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
            rainbow_hexes = [MINECRAFT_CODE_TO_HEX.get(c, '#FFFFFF') for c in RAINBOW_CODES]
            
            for i, ch in enumerate(str(level)):
                col = rainbow_hexes[i % len(rainbow_hexes)]
                segments.append((col, ch))
                
            segments.append((bracket_col, f"{icon}]"))
        else:
            color = get_prestige_color(level)
            hexcol = '#{:02x}{:02x}{:02x}'.format(*color)
            segments = [(hexcol, f"[{level}{icon}]")]
    
    return segments

def _safe_guild_tag(guild_tag: str) -> Optional[str]:
    """Try to return guild tag, but return None if it contains problematic unicode."""
    if not guild_tag:
        return None
    # Only allow ASCII characters to prevent rendering issues
    try:
        guild_tag.encode('ascii')
        return guild_tag
    except UnicodeEncodeError:
        # Filter out non-ASCII
        cleaned = "".join(c for c in guild_tag if ord(c) < 128)
        return cleaned if cleaned else None

def render_prestige_with_text(level: int, icon: str, ign: str, suffix: str, ign_color: str = None, guild_tag: str = None, guild_color: str = None, two_line: bool = False) -> io.BytesIO:
    """Render a prestige prefix with IGN, optional guild tag, and optional suffix text.
    
    Returns a BytesIO containing the rendered PNG image.
    If Pillow is not available, raises RuntimeError.
    ign_color: Hex color code for the IGN (e.g., '#FF5555')
    guild_tag: Guild tag to display after username (e.g., 'QUEBEC')
    guild_color: Color name from Hypixel API (e.g., 'DARK_AQUA')
    two_line: If True, formats as [level icon] username [guild] on first line, suffix on second line
    """
    if Image is None:
        raise RuntimeError("Pillow not available")
    
    segments = get_prestige_segments(level, icon)
    
    # Add IGN with custom color if specified
    ign_hex = ign_color if ign_color else MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF')
    segments.append((ign_hex, f" {ign}"))
    
    # Add guild tag if provided (with safety check)
    safe_tag = _safe_guild_tag(guild_tag)
    if safe_tag and guild_color:
        if guild_color.startswith('#'):
            guild_hex = guild_color
        else:
            guild_hex = MINECRAFT_NAME_TO_HEX.get(guild_color.upper(), '#FFFFFF')
        segments.append((guild_hex, f" [{safe_tag}]"))
    elif safe_tag:
        segments.append((MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), f" [{safe_tag}]"))
    
    if two_line and suffix:
        # Two-line format: first line ends after guild tag, second line is the suffix
        return _render_text_segments_to_image_multiline([segments, [(MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), suffix)]])
    elif suffix:
        # Single line format: append suffix with " - " prefix
        segments.append((MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), suffix))
    
    return _render_text_segments_to_image(segments)


def _render_text_segments_to_image(segments: list, font=None, padding=(8,6)) -> io.BytesIO:
    """Render colored text segments to a PNG and return a BytesIO."""
    if Image is None:
        raise RuntimeError("Pillow not available")
    if font is None:
        font = _load_font("DejaVuSans.ttf", 18)

    # Measure total size and vertical bounds
    total_w = 0
    min_y = float('inf')
    max_y = float('-inf')
    
    draw_dummy = ImageDraw.Draw(Image.new('RGBA', (1,1)))
    
    # Handle empty segments case
    if not segments:
        min_y, max_y = 0, 0

    for color_hex, text in segments:
        bbox = draw_dummy.textbbox((0, 0), text, font=font)
        # bbox is (left, top, right, bottom)
        w = bbox[2] - bbox[0]
        total_w += w
        
        if bbox[1] < min_y: min_y = bbox[1]
        if bbox[3] > max_y: max_y = bbox[3]

    if min_y == float('inf'): min_y = 0
    if max_y == float('-inf'): max_y = 0
    
    content_h = max_y - min_y
    # Ensure minimal height
    if content_h <= 0: content_h = 10

    img_w = int(total_w + padding[0]*2)
    img_h = int(content_h + padding[1]*2)
    
    img = Image.new('RGBA', (img_w, img_h), (0,0,0,0))
    draw = ImageDraw.Draw(img)

    x = padding[0]
    # Shift drawing so the top of the ink (min_y) aligns with padding[1]
    y_draw = padding[1] - min_y
    
    for color_hex, text in segments:
        try:
            color = tuple(int(color_hex.lstrip('#')[i:i+2], 16) for i in (0,2,4))
        except Exception:
            color = (255,255,255)
        draw.text((x, y_draw), text, font=font, fill=color)
        bbox = draw.textbbox((x, y_draw), text, font=font)
        w = bbox[2] - bbox[0]
        x += w

    out = io.BytesIO()
    img.save(out, format='PNG')
    out.seek(0)
    return out


def _render_text_segments_to_image_multiline(lines: list, font=None, padding=(8,6), line_spacing=2) -> io.BytesIO:
    """Render multiple lines of colored text segments to a PNG.
    
    Args:
        lines: List of segment lists, where each segment list is [(color_hex, text), ...]
        font: Font to use
        padding: Horizontal and vertical padding
        line_spacing: Additional vertical space between lines
    """
    if Image is None:
        raise RuntimeError("Pillow not available")
    if font is None:
        font = _load_font("DejaVuSans.ttf", 26)

    draw_dummy = ImageDraw.Draw(Image.new('RGBA', (1,1)))
    
    # Measure each line
    line_widths = []
    line_heights = []
    for segments in lines:
        line_w = 0
        line_h = 0
        for color_hex, text in segments:
            bbox = draw_dummy.textbbox((0, 0), text, font=font)
            w = bbox[2] - bbox[0]
            h = bbox[3] - bbox[1]
            line_w += w
            line_h = max(line_h, h)
        line_widths.append(line_w)
        line_heights.append(line_h)
    
    # Calculate total image size
    max_w = max(line_widths) if line_widths else 0
    total_h = sum(line_heights) + (len(lines) - 1) * line_spacing if lines else 0
    
    img_w = max_w + padding[0] * 2
    img_h = total_h + padding[1] * 2
    img = Image.new('RGBA', (img_w, img_h), (0,0,0,0))
    draw = ImageDraw.Draw(img)
    
    # Draw each line (center each line horizontally)
    y = padding[1]
    for line_idx, segments in enumerate(lines):
        # Calculate starting x position to center this line
        line_width = line_widths[line_idx]
        x = (img_w - line_width) // 2
        
        for color_hex, text in segments:
            try:
                color = tuple(int(color_hex.lstrip('#')[i:i+2], 16) for i in (0,2,4))
            except Exception:
                color = (255,255,255)
            draw.text((x, y), text, font=font, fill=color)
            bbox = draw.textbbox((x, y), text, font=font)
            w = bbox[2] - bbox[0]
            x += w
        y += line_heights[line_idx] + line_spacing
    
    out = io.BytesIO()
    img.save(out, format='PNG')
    out.seek(0)
    return out


def render_stat_box(label: str, value: str, width: int = 200, height: int = 80):
    """Render a single stat box with label and value using modern card style."""
    if Image is None:
        raise RuntimeError("Pillow not available")
    
    # Determine color based on label content - updated to match new layout
    color = (255, 255, 255)
    l = label.lower()
    
    # Row 1 colors
    if "wins/hour" in l or "kills/hour" in l:
        color = (140, 100, 200)  # Softer purple
    elif "playtime" in l:
        color = (255, 85, 255)  # Magenta
    elif "exp/hour" in l or "exp/game" in l:
        color = (100, 200, 200)  # Muted cyan
    # Row 2+ colors - specific order matters
    elif l == "wins" or l == "kills" or l == "void kills" or l == "explosive kills" or l == "bow kills" or l == "melee kills":
        color = (85, 255, 85)  # Green
    elif l == "losses" or l == "deaths" or l == "void deaths" or l == "explosive deaths" or l == "bow deaths" or l == "melee deaths":
        color = (255, 85, 85)  # Red
    elif "kdr" in l or l == "wlr":
        color = (255, 170, 0)  # Orange for ratios
    elif "coins" in l or l == "layers":
        color = (255, 215, 0)  # Yellow
    elif l == "damage dealt" or l == "sheep thrown" or l == "games played" or l == "magic wool hit":
        color = (85, 170, 255)  # Blue
    elif "/game" in l or "/sheep" in l or "survival rate" in l:
        color = (255, 170, 0)  # Orange for all per-game/per-sheep stats and survival rate
        
    return render_modern_card(label, value, width, height, color=color)


def create_stats_composite_image(level, icon, ign, tab_name, wins, losses, wl_ratio, kills, deaths, kd_ratio, 
                                  ign_color=None, guild_tag=None, guild_hex=None, playtime_seconds=0,
                                  status_text="Online", status_color=(85, 255, 85), skin_image=None):
    canvas_w, canvas_h = 1200, 650
    margin, spacing = 40, 15
    composite = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    
    formatted_playtime = format_playtime(playtime_seconds)
    skin_w, skin_h = 240, 285
    header_card_w = (canvas_w - (margin * 2) - skin_w - (spacing * 2)) // 2
    
    skin_card = Image.new('RGBA', (skin_w, skin_h), (0, 0, 0, 0))
    ImageDraw.Draw(skin_card).rounded_rectangle([0, 0, skin_w-1, skin_h-1], radius=15, fill=(35, 30, 45, 240))
    if skin_image:
        skin = skin_image
    else:
        skin = get_player_body(ign)
    if skin:
        skin.thumbnail((220, 260), Image.Resampling.LANCZOS)
        skin_card.paste(skin, ((skin_w - skin.width)//2, (skin_h - skin.height)//2), skin)
    composite.paste(skin_card, (margin, margin), skin_card)

    col1_x = margin + skin_w + spacing
    col2_x = col1_x + header_card_w + spacing
    
    ign_rgb = (85, 255, 255)
    if ign_color:
        try:
            ign_rgb = tuple(int(str(ign_color).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass

    c1 = render_modern_card("IGN", ign, header_card_w, 85, is_header=True, color=ign_rgb)
    
    # Render Level card with multi-color support
    c2 = render_modern_card("Prestige", "", header_card_w, 85)
    segs = get_prestige_segments(level, icon)
    font_lvl = _load_font("DejaVuSans-Bold.ttf", 24)
    txt_io = _render_text_segments_to_image(segs, font=font_lvl, padding=(0,0))
    txt_img = Image.open(txt_io).convert("RGBA")
    # Align vertically with other cards (which draw text centered at height * 0.6)
    c2.paste(txt_img, ((c2.width - txt_img.width) // 2, int(c2.height * 0.6 - txt_img.height / 2) + 4), txt_img)

    c3 = render_modern_card("Mode", tab_name.upper(), header_card_w, 85)
    c4 = render_modern_card("Playtime", formatted_playtime, header_card_w, 85, is_header=True, color=(255, 85, 255))
    
    g_rgb = (255, 255, 255)
    if guild_hex:
        # Handle Minecraft color names (e.g. "DARK_AQUA")
        if str(guild_hex).upper() in MINECRAFT_NAME_TO_HEX:
            guild_hex = MINECRAFT_NAME_TO_HEX[str(guild_hex).upper()]
            
        try:
            g_rgb = tuple(int(str(guild_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            g_rgb = (170, 170, 170)
    safe_tag = _safe_guild_tag(guild_tag)
    c5 = render_modern_card("Guild", f"{safe_tag if safe_tag else 'None'}", header_card_w, 85, color=g_rgb)
    c6 = render_modern_card("Status", status_text, header_card_w, 85, color=status_color)

    for i, card in enumerate([c1, c2, c3]):
        composite.paste(card, (col1_x, margin + i*(85+spacing)), card)
    for i, card in enumerate([c4, c5, c6]):
        composite.paste(card, (col2_x, margin + i*(85+spacing)), card)

    line_y = margin + skin_h + 25
    ImageDraw.Draw(composite).line([margin, line_y, canvas_w - margin, line_y], fill=(60, 60, 80), width=2)
    
    grid_y = line_y + 25
    cols = 3
    grid_card_w = (canvas_w - (margin * 2) - (spacing * (cols - 1))) // cols
    grid_card_h = 110
    
    stats_data = [
        ("Wins", f"{int(wins):,}", (85, 255, 85)), ("Losses", f"{int(losses):,}", (255, 85, 85)), ("WLR", f"{wl_ratio:.2f}", (85, 255, 85)),
        ("Kills", f"{int(kills):,}", (255, 255, 255)), ("Deaths", f"{int(deaths):,}", (255, 255, 255)), ("KDR", f"{kd_ratio:.2f}", (85, 255, 85))
    ]

    for i, (label, val, color) in enumerate(stats_data):
        row, col = divmod(i, cols)
        card = render_modern_card(label, val, grid_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + col * (grid_card_w + spacing)), int(grid_y + row * (grid_card_h + spacing))), card)

    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.convert("RGB").save(out, format='PNG')
    out.seek(0)
    return out


def create_ww_stats_composite_image(level, icon, ign, tab_name, wins, losses, wl_ratio, kills, deaths, kd_ratio, assists,
                                     ign_color=None, guild_tag=None, guild_hex=None, playtime_seconds=0,
                                     status_text="Online", status_color=(85, 255, 85), skin_image=None, class_mode="overall"):
    """Create Wool Wars stats image with optional class filtering."""
    canvas_w = 1200
    # Adjust height: 650 for overall (2 rows), 540 for class-specific (1 row)
    canvas_h = 650 if class_mode == "overall" else 540
    margin, spacing = 40, 15
    composite = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    
    formatted_playtime = format_playtime(playtime_seconds)
    skin_w, skin_h = 240, 285
    header_card_w = (canvas_w - (margin * 2) - skin_w - (spacing * 2)) // 2
    
    skin_card = Image.new('RGBA', (skin_w, skin_h), (0, 0, 0, 0))
    ImageDraw.Draw(skin_card).rounded_rectangle([0, 0, skin_w-1, skin_h-1], radius=15, fill=(35, 30, 45, 240))
    if skin_image:
        skin = skin_image
    else:
        skin = get_player_body(ign)
    if skin:
        skin.thumbnail((220, 260), Image.Resampling.LANCZOS)
        skin_card.paste(skin, ((skin_w - skin.width)//2, (skin_h - skin.height)//2), skin)
    composite.paste(skin_card, (margin, margin), skin_card)

    col1_x = margin + skin_w + spacing
    col2_x = col1_x + header_card_w + spacing
    
    ign_rgb = (85, 255, 255)
    if ign_color:
        try:
            ign_rgb = tuple(int(str(ign_color).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass

    c1 = render_modern_card("IGN", ign, header_card_w, 85, is_header=True, color=ign_rgb)
    
    # Render Level card with multi-color support
    c2 = render_modern_card("Prestige", "", header_card_w, 85)
    segs = get_prestige_segments(level, icon)
    font_lvl = _load_font("DejaVuSans-Bold.ttf", 24)
    txt_io = _render_text_segments_to_image(segs, font=font_lvl, padding=(0,0))
    txt_img = Image.open(txt_io).convert("RGBA")
    c2.paste(txt_img, ((c2.width - txt_img.width) // 2, int(c2.height * 0.6 - txt_img.height / 2) + 4), txt_img)

    c3 = render_modern_card("Mode", tab_name.upper(), header_card_w, 85)
    c4 = render_modern_card("Playtime", formatted_playtime, header_card_w, 85, is_header=True, color=(255, 85, 255))
    
    g_rgb = (255, 255, 255)
    if guild_hex:
        if str(guild_hex).upper() in MINECRAFT_NAME_TO_HEX:
            guild_hex = MINECRAFT_NAME_TO_HEX[str(guild_hex).upper()]
        try:
            g_rgb = tuple(int(str(guild_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            g_rgb = (170, 170, 170)
    safe_tag = _safe_guild_tag(guild_tag)
    c5 = render_modern_card("Guild", f"{safe_tag if safe_tag else 'None'}", header_card_w, 85, color=g_rgb)
    c6 = render_modern_card("Status", status_text, header_card_w, 85, color=status_color)

    for i, card in enumerate([c1, c2, c3]):
        composite.paste(card, (col1_x, margin + i*(85+spacing)), card)
    for i, card in enumerate([c4, c5, c6]):
        composite.paste(card, (col2_x, margin + i*(85+spacing)), card)

    line_y = margin + skin_h + 25
    ImageDraw.Draw(composite).line([margin, line_y, canvas_w - margin, line_y], fill=(60, 60, 80), width=2)
    
    grid_y = line_y + 25
    cols = 4
    grid_card_w = (canvas_w - (margin * 2) - (spacing * (cols - 1))) // cols
    grid_card_h = 110
    
    # Different layouts for overall vs class-specific
    if class_mode == "overall":
        # 2 rows: Wins/Losses/WLR/empty, Kills/Deaths/KDR/Assists
        stats_data = [
            ("Wins", f"{int(wins):,}", (85, 255, 85)), 
            ("Losses", f"{int(losses):,}", (255, 85, 85)), 
            ("WLR", f"{wl_ratio:.2f}", (85, 255, 85)),
            ("Games", f"{int(wins + losses):,}", (255, 255, 255)),
            ("Kills", f"{int(kills):,}", (255, 255, 255)), 
            ("Deaths", f"{int(deaths):,}", (255, 255, 255)), 
            ("KDR", f"{kd_ratio:.2f}", (85, 255, 85)),
            ("Assists", f"{int(assists):,}", (85, 255, 255))
        ]
    else:
        # 1 row for class: Kills/Deaths/KDR/Assists
        stats_data = [
            ("Kills", f"{int(kills):,}", (255, 255, 255)), 
            ("Deaths", f"{int(deaths):,}", (255, 255, 255)), 
            ("KDR", f"{kd_ratio:.2f}", (85, 255, 85)),
            ("Assists", f"{int(assists):,}", (85, 255, 255))
        ]

    for i, (label, val, color) in enumerate(stats_data):
        row, col = divmod(i, cols)
        card = render_modern_card(label, val, grid_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + col * (grid_card_w + spacing)), int(grid_y + row * (grid_card_h + spacing))), card)

    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.convert("RGB").save(out, format='PNG')
    out.seek(0)
    return out


def create_ctw_stats_composite_image(level, icon, ign, tab_name, wins, losses, wl_ratio, draws,
                                      kills, deaths, kd_ratio, assists,
                                      wools_picked, wools_captured,
                                      kills_on_wh, deaths_to_wh, kills_as_wh, deaths_as_wh,
                                      gold_earned, gold_spent,
                                      ign_color=None, guild_tag=None, guild_hex=None, playtime_seconds=0,
                                      status_text="Online", status_color=(85, 255, 85), skin_image=None):
    """Create CTW stats image with custom layout."""
    canvas_w, canvas_h = 1200, 1030
    margin, spacing = 40, 15
    composite = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    
    formatted_playtime = format_playtime(playtime_seconds)
    skin_w, skin_h = 240, 285
    header_card_w = (canvas_w - (margin * 2) - skin_w - (spacing * 2)) // 2
    
    skin_card = Image.new('RGBA', (skin_w, skin_h), (0, 0, 0, 0))
    ImageDraw.Draw(skin_card).rounded_rectangle([0, 0, skin_w-1, skin_h-1], radius=15, fill=(35, 30, 45, 240))
    if skin_image:
        skin = skin_image
    else:
        skin = get_player_body(ign)
    if skin:
        skin.thumbnail((220, 260), Image.Resampling.LANCZOS)
        skin_card.paste(skin, ((skin_w - skin.width)//2, (skin_h - skin.height)//2), skin)
    composite.paste(skin_card, (margin, margin), skin_card)

    col1_x = margin + skin_w + spacing
    col2_x = col1_x + header_card_w + spacing
    
    ign_rgb = (85, 255, 255)
    if ign_color:
        try:
            ign_rgb = tuple(int(str(ign_color).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass

    c1 = render_modern_card("IGN", ign, header_card_w, 85, is_header=True, color=ign_rgb)
    
    # Render Level card with multi-color support
    c2 = render_modern_card("Prestige", "", header_card_w, 85)
    segs = get_prestige_segments(level, icon)
    font_lvl = _load_font("DejaVuSans-Bold.ttf", 24)
    txt_io = _render_text_segments_to_image(segs, font=font_lvl, padding=(0,0))
    txt_img = Image.open(txt_io).convert("RGBA")
    c2.paste(txt_img, ((c2.width - txt_img.width) // 2, int(c2.height * 0.6 - txt_img.height / 2) + 4), txt_img)

    c3 = render_modern_card("Mode", tab_name.upper(), header_card_w, 85)
    c4 = render_modern_card("Playtime", formatted_playtime, header_card_w, 85, is_header=True, color=(255, 85, 255))
    
    g_rgb = (255, 255, 255)
    if guild_hex:
        if str(guild_hex).upper() in MINECRAFT_NAME_TO_HEX:
            guild_hex = MINECRAFT_NAME_TO_HEX[str(guild_hex).upper()]
        try:
            g_rgb = tuple(int(str(guild_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            g_rgb = (170, 170, 170)
    safe_tag = _safe_guild_tag(guild_tag)
    c5 = render_modern_card("Guild", f"{safe_tag if safe_tag else 'None'}", header_card_w, 85, color=g_rgb)
    c6 = render_modern_card("Status", status_text, header_card_w, 85, color=status_color)

    for i, card in enumerate([c1, c2, c3]):
        composite.paste(card, (col1_x, margin + i*(85+spacing)), card)
    for i, card in enumerate([c4, c5, c6]):
        composite.paste(card, (col2_x, margin + i*(85+spacing)), card)

    line_y = margin + skin_h + 25
    ImageDraw.Draw(composite).line([margin, line_y, canvas_w - margin, line_y], fill=(60, 60, 80), width=2)
    
    grid_y = line_y + 25
    cols = 4
    grid_card_w = (canvas_w - (margin * 2) - (spacing * (cols - 1))) // cols
    grid_card_h = 110
    
    # Row 1: Wins, Losses, WLR, Draws (4 columns)
    row1_data = [
        ("Wins", f"{int(wins):,}", (85, 255, 85)),
        ("Losses", f"{int(losses):,}", (255, 85, 85)),
        ("WLR", f"{wl_ratio:.2f}", (85, 255, 85)),
        ("Draws", f"{int(draws):,}", (255, 255, 255))
    ]
    
    for i, (label, val, color) in enumerate(row1_data):
        card = render_modern_card(label, val, grid_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + i * (grid_card_w + spacing)), int(grid_y)), card)
    
    # Row 2: Kills, Deaths, KDR, Assists (4 columns)
    row2_y = grid_y + grid_card_h + spacing
    row2_data = [
        ("Kills", f"{int(kills):,}", (255, 255, 255)),
        ("Deaths", f"{int(deaths):,}", (255, 255, 255)),
        ("KDR", f"{kd_ratio:.2f}", (85, 255, 85)),
        ("Assists", f"{int(assists):,}", (85, 255, 255))
    ]
    
    for i, (label, val, color) in enumerate(row2_data):
        card = render_modern_card(label, val, grid_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + i * (grid_card_w + spacing)), int(row2_y)), card)
    
    # Row 3: Wools picked up, Wools captured (2 columns, each 2x wide)
    row3_y = row2_y + grid_card_h + spacing
    wide_card_w = 2 * grid_card_w + spacing
    row3_data = [
        ("Wools Picked Up", f"{int(wools_picked):,}", (255, 255, 85)),
        ("Wools Captured", f"{int(wools_captured):,}", (85, 255, 85))
    ]
    
    for i, (label, val, color) in enumerate(row3_data):
        card = render_modern_card(label, val, wide_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + i * (wide_card_w + spacing)), int(row3_y)), card)
    
    # Row 4: Kills on WH, Deaths to WH, Kills as WH, Deaths as WH (4 columns)
    row4_y = row3_y + grid_card_h + spacing
    row4_data = [
        ("Kills on WH", f"{int(kills_on_wh):,}", (255, 170, 85)),
        ("Deaths to WH", f"{int(deaths_to_wh):,}", (255, 100, 100)),
        ("Kills as WH", f"{int(kills_as_wh):,}", (85, 200, 255)),
        ("Deaths as WH", f"{int(deaths_as_wh):,}", (200, 100, 255))
    ]
    
    for i, (label, val, color) in enumerate(row4_data):
        card = render_modern_card(label, val, grid_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + i * (grid_card_w + spacing)), int(row4_y)), card)
    
    # Row 5: Gold earned, Gold spent (2 columns, each 2x wide)
    row5_y = row4_y + grid_card_h + spacing
    row5_data = [
        ("Gold Earned", f"{int(gold_earned):,}", (255, 215, 0)),
        ("Gold Spent", f"{int(abs(gold_spent)):,}", (255, 140, 0))
    ]
    
    for i, (label, val, color) in enumerate(row5_data):
        card = render_modern_card(label, val, wide_card_w, grid_card_h, color=color)
        composite.paste(card, (int(margin + i * (wide_card_w + spacing)), int(row5_y)), card)

    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.convert("RGB").save(out, format='PNG')
    out.seek(0)
    return out


def create_full_stats_image(ign: str, tab_name: str, level: int, icon: str, stats: dict,
                             ign_color: str = None, guild_tag: str = None, guild_color: str = None) -> io.BytesIO:
    """Render the full stats layout defined in Template.xlsx.

    Layout rules:
    - First line: 5 boxes (Wins/Hour, EXP/Hour, Playtime, EXP/Game, Kills/Hour)
    - Remaining lines: 5 boxes each
    """
    if Image is None:
        raise RuntimeError("Pillow not available")

    # Title image with prestige icon and tab name
    title_io = render_prestige_with_text(level, icon, ign, f"{tab_name.title()} Stats", ign_color, guild_tag, guild_color, two_line=True)
    title_img = Image.open(title_io)

    box_width = 200
    box_height = 80
    spacing = 10
    max_boxes = 5
    line_width_max = box_width * max_boxes + spacing * (max_boxes - 1)

    # Build lines from the template-driven order
    lines = [
        [
            ("Wins/Hour", stats.get("wins_per_hour", "0")),
            ("Kills/Hour", stats.get("kills_per_hour", "0")),
            ("Playtime", stats.get("playtime", "0")),
            ("Exp/Game", stats.get("exp_per_game", "0")),
            ("Exp/Hour", stats.get("exp_per_hour", "0")),
        ],
        [
            ("Wins", stats.get("wins", "0")),
            ("Losses", stats.get("losses", "0")),
            ("WLR", stats.get("wlr", "0")),
            ("Coins (Wool)", stats.get("coins", "0")),
            ("Levels Gained", stats.get("layers", "0")),
        ],
        [
            ("Kills", stats.get("kills", "0")),
            ("Deaths", stats.get("deaths", "0")),
            ("KDR", stats.get("kdr", "0")),
            ("Damage dealt", stats.get("damage", "0")),
            ("Sheep thrown", stats.get("sheeps_thrown", "0")),
        ],
        [
            ("Void kills", stats.get("void_kills", "0")),
            ("Void deaths", stats.get("void_deaths", "0")),
            ("Void KDR", stats.get("void_kdr", "0")),
            ("Games played", stats.get("games_played", "0")),
            ("Magic wool hit", stats.get("magic_wools", "0")),
        ],
        [
            ("Explosive kills", stats.get("explosive_kills", "0")),
            ("Explosive deaths", stats.get("explosive_deaths", "0")),
            ("Explosive KDR", stats.get("explosive_kdr", "0")),
            ("Damage/Game", stats.get("damage_per_game", "0")),
            ("Sheeps/Game", stats.get("sheeps_per_game", "0")),
        ],
        [
            ("Bow kills", stats.get("bow_kills", "0")),
            ("Bow deaths", stats.get("bow_deaths", "0")),
            ("Bow KDR", stats.get("bow_kdr", "0")),
            ("Kill/Game", stats.get("kills_per_game", "0")),
            ("Wools/Game", stats.get("wools_per_game", "0")),
        ],
        [
            ("Melee kills", stats.get("melee_kills", "0")),
            ("Melee deaths", stats.get("melee_deaths", "0")),
            ("Melee KDR", stats.get("melee_kdr", "0")),
            ("Damage/Sheep", stats.get("damage_per_sheep", "0")),
            ("Survival rate", stats.get("survival_rate", "0")),
        ],
    ]

    # Render all boxes
    rendered_lines = []
    for line_idx, line in enumerate(lines):
        rendered = []
        for col_idx, (label, value) in enumerate(line):
            try:
                rendered.append(render_stat_box(label, str(value), width=box_width, height=box_height))
            except Exception as e:
                print(f"[WARNING] Failed to render box {label}: {e}")
        rendered_lines.append(rendered)

    # Compute overall dimensions
    line_heights = []
    line_widths = []
    for line in rendered_lines:
        line_height = box_height
        # Calculate width
        line_width = 0
        for i, box in enumerate(line):
            line_width += box.width
            if i < len(line) - 1:
                line_width += spacing
        line_heights.append(line_height)
        line_widths.append(line_width)

    grid_height = sum(line_heights) + spacing * (len(rendered_lines) - 1)
    grid_width = line_width_max

    margin_x = 40

    # Scale title if too wide
    title_width = title_img.width
    title_height = title_img.height
    if title_width > grid_width:
        scale_factor = grid_width / title_width
        title_width = grid_width
        title_height = int(title_img.height * scale_factor)
        title_img = title_img.resize((title_width, title_height), Image.LANCZOS)

    composite_width = grid_width + (margin_x * 2)
    title_x_offset = (composite_width - title_width) // 2
    bottom_padding = 40
    composite_height = title_height + spacing + grid_height + bottom_padding

    composite = Image.new('RGBA', (composite_width, composite_height), (18, 18, 20, 255))
    composite.paste(title_img, (title_x_offset, 0), title_img if title_img.mode == 'RGBA' else None)

    # Paste lines centered horizontally
    y_offset = title_height + spacing
    for idx, line in enumerate(rendered_lines):
        line_width = line_widths[idx]
        x_start = margin_x + (grid_width - line_width) // 2 if line_width > 0 else margin_x
        x = x_start
        for box in line:
            composite.paste(box, (x, y_offset), box)
            x += box.width + spacing
        y_offset += line_heights[idx] + spacing

    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((composite_width - text_w) // 2, composite_height - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.save(out, format='PNG')
    out.seek(0)
    return out



def create_streaks_image(ign: str, level: int, icon: str, ign_color: str, guild_tag: str, guild_color: str, winstreak: int, killstreak: int) -> io.BytesIO:
    if Image is None:
        raise RuntimeError("Pillow not available")

    title_io = render_prestige_with_text(level, icon, ign, "", ign_color, guild_tag, guild_color, two_line=False)
    title_img = Image.open(title_io)

    box_width = 300
    box_height = 120
    spacing = 20

    boxes = [
        render_stat_box("Current Winstreak", f"{int(winstreak):,}", width=box_width, height=box_height),
        render_stat_box("Current Killstreak", f"{int(killstreak):,}", width=box_width, height=box_height),
    ]

    line_width = boxes[0].width + boxes[1].width + spacing
    grid_height = box_height
    margin_x = 40
    margin_y = 40

    title_width = title_img.width
    title_height = title_img.height
    # Enforce minimum width and even width for symmetry
    content_width = max(title_width, line_width, 800)

    composite_width = content_width + margin_x * 2
    if composite_width % 2 != 0:
        composite_width += 1
    
    # Adjust bottom margin to match visual top margin of text (title image has ~6px top padding)
    visual_top_margin = margin_y + 6
    margin_bottom = visual_top_margin
    composite_height = title_height + spacing + grid_height + margin_y + margin_bottom

    composite = Image.new('RGBA', (composite_width, composite_height), (18, 18, 20, 255))
    composite.paste(title_img, ((composite_width - title_width) // 2, margin_y), title_img if title_img.mode == 'RGBA' else None)

    y_offset = margin_y + title_height + spacing
    x_start = (composite_width - line_width) // 2
    x = x_start
    for box in boxes:
        composite.paste(box, (x, y_offset), box)
        x += box.width + spacing

    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((composite_width - text_w) // 2, composite_height - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    composite.save(out, format='PNG')
    out.seek(0)
    return out


def calculate_stat_winners(stats1: dict, stats2: dict) -> tuple:
    """Calculate winners for important and non-important stats.
    
    Returns:
        tuple: (important_wins1, important_wins2, nonimportant_wins1, nonimportant_wins2)
    """
    # Important stats: KDR, WLR, survival rate, and all types of KDR
    # Lower is better for losses and death stats
    important_stats = [
        ('kdr', False),           # Higher is better
        ('wlr', False),           # Higher is better
        ('survival_rate', False), # Higher is better
        ('void_kdr', False),      # Higher is better
        ('explosive_kdr', False), # Higher is better
        ('bow_kdr', False),       # Higher is better
        ('melee_kdr', False),     # Higher is better
    ]
    
    # Non-important stats: everything else
    # Higher is better except for losses and deaths
    nonimportant_stats = [
        ('wins', False),
        ('losses', True),          # Lower is better
        ('kills', False),
        ('deaths', True),          # Lower is better
        ('void_kills', False),
        ('void_deaths', True),     # Lower is better
        ('explosive_kills', False),
        ('explosive_deaths', True), # Lower is better
        ('bow_kills', False),
        ('bow_deaths', True),      # Lower is better
        ('melee_kills', False),
        ('melee_deaths', True),    # Lower is better
        ('damage', False),
        ('magic_wools', False),
        ('sheeps_thrown', False),
        ('kills_per_game', False),
        ('kills_per_win', False),
        ('damage_per_game', False),
        ('damage_per_sheep', False),
        ('damage_per_kill', False),
        ('wools_per_game', False),
        ('sheeps_per_game', False),
        ('exp_per_hour', False),
        ('exp_per_game', False),
        ('wins_per_hour', False),
        ('kills_per_hour', False),
        ('games_played', False),
        ('layers', False),
        ('coins', False),
    ]
    
    def parse_value(val_str):
        """Parse stat value from formatted string."""
        try:
            # Remove commas and convert to float
            return float(str(val_str).replace(',', ''))
        except (ValueError, AttributeError):
            return 0.0
    
    def compare_stat(stat_key, lower_is_better, stats1, stats2):
        """Compare a single stat between two players.
        
        Returns: 1 if player1 wins, 2 if player2 wins, 0 if tie
        """
        val1 = parse_value(stats1.get(stat_key, '0'))
        val2 = parse_value(stats2.get(stat_key, '0'))
        
        if val1 == val2:
            return 0
        
        if lower_is_better:
            return 1 if val1 < val2 else 2
        else:
            return 1 if val1 > val2 else 2
    
    important_wins1 = 0
    important_wins2 = 0
    nonimportant_wins1 = 0
    nonimportant_wins2 = 0
    
    # Calculate important stats
    for stat_key, lower_is_better in important_stats:
        winner = compare_stat(stat_key, lower_is_better, stats1, stats2)
        if winner == 1:
            important_wins1 += 1
        elif winner == 2:
            important_wins2 += 1
    
    # Calculate non-important stats
    for stat_key, lower_is_better in nonimportant_stats:
        winner = compare_stat(stat_key, lower_is_better, stats1, stats2)
        if winner == 1:
            nonimportant_wins1 += 1
        elif winner == 2:
            nonimportant_wins2 += 1
    
    return (important_wins1, important_wins2, nonimportant_wins1, nonimportant_wins2)


def create_compare_stats_image(ign1: str, ign2: str, tab_name: str, stats1: dict, stats2: dict, 
                                level1: int, level2: int, icon1: str, icon2: str,
                                ign_color1: str = None, ign_color2: str = None,
                                guild_tag1: str = None, guild_tag2: str = None,
                                guild_color1: str = None, guild_color2: str = None) -> io.BytesIO:
    """Render comparison stats with ign1 in blue, ign2 in red, and stat labels in magenta."""
    if Image is None:
        raise RuntimeError("Pillow not available")

    # Force blue and red colors for comparison
    blue_color = "#55AAFF"  # Blue for ign1
    red_color = "#FF5555"   # Red for ign2

    # Create title with both players using forced colors
    title_io1 = render_prestige_with_text(level1, icon1, ign1, "", blue_color, guild_tag1, guild_color1, two_line=False)
    title_img1 = Image.open(title_io1)
    
    title_io2 = render_prestige_with_text(level2, icon2, ign2, "", red_color, guild_tag2, guild_color2, two_line=False)
    title_img2 = Image.open(title_io2)
    
    # Scale up usernames to match VS text size (1.5x larger)
    scale_factor = 1.5
    new_width1 = int(title_img1.width * scale_factor)
    new_height1 = int(title_img1.height * scale_factor)
    title_img1 = title_img1.resize((new_width1, new_height1), Image.LANCZOS)
    
    new_width2 = int(title_img2.width * scale_factor)
    new_height2 = int(title_img2.height * scale_factor)
    title_img2 = title_img2.resize((new_width2, new_height2), Image.LANCZOS)
    
    # Title section with VS text - adjusted spacing
    spacing_title = 15
    vs_width = 60  # Reduced from 80
    title_width = title_img1.width + vs_width + spacing_title * 2 + title_img2.width
    title_height = max(title_img1.height, title_img2.height)
    
    title_composite = Image.new('RGBA', (title_width, title_height), (0, 0, 0, 0))
    title_composite.paste(title_img1, (0, (title_height - title_img1.height) // 2), title_img1 if title_img1.mode == 'RGBA' else None)
    
    # Draw "VS" text with smaller font
    draw_vs = ImageDraw.Draw(title_composite)
    font_vs = _load_font("DejaVuSans-Bold.ttf", 24)  # Reduced from 36
    vs_text = "VS"
    vs_bbox = draw_vs.textbbox((0, 0), vs_text, font=font_vs)
    vs_text_w = vs_bbox[2] - vs_bbox[0]
    vs_text_h = vs_bbox[3] - vs_bbox[1]
    vs_x = title_img1.width + spacing_title + (vs_width - vs_text_w) // 2
    vs_y = (title_height - vs_text_h) // 2
    draw_vs.text((vs_x, vs_y), vs_text, font=font_vs, fill=(200, 200, 200))
    
    title_composite.paste(title_img2, (title_img1.width + vs_width + spacing_title * 2, (title_height - title_img2.height) // 2), title_img2 if title_img2.mode == 'RGBA' else None)

    # Tab name label
    tab_label_height = 40
    tab_label = Image.new('RGBA', (title_width, tab_label_height), (0, 0, 0, 0))
    draw_tab = ImageDraw.Draw(tab_label)
    font_tab = _load_font("DejaVuSans-Bold.ttf", 24)
    tab_text = f"{tab_name.title()} Stats Comparison"
    tab_bbox = draw_tab.textbbox((0, 0), tab_text, font=font_tab)
    tab_text_w = tab_bbox[2] - tab_bbox[0]
    draw_tab.text(((title_width - tab_text_w) // 2, 5), tab_text, font=font_tab, fill=(255, 255, 255))

    # Stats layout - stat name in magenta, then ign1 value (blue), then ign2 value (red)
    box_width = 200
    box_height = 80
    spacing = 10
    
    # Colors
    magenta = (255, 85, 255)  # magenta for stat labels
    blue = (85, 170, 255)      # blue for ign1
    red = (255, 85, 85)        # red for ign2
    
    # Build lines from stats (same layout as full stats)
    stat_keys = [
        [("Wins/Hour", "wins_per_hour"), ("Kills/Hour", "kills_per_hour"), ("Playtime", "playtime"), ("Exp/Game", "exp_per_game"), ("Exp/Hour", "exp_per_hour")],
        [("Wins", "wins"), ("Losses", "losses"), ("WLR", "wlr"), ("Coins (Wool)", "coins"), ("Layers", "layers")],
        [("Kills", "kills"), ("Deaths", "deaths"), ("KDR", "kdr"), ("Damage dealt", "damage"), ("Sheep thrown", "sheeps_thrown")],
        [("Void kills", "void_kills"), ("Void deaths", "void_deaths"), ("Void KDR", "void_kdr"), ("Games played", "games_played"), ("Magic wool hit", "magic_wools")],
        [("Explosive kills", "explosive_kills"), ("Explosive deaths", "explosive_deaths"), ("Explosive KDR", "explosive_kdr"), ("Damage/Game", "damage_per_game"), ("Sheeps/Game", "sheeps_per_game")],
        [("Bow kills", "bow_kills"), ("Bow deaths", "bow_deaths"), ("Bow KDR", "bow_kdr"), ("Kill/Game", "kills_per_game"), ("Wools/Game", "wools_per_game")],
        [("Melee kills", "melee_kills"), ("Melee deaths", "melee_deaths"), ("Melee KDR", "melee_kdr"), ("Damage/Sheep", "damage_per_sheep"), ("Survival rate", "survival_rate")],
    ]
    
    # Render all comparison boxes - single unified box with 3 text lines
    rendered_lines = []
    for line_idx, line in enumerate(stat_keys):
        rendered = []
        for col_idx, (label, key) in enumerate(line):
            value1 = stats1.get(key, "0")
            value2 = stats2.get(key, "0")
            try:
                # Create a single unified box
                img = Image.new('RGBA', (int(box_width), int(box_height)), (0, 0, 0, 0))
                draw = ImageDraw.Draw(img)
                card_bg = (35, 30, 45, 240) 
                draw.rounded_rectangle([0, 0, box_width-1, box_height-1], radius=15, fill=card_bg)
                
                # Font sizes
                font_label = _load_font("DejaVuSans-Bold.ttf", 13)
                font_value = _load_font("DejaVuSans-Bold.ttf", 18)
                
                # Calculate vertical spacing for 3 lines
                line_height = box_height / 3
                
                # Draw label in white (top third)
                l_text = f"{label.upper()}"
                l_bbox = draw.textbbox((0, 0), l_text, font=font_label)
                l_w = l_bbox[2] - l_bbox[0]
                draw.text(((box_width - l_w) // 2, int(line_height * 0.5) - 9), l_text, font=font_label, fill=(255, 255, 255))
                
                # Draw ign1 value in blue (middle third)
                v1_text = str(value1)
                draw.text((box_width // 2, int(line_height * 1.5)), v1_text, font=font_value, fill=blue, anchor="mm")
                
                # Draw ign2 value in red (bottom third)
                v2_text = str(value2)
                draw.text((box_width // 2, int(line_height * 2.5)), v2_text, font=font_value, fill=red, anchor="mm")
                
                rendered.append(img)
            except Exception as e:
                print(f"[WARNING] Failed to render comparison box {label}: {e}")
        rendered_lines.append(rendered)
    
    # Compute dimensions
    max_boxes = 5
    line_width_max = box_width * max_boxes + spacing * (max_boxes - 1)
    line_heights = [box_height for _ in rendered_lines]
    
    line_widths = []
    for line in rendered_lines:
        line_width = 0
        for i, box in enumerate(line):
            line_width += box.width
            if i < len(line) - 1:
                line_width += spacing
        line_widths.append(line_width)
    
    grid_height = sum(line_heights) + spacing * (len(rendered_lines) - 1)
    grid_width = line_width_max
    
    margin_x = 40
    
    # Scale title if too wide
    if title_width > grid_width:
        scale_factor = grid_width / title_width
        new_title_width = grid_width
        new_title_height = int(title_height * scale_factor)
        title_composite = title_composite.resize((new_title_width, new_title_height), Image.LANCZOS)
        title_width = new_title_width
        title_height = new_title_height
        
        # Also scale tab label
        tab_label = tab_label.resize((new_title_width, int(tab_label_height * scale_factor)), Image.LANCZOS)
        tab_label_height = int(tab_label_height * scale_factor)
    
    composite_width = grid_width + (margin_x * 2)
    title_x_offset = (composite_width - title_width) // 2
    bottom_padding = 40
    winner_row_height = 100  # Height for winner row
    composite_height = title_height + tab_label_height + spacing + grid_height + spacing + winner_row_height + bottom_padding
    
    composite = Image.new('RGBA', (composite_width, composite_height), (18, 18, 20, 255))
    composite.paste(title_composite, (title_x_offset, 0), title_composite if title_composite.mode == 'RGBA' else None)
    composite.paste(tab_label, (title_x_offset, title_height), tab_label if tab_label.mode == 'RGBA' else None)
    
    # Paste lines centered horizontally
    y_offset = title_height + tab_label_height + spacing
    for idx, line in enumerate(rendered_lines):
        line_width = line_widths[idx]
        x_start = margin_x + (grid_width - line_width) // 2 if line_width > 0 else margin_x
        x = x_start
        for box in line:
            composite.paste(box, (x, y_offset), box)
            x += box.width + spacing
        y_offset += line_heights[idx] + spacing
    
    # Calculate winners
    imp_wins1, imp_wins2, nonimp_wins1, nonimp_wins2 = calculate_stat_winners(stats1, stats2)
    
    # Winner row - two boxes side by side
    winner_box_width = (grid_width - spacing) // 2
    winner_box_height = 100
    winner_row_y = y_offset
    
    # Important stats box
    imp_box = Image.new('RGBA', (winner_box_width, winner_box_height), (0, 0, 0, 0))
    draw_imp = ImageDraw.Draw(imp_box)
    draw_imp.rounded_rectangle([0, 0, winner_box_width-1, winner_box_height-1], radius=15, fill=(35, 30, 45, 240))
    
    font_winner_title = _load_font("DejaVuSans-Bold.ttf", 14)
    font_winner_score = _load_font("DejaVuSans-Bold.ttf", 20)
    font_winner_text = _load_font("DejaVuSans-Bold.ttf", 16)
    
    # Important stats title
    imp_title = "IMPORTANT STATS"
    imp_title_bbox = draw_imp.textbbox((0, 0), imp_title, font=font_winner_title)
    imp_title_w = imp_title_bbox[2] - imp_title_bbox[0]
    draw_imp.text(((winner_box_width - imp_title_w) // 2, 10), imp_title, font=font_winner_title, fill=(255, 255, 255))
    
    # Score
    imp_score_text = f"{imp_wins1} - {imp_wins2}"
    draw_imp.text((winner_box_width // 2, 40), imp_score_text, font=font_winner_score, fill=(255, 255, 255), anchor="mm")
    
    # Winner
    if imp_wins1 > imp_wins2:
        imp_winner_text = f"Winner: {ign1}"
        imp_winner_color = blue
    elif imp_wins2 > imp_wins1:
        imp_winner_text = f"Winner: {ign2}"
        imp_winner_color = red
    else:
        imp_winner_text = "Winner: Tie"
        imp_winner_color = (200, 200, 200)
    
    draw_imp.text((winner_box_width // 2, 70), imp_winner_text, font=font_winner_text, fill=imp_winner_color, anchor="mm")
    
    # Non-important stats box
    nonimp_box = Image.new('RGBA', (winner_box_width, winner_box_height), (0, 0, 0, 0))
    draw_nonimp = ImageDraw.Draw(nonimp_box)
    draw_nonimp.rounded_rectangle([0, 0, winner_box_width-1, winner_box_height-1], radius=15, fill=(35, 30, 45, 240))
    
    # Non-important stats title
    nonimp_title = "NON-IMPORTANT STATS"
    nonimp_title_bbox = draw_nonimp.textbbox((0, 0), nonimp_title, font=font_winner_title)
    nonimp_title_w = nonimp_title_bbox[2] - nonimp_title_bbox[0]
    draw_nonimp.text(((winner_box_width - nonimp_title_w) // 2, 10), nonimp_title, font=font_winner_title, fill=(255, 255, 255))
    
    # Score
    nonimp_score_text = f"{nonimp_wins1} - {nonimp_wins2}"
    draw_nonimp.text((winner_box_width // 2, 40), nonimp_score_text, font=font_winner_score, fill=(255, 255, 255), anchor="mm")
    
    # Winner
    if nonimp_wins1 > nonimp_wins2:
        nonimp_winner_text = f"Winner: {ign1}"
        nonimp_winner_color = blue
    elif nonimp_wins2 > nonimp_wins1:
        nonimp_winner_text = f"Winner: {ign2}"
        nonimp_winner_color = red
    else:
        nonimp_winner_text = "Winner: Tie"
        nonimp_winner_color = (200, 200, 200)
    
    draw_nonimp.text((winner_box_width // 2, 70), nonimp_winner_text, font=font_winner_text, fill=nonimp_winner_color, anchor="mm")
    
    # Paste winner boxes
    composite.paste(imp_box, (margin_x, winner_row_y), imp_box)
    composite.paste(nonimp_box, (margin_x + winner_box_width + spacing, winner_row_y), nonimp_box)
    
    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((composite_width - text_w) // 2, composite_height - 30), footer_text, font=font_footer, fill=(60, 60, 65))
    
    out = io.BytesIO()
    composite.save(out, format='PNG')
    out.seek(0)
    return out


def create_single_stat_compare_image(ign1: str, ign2: str, tab_name: str, stat_key: str,
                                      stats1: dict, stats2: dict,
                                      level1: int, level2: int, icon1: str, icon2: str,
                                      ign_color1: str = None, ign_color2: str = None,
                                      guild_tag1: str = None, guild_tag2: str = None,
                                      guild_color1: str = None, guild_color2: str = None) -> io.BytesIO:
    """Generate comparison image for a single stat between two players.
    
    Note: stats1 and stats2 should already be calculated/formatted stats from _collect_stats,
    which includes both raw and derived stats.
    """
    if Image is None:
        raise RuntimeError("Pillow not available")

    # Get stat display name
    stat_display_name = _get_stat_display_name_for_image(stat_key)
    
    # Create title showing tab and stat name
    title_height = 60
    title_width = 800
    title_composite = Image.new('RGBA', (title_width, title_height), (0, 0, 0, 0))
    draw_title = ImageDraw.Draw(title_composite)
    font_title = _load_font("DejaVuSans-Bold.ttf", 32)
    title_text = f"{tab_name.title()} {stat_display_name}"
    title_bbox = draw_title.textbbox((0, 0), title_text, font=font_title)
    title_text_w = title_bbox[2] - title_bbox[0]
    draw_title.text(((title_width - title_text_w) // 2, 10), title_text, font=font_title, fill=(255, 255, 255))

    # Get stat values (already formatted from _collect_stats)
    value1 = stats1.get(stat_key, "0")
    value2 = stats2.get(stat_key, "0")
    
    # Determine winner
    def parse_value(val_str):
        try:
            return float(str(val_str).replace(',', ''))
        except (ValueError, AttributeError):
            return 0.0
    
    val1_num = parse_value(value1)
    val2_num = parse_value(value2)
    
    if val1_num > val2_num:
        winner_text = f"Winner: {ign1}"
        winner_color = (85, 170, 255)  # blue
    elif val2_num > val1_num:
        winner_text = f"Winner: {ign2}"
        winner_color = (255, 85, 85)  # red
    else:
        winner_text = "Winner: Tie"
        winner_color = (200, 200, 200)
    
    # Create main stat comparison box
    box_width = 800
    box_height = 200
    stat_box = Image.new('RGBA', (box_width, box_height), (0, 0, 0, 0))
    draw_box = ImageDraw.Draw(stat_box)
    draw_box.rounded_rectangle([0, 0, box_width - 1, box_height - 1], radius=15, fill=(35, 30, 45, 240))
    
    # Fonts
    font_player_label = _load_font("DejaVuSans-Bold.ttf", 24)
    font_value = _load_font("DejaVuSans-Bold.ttf", 42)
    font_winner = _load_font("DejaVuSans-Bold.ttf", 26)
    
    # Player 1 data (left side)
    p1_x = box_width // 4
    p1_label_y = 50
    p1_value_y = 100
    
    draw_box.text((p1_x, p1_label_y), ign1, font=font_player_label, fill=(85, 170, 255), anchor="mm")
    draw_box.text((p1_x, p1_value_y), str(value1), font=font_value, fill=(85, 170, 255), anchor="mm")
    
    # Player 2 data (right side)
    p2_x = (box_width * 3) // 4
    p2_label_y = 50
    p2_value_y = 100
    
    draw_box.text((p2_x, p2_label_y), ign2, font=font_player_label, fill=(255, 85, 85), anchor="mm")
    draw_box.text((p2_x, p2_value_y), str(value2), font=font_value, fill=(255, 85, 85), anchor="mm")
    
    # Draw winner at bottom
    winner_y = 160
    draw_box.text((box_width // 2, winner_y), winner_text, font=font_winner, fill=winner_color, anchor="mm")
    
    # Compute canvas dimensions
    margin_x = 60
    spacing = 20
    
    # Scale title if needed
    content_width = box_width
    if title_width > content_width:
        scale_factor = content_width / title_width
        new_title_width = content_width
        new_title_height = int(title_height * scale_factor)
        title_composite = title_composite.resize((new_title_width, new_title_height), Image.LANCZOS)
        title_width = new_title_width
        title_height = new_title_height
    
    composite_width = content_width + (margin_x * 2)
    title_x_offset = (composite_width - title_width) // 2
    bottom_padding = 40
    composite_height = title_height + spacing + box_height + bottom_padding
    
    composite = Image.new('RGBA', (composite_width, composite_height), (18, 18, 20, 255))
    composite.paste(title_composite, (title_x_offset, 0), title_composite if title_composite.mode == 'RGBA' else None)
    
    # Paste stat box centered
    box_x = (composite_width - box_width) // 2
    box_y = title_height + spacing
    composite.paste(stat_box, (box_x, box_y), stat_box)
    
    # Footer
    draw = ImageDraw.Draw(composite)
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((composite_width - text_w) // 2, composite_height - 30), footer_text, font=font_footer, fill=(60, 60, 65))
    
    out = io.BytesIO()
    composite.save(out, format='PNG')
    out.seek(0)
    return out


def _get_stat_display_name_for_image(stat_key: str) -> str:
    """Convert stat key to display name for images."""
    stat_names = {
        'kills': 'Total Kills',
        'deaths': 'Total Deaths',
        'kdr': 'K/D Ratio',
        'wins': 'Wins',
        'losses': 'Losses',
        'wlr': 'W/L Ratio',
        'damage_dealt': 'Damage Dealt',
        'damage': 'Damage Dealt',
        'games_played': 'Games Played',
        'sheep_thrown': 'Sheep Thrown',
        'sheeps_thrown': 'Sheep Thrown',
        'magic_wool_hit': 'Magic Wool Hit',
        'magic_wools': 'Magic Wool Hit',
        'kills_void': 'Void Kills',
        'void_kills': 'Void Kills',
        'deaths_void': 'Void Deaths',
        'void_deaths': 'Void Deaths',
        'kills_explosive': 'Explosive Kills',
        'explosive_kills': 'Explosive Kills',
        'deaths_explosive': 'Explosive Deaths',
        'explosive_deaths': 'Explosive Deaths',
        'kills_bow': 'Bow Kills',
        'bow_kills': 'Bow Kills',
        'deaths_bow': 'Bow Deaths',
        'bow_deaths': 'Bow Deaths',
        'kills_melee': 'Melee Kills',
        'melee_kills': 'Melee Kills',
        'deaths_melee': 'Melee Deaths',
        'melee_deaths': 'Melee Deaths',
        'void_kdr': 'Void K/D Ratio',
        'explosive_kdr': 'Explosive K/D Ratio',
        'bow_kdr': 'Bow K/D Ratio',
        'melee_kdr': 'Melee K/D Ratio',
        'kills_per_game': 'Kills per Game',
        'kills_per_win': 'Kills per Win',
        'kills_per_hour': 'Kills per Hour',
        'damage_per_game': 'Damage per Game',
        'damage_per_sheep': 'Damage per Sheep',
        'wools_per_game': 'Wools per Game',
        'sheeps_per_game': 'Sheeps per Game',
    }
    return stat_names.get(stat_key, stat_key.replace('_', ' ').title())


def create_rankings_image(username: str, category_display: str, period: str, rankings_data: list, page: int = 0, total_pages: int = 1) -> io.BytesIO:
    """Generate an image showing all rankings for a user.
    
    Args:
        username: The player's username
        category_display: Display name for the category (e.g., "Wool Games")
        period: Time period (e.g., "lifetime", "daily")
        rankings_data: List of tuples (rank, metric_label, value)
        page: Current page number (0-indexed)
        total_pages: Total number of pages
    
    Returns:
        BytesIO object containing the PNG image
    """
    # Design constants
    canvas_w = 1200
    margin = 40
    spacing = 10
    row_height = 60
    header_height = 80
    
    content_height = header_height + spacing + (len(rankings_data) * (row_height + spacing))
    canvas_h = margin + content_height + margin
    
    img = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    draw = ImageDraw.Draw(img)
    
    font_header = _load_font("DejaVuSans-Bold.ttf", 32)
    font_rank = _load_font("DejaVuSans-Bold.ttf", 24)
    font_name = _load_font("DejaVuSans-Bold.ttf", 24)
    font_val = _load_font("DejaVuSans-Bold.ttf", 24)
    font_small = _load_font("DejaVuSans-Bold.ttf", 16)
    
    # Header Card
    draw.rounded_rectangle([margin, margin, canvas_w - margin, margin + header_height], radius=15, fill=(35, 30, 45, 240))
    
    title_text = f"{period.title()} {category_display} Rankings for {username}"
    page_text = f"Page {page + 1}/{total_pages}"
    
    bbox = draw.textbbox((0, 0), title_text, font=font_header)
    draw.text((margin + (canvas_w - margin*2 - (bbox[2]-bbox[0]))//2, margin + (header_height - (bbox[3]-bbox[1]))//2 - 5), title_text, font=font_header, fill=(255, 255, 255))
    
    bbox_p = draw.textbbox((0, 0), page_text, font=font_small)
    draw.text((canvas_w - margin - (bbox_p[2]-bbox_p[0]) - 20, margin + (header_height - (bbox_p[3]-bbox_p[1]))//2), page_text, font=font_small, fill=(180, 180, 200))

    y = margin + header_height + spacing
    
    for rank, metric_label, value in rankings_data:
        # Row Card
        draw.rounded_rectangle([margin, y, canvas_w - margin, y + row_height], radius=15, fill=(35, 30, 45, 240))
        
        # Rank color
        r_col = (180, 180, 200)
        if rank == 1: r_col = (255, 215, 0)
        elif rank == 2: r_col = (192, 192, 192)
        elif rank == 3: r_col = (205, 127, 50)
        
        draw.text((margin + 20, y + 15), f"#{rank}", font=font_rank, fill=r_col)
        
        # Stat name
        rank_w = draw.textbbox((0,0), f"#{rank}", font=font_rank)[2] - draw.textbbox((0,0), f"#{rank}", font=font_rank)[0]
        stat_x = margin + 20 + rank_w + 30
        
        draw.text((stat_x, y + 15), metric_label, font=font_name, fill=(255, 255, 255))
        
        # Value (right-aligned)
        v_w = draw.textbbox((0,0), value, font=font_val)[2] - draw.textbbox((0,0), value, font=font_val)[0]
        draw.text((canvas_w - margin - 20 - v_w, y + 15), value, font=font_val, fill=(85, 255, 255))
        
        y += row_height + spacing

    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    img.save(out, format='PNG')
    out.seek(0)
    return out


def create_leaderboard_image(tab_name: str, metric_label: str, leaderboard_data: list, page: int = 0, total_pages: int = 1) -> io.BytesIO:
    # Design constants matching sheepwars command
    canvas_w = 1200
    margin = 40
    spacing = 10
    row_height = 60
    header_height = 80
    
    content_height = header_height + spacing + (len(leaderboard_data) * (row_height + spacing))
    canvas_h = margin + content_height + margin
    
    img = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    draw = ImageDraw.Draw(img)
    
    font_header = _load_font("DejaVuSans-Bold.ttf", 32)
    font_rank = _load_font("DejaVuSans-Bold.ttf", 24)
    font_name = _load_font("DejaVuSans-Bold.ttf", 24)
    font_val = _load_font("DejaVuSans-Bold.ttf", 24)
    font_small = _load_font("DejaVuSans-Bold.ttf", 16)
    
    # Header Card
    draw.rounded_rectangle([margin, margin, canvas_w - margin, margin + header_height], radius=15, fill=(35, 30, 45, 240))
    
    title_text = f"{tab_name} {metric_label} Leaderboard"
    page_text = f"Page {page + 1}/{total_pages}"
    
    bbox = draw.textbbox((0, 0), title_text, font=font_header)
    draw.text((margin + (canvas_w - margin*2 - (bbox[2]-bbox[0]))//2, margin + (header_height - (bbox[3]-bbox[1]))//2 - 5), title_text, font=font_header, fill=(255, 255, 255))
    
    bbox_p = draw.textbbox((0, 0), page_text, font=font_small)
    draw.text((canvas_w - margin - (bbox_p[2]-bbox_p[0]) - 20, margin + (header_height - (bbox_p[3]-bbox_p[1]))//2), page_text, font=font_small, fill=(180, 180, 200))

    y = margin + header_height + spacing
    
    for entry in leaderboard_data:
        rank, player, level, icon, p_hex, g_tag, g_hex, value, is_playtime = entry
        player = str(player)
        
        # Row Card
        draw.rounded_rectangle([margin, y, canvas_w - margin, y + row_height], radius=15, fill=(35, 30, 45, 240))
        
        # Rank color
        r_col = (180, 180, 200)
        if rank == 1: r_col = (255, 215, 0)
        elif rank == 2: r_col = (192, 192, 192)
        elif rank == 3: r_col = (205, 127, 50)
        
        draw.text((margin + 20, y + 15), f"#{rank}", font=font_rank, fill=r_col)
        
        # Prestige (skip for guilds - when level=0 and icon is empty)
        rank_w = draw.textbbox((0,0), f"#{rank}", font=font_rank)[2] - draw.textbbox((0,0), f"#{rank}", font=font_rank)[0]
        p_x = margin + 20 + rank_w + 15
        
        # Check if this is a guild entry (no prestige)
        is_guild = (level == 0 and icon == "")
        
        if not is_guild:
            # Render prestige for players
            segments = get_prestige_segments(level, icon)
            current_x = p_x
            for hex_color, text in segments:
                try:
                    rgb = tuple(int(hex_color.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
                except:
                    rgb = (255, 255, 255)
                draw.text((current_x, y + 15), text, font=font_name, fill=rgb)
                seg_w = draw.textbbox((0,0), text, font=font_name)[2] - draw.textbbox((0,0), text, font=font_name)[0]
                current_x += seg_w
            n_x = current_x + 10
        else:
            # No prestige for guilds, start name directly after rank
            n_x = p_x
        
        # Name - for guilds, use guild tag color instead of player color
        if is_guild and g_hex:
            # Use guild tag color for guild name
            try:
                name_rgb = tuple(int(str(g_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
            except:
                name_rgb = (255, 255, 255)
        else:
            # Use player color for player names
            try:
                name_rgb = tuple(int(str(p_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
            except:
                name_rgb = (255, 255, 255)
        draw.text((n_x, y + 15), player, font=font_name, fill=name_rgb)
        
        # Guild
        safe_tag = _safe_guild_tag(g_tag)
        if safe_tag:
            n_w = draw.textbbox((0,0), player, font=font_name)[2] - draw.textbbox((0,0), player, font=font_name)[0]
            g_x = n_x + n_w + 10
            try:
                g_rgb = tuple(int(str(g_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
            except:
                g_rgb = (170, 170, 170)
            draw.text((g_x, y + 15), f"[{safe_tag}]", font=font_name, fill=g_rgb)
        
        # Value
        if is_playtime:
            val_str = format_playtime(int(value))
        elif metric_label == "Level":
            val_str = f"{float(value):,.2f}"
        elif "Ratio" in metric_label or "/" in metric_label or "Per" in metric_label or "Rate" in metric_label or "Score" in metric_label:
            val_str = f"{float(value):,.2f}"
        else:
            val_str = f"{int(value):,}"
        v_w = draw.textbbox((0,0), val_str, font=font_val)[2] - draw.textbbox((0,0), val_str, font=font_val)[0]
        draw.text((canvas_w - margin - 20 - v_w, y + 15), val_str, font=font_val, fill=(85, 255, 255))
        
        y += row_height + spacing

    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    img.save(out, format='PNG')
    out.seek(0)
    return out


def create_distribution_pie(title: str, slices: list) -> io.BytesIO:
    """Render a pie chart with a subtle 3D tilt and legend."""
    if Image is None:
        raise RuntimeError("Pillow not available")

    total = sum(v for _, v, _ in slices) # Calculate total value for percentages
    if total <= 0:
        total = 1

    width, height = 1032, 672
    padding = 45
    legend_height = 220
    pie_top = 85
    depth = 45  # vertical extrusion to fake 3D
    usable_height = height - legend_height - padding - pie_top
    pie_height = max(160, usable_height - depth)

    img = Image.new("RGBA", (width, height), (18, 18, 20, 255))
    draw = ImageDraw.Draw(img)

    try:
        title_font = _load_font("DejaVuSans-Bold.ttf", 26)
        legend_font = _load_font("DejaVuSans.ttf", 17)
    except Exception:
        title_font = ImageFont.load_default()
        legend_font = ImageFont.load_default()

    bbox = draw.textbbox((0, 0), title, font=title_font)
    title_width = bbox[2] - bbox[0]
    draw.text(((width - title_width) // 2, 20), title, font=title_font, fill=(230, 230, 230))

    top_bbox = (padding, pie_top, width - padding, pie_top + pie_height)
    outline_dark = (18, 18, 24)

    def _shade(color, factor: float):
        return tuple(max(0, min(255, int(channel * factor))) for channel in color)

    # Precompute slice angles so we can reuse them for the depth and top faces
    slice_angles = []
    start_angle = 90  # start at 90 degrees (middle-right position)
    for _, value, color in slices:
        extent = 360 * (value / total)
        end_angle = start_angle + extent
        if extent > 0:
            slice_angles.append((start_angle, end_angle, color))
        start_angle = end_angle

    # Draw depth layers from back to front, one z-level at a time
    # This ensures all slices are visible at each depth level
    for z in range(depth, -1, -1):  # Include z=0 to eliminate gap
        for start_angle, end_angle, color in slice_angles:
            # Shade the sides to be slightly darker
            side_color = _shade(color, 0.8)
            offset_bbox = (
                top_bbox[0],
                top_bbox[1] + z,
                top_bbox[2],
                top_bbox[3] + z,
            )
            # Use side_color for both fill and outline to eliminate any gaps between layers
            draw.pieslice(offset_bbox, start=start_angle, end=end_angle, fill=side_color, outline=side_color, width=2)

    # Draw separator lines on the top face only (no fill, just outline)
    separator_color = (20, 20, 25)  # Dark separator between slices

    # Draw vertical separators for the visible sides (front face)
    cx = (top_bbox[0] + top_bbox[2]) / 2
    cy = (top_bbox[1] + top_bbox[3]) / 2
    rx = (top_bbox[2] - top_bbox[0]) / 2
    ry = (top_bbox[3] - top_bbox[1]) / 2

    boundaries = set()
    for s, e, _ in slice_angles:
        boundaries.add(s % 360)
        boundaries.add(e % 360)

    for angle in boundaries:
        # Only draw separators on the front face (0 to 180 degrees)
        if 0 <= angle <= 180:
            rad = math.radians(angle)
            x = cx + rx * math.cos(rad)
            y = cy + ry * math.sin(rad)
            draw.line([(x, y), (x, y + depth)], fill=separator_color, width=2)

    for start_angle, end_angle, color in slice_angles:
        draw.pieslice(top_bbox, start=start_angle, end=end_angle, fill=None, outline=separator_color, width=2)

    legend_x = padding + 10
    legend_y = top_bbox[3] + depth + 24
    box_size = 20
    line_spacing = 28
    for idx, (label, value, color) in enumerate(slices):
        percent = (value / total * 100) if total else 0
        y = legend_y + idx * line_spacing
        draw.rectangle([legend_x, y, legend_x + box_size, y + box_size], fill=color, outline=(240, 240, 240))
        text = f"{label}: {value} ({percent:.2f}%)"
        draw.text((legend_x + box_size + 10, y - 2), text, font=legend_font, fill=(220, 220, 220))

    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((width - text_w) // 2, height - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    buf = io.BytesIO()
    img.save(buf, format="PNG")
    buf.seek(0)
    return buf


def render_prestige_range_image(base: int, end_display: int) -> io.BytesIO:
    """Render an image showing the colored prestige bracket for the base level only."""
    raw = PRESTIGE_RAW_PATTERNS.get(base)
    if not raw:
        # Fallback to simple text
        parts = [(MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), f'[{base}]')]
        return _render_text_segments_to_image(parts, padding=(6,4))

    parts = _parse_raw_pattern(raw)

    def _build_replaced_segments(parts, replacement_str, rainbow=False):
        """Replace the first numeric span in the concatenated parts with replacement_str once, preserving color segments.

        If `rainbow` is True, expand the replacement into per-character colored segments cycling a rainbow palette.
        """
        concat = ''.join(t for (_, t) in parts)
        m = re.search(r"\d+", concat)
        if not m:
            return [(MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF'), text) for code, text in parts]

        num_start, num_end = m.start(), m.end()
        out_parts = []
        pos = 0
        replaced = False
        for code, text in parts:
            part_start = pos
            part_end = pos + len(text)
            pos = part_end
            hexcol = MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF')

            if part_end <= num_start or part_start >= num_end:
                out_parts.append((hexcol, text))
                continue

            # prefix
            prefix_len = max(0, num_start - part_start)
            if prefix_len > 0:
                prefix = text[:prefix_len]
                out_parts.append((hexcol, prefix))

            # replacement
            if not replaced:
                if rainbow:
                    # Build the original color sequence that covered the numeric span
                    colors_in_span = []
                    span_pos = 0
                    # Re-iterate to collect per-char colors within the numeric span
                    pos2 = 0
                    for c_code, c_text in parts:
                        part_s = pos2
                        part_e = pos2 + len(c_text)
                        pos2 = part_e
                        overlap_s = max(part_s, num_start)
                        overlap_e = min(part_e, num_end)
                        if overlap_e > overlap_s:
                            hex_here = MINECRAFT_CODE_TO_HEX.get(c_code.lower(), '#FFFFFF')
                            # number of covered chars in original
                            count = overlap_e - overlap_s
                            colors_in_span.extend([hex_here] * count)

                    if not colors_in_span:
                        # fallback rainbow cycle
                        RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
                        colors_in_span = [MINECRAFT_CODE_TO_HEX.get(code, '#FFFFFF') for code in RAINBOW_CODES]

                    # Apply colors across the replacement string, repeating as needed
                    repl = str(replacement_str)
                    for i, ch in enumerate(repl):
                        col = colors_in_span[i % len(colors_in_span)]
                        out_parts.append((col, ch))
                else:
                    out_parts.append((hexcol, replacement_str))
                replaced = True

            # suffix
            suffix_start_in_part = max(0, num_end - part_start)
            if suffix_start_in_part < len(text):
                suffix = text[suffix_start_in_part:]
                out_parts.append((hexcol, suffix))

        return out_parts

    # Choose fallback icons for bases where emoji fonts may be missing
    # 1200 = Arabic character ۞, 3000 = Japanese character ツ
    BAD_ICON_BASES = {800, 1200, 1800, 2800, 3000, 3800}

    # Determine if this prestige base should be rainbow (PRESTIGE_COLORS maps to None)
    rainbow_bases = {k for k, v in PRESTIGE_COLORS.items() if v is None}

    segments = _build_replaced_segments(parts, str(base), rainbow=(base in rainbow_bases))

    # If problematic base, replace any non-ascii icon with fallback from PRESTIGE_ICONS
    if base in BAD_ICON_BASES:
        def _replace_bad_icons(segments, base_val):
            res = []
            for col, txt in segments:
                # replace any non-basic symbol at end inside brackets with fallback
                newtxt = re.sub(r"\[(\s*\d+)([^\d\]]+)\]", lambda m: f"[{m.group(1)}{PRESTIGE_ICONS[(base_val//100) % len(PRESTIGE_ICONS)]}]", txt)
                res.append((col, newtxt))
            return res
        segments = _replace_bad_icons(segments, base)

    return _render_text_segments_to_image(segments, padding=(6,4))


def render_all_prestiges_combined(spacing: int = 20) -> io.BytesIO:
    """Render all prestiges as individual images and combine them vertically into one PNG."""
    if Image is None:
        raise RuntimeError("Pillow not available")

    # Build a 4-column layout where columns are offsets [0,1000,2000,3000]
    offsets = [0, 1000, 2000, 3000]

    # Rows are the base mods 0,100,...,900 (we limit to prestiges up to 4000)
    base_mods = [i * 100 for i in range(0, 10)]

    # Prepare grid of images (rows x cols). Use placeholder transparent images for missing cells.
    grid = []
    for base_mod in base_mods:
        row_imgs = []
        for off in offsets:
            key = base_mod + off
            if key in PRESTIGE_RAW_PATTERNS:
                try:
                    imgio = render_prestige_range_image(key, key + 99)
                    imgio.seek(0)
                    im = Image.open(imgio).convert('RGBA')
                except Exception:
                    im = Image.new('RGBA', (120, 30), (0,0,0,0))
            else:
                im = Image.new('RGBA', (120, 30), (0,0,0,0))
            row_imgs.append(im)
        grid.append(row_imgs)

    # Compute uniform cell size
    max_w = max((im.width for row in grid for im in row), default=120) + 20
    max_h = max((im.height for row in grid for im in row), default=30) + 12

    # Optional title at the top
    title_text = "Wool Games Prestiges"
    try:
        title_font = _load_font("DejaVuSans-Bold.ttf", 32)
    except Exception:
        title_font = ImageFont.load_default()
    
    draw_dummy = ImageDraw.Draw(Image.new('RGBA', (1,1)))
    tb = draw_dummy.textbbox((0,0), title_text, font=title_font)
    title_h = tb[3] - tb[1] + 40

    cols = len(offsets)
    rows = len(grid)

    margin = 30
    spacing = 12
    total_w = margin * 2 + cols * max_w + spacing * (cols - 1)
    total_h = margin * 2 + title_h + rows * max_h + spacing * (rows - 1) + 30

    combined = Image.new('RGBA', (total_w, total_h), (18, 18, 20, 255))
    draw = ImageDraw.Draw(combined)

    # Draw title centered
    title_x = total_w // 2
    title_y = margin + (title_h // 2) - 10
    draw.text((title_x, title_y), title_text, font=title_font, fill=(255, 255, 255), anchor='mm')

    start_y = margin + title_h
    for r, row in enumerate(grid):
        y = start_y + r * (max_h + spacing)
        base_mod = base_mods[r]
        for c, im in enumerate(row):
            x = margin + c * (max_w + spacing)
            offset = offsets[c]
            level = base_mod + offset
            
            # Determine background color based on text brightness
            p_color = get_prestige_color(level)
            #lum = (0.299 * p_color[0] + 0.587 * p_color[1] + 0.114 * p_color[2])
            bg_color = (35, 30, 45, 255) #if lum < 90 else (35, 30, 45, 255)

            # Draw card background
            draw.rounded_rectangle([x, y, x + max_w, y + max_h], radius=8, fill=bg_color)
            
            # center each image within its cell
            paste_x = x + (max_w - im.width) // 2
            paste_y = y + (max_h - im.height) // 2
            combined.paste(im, (paste_x, paste_y), im)

    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 14)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((total_w - text_w) // 2, total_h - 30), footer_text, font=font_footer, fill=(60, 60, 65))

    out = io.BytesIO()
    combined.save(out, format='PNG')
    out.seek(0)
    return out



def format_prestige_ansi(level: int, icon: str) -> str:
    """Return an ANSI-colored prestige bracket+level+icon string.

    If a multi-color pattern exists for the prestige base (e.g. 1900), use it;
    otherwise color the whole bracket using the single prestige color.
    """
    reset = "\u001b[0m"
    try:
        lvl = int(level)
    except Exception:
        lvl = 0
    base = (lvl // 100) * 100
    # If a raw pattern exists, parse it into (code, text) pieces
    if base in PRESTIGE_RAW_PATTERNS:
        raw = PRESTIGE_RAW_PATTERNS[base]
        parts = []
        cur_code = None
        buf = ''
        i = 0
        while i < len(raw):
            ch = raw[i]
            if ch == '&' and i + 1 < len(raw):
                # flush buf
                if buf:
                    parts.append((cur_code or 'f', buf))
                    buf = ''
                cur_code = raw[i+1].lower()
                i += 2
                continue
            else:
                buf += ch
                i += 1
        if buf:
            parts.append((cur_code or 'f', buf))

        out = []
        for code, text in parts:
            # Use chosen xterm-256 SGR for inline/code-block rendering
            sgr = MINECRAFT_CODE_TO_ANSI_SGR.get(code.lower(), "\u001b[37m")
            out.append(make_bold_ansi(sgr) + text)

        joined = ''.join(out) + reset
        # When a raw pattern exists we trust it includes the correct icon and colors.
        return joined

    # Fallback: color whole bracket with single color for the level
    ansi = get_ansi_color_code(level)
    bold = make_bold_ansi(ansi)
    return f"{bold}[{level}{icon}]{reset}"


async def _send_paged_ansi_followups(interaction: discord.Interaction, lines: list[str], block: str = 'ansi'):
    """Send potentially-long ANSI lines as one or more followup messages, each <= 2000 chars.

    Splits `lines` into code-block chunks and sends them via `interaction.followup.send`.
    Falls back to sanitized plain text if sending fails.
    """
    wrapper_open = f"```{block}\n"
    wrapper_close = "\n```"
    max_len = 2000

    chunks = []
    cur_lines = []
    # start with the wrapper overhead
    cur_len = len(wrapper_open) + len(wrapper_close)

    for ln in lines:
        ln_with_nl = ln + "\n"
        lnlen = len(ln_with_nl)
        if cur_len + lnlen > max_len:
            # flush current chunk
            if cur_lines:
                chunks.append("".join(cur_lines).rstrip('\n'))
            cur_lines = [ln_with_nl]
            cur_len = len(wrapper_open) + len(wrapper_close) + lnlen
        else:
            cur_lines.append(ln_with_nl)
            cur_len += lnlen

    if cur_lines:
        chunks.append("".join(cur_lines).rstrip('\n'))

    # Send chunks as followups
    for chunk in chunks:
        content = wrapper_open + chunk + wrapper_close
        try:
            await interaction.followup.send(content)
        except Exception:
            # fallback: send sanitized text without ANSI wrapper
            try:
                await interaction.followup.send(sanitize_output(chunk))
            except Exception:
                # give up silently
                pass

def load_tracked_users():
    return get_tracked_users()


def load_tracked_streaks() -> dict:
    try:
        return get_all_tracked_streaks()
    except Exception:
        pass
    return {}


def save_tracked_streaks(data: dict):
    try:
        for username, streak_data in data.items():
            update_tracked_streaks(username, streak_data)
    except Exception as e:
        print(f"[STREAK] Failed to save streaks to database: {e}")


def load_user_colors() -> dict:
    """Load user colors and metadata from database"""
    try:
        result = {}
        # Use optimized bulk fetch
        all_meta = get_all_user_meta()
        
        for username, meta in all_meta.items():
            # Convert database format to expected format
            result[username.lower()] = {
                'color': meta.get('ign_color'),
                'guild_tag': meta.get('guild_tag'),
                'guild_color': meta.get('guild_hex'),
                'icon': meta.get('icon'),
                'rank': meta.get('rank'),
            }
        # print(f"[DEBUG] Loaded colors/meta for {len(result)} users")
        if result:
            sample_user = next(iter(result))
            # print(f"[DEBUG] Sample meta for {sample_user}: {result[sample_user]}")
        return result
    except Exception as e:
        print(f"[ERROR] Failed to load user colors: {e}")
    return {}


def _get_lifetime_value(stats: dict, key: str) -> int:
    try:
        return int(stats.get(key, {}).get("lifetime", 0))
    except Exception:
        return 0


def update_streaks_from_stats(username: str, processed_stats: dict) -> bool:
    streaks = load_tracked_streaks()
    entry = streaks.get(username)
    if not entry:
        return False

    wins = _get_lifetime_value(processed_stats, "wins")
    losses = _get_lifetime_value(processed_stats, "losses")
    kills = _get_lifetime_value(processed_stats, "kills")
    deaths = _get_lifetime_value(processed_stats, "deaths")

    last_wins = int(entry.get("last_wins", wins))
    last_losses = int(entry.get("last_losses", losses))
    last_kills = int(entry.get("last_kills", kills))
    last_deaths = int(entry.get("last_deaths", deaths))

    winstreak = int(entry.get("winstreak", 0))
    killstreak = int(entry.get("killstreak", 0))

    win_delta = wins - last_wins
    loss_delta = losses - last_losses
    kill_delta = kills - last_kills
    death_delta = deaths - last_deaths

    if loss_delta > 0:
        winstreak = 0
    elif win_delta > 0 and loss_delta <= 0:
        winstreak = max(0, winstreak) + win_delta

    if death_delta > 0:
        killstreak = 0
    elif kill_delta > 0 and death_delta <= 0:
        killstreak = max(0, killstreak) + kill_delta

    entry.update({
        "winstreak": winstreak,
        "killstreak": killstreak,
        "last_wins": wins,
        "last_losses": losses,
        "last_kills": kills,
        "last_deaths": deaths,
    })
    streaks[username] = entry
    save_tracked_streaks(streaks)
    return True


def initialize_streak_entry(username: str, processed_stats: dict):
    streaks = load_tracked_streaks()
    wins = _get_lifetime_value(processed_stats, "wins")
    losses = _get_lifetime_value(processed_stats, "losses")
    kills = _get_lifetime_value(processed_stats, "kills")
    deaths = _get_lifetime_value(processed_stats, "deaths")

    streaks[username] = {
        "winstreak": 0,
        "killstreak": 0,
        "last_wins": wins,
        "last_losses": losses,
        "last_kills": kills,
        "last_deaths": deaths,
    }
    save_tracked_streaks(streaks)

def load_user_links():
    """Load username -> Discord user ID mappings from database"""
    try:
        return get_all_user_links()
    except Exception:
        return {}

def save_user_links(links: dict):
    """Save username -> Discord user ID mappings to database"""
    for username, discord_id in links.items():
        set_discord_link(username, discord_id)

def link_user_to_ign(discord_user_id: int, ign: str):
    """Link a Discord user ID to a Minecraft username (case-insensitive)"""
    links = load_user_links()
    # Store with original case but search case-insensitively
    links[ign.casefold()] = str(discord_user_id)
    save_user_links(links)

def is_user_authorized(discord_user_id: int, ign: str) -> bool:
    """Check if a Discord user is authorized to manage a username"""
    links = load_user_links()
    key = ign.casefold()
    return links.get(key) == str(discord_user_id)

def is_admin(user: Union[discord.User, discord.Member]) -> bool:
    """Check if user is a bot admin."""
    if str(user.id) in ADMIN_IDS:
        return True
    if user.name.casefold() in [n.casefold() for n in ADMIN_NAMES]:
        return True
    return False

def unlink_user_from_ign(ign: str) -> bool:
    """Remove username -> Discord user ID link"""
    links = load_user_links()
    key = ign.casefold()
    if key in links:
        del links[key]
        save_user_links(links)
        return True
    return False

def remove_user_color(ign: str) -> bool:
    """Remove username color from database"""
    try:
        meta = get_user_meta(ign)
        if meta and meta.get('ign_color'):
            update_user_meta(ign, ign_color="")
            return True
        return False
    except Exception as e:
        print(f"[ERROR] Failed to remove user color for {ign}: {e}")
        return False

def delete_user_sheet(ign: str) -> bool:
    """Delete user's data from database."""
    try:
        from db_helper import delete_user, user_exists
        
        if not user_exists(ign):
            print(f"[INFO] User {ign} not found in database")
            return False
        
        delete_user(ign)
        print(f"[INFO] Deleted user {ign} from database")
        return True
        
    except Exception as e:
        print(f"[ERROR] Failed to delete user {ign}: {e}")
        return False

def render_modern_card(label, value, width, height, color=(255, 255, 255), is_header=False):
    img = Image.new('RGBA', (int(width), int(height)), (0, 0, 0, 0))
    draw = ImageDraw.Draw(img)
    card_bg = (35, 30, 45, 240) 
    draw.rounded_rectangle([0, 0, width-1, height-1], radius=15, fill=card_bg)
    font_label = _load_font("DejaVuSans-Bold.ttf", 14)
    font_value = _load_font("DejaVuSans-Bold.ttf", 28 if is_header else 24)
    l_text = f"{label.upper()}:"
    l_bbox = draw.textbbox((0, 0), l_text, font=font_label)
    draw.text(((width - (l_bbox[2]-l_bbox[0])) // 2, height * 0.2), l_text, font=font_label, fill=(180, 180, 200))
    v_text = str(value)
    draw.text((width // 2, int(height * 0.6)), v_text, font=font_value, fill=color, anchor="mm")
    return img

_UUID_CACHE = {}

def get_uuid(ign: str) -> Optional[str]:
    ign_lower = ign.lower()
    if ign_lower in _UUID_CACHE:
        return _UUID_CACHE[ign_lower]
    
    headers = {"User-Agent": "SheepWarsBot/1.0"}
    
    # Try Mojang
    try:
        r = requests.get(f"https://api.mojang.com/users/profiles/minecraft/{ign}", headers=headers, timeout=2)
        if r.status_code == 200:
            data = r.json()
            uuid = data.get('id')
            if uuid:
                _UUID_CACHE[ign_lower] = uuid
                return uuid
    except:
        pass
        
    # Try PlayerDB
    try:
        r = requests.get(f"https://playerdb.co/api/player/minecraft/{ign}", headers=headers, timeout=2)
        if r.status_code == 200:
            data = r.json()
            if data.get('success'):
                uuid = data.get('data', {}).get('player', {}).get('raw_id')
                if uuid:
                    _UUID_CACHE[ign_lower] = uuid
                    return uuid
    except:
        pass
        
    return None

def get_player_body(ign):
    # Resolve UUID for better API support
    uuid = get_uuid(ign)
    identifier = uuid if uuid else ign

    # Try multiple providers to find one that works/updates
    # Using random param to bypass edge caching where possible
    ts = random.randint(0, 10000)
    providers = [
        f"https://api.mineatar.io/body/full/{identifier}?scale=10&ts={ts}"
    ]
    for url in providers:
        try:
            r = requests.get(url, headers={"User-Agent": "SheepWarsBot/1.0"}, timeout=5)
            if r.status_code == 200:
                return Image.open(io.BytesIO(r.content)).convert("RGBA")
        except Exception:
            continue
    return None

def get_api_key():
    try:
        with open(os.path.join(BOT_DIR, "API_KEY.txt"), "r") as f:
            return f.read().strip()
    except:
        return None

def verify_api_key():
    """Verify Hypixel API key validity on startup."""
    key = get_api_key()
    if not key:
        print("[STARTUP] [ERROR] API_KEY.txt not found or empty!")
        return

    print("[STARTUP] Verifying Hypixel API key...")
    try:
        r = requests.get("https://api.hypixel.net/v2/counts", headers={"API-Key": key}, timeout=10)
        if r.status_code == 200:
            data = r.json()
            if data.get('success'):
                playerCount = data.get('playerCount')
                print(f"[STARTUP] [OK] API Key verified! Hypixel player count: {playerCount}")
                return
        if r.status_code == 403:
            data=r.json()
            cause = data.get('cause')
            print(f"[STARTUP] [ERROR] API Key verification failed! (Status: {r.status_code}) Reason: {cause}")
            return
        if r.status_code == 429:
            data=r.json()
            cause = data.get('cause')
            isThrottle = data.get('throttle')
            isGlobal = data.get('global')
            print(f"[STARTUP] [ERROR] API Key verification failed! (Status: {r.status_code}) Reason: {cause}. Global throttle: {isGlobal}")
            return

        print(f"[STARTUP] [WARNING] API Key verification failed. Status: {r.status_code}")
    except Exception as e:
        print(f"[STARTUP] [ERROR] Failed to connect to Hypixel API: {e}")

async def check_legacy_migration():
    """Check if legacy files exist and database is empty, then migrate."""
    # Check if we need to migrate stats.xlsx
    db_stats = get_database_stats()
    if db_stats['users'] == 0:
        excel_file = BOT_DIR / "stats.xlsx"
        if excel_file.exists():
            print("[MIGRATION] Database empty but stats.xlsx found. Running conversion...")
            try:
                # Run convert_to_db.py with --force to skip prompt
                subprocess.run([sys.executable, "convert_to_db.py", "--force"], cwd=str(BOT_DIR), check=True)
                print("[MIGRATION] Conversion script finished.")
                # Force cache refresh after migration
                await STATS_CACHE.refresh()
            except Exception as e:
                print(f"[ERROR] Migration failed: {e}")

    # Check for JSON files migration (user_links, default_users, tracked_streaks)
    json_files = ["user_links.json", "default_users.json", "tracked_streaks.json"]
    should_migrate_json = False
    
    if (BOT_DIR / "user_links.json").exists() and not get_all_user_links():
        should_migrate_json = True
    elif (BOT_DIR / "default_users.json").exists() and not get_all_default_users():
        should_migrate_json = True
    elif (BOT_DIR / "tracked_streaks.json").exists() and not get_all_tracked_streaks():
        should_migrate_json = True
        
    if should_migrate_json:
        print("[MIGRATION] Legacy JSON files found and tables empty. Running conversion...")
        try:
            subprocess.run([sys.executable, "convert_to_db.py", "--force"], cwd=str(BOT_DIR), check=True)
            print("[MIGRATION] JSON conversion finished.")
        except Exception as e:
            print(f"[ERROR] JSON migration failed: {e}")

    # Check if we need to migrate tracked_users.txt (if not done by convert_to_db)
    tracked_users = get_tracked_users()
    if not tracked_users:
        tracked_file = BOT_DIR / "tracked_users.txt"
        if tracked_file.exists():
            print("[MIGRATION] tracked_users table empty but tracked_users.txt found. Migrating...")
            try:
                # We can re-run convert_to_db.py as it handles tracked_users now, or do it manually.
                # Running convert_to_db.py is safer as it centralizes logic.
                subprocess.run([sys.executable, "convert_to_db.py", "--force"], cwd=str(BOT_DIR), check=True)
            except Exception as e:
                print(f"[ERROR] Tracked users migration failed: {e}")

def get_player_status(ign):
    """Fetch player online status from Hypixel API."""
    api_key = get_api_key()
    if not api_key:
        return "Unknown", (170, 170, 170) # Gray
    
    # Get UUID
    uuid = get_uuid(ign)
    if not uuid:
        return "Unknown", (170, 170, 170)

    try:
        headers = {"API-Key": api_key, "User-Agent": "SheepWarsBot/1.0"}
        r = requests.get("https://api.hypixel.net/status", params={"uuid": uuid}, headers=headers, timeout=3)
        if r.status_code == 200:
            data = r.json()
            session = data.get('session')
            if data.get('success') and session and isinstance(session, dict) and session.get('online'):
                return "Online", (85, 255, 85) # Green
            else:
                return "Offline", (255, 85, 85) # Red
        else:
            print(f"[WARNING] Hypixel status check failed: {r.status_code}")
    except Exception as e:
        print(f"[WARNING] Status check error: {e}")
    
    return "Unknown", (170, 170, 170)

# ---- Default IGN helpers ----
def load_default_users() -> dict:
    try:
        return get_all_default_users()
    except Exception:
        return {}

def save_default_users(defaults: dict):
    for discord_id, username in defaults.items():
        set_default_username(discord_id, username)

def set_default_user(discord_user_id: int, ign: str):
    defaults = load_default_users()
    defaults[str(discord_user_id)] = ign
    save_default_users(defaults)

def remove_default_user(discord_user_id: int) -> bool:
    defaults = load_default_users()
    key = str(discord_user_id)
    if key in defaults:
        del defaults[key]
        save_default_users(defaults)
        return True
    return False

def get_default_user(discord_user_id: int) -> Optional[str]:
    defaults = load_default_users()
    return defaults.get(str(discord_user_id))

# DEPRECATED: This function is no longer used. Users are now registered permanently in the database
# for leaderboard accuracy, rather than being cleaned up after queries.
# async def cleanup_untracked_user_delayed(ign: str, delay_seconds: int = 60):
#     """Schedule cleanup of untracked user data after a delay.
#     
#     Waits for delay_seconds, then checks if the user is still untracked.
#     If they're still untracked, removes their color data and sheet.
#     """
#     try:
#         print(f"[CLEANUP] Scheduled cleanup for '{ign}' in {delay_seconds} seconds")
#         await asyncio.sleep(delay_seconds)
#         
#         # Check if user is now tracked
#         tracked_users = load_tracked_users()
#         print(f"[CLEANUP] After {delay_seconds}s delay, checking if '{ign}' is tracked")
#         print(f"[CLEANUP] Tracked users list: {tracked_users}")
#         
#         key = ign.casefold()
#         for tracked_user in tracked_users:
#             if tracked_user.casefold() == key:
#                 # User is now tracked, don't clean up
#                 print(f"[CLEANUP] SKIPPING cleanup for '{ign}' - found in tracked users database as '{tracked_user}'")
#                 return
#         
#         # User is still untracked, proceed with cleanup
#         print(f"[CLEANUP] User '{ign}' NOT FOUND in tracked users database")
#         print(f"[CLEANUP] Reason: User was queried via /sheepwars but is not in tracked list")
#         print(f"[CLEANUP] Proceeding with cleanup: removing color data and deleting sheet")
#         
#         color_removed = remove_user_color(ign)
#         sheet_deleted = delete_user_sheet(ign)
#         
#         print(f"[CLEANUP] Cleanup complete for '{ign}' - color_removed={color_removed}, sheet_deleted={sheet_deleted}")
#     except asyncio.CancelledError:
#         print(f"[CLEANUP] Cleanup task cancelled for '{ign}'")
#     except Exception as e:
#         print(f"[CLEANUP] ERROR during cleanup for '{ign}': {e}")
#         import traceback
#         traceback.print_exc()

async def send_fetch_message(message: str):
    # DM the creator (prefer explicit ID if set)
    user = None
    if CREATOR_ID is not None:
        try:
            uid = int(CREATOR_ID)
            user = bot.get_user(uid) or await bot.fetch_user(uid)
        except Exception:
            user = None
    if user is None:
        # fallback to name/display name search across guilds
        for guild in bot.guilds:
            for member in guild.members:
                if member.bot:
                    continue
                name_match = member.name.casefold() == CREATOR_NAME.casefold()
                display_match = member.display_name.casefold() == CREATOR_NAME.casefold()
                if name_match or display_match:
                    user = member
                    break
            if user:
                break
    if user:
        try:
            await user.send(message)
            return
        except Exception as e:
            # Common cause: user has DMs disabled (Discord error 50007). Fall back to channel.
            print(f"[WARNING] Could not DM creator: {e}")
    # fallback: send to system channel or first writable channel
    for guild in bot.guilds:
        channel = None
        if guild.system_channel and guild.system_channel.permissions_for(guild.me).send_messages:
            channel = guild.system_channel
        else:
            for ch in guild.text_channels:
                if ch.permissions_for(guild.me).send_messages:
                    channel = ch
                    break
        if channel:
            try:
                await channel.send(message)
                break
            except Exception:
                continue

async def _delayed_refresh_user(username: str, delay: float):
    """Sleep for `delay` seconds then run api_get.py for the given username."""
    try:
        await asyncio.sleep(delay)
        result = await asyncio.to_thread(run_script, "api_get.py", ["-ign", username])

        # Track API calls
        if result and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                # Log API calls to tracker
                if json_data and 'api_calls' in json_data:
                    api_calls = json_data['api_calls']
                    if api_calls.get('player', 0) > 0:
                        await API_TRACKER.log_request('player')
                    if api_calls.get('guild', 0) > 0:
                        await API_TRACKER.log_request('guild')
            except Exception:
                pass

        # Try to update cache/streaks from stdout JSON
        if result and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
            except Exception as parse_err:
                print(f"[REFRESH] Failed to parse refresh output for {username}: {parse_err}")
    except asyncio.CancelledError:
        return
    except Exception as e:
        print(f"[REFRESH] Error refreshing {username}: {e}")


async def staggered_stats_refresher(interval_minutes: int = 10):
    """Background task that refreshes every tracked user's stats once per `interval_minutes`.

    Each user's refresh is scheduled at a random point during the interval to spread load.
    """
    interval = interval_minutes * 60
    buffer = 5  # seconds buffer to avoid scheduling at the very end
    while True:
        try:
            # Run database query in thread to avoid blocking event loop
            users = await asyncio.to_thread(load_tracked_users)
            
            # Identify users who are being tracked for streaks (they get their own fast refresh loop)
            streak_data = await asyncio.to_thread(load_tracked_streaks)
            streak_users = {u.lower() for u in streak_data.keys()}
            
            # Only process users NOT in the streak list
            users_to_refresh = [u for u in users if u.lower() not in streak_users]
            
            if not users_to_refresh:
                await asyncio.sleep(interval)
                continue

            # assign a random delay in [0, interval-buffer) to each user, then schedule
            tasks = []
            for u in users_to_refresh:
                d = random.uniform(0, max(0, interval - buffer))
                tasks.append(asyncio.create_task(_delayed_refresh_user(u, d)))

            # wait for the interval to elapse; leave any straggling tasks to finish in background
            await asyncio.sleep(interval)

            # optionally gather any finished tasks and suppress exceptions
            for t in tasks:
                if t.done():
                    try:
                        t.result()
                    except Exception:
                        pass

        except Exception as e:
            print(f"[REFRESH] Staggered refresher error: {e}")
            await asyncio.sleep(interval)


async def streak_stats_refresher(interval_seconds: int = 60):
    """Background task that refreshes streak-tracked users every minute."""
    while True:
        try:
            # Get users with active streak tracking (run in thread to avoid blocking)
            streak_data = await asyncio.to_thread(load_tracked_streaks)
            streak_users = list(streak_data.keys())
            
            if not streak_users:
                await asyncio.sleep(interval_seconds)
                continue
            
            # Spread updates over the interval to avoid spikes
            tasks = []
            for u in streak_users:
                # Random delay within the interval (leaving 5s buffer)
                d = random.uniform(0, max(0, interval_seconds - 5))
                tasks.append(asyncio.create_task(_delayed_refresh_user(u, d)))
            
            await asyncio.sleep(interval_seconds)
            
            # Cleanup tasks
            for t in tasks:
                if t.done():
                    try: t.result() 
                    except: pass
                    
        except Exception as e:
            print(f"[REFRESH] Streak refresher error: {e}")
            await asyncio.sleep(interval_seconds)


async def guild_updater_hourly():
    """Background task that updates guild data for all tracked users' guilds every hour.
    
    This runs on a 1-hour interval to fetch guild tag/color/exp data for guilds
    of tracked users, reducing API pressure while keeping guild data reasonably fresh.
    """
    from api_get import api_update_guild_database
    
    # Wait 5 minutes before first run to let the bot stabilize
    await asyncio.sleep(300)
    
    while True:
        try:
            print(f"[GUILD_UPDATE] Starting hourly guild update cycle")
            
            # Get all tracked users and extract their guilds
            tracked_users = await asyncio.to_thread(load_tracked_users)
            user_colors = await asyncio.to_thread(load_user_colors)
            
            # Always include explicitly tracked guilds
            guilds_to_update = set(await asyncio.to_thread(get_tracked_guilds))
            
            # Also include guilds of tracked users
            for username in tracked_users:
                user_meta = user_colors.get(username.lower(), {})
                guild_name = user_meta.get('guild_name')
                if guild_name:
                    guilds_to_update.add(guild_name)
            
            if guilds_to_update:
                print(f"[GUILD_UPDATE] Found {len(guilds_to_update)} guilds to update")
                
                # Spread guild updates over 30 minutes to avoid API spikes
                interval_per_guild = min(1800 / len(guilds_to_update), 60)  # Max 30 min, min 1 sec between
                
                for guild_name in guilds_to_update:
                    try:
                        # Update guild data (without resetting snapshots - periodic updates maintain existing deltas)
                        result = await asyncio.to_thread(
                            api_update_guild_database, guild_name, None, None
                        )
                        
                        if 'error' not in result:
                            await API_TRACKER.log_request('guild')
                            print(f"[GUILD_UPDATE] Successfully updated guild: {guild_name}")
                        else:
                            print(f"[GUILD_UPDATE] Failed to update guild {guild_name}: {result.get('error', 'Unknown error')}")
                        
                        # Wait before next guild
                        await asyncio.sleep(interval_per_guild)
                        
                    except Exception as e:
                        print(f"[GUILD_UPDATE] Error updating guild {guild_name}: {e}")
                        await asyncio.sleep(5)  # Brief pause on error
                
                print(f"[GUILD_UPDATE] Hourly guild update cycle complete")
            else:
                print(f"[GUILD_UPDATE] No guilds to update")
            
            # Wait 1 hour until next cycle
            await asyncio.sleep(3600)
            
        except Exception as e:
            print(f"[GUILD_UPDATE] Hourly guild updater error: {e}")
            await asyncio.sleep(3600)  # Wait an hour before retrying on error


# Track last known player count for Sheep Wars to calculate delta
_sheep_wars_last_players = None


async def _get_wool_games_status() -> str | None:
    """Fetch Sheep/Wool Wars player status via HyTrack's socket.io feed.

    Returns a string like "Players: 12 (+1)" or None on failure/timeout.
    """
    global _sheep_wars_last_players
    
    try:
        import socketio  # python-socketio
    except Exception as e:
        print(f"[PRESENCE] socketio import failed: {e}")
        return None

    target_key = "WOOL_GAMES__sheep_wars_two_six"
    status_box = {"value": None}
    status_event = asyncio.Event()

    def _set_status(entry):
        global _sheep_wars_last_players
        if not entry:
            return
        if isinstance(entry, list):
            for item in entry:
                _set_status(item)
            return
        if not isinstance(entry, dict):
            return
        info = entry.get("info", {})
        if info.get("key") != target_key:
            return
        players = entry.get("players")
        if players is None:
            return
        
        # Calculate delta by comparing to last known value
        if _sheep_wars_last_players is not None:
            delta = players - _sheep_wars_last_players
        else:
            delta = 0
        
        # Update last known value
        _sheep_wars_last_players = players
        
        status_box["value"] = f"Sheepers: {players} ({delta:+d})"
        status_event.set()

    sio = socketio.AsyncClient(reconnection=False, logger=False, engineio_logger=False)

    @sio.event
    async def connect():
        try:
            await sio.emit("requestListing", "WOOL_GAMES")
        except Exception as e:
            print(f"[PRESENCE] emit requestListing failed: {e}")

    @sio.on("add")
    async def on_add(entries):
        try:
            _set_status(entries)
        except Exception as e:
            pass

    @sio.on("update")
    async def on_update(update):
        try:
            _set_status(update)
        except Exception as e:
            pass

    try:
        await sio.connect("https://hytrack.me", transports=["websocket", "polling"], wait_timeout=5)
        try:
            await asyncio.wait_for(status_event.wait(), timeout=8)
        except asyncio.TimeoutError:
            pass
    except Exception as e:
        print(f"[PRESENCE] socket connect failed: {e}")
    finally:
        try:
            await sio.disconnect()
        except Exception:
            pass

    return status_box["value"]


async def presence_updater_loop(interval_seconds: int = 5):
    """Background loop: poll site and update bot presence to show current players."""
    last = None
    while True:
        try:
            status = await _get_wool_games_status()
            if status:
                # If status changed, update presence
                if status != last:
                    try:
                        await bot.change_presence(activity=discord.Game(name=status))
                        #print(f"[PRESENCE] Updated presence to: {status}")
                        last = status
                    except Exception as e:
                        print(f"[PRESENCE] Failed to change presence: {e}")
            else:
                # If no status, optionally clear presence
                pass
        except Exception as e:
            print(f"[PRESENCE] Loop error: {e}")
        await asyncio.sleep(interval_seconds)


def inline_backup_fallback():
    """Inline backup fallback when backup_hourly.py script fails."""
    import shutil
    from datetime import datetime
    from db_helper import backup_database
    
    try:
        db_file = DB_FILE
        backup_dir = BOT_DIR / "backups"
        
        # Try primary backup directory
        if not backup_dir.exists():
            try:
                backup_dir.mkdir(exist_ok=True, mode=0o755)
            except:
                # Fallback to home directory
                from pathlib import Path
                backup_dir = Path.home() / "backup_api_backups"
                backup_dir.mkdir(exist_ok=True, mode=0o755)
                print(f"[FALLBACK] Using alternate directory: {backup_dir}")
        
        if not db_file.exists():
            print(f"[FALLBACK] Database file not found: {db_file}")
            return False
        
        timestamp = datetime.now().strftime("%Y-%m-%d_%H-00-00")
        backup_path = backup_dir / f"stats_{timestamp}.db"
        
        if backup_path.exists():
            print(f"[FALLBACK] Backup already exists: {backup_path.name}")
            return True
        
        # Use database helper to backup
        success = backup_database(backup_path)
        
        if success and backup_path.exists():
            size = backup_path.stat().st_size
            print(f"[FALLBACK] Backup created: {backup_path.name} ({size:,} bytes)")
            return True
        else:
            print(f"[FALLBACK] Backup was not created")
            return False
            
    except Exception as e:
        print(f"[FALLBACK] Inline backup error: {e}")
        import traceback
        traceback.print_exc()
        return False


async def scheduler_loop():
    """Automatic scheduler for daily and monthly snapshots, plus hourly backups"""
    last_snapshot_run = None
    last_backup_hour = None
    
    while True:
        now = datetime.datetime.now(tz=CREATOR_TZ)
        
        # Hourly backup - runs at :00 minutes every hour
        if now.minute == 0:
            current_hour = (now.date(), now.hour)
            if last_backup_hour != current_hour:
                try:
                    print(f"[SCHEDULER] Running hourly backup at {now.strftime('%I:%M %p')}")
                    
                    def run_hourly_backup():
                        import subprocess
                        script_path = BOT_DIR / "backup_hourly.py"
                        # Ensure script is executable on Linux
                        if not script_path.exists():
                            raise FileNotFoundError(f"Backup script not found: {script_path}")
                        
                        print(f"[SCHEDULER] Backup script path: {script_path}")
                        print(f"[SCHEDULER] Python executable: {sys.executable}")
                        print(f"[SCHEDULER] Working directory: {BOT_DIR}")
                        
                        return subprocess.run(
                            [sys.executable, str(script_path)],
                            cwd=str(BOT_DIR),
                            capture_output=True,
                            text=True,
                            timeout=120
                        )
                    
                    backup_result = await asyncio.to_thread(run_hourly_backup)
                    if backup_result.returncode == 0:
                        print(f"[SCHEDULER] Hourly backup completed successfully")
                        # Show output even on success for debugging
                        if backup_result.stdout:
                            print(f"[SCHEDULER] Backup output:\n{backup_result.stdout[:500]}")
                    else:
                        print(f"[SCHEDULER] Hourly backup failed with exit code {backup_result.returncode}")
                        if backup_result.stdout:
                            print(f"[SCHEDULER] Backup stdout:\n{backup_result.stdout[:500]}")
                        if backup_result.stderr:
                            print(f"[SCHEDULER] Backup stderr:\n{backup_result.stderr[:500]}")
                        
                        # FALLBACK: Try inline backup
                        print(f"[FALLBACK] Attempting inline backup...")
                        try:
                            await asyncio.to_thread(inline_backup_fallback)
                            print(f"[FALLBACK] Inline backup completed")
                        except Exception as fallback_error:
                            print(f"[FALLBACK] Inline backup also failed: {fallback_error}")
                except Exception as e:
                    print(f"[SCHEDULER] Hourly backup error: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    
                    # FALLBACK: Try inline backup
                    print(f"[FALLBACK] Attempting inline backup after exception...")
                    try:
                        await asyncio.to_thread(inline_backup_fallback)
                        print(f"[FALLBACK] Inline backup completed")
                    except Exception as fallback_error:
                        print(f"[FALLBACK] Inline backup also failed: {fallback_error}")
                
                last_backup_hour = current_hour
        
        # Run snapshot updates at 9:30 AM
        if now.hour == 9 and now.minute == 30:
            today = now.date()
            if last_snapshot_run != today:
                try:
                    # Step 1: Run yesterday snapshot (before daily overwrites it)
                    def run_yesterday():
                        return run_script_batch("batch_update.py", ["-schedule", "yesterday"])
                    
                    yesterday_result = await asyncio.to_thread(run_yesterday)
                    if yesterday_result.returncode != 0:
                        error_detail = yesterday_result.stderr or yesterday_result.stdout or "Unknown error"
                        print(f"[SCHEDULER] Yesterday snapshot FAILED - returncode: {yesterday_result.returncode}")
                        print(f"[SCHEDULER] Full stdout:\n{yesterday_result.stdout}")
                        print(f"[SCHEDULER] Full stderr:\n{yesterday_result.stderr}")
                        await send_fetch_message(f"Warning: Yesterday snapshot failed at {now.strftime('%I:%M %p')}\nError: {error_detail[:500]}")
                    else:
                        print(f"[SCHEDULER] Yesterday snapshot completed successfully")
                    
                    # Step 1.5: Run weekly reset on Mondays (weekday() returns 0 for Monday)
                    if now.weekday() == 0:
                        print(f"[SCHEDULER] Running weekly reset (Monday)")
                        def run_weekly():
                            return run_script_batch("batch_update.py", ["-schedule", "weekly"])
                        
                        weekly_result = await asyncio.to_thread(run_weekly)
                        if weekly_result.returncode != 0:
                            error_detail = weekly_result.stderr or weekly_result.stdout or "Unknown error"
                            print(f"[SCHEDULER] Weekly reset FAILED - returncode: {weekly_result.returncode}")
                            print(f"[SCHEDULER] Full stdout:\n{weekly_result.stdout}")
                            print(f"[SCHEDULER] Full stderr:\n{weekly_result.stderr}")
                            await send_fetch_message(f"Warning: Weekly reset failed at {now.strftime('%I:%M %p')}\nError: {error_detail[:500]}")
                        else:
                            print(f"[SCHEDULER] Weekly reset completed successfully")
                    
                    # Step 2: Determine which snapshots to take
                    # Daily: always
                    # Monthly: only on 1st of month
                    if now.day == 1:
                        schedule = "all"  # daily + monthly
                    else:
                        schedule = "daily"
                    
                    # Step 3: Run batch_update.py for daily (and monthly if 1st)
                    def run_batch():
                        return run_script_batch("batch_update.py", ["-schedule", schedule])
                    
                    result = await asyncio.to_thread(run_batch)
                    if result.returncode == 0:
                        msg = f"Daily snapshot completed at {now.strftime('%I:%M %p')}"
                        if now.day == 1:
                            msg += " (including monthly snapshots)"
                        if now.weekday() == 0:
                            msg += " + Weekly reset"
                        await send_fetch_message(msg)
                    else:
                        error_msg = result.stderr or result.stdout or "Unknown error"
                        print(f"[SCHEDULER] Daily snapshot FAILED - returncode: {result.returncode}")
                        print(f"[SCHEDULER] Full stdout:\n{result.stdout}")
                        print(f"[SCHEDULER] Full stderr:\n{result.stderr}")
                        await send_fetch_message(f"Daily snapshot failed: {error_msg[:500]}")
                except Exception as e:
                    print(f"[SCHEDULER] Snapshot update exception: {str(e)}")
                    import traceback
                    traceback.print_exc()
                    await send_fetch_message(f"Snapshot update error: {str(e)}")
                
                last_snapshot_run = today
        
        await asyncio.sleep(20)

# Helper class for stats tab view
class StatsTabView(discord.ui.View):
    def __init__(self, data_dict, ign, level_value: int, prestige_icon: str, 
                 ign_color: str = None, guild_tag: str = None, guild_hex: str = None,
                 status_text="Online", status_color=(85, 255, 85), skin_image=None):
        super().__init__(timeout=180)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.data = data_dict 
        self.ign = ign
        self.level_value = level_value
        self.prestige_icon = prestige_icon
        self.status_text = status_text
        self.status_color = status_color
        self.skin_image = skin_image
        self.current_tab = "all-time"
        self.message = None  # Store message reference for timeout handling
        
        self.ign_color = ign_color
        self.guild_tag = guild_tag
        self.guild_hex = guild_hex
        
        if self.ign_color is None or self.guild_tag is None:
            self._load_color()
            
        self.update_button_styles()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible

    def _load_color(self):
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                if not self.ign_color:
                    self.ign_color = get_rank_color_hex(meta.get('rank'))
                self.guild_tag = meta.get('guild_tag')
                self.guild_hex = meta.get('guild_hex') or "#AAAAAA"
        except: pass

    def update_button_styles(self):
        """Setzt den aktiven Button auf Blau (Primary) und andere auf Grau (Secondary)."""
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_tab:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary

    def generate_composite_image(self, tab_name):
        tab_data = self.data[tab_name]
        
        # Rendert das Bild mit den gespeicherten Daten
        img_io = create_stats_composite_image(
            self.level_value, self.prestige_icon, self.ign, tab_name,
            tab_data['wins'], tab_data['losses'], tab_data['wlr'], 
            tab_data['kills'], tab_data['deaths'], tab_data['kdr'],
            self.ign_color, self.guild_tag, self.guild_hex, 
            playtime_seconds=tab_data['playtime'],
            status_text=self.status_text, status_color=self.status_color,
            skin_image=self.skin_image
        )
        return discord.File(img_io, filename=f"{self.ign}_{tab_name}.png")

    async def handle_tab_click(self, interaction: discord.Interaction, tab_name: str):
        self.current_tab = tab_name
        self.update_button_styles()
        file = self.generate_composite_image(tab_name)
        # Wichtig: View=self mitgeben, damit die Styles aktualisiert werden
        await interaction.response.edit_message(attachments=[file], view=self)

    @discord.ui.button(label="All-time", custom_id="all-time")
    async def all_time(self, interaction, button):
        await self.handle_tab_click(interaction, "all-time")

    @discord.ui.button(label="Session", custom_id="session")
    async def session(self, interaction, button):
        await self.handle_tab_click(interaction, "session")

    @discord.ui.button(label="Daily", custom_id="daily")
    async def daily(self, interaction, button):
        await self.handle_tab_click(interaction, "daily")

    @discord.ui.button(label="Yesterday", custom_id="yesterday")
    async def yesterday(self, interaction, button):
        await self.handle_tab_click(interaction, "yesterday")

    @discord.ui.button(label="Weekly", custom_id="weekly")
    async def weekly(self, interaction, button):
        await self.handle_tab_click(interaction, "weekly")

    @discord.ui.button(label="Monthly", custom_id="monthly")
    async def monthly(self, interaction, button):
        await self.handle_tab_click(interaction, "monthly")

class WWStatsView(discord.ui.View):
    """View for Wool Wars stats with class selection dropdown."""
    def __init__(self, data_dict, ign, level_value: int, prestige_icon: str, 
                 ign_color: str = None, guild_tag: str = None, guild_hex: str = None,
                 status_text="Online", status_color=(85, 255, 85), skin_image=None, 
                 show_period_buttons: bool = True):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.data = data_dict 
        self.ign = ign
        self.level_value = level_value
        self.prestige_icon = prestige_icon
        self.status_text = status_text
        self.status_color = status_color
        self.skin_image = skin_image
        self.current_tab = "all-time"
        self.current_class = "overall"
        self.show_period_buttons = show_period_buttons
        self.message = None  # Store message reference for timeout handling
        
        self.ign_color = ign_color
        self.guild_tag = guild_tag
        self.guild_hex = guild_hex
        
        if self.ign_color is None or self.guild_tag is None:
            self._load_color()
        
        # Remove period buttons if not tracked (keep only class dropdown)
        if not show_period_buttons:
            # Remove all button items from children (buttons are added by decorators)
            items_to_remove = [item for item in self.children if isinstance(item, discord.ui.Button)]
            for item in items_to_remove:
                self.remove_item(item)
        
        # Add class selector
        self.class_selector = WWClassSelect(self)
        self.add_item(self.class_selector)
        
        self.update_button_styles()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible

    def _load_color(self):
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                if not self.ign_color:
                    self.ign_color = get_rank_color_hex(meta.get('rank'))
                self.guild_tag = meta.get('guild_tag')
                self.guild_hex = meta.get('guild_hex') or "#AAAAAA"
        except: pass

    def update_button_styles(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_tab:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary

    def generate_composite_image(self, tab_name, class_mode="overall"):
        tab_data = self.data[tab_name]
        
        # Get class-specific or overall stats
        if class_mode == "overall":
            wins = tab_data.get('wins', 0)
            losses = tab_data.get('games', 0) - wins
            kills = tab_data.get('kills', 0)
            deaths = tab_data.get('deaths', 0)
            assists = tab_data.get('assists', 0)
            playtime = tab_data.get('playtime', 0)
        else:
            # Class-specific stats (no wins/losses for classes)
            kills = tab_data.get(f'{class_mode}_kills', 0)
            deaths = tab_data.get(f'{class_mode}_deaths', 0)
            assists = tab_data.get(f'{class_mode}_assists', 0)
            playtime = tab_data.get('playtime', 0)
            wins = 0
            losses = 0
        
        wlr = wins / losses if losses > 0 else wins
        kdr = kills / deaths if deaths > 0 else kills
        
        img_io = create_ww_stats_composite_image(
            self.level_value, self.prestige_icon, self.ign, tab_name,
            wins, losses, wlr,
            kills, deaths, kdr, assists,
            self.ign_color, self.guild_tag, self.guild_hex, 
            playtime_seconds=playtime,
            status_text=self.status_text, status_color=self.status_color,
            skin_image=self.skin_image,
            class_mode=class_mode
        )
        return discord.File(img_io, filename=f"{self.ign}_ww_{tab_name}_{class_mode}.png")

    async def handle_tab_click(self, interaction: discord.Interaction, tab_name: str):
        self.current_tab = tab_name
        self.update_button_styles()
        file = self.generate_composite_image(tab_name, self.current_class)
        await interaction.response.edit_message(attachments=[file], view=self)

    async def handle_class_change(self, interaction: discord.Interaction, class_mode: str):
        self.current_class = class_mode
        file = self.generate_composite_image(self.current_tab, class_mode)
        await interaction.response.edit_message(attachments=[file], view=self)

    @discord.ui.button(label="All-time", custom_id="all-time")
    async def all_time(self, interaction, button):
        await self.handle_tab_click(interaction, "all-time")

    @discord.ui.button(label="Session", custom_id="session")
    async def session(self, interaction, button):
        await self.handle_tab_click(interaction, "session")

    @discord.ui.button(label="Daily", custom_id="daily")
    async def daily(self, interaction, button):
        await self.handle_tab_click(interaction, "daily")

    @discord.ui.button(label="Yesterday", custom_id="yesterday")
    async def yesterday(self, interaction, button):
        await self.handle_tab_click(interaction, "yesterday")

    @discord.ui.button(label="Weekly", custom_id="weekly")
    async def weekly(self, interaction, button):
        await self.handle_tab_click(interaction, "weekly")

    @discord.ui.button(label="Monthly", custom_id="monthly")
    async def monthly(self, interaction, button):
        await self.handle_tab_click(interaction, "monthly")


class WWClassSelect(discord.ui.Select):
    """Dropdown for selecting Wool Wars class."""
    def __init__(self, view: WWStatsView):
        options = [
            discord.SelectOption(label="Overall", value="overall", default=True),
            discord.SelectOption(label="Tank", value="tank"),
            discord.SelectOption(label="Assault", value="assault"),
            discord.SelectOption(label="Golem", value="golem"),
            discord.SelectOption(label="Swordsman", value="swordsman"),
            discord.SelectOption(label="Archer", value="archer"),
            discord.SelectOption(label="Engineer", value="engineer"),
        ]
        super().__init__(
            placeholder="Select class...",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="ww_class_select",
        )
        self.view_ref = view

    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref.handle_class_change(interaction, selected)


class CTWStatsView(discord.ui.View):
    """View for CTW stats display."""
    def __init__(self, data_dict, ign, level_value: int, prestige_icon: str, 
                 ign_color: str = None, guild_tag: str = None, guild_hex: str = None,
                 status_text="Online", status_color=(85, 255, 85), skin_image=None):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.data = data_dict 
        self.ign = ign
        self.level_value = level_value
        self.prestige_icon = prestige_icon
        self.status_text = status_text
        self.status_color = status_color
        self.skin_image = skin_image
        self.current_tab = "all-time"
        self.message = None  # Store message reference for timeout handling
        
        self.ign_color = ign_color
        self.guild_tag = guild_tag
        self.guild_hex = guild_hex
        
        if self.ign_color is None or self.guild_tag is None:
            self._load_color()
            
        self.update_button_styles()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible

    def _load_color(self):
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                if not self.ign_color:
                    self.ign_color = get_rank_color_hex(meta.get('rank'))
                self.guild_tag = meta.get('guild_tag')
                self.guild_hex = meta.get('guild_hex') or "#AAAAAA"
        except: pass

    def update_button_styles(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_tab:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary

    def generate_composite_image(self, tab_name):
        tab_data = self.data[tab_name]
        
        wins = tab_data.get('wins', 0)
        losses = tab_data.get('losses', 0)
        draws = tab_data.get('draws', 0)
        kills = tab_data.get('kills', 0)
        deaths = tab_data.get('deaths', 0)
        assists = tab_data.get('assists', 0)
        wools_picked = tab_data.get('wools_picked', 0)
        wools_captured = tab_data.get('wools_captured', 0)
        kills_on_wh = tab_data.get('kills_on_wh', 0)
        deaths_to_wh = tab_data.get('deaths_to_wh', 0)
        kills_as_wh = tab_data.get('kills_as_wh', 0)
        deaths_as_wh = tab_data.get('deaths_as_wh', 0)
        gold_earned = tab_data.get('gold_earned', 0)
        gold_spent = tab_data.get('gold_spent', 0)
        playtime = tab_data.get('playtime', 0)
        
        wlr = wins / losses if losses > 0 else wins
        kdr = kills / deaths if deaths > 0 else kills
        
        img_io = create_ctw_stats_composite_image(
            self.level_value, self.prestige_icon, self.ign, tab_name,
            wins, losses, wlr, draws,
            kills, deaths, kdr, assists,
            wools_picked, wools_captured,
            kills_on_wh, deaths_to_wh, kills_as_wh, deaths_as_wh,
            gold_earned, gold_spent,
            self.ign_color, self.guild_tag, self.guild_hex, 
            playtime_seconds=playtime,
            status_text=self.status_text, status_color=self.status_color,
            skin_image=self.skin_image
        )
        return discord.File(img_io, filename=f"{self.ign}_ctw_{tab_name}.png")

    async def handle_tab_click(self, interaction: discord.Interaction, tab_name: str):
        self.current_tab = tab_name
        self.update_button_styles()
        file = self.generate_composite_image(tab_name)
        await interaction.response.edit_message(attachments=[file], view=self)

    @discord.ui.button(label="All-time", custom_id="all-time")
    async def all_time(self, interaction, button):
        await self.handle_tab_click(interaction, "all-time")

    @discord.ui.button(label="Session", custom_id="session")
    async def session(self, interaction, button):
        await self.handle_tab_click(interaction, "session")

    @discord.ui.button(label="Daily", custom_id="daily")
    async def daily(self, interaction, button):
        await self.handle_tab_click(interaction, "daily")

    @discord.ui.button(label="Yesterday", custom_id="yesterday")
    async def yesterday(self, interaction, button):
        await self.handle_tab_click(interaction, "yesterday")

    @discord.ui.button(label="Weekly", custom_id="weekly")
    async def weekly(self, interaction, button):
        await self.handle_tab_click(interaction, "weekly")

    @discord.ui.button(label="Monthly", custom_id="monthly")
    async def monthly(self, interaction, button):
        await self.handle_tab_click(interaction, "monthly")


# Extended stats view (Template.xlsx layout)
class StatsFullView(discord.ui.View):
    def __init__(self, user_data, ign: str):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.ign = ign
        self.user_data = user_data
        self.meta = user_data.get("meta", {})
        self.current_tab = "all-time"
        self.message = None  # Store message reference for timeout handling
        self._load_color()
        
        self.update_buttons()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible

    def _load_color(self):
        """Load or reload the color and guild info for this username from database"""
        self.ign_color = None
        self.guild_tag = None
        self.guild_color = None
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                if not self.ign_color:
                    self.ign_color = get_rank_color_hex(meta.get('rank'))
                self.guild_tag = meta.get('guild_tag')
                self.guild_color = meta.get('guild_hex')
                print(f"[DEBUG] Loaded color for {self.ign}: {self.ign_color}, guild: [{self.guild_tag}] ({self.guild_color})")
        except Exception as e:
            print(f"[WARNING] Failed to load color for {self.ign}: {e}")

    def _get_value(self, stat_key: str, tab_name: str) -> float:
        # Map tab names to cache keys
        key_map = {"all-time": "lifetime"}
        cache_key = key_map.get(tab_name, tab_name)
        return self.user_data.get("stats", {}).get(stat_key.lower(), {}).get(cache_key, 0)

    def _collect_stats(self, tab_name: str) -> dict:
        def safe_div(n, d):
            return n / d if d else 0
        def fmt_int(v):
            return f"{int(round(v)):,}"
        def fmt_ratio(v):
            return f"{v:.2f}"

        # Base values
        experience = self._get_value('experience', tab_name)
        playtime_seconds = self._get_value('playtime', tab_name)
        games = self._get_value('games_played', tab_name)
        wins = self._get_value('wins', tab_name)
        losses = self._get_value('losses', tab_name)
        kills = self._get_value('kills', tab_name)
        deaths = self._get_value('deaths', tab_name)
        coins = self._get_value('coins', tab_name)
        level_delta = self._get_value('level', tab_name)
        damage = self._get_value('damage_dealt', tab_name)
        kills_void = self._get_value('kills_void', tab_name)
        deaths_void = self._get_value('deaths_void', tab_name)
        magic_wools = self._get_value('magic_wool_hit', tab_name)
        kills_explosive = self._get_value('kills_explosive', tab_name)
        deaths_explosive = self._get_value('deaths_explosive', tab_name)
        sheep_thrown = self._get_value('sheep_thrown', tab_name)
        kills_bow = self._get_value('kills_bow', tab_name)
        deaths_bow = self._get_value('deaths_bow', tab_name)
        kills_melee = self._get_value('kills_melee', tab_name)
        deaths_melee = self._get_value('deaths_melee', tab_name)

        # Derived values
        playtime_hours = playtime_seconds / 3600 if playtime_seconds else 0
        exp_per_hour = safe_div(experience, playtime_hours)
        exp_per_game = safe_div(experience, games)
        wins_per_hour = safe_div(wins, playtime_hours)
        kills_per_hour = safe_div(kills, playtime_hours)
        kdr = safe_div(kills, deaths) if deaths else kills
        wlr = safe_div(wins, losses) if losses else wins
        kills_per_game = safe_div(kills, games)
        kills_per_win = safe_div(kills, wins)
        damage_per_game = safe_div(damage, games)
        damage_per_sheep = safe_div(damage, sheep_thrown)
        damage_per_kill = safe_div(damage, kills)
        void_kdr = safe_div(kills_void, deaths_void) if deaths_void else kills_void
        wools_per_game = safe_div(magic_wools, games)
        explosive_kdr = safe_div(kills_explosive, deaths_explosive) if deaths_explosive else kills_explosive
        sheeps_per_game = safe_div(sheep_thrown, games)
        bow_kdr = safe_div(kills_bow, deaths_bow) if deaths_bow else kills_bow
        melee_kdr = safe_div(kills_melee, deaths_melee) if deaths_melee else kills_melee
        survival_rate = safe_div(games - deaths, games) if games else 0

        stats = {
            "username": self.ign,
            "guild": f"[{_safe_guild_tag(self.guild_tag)}]" if _safe_guild_tag(self.guild_tag) else "N/A",
            "playtime": format_playtime(int(playtime_seconds)) if playtime_seconds else "0s",
            "level": fmt_int(self._get_value('level', tab_name)),
            "exp_per_hour": fmt_ratio(exp_per_hour),
            "exp_per_game": fmt_ratio(exp_per_game),
            "wins_per_hour": fmt_ratio(wins_per_hour),
            "kills_per_hour": fmt_ratio(kills_per_hour),
            "sheepwars_label": "",
            "wins": fmt_int(wins),
            "losses": fmt_int(losses),
            "wlr": fmt_ratio(wlr),
            "layers": fmt_ratio(level_delta),
            "coins": fmt_int(coins),
            "kills": fmt_int(kills),
            "deaths": fmt_int(deaths),
            "kdr": fmt_ratio(kdr),
            "kills_per_game": fmt_ratio(kills_per_game),
            "kills_per_win": fmt_ratio(kills_per_win),
            "damage": fmt_int(damage),
            "damage_per_game": fmt_ratio(damage_per_game),
            "damage_per_kill": fmt_ratio(damage_per_kill),
            "damage_per_sheep": fmt_ratio(damage_per_sheep),
            "void_kills": fmt_int(kills_void),
            "void_deaths": fmt_int(deaths_void),
            "void_kdr": fmt_ratio(void_kdr),
            "magic_wools": fmt_int(magic_wools),
            "wools_per_game": fmt_ratio(wools_per_game),
            "explosive_kills": fmt_int(kills_explosive),
            "explosive_deaths": fmt_int(deaths_explosive),
            "explosive_kdr": fmt_ratio(explosive_kdr),
            "sheeps_thrown": fmt_int(sheep_thrown),
            "sheeps_per_game": fmt_ratio(sheeps_per_game),
            "bow_kills": fmt_int(kills_bow),
            "bow_deaths": fmt_int(deaths_bow),
            "bow_kdr": fmt_ratio(bow_kdr),
            "games_played": fmt_int(games),
            "melee_kills": fmt_int(kills_melee),
            "melee_deaths": fmt_int(deaths_melee),
            "melee_kdr": fmt_ratio(melee_kdr),
            "survival_rate": fmt_ratio(survival_rate),
            # Add raw stat key aliases for single stat comparison
            "damage_dealt": fmt_int(damage),
            "sheep_thrown": fmt_int(sheep_thrown),
            "magic_wool_hit": fmt_int(magic_wools),
            "kills_void": fmt_int(kills_void),
            "deaths_void": fmt_int(deaths_void),
            "kills_explosive": fmt_int(kills_explosive),
            "deaths_explosive": fmt_int(deaths_explosive),
            "kills_bow": fmt_int(kills_bow),
            "deaths_bow": fmt_int(deaths_bow),
            "kills_melee": fmt_int(kills_melee),
            "deaths_melee": fmt_int(deaths_melee),
        }

        ordered_fields = [
            ("Wins", stats["wins"]), ("Losses", stats["losses"]), ("WLR", stats["wlr"]), ("Damage/Game", stats["damage_per_game"]), ("Coins", stats["coins"]),
            ("Kills", stats["kills"]), ("Deaths", stats["deaths"]), ("KDR", stats["kdr"]), ("Kill/Game", stats["kills_per_game"]), ("Kill/Win", stats["kills_per_win"]),
            ("Damage dealt", stats["damage"]), ("Damage/Kill", stats["damage_per_kill"]), ("Void kills", stats["void_kills"]), ("Void deaths", stats["void_deaths"]), ("Void KDR", stats["void_kdr"]),
            ("Magic wools", stats["magic_wools"]), ("Wools/Game", stats["wools_per_game"]), ("Explosive kills", stats["explosive_kills"]), ("Explosive deaths", stats["explosive_deaths"]), ("Explosive KDR", stats["explosive_kdr"]),
            ("Sheeps thrown", stats["sheeps_thrown"]), ("Sheeps thrown/Game", stats["sheeps_per_game"]), ("Bow kills", stats["bow_kills"]), ("Bow deaths", stats["bow_deaths"]), ("Bow KDR", stats["bow_kdr"]),
            ("Games Played", stats["games_played"]), ("Damage/Sheep", stats["damage_per_sheep"]), ("Meelee kills", stats["melee_kills"]), ("Meelee Deaths", stats["melee_deaths"]), ("Meelee KDR", stats["melee_kdr"]),
        ]
        stats["ordered_fields"] = ordered_fields
        return stats

    def update_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                child.style = discord.ButtonStyle.primary if child.custom_id == self.current_tab else discord.ButtonStyle.secondary

    def generate_full_image(self, tab_name: str):
        stats = self._collect_stats(tab_name)
        if Image is not None:
            try:
                img_io = create_full_stats_image(self.ign, tab_name, self.meta.get("level", 0), self.meta.get("icon", ""), stats, self.ign_color, self.guild_tag, self.guild_color)
                filename = f"{self.ign}_{tab_name}_stats_full.png"
                return None, discord.File(img_io, filename=filename)
            except Exception as e:
                print(f"[WARNING] Full stats image generation failed: {e}")

        embed = discord.Embed(title=f"{self.ign} - {tab_name.title()} stats")
        for label, value in stats.get("ordered_fields", [])[:25]:
            embed.add_field(name=label, value=f"```{value}```", inline=True)
        return embed, None

    @discord.ui.button(label="All-time", custom_id="all-time", style=discord.ButtonStyle.primary)
    async def full_all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "all-time"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Session", custom_id="session", style=discord.ButtonStyle.secondary)
    async def full_session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "session"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Daily", custom_id="daily", style=discord.ButtonStyle.secondary)
    async def full_daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "daily"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Yesterday", custom_id="yesterday", style=discord.ButtonStyle.secondary)
    async def full_yesterday_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "yesterday"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Weekly", custom_id="weekly", style=discord.ButtonStyle.secondary)
    async def full_weekly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "weekly"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Monthly", custom_id="monthly", style=discord.ButtonStyle.secondary)
    async def full_monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "monthly"
        self.update_buttons()
        embed, file = self.generate_full_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)


class CompareView(discord.ui.View):
    def __init__(self, user_data1, user_data2, ign1: str, ign2: str, stat: str = None):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.ign1 = ign1
        self.ign2 = ign2
        self.user_data1 = user_data1
        self.user_data2 = user_data2
        self.meta1 = user_data1.get("meta", {})
        self.meta2 = user_data2.get("meta", {})
        self.current_tab = "all-time"
        self.stat = stat  # Single stat to compare, if specified
        self.message = None  # Store message reference for timeout handling
        self._load_colors()
        
        self.update_buttons()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible

    def _load_colors(self):
        """Load colors and guild info for both usernames from database"""
        self.ign_color1 = None
        self.guild_tag1 = None
        self.guild_color1 = None
        self.ign_color2 = None
        self.guild_tag2 = None
        self.guild_color2 = None
        
        try:
            meta1 = get_user_meta(self.ign1)
            if meta1:
                self.ign_color1 = meta1.get('ign_color')
                if not self.ign_color1:
                    self.ign_color1 = get_rank_color_hex(meta1.get('rank'))
                self.guild_tag1 = meta1.get('guild_tag')
                self.guild_color1 = meta1.get('guild_hex')
        except Exception as e:
            print(f"[WARNING] Failed to load color for {self.ign1}: {e}")
        
        try:
            meta2 = get_user_meta(self.ign2)
            if meta2:
                self.ign_color2 = meta2.get('ign_color')
                if not self.ign_color2:
                    self.ign_color2 = get_rank_color_hex(meta2.get('rank'))
                self.guild_tag2 = meta2.get('guild_tag')
                self.guild_color2 = meta2.get('guild_hex')
        except Exception as e:
            print(f"[WARNING] Failed to load color for {self.ign2}: {e}")

    def _get_value(self, user_data, stat_key: str, tab_name: str) -> float:
        # Map tab names to cache keys
        key_map = {"all-time": "lifetime"}
        cache_key = key_map.get(tab_name, tab_name)
        stats = user_data.get("stats", {})
        stat_data = stats.get(stat_key.lower(), {})
        value = stat_data.get(cache_key, 0)
        # Debug: if we're getting 0 and we're on all-time, check if data exists at all
        if value == 0 and cache_key == "lifetime" and stat_data:
            print(f"[DEBUG] Stat {stat_key} has data {stat_data} but returning 0 for {cache_key}")
        return value

    def _get_stat_display_name(self, stat_key: str) -> str:
        """Convert stat key to display name."""
        stat_names = {
            'kills': 'Total Kills',
            'deaths': 'Total Deaths',
            'kdr': 'K/D Ratio',
            'wins': 'Wins',
            'losses': 'Losses',
            'wlr': 'W/L Ratio',
            'damage_dealt': 'Damage Dealt',
            'damage': 'Damage Dealt',
            'games_played': 'Games Played',
            'sheep_thrown': 'Sheep Thrown',
            'sheeps_thrown': 'Sheep Thrown',
            'magic_wool_hit': 'Magic Wool Hit',
            'magic_wools': 'Magic Wool Hit',
            'kills_void': 'Void Kills',
            'void_kills': 'Void Kills',
            'deaths_void': 'Void Deaths',
            'void_deaths': 'Void Deaths',
            'kills_explosive': 'Explosive Kills',
            'explosive_kills': 'Explosive Kills',
            'deaths_explosive': 'Explosive Deaths',
            'explosive_deaths': 'Explosive Deaths',
            'kills_bow': 'Bow Kills',
            'bow_kills': 'Bow Kills',
            'deaths_bow': 'Bow Deaths',
            'bow_deaths': 'Bow Deaths',
            'kills_melee': 'Melee Kills',
            'melee_kills': 'Melee Kills',
            'deaths_melee': 'Melee Deaths',
            'melee_deaths': 'Melee Deaths',
            'void_kdr': 'Void K/D Ratio',
            'explosive_kdr': 'Explosive K/D Ratio',
            'bow_kdr': 'Bow K/D Ratio',
            'melee_kdr': 'Melee K/D Ratio',
            'kills_per_game': 'Kills per Game',
            'kills_per_win': 'Kills per Win',
            'kills_per_hour': 'Kills per Hour',
            'damage_per_game': 'Damage per Game',
            'damage_per_sheep': 'Damage per Sheep',
            'wools_per_game': 'Wools per Game',
            'sheeps_per_game': 'Sheeps per Game',
        }
        return stat_names.get(stat_key, stat_key.replace('_', ' ').title())

    def _collect_stats(self, user_data, tab_name: str) -> dict:
        def safe_div(n, d):
            return n / d if d else 0
        def fmt_int(v):
            return f"{int(round(v)):,}"
        def fmt_ratio(v):
            return f"{v:.2f}"

        # Debug: Check what stats we have
        stats_available = user_data.get("stats", {})
        print(f"[DEBUG] _collect_stats for tab {tab_name}. User has {len(stats_available)} stat types. Sample: {list(stats_available.keys())[:5]}")

        # Base values
        experience = self._get_value(user_data, 'experience', tab_name)
        playtime_seconds = self._get_value(user_data, 'playtime', tab_name)
        games = self._get_value(user_data, 'games_played', tab_name)
        wins = self._get_value(user_data, 'wins', tab_name)
        print(f"[DEBUG] Basic stats - exp:{experience}, playtime:{playtime_seconds}, games:{games}, wins:{wins}")
        losses = self._get_value(user_data, 'losses', tab_name)
        kills = self._get_value(user_data, 'kills', tab_name)
        deaths = self._get_value(user_data, 'deaths', tab_name)
        coins = self._get_value(user_data, 'coins', tab_name)
        layers = self._get_value(user_data, 'available_layers', tab_name)
        damage = self._get_value(user_data, 'damage_dealt', tab_name)
        kills_void = self._get_value(user_data, 'kills_void', tab_name)
        deaths_void = self._get_value(user_data, 'deaths_void', tab_name)
        magic_wools = self._get_value(user_data, 'magic_wool_hit', tab_name)
        kills_explosive = self._get_value(user_data, 'kills_explosive', tab_name)
        deaths_explosive = self._get_value(user_data, 'deaths_explosive', tab_name)
        sheep_thrown = self._get_value(user_data, 'sheep_thrown', tab_name)
        kills_bow = self._get_value(user_data, 'kills_bow', tab_name)
        deaths_bow = self._get_value(user_data, 'deaths_bow', tab_name)
        kills_melee = self._get_value(user_data, 'kills_melee', tab_name)
        deaths_melee = self._get_value(user_data, 'deaths_melee', tab_name)

        # Derived values
        playtime_hours = playtime_seconds / 3600 if playtime_seconds else 0
        exp_per_hour = safe_div(experience, playtime_hours)
        exp_per_game = safe_div(experience, games)
        wins_per_hour = safe_div(wins, playtime_hours)
        kills_per_hour = safe_div(kills, playtime_hours)
        kdr = safe_div(kills, deaths) if deaths else kills
        wlr = safe_div(wins, losses) if losses else wins
        kills_per_game = safe_div(kills, games)
        kills_per_win = safe_div(kills, wins)
        damage_per_game = safe_div(damage, games)
        damage_per_sheep = safe_div(damage, sheep_thrown)
        damage_per_kill = safe_div(damage, kills)
        void_kdr = safe_div(kills_void, deaths_void) if deaths_void else kills_void
        wools_per_game = safe_div(magic_wools, games)
        explosive_kdr = safe_div(kills_explosive, deaths_explosive) if deaths_explosive else kills_explosive
        sheeps_per_game = safe_div(sheep_thrown, games)
        bow_kdr = safe_div(kills_bow, deaths_bow) if deaths_bow else kills_bow
        melee_kdr = safe_div(kills_melee, deaths_melee) if deaths_melee else kills_melee
        survival_rate = safe_div(games - deaths, games) if games else 0

        return {
            "playtime": format_playtime(int(playtime_seconds)) if playtime_seconds else "0s",
            "exp_per_hour": fmt_ratio(exp_per_hour),
            "exp_per_game": fmt_ratio(exp_per_game),
            "wins_per_hour": fmt_ratio(wins_per_hour),
            "kills_per_hour": fmt_ratio(kills_per_hour),
            "wins": fmt_int(wins),
            "losses": fmt_int(losses),
            "wlr": fmt_ratio(wlr),
            "layers": fmt_int(layers),
            "coins": fmt_int(coins),
            "kills": fmt_int(kills),
            "deaths": fmt_int(deaths),
            "kdr": fmt_ratio(kdr),
            "kills_per_game": fmt_ratio(kills_per_game),
            "kills_per_win": fmt_ratio(kills_per_win),
            "damage": fmt_int(damage),
            "damage_per_game": fmt_ratio(damage_per_game),
            "damage_per_kill": fmt_ratio(damage_per_kill),
            "damage_per_sheep": fmt_ratio(damage_per_sheep),
            "void_kills": fmt_int(kills_void),
            "void_deaths": fmt_int(deaths_void),
            "void_kdr": fmt_ratio(void_kdr),
            "magic_wools": fmt_int(magic_wools),
            "wools_per_game": fmt_ratio(wools_per_game),
            "explosive_kills": fmt_int(kills_explosive),
            "explosive_deaths": fmt_int(deaths_explosive),
            "explosive_kdr": fmt_ratio(explosive_kdr),
            "sheeps_thrown": fmt_int(sheep_thrown),
            "sheeps_per_game": fmt_ratio(sheeps_per_game),
            "bow_kills": fmt_int(kills_bow),
            "bow_deaths": fmt_int(deaths_bow),
            "bow_kdr": fmt_ratio(bow_kdr),
            "games_played": fmt_int(games),
            "melee_kills": fmt_int(kills_melee),
            "melee_deaths": fmt_int(deaths_melee),
            "melee_kdr": fmt_ratio(melee_kdr),
            "survival_rate": fmt_ratio(survival_rate),
        }

    def update_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                child.style = discord.ButtonStyle.primary if child.custom_id == self.current_tab else discord.ButtonStyle.secondary

    def generate_compare_image(self, tab_name: str):
        stats1 = self._collect_stats(self.user_data1, tab_name)
        stats2 = self._collect_stats(self.user_data2, tab_name)
        
        if Image is not None:
            try:
                # Check if we're comparing a single stat
                if self.stat:
                    img_io = create_single_stat_compare_image(
                        self.ign1, self.ign2, tab_name, self.stat,
                        stats1, stats2,
                        self.meta1.get("level", 0), self.meta2.get("level", 0),
                        self.meta1.get("icon", ""), self.meta2.get("icon", ""),
                        self.ign_color1, self.ign_color2,
                        self.guild_tag1, self.guild_tag2,
                        self.guild_color1, self.guild_color2
                    )
                    filename = f"{self.ign1}_vs_{self.ign2}_{self.stat}_{tab_name}_compare.png"
                else:
                    img_io = create_compare_stats_image(
                        self.ign1, self.ign2, tab_name,
                        stats1, stats2,
                        self.meta1.get("level", 0), self.meta2.get("level", 0),
                        self.meta1.get("icon", ""), self.meta2.get("icon", ""),
                        self.ign_color1, self.ign_color2,
                        self.guild_tag1, self.guild_tag2,
                        self.guild_color1, self.guild_color2
                    )
                    filename = f"{self.ign1}_vs_{self.ign2}_{tab_name}_compare.png"
                return None, discord.File(img_io, filename=filename)
            except Exception as e:
                print(f"[WARNING] Compare stats image generation failed: {e}")
        
        # Fallback to embed
        if self.stat:
            embed = discord.Embed(title=f"{self.ign1} vs {self.ign2} - {self.stat.title()} - {tab_name.title()}")
            stat_label = self._get_stat_display_name(self.stat)
            embed.add_field(name=f"{self.ign1} {stat_label}", value=f"```{stats1.get(self.stat, '0')}```", inline=True)
            embed.add_field(name=f"{self.ign2} {stat_label}", value=f"```{stats2.get(self.stat, '0')}```", inline=True)
            embed.add_field(name="", value="", inline=True)
        else:
            embed = discord.Embed(title=f"{self.ign1} vs {self.ign2} - {tab_name.title()} Comparison")
            embed.add_field(name=f"{self.ign1} Wins", value=f"```{stats1['wins']}```", inline=True)
            embed.add_field(name=f"{self.ign2} Wins", value=f"```{stats2['wins']}```", inline=True)
            embed.add_field(name="", value="", inline=True)
        return embed, None

    @discord.ui.button(label="All-time", custom_id="all-time", style=discord.ButtonStyle.primary)
    async def compare_all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "all-time"
        self.update_buttons()
        embed, file = self.generate_compare_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Session", custom_id="session", style=discord.ButtonStyle.secondary)
    async def compare_session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "session"
        self.update_buttons()
        embed, file = self.generate_compare_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Daily", custom_id="daily", style=discord.ButtonStyle.secondary)
    async def compare_daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "daily"
        self.update_buttons()
        embed, file = self.generate_compare_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Yesterday", custom_id="yesterday", style=discord.ButtonStyle.secondary)
    async def compare_yesterday_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "yesterday"
        self.update_buttons()
        embed, file = self.generate_compare_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Weekly", custom_id="weekly", style=discord.ButtonStyle.secondary)
    async def compare_weekly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "weekly"
        self.update_buttons()
        embed, file = self.generate_compare_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Monthly", custom_id="monthly", style=discord.ButtonStyle.secondary)
    async def compare_monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "monthly"
        self.update_buttons()
        embed, file = self.generate_compare_image(self.current_tab)
        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)


class DistributionView(discord.ui.View):
    def __init__(self, user_data, ign: str, mode: str):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.ign = ign
        self.user_data = user_data
        self.mode = mode  # 'kill' or 'death'
        self.current_tab = "all-time"
        self.message = None  # Store message reference for timeout handling
        # Colors for legend slices
        self.slice_colors = {
            "void": (90, 155, 255),        # blue
            "explosive": (255, 119, 84),   # orange-red
            "bow": (255, 214, 102),        # golden
            "melee": (126, 217, 126),      # green
        }
        self.update_buttons()
        
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible

    def update_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                child.style = discord.ButtonStyle.primary if child.custom_id.endswith(self.current_tab) else discord.ButtonStyle.secondary

    def _get_counts(self, tab_name: str):
        if self.mode == "kill":
            keys = [
                ("Melee Kills", "kills_melee", "melee"),
                ("Bow Kills", "kills_bow", "bow"),
                ("Explosive Kills", "kills_explosive", "explosive"),
                ("Void Kills", "kills_void", "void"),
            ]
        else:
            keys = [
                ("Melee Deaths", "deaths_melee", "melee"),
                ("Bow Deaths", "deaths_bow", "bow"),
                ("Explosive Deaths", "deaths_explosive", "explosive"),
                ("Void Deaths", "deaths_void", "void"),
            ]

        key_map = {"all-time": "lifetime"}
        cache_key = key_map.get(tab_name, tab_name)
        counts = []
        stats = self.user_data.get("stats", {})
        for label, key, color_key in keys:
            val = stats.get(key, {}).get(cache_key, 0)
            counts.append((label, max(0, float(val)), color_key))
        return counts

    def generate_distribution(self, tab_name: str):
        counts = self._get_counts(tab_name)
        total = sum(v for _, v, _ in counts)
        metric_label = "Kill" if self.mode == "kill" else "Death"

        if total <= 0:
            embed = discord.Embed(
                title=f"{self.ign} - {tab_name.title()} {metric_label} Distribution",
                description="No data for this period.",
                color=discord.Color.from_rgb(54, 57, 63),
            )
            return embed, None

        slice_payload = []
        for label, value, color_key in counts:
            color = self.slice_colors.get(color_key, (180, 180, 180))
            slice_payload.append((label, value, color))

        if Image is not None:
            try:
                title = f"{self.ign} - {tab_name.title()} {metric_label} Distribution"
                img_io = create_distribution_pie(title, slice_payload)
                filename = f"{self.ign}_{self.mode}_{tab_name}_distribution.png"
                return None, discord.File(img_io, filename=filename)
            except Exception as e:
                print(f"[WARNING] Distribution image generation failed: {e}")

        # Fallback to embed if Pillow is missing or image failed
        embed = discord.Embed(
            title=f"{self.ign} - {tab_name.title()} {metric_label} Distribution",
            color=discord.Color.from_rgb(54, 57, 63),
        )
        lines = []
        for label, value, _ in counts:
            percent = (value / total * 100) if total else 0
            lines.append(f"{label}: {value} ({percent:.1f}%)")
        embed.description = "\n".join(lines)
        return embed, None

    @discord.ui.button(label="All-time", custom_id="dist-all-time", style=discord.ButtonStyle.primary)
    async def dist_all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "all-time"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Session", custom_id="dist-session", style=discord.ButtonStyle.secondary)
    async def dist_session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "session"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Daily", custom_id="dist-daily", style=discord.ButtonStyle.secondary)
    async def dist_daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "daily"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Yesterday", custom_id="dist-yesterday", style=discord.ButtonStyle.secondary)
    async def dist_yesterday_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "yesterday"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Weekly", custom_id="dist-weekly", style=discord.ButtonStyle.secondary)
    async def dist_weekly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "weekly"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])

    @discord.ui.button(label="Monthly", custom_id="dist-monthly", style=discord.ButtonStyle.secondary)
    async def dist_monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_tab = "monthly"
        self.update_buttons()
        embed, file = self.generate_distribution(self.current_tab)
        if file:
            await interaction.response.edit_message(embed=None, view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self, attachments=[])


# ==============================
# RATIOS MILESTONE PREDICTION
# ==============================

def calculate_next_milestone(current_value: float) -> int:
    """Calculate the next round number milestone for a ratio.
    
    For example:
    - 3.30 -> 4
    - 4.00 -> 5
    - 0.85 -> 1
    - 10.5 -> 11
    """
    return math.ceil(current_value)


def calculate_wins_needed_for_wlr(current_wins: int, current_losses: int, target_wlr: float) -> int:
    """Calculate how many additional wins are needed to reach a target WLR.
    
    Formula: target_wlr = (current_wins + x) / current_losses
    Solving for x: x = (target_wlr * current_losses) - current_wins
    """
    if current_losses == 0:
        return 0  # Already at infinity WLR
    
    wins_needed = (target_wlr * current_losses) - current_wins
    return max(0, math.ceil(wins_needed))


def calculate_kills_needed_for_kdr(current_kills: int, current_deaths: int, target_kdr: float) -> int:
    """Calculate how many additional kills are needed to reach a target KDR.
    
    Formula: target_kdr = (current_kills + x) / current_deaths
    Solving for x: x = (target_kdr * current_deaths) - current_kills
    """
    if current_deaths == 0:
        return 0  # Already at infinity KDR
    
    kills_needed = (target_kdr * current_deaths) - current_kills
    return max(0, math.ceil(kills_needed))


def format_time_estimate(remaining: int, rate_per_hour: float) -> str:
    """Format time estimation based on remaining items and rate per hour.
    
    Args:
        remaining: Number of items remaining (wins or kills)
        rate_per_hour: Rate per hour (wins/hour or kills/hour)
    
    Returns:
        Formatted string like "5h 30m", "2d 3h", "Never" (if rate is 0), or "Complete" (if remaining is 0)
    """
    if remaining <= 0:
        return "Complete"
    
    if rate_per_hour <= 0:
        return "Never"
    
    hours_needed = remaining / rate_per_hour
    
    if hours_needed < 1:
        minutes = int(hours_needed * 60)
        return f"{minutes}m"
    elif hours_needed < 24:
        hours = int(hours_needed)
        minutes = int((hours_needed - hours) * 60)
        if minutes > 0:
            return f"{hours}h {minutes}m"
        return f"{hours}h"
    else:
        days = int(hours_needed / 24)
        remaining_hours = int(hours_needed % 24)
        if remaining_hours > 0:
            return f"{days}d {remaining_hours}h"
        return f"{days}d"


def create_ratios_image(ign: str, level: int, icon: str, tab_name: str, 
                        milestone_wlr: int, wins_remaining: int, wins_per_hour: float, wlr_estimate: str,
                        milestone_kdr: int, kills_remaining: int, kills_per_hour: float, kdr_estimate: str,
                        ign_color: str = None, guild_tag: str = None, guild_hex: str = None) -> io.BytesIO:
    """Create an image showing milestone predictions for WLR and KDR."""
    
    if Image is None:
        raise RuntimeError("Pillow not available")
    
    # Canvas dimensions
    canvas_w, canvas_h = 950, 340
    margin = 35
    composite = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    draw = ImageDraw.Draw(composite)
    
    # Render prestige and username - CENTERED
    ign_rgb = (85, 255, 255)
    if ign_color:
        try:
            ign_rgb = tuple(int(str(ign_color).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass
    
    # Get guild color
    g_rgb = (170, 170, 170)
    if guild_tag and guild_hex:
        if str(guild_hex).upper() in MINECRAFT_NAME_TO_HEX:
            guild_hex = MINECRAFT_NAME_TO_HEX[str(guild_hex).upper()]
        try:
            g_rgb = tuple(int(str(guild_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass
    
    # Get prestige segments and render
    segs = get_prestige_segments(level, icon)
    font_title = _load_font("DejaVuSans-Bold.ttf", 24)
    
    # Render prestige
    txt_io = _render_text_segments_to_image(segs, font=font_title, padding=(0, 0))
    prestige_img = Image.open(txt_io).convert("RGBA")
    
    # Calculate total width for centering
    temp_draw = ImageDraw.Draw(Image.new('RGBA', (1, 1)))
    username_text = f" {ign}"
    username_bbox = temp_draw.textbbox((0, 0), username_text, font=font_title)
    username_width = username_bbox[2] - username_bbox[0]
    
    guild_width = 0
    if guild_tag:
        safe_tag = _safe_guild_tag(guild_tag)
        guild_text = f" [{safe_tag}]"
        guild_bbox = temp_draw.textbbox((0, 0), guild_text, font=font_title)
        guild_width = guild_bbox[2] - guild_bbox[0]
    
    total_width = prestige_img.width + username_width + guild_width
    
    # Start position for centered title
    title_y = 25
    x_pos = (canvas_w - total_width) // 2
    
    # Paste prestige
    composite.paste(prestige_img, (x_pos, title_y), prestige_img)
    x_pos += prestige_img.width
    
    # Draw username
    draw.text((x_pos, title_y + prestige_img.height // 2), username_text, 
              font=font_title, fill=ign_rgb, anchor="lm")
    x_pos += username_width
    
    # Draw guild tag if present
    if guild_tag:
        draw.text((x_pos, title_y + prestige_img.height // 2), guild_text, 
                  font=font_title, fill=g_rgb, anchor="lm")
    
    # Subtitle (tab name)
    font_subtitle = _load_font("DejaVuSans.ttf", 16)
    subtitle_text = f"Milestone Predictions - {tab_name.title()}"
    subtitle_bbox = draw.textbbox((0, 0), subtitle_text, font=font_subtitle)
    subtitle_w = subtitle_bbox[2] - subtitle_bbox[0]
    draw.text(((canvas_w - subtitle_w) // 2, title_y + 38), subtitle_text, 
              font=font_subtitle, fill=(180, 180, 200))
    
    # Separator line
    line_y = title_y + 70
    draw.line([margin, line_y, canvas_w - margin, line_y], fill=(60, 60, 80), width=2)
    
    # Table headers
    table_y = line_y + 25
    col_widths = [190, 165, 270, 200]
    col_spacing = 20
    col_x = [
        margin,
        margin + col_widths[0] + col_spacing,
        margin + col_widths[0] + col_widths[1] + col_spacing * 2,
        margin + col_widths[0] + col_widths[1] + col_widths[2] + col_spacing * 3
    ]
    
    font_header = _load_font("DejaVuSans-Bold.ttf", 15)
    headers = ["Milestone", "Missing", "Current Rate", "Estimation"]
    
    for i, (x, header) in enumerate(zip(col_x, headers)):
        draw.text((x, table_y), header, font=font_header, fill=(255, 255, 255))
    
    # Header underline
    draw.line([margin, table_y + 26, canvas_w - margin, table_y + 26], fill=(80, 80, 100), width=1)
    
    # Table rows
    row_height = 52
    font_value = _load_font("DejaVuSans.ttf", 16)
    font_label = _load_font("DejaVuSans-Bold.ttf", 16)
    
    # Row 1: WLR
    row_y = table_y + 42
    wlr_color = (85, 255, 85)
    draw.text((col_x[0], row_y), f"{milestone_wlr}.00 WLR", font=font_label, fill=wlr_color)
    draw.text((col_x[1], row_y), f"{wins_remaining:,} Wins", font=font_value, fill=(255, 255, 255))
    draw.text((col_x[2], row_y), f"{wins_per_hour:.2f} Wins/Hour", font=font_value, fill=(255, 255, 255))
    draw.text((col_x[3], row_y), wlr_estimate, font=font_value, fill=(255, 170, 0))
    
    # Row separator
    draw.line([margin, row_y + 28, canvas_w - margin, row_y + 28], fill=(50, 50, 60), width=1)
    
    # Row 2: KDR
    row_y += row_height
    kdr_color = (85, 255, 255)
    draw.text((col_x[0], row_y), f"{milestone_kdr}.00 KDR", font=font_label, fill=kdr_color)
    draw.text((col_x[1], row_y), f"{kills_remaining:,} Kills", font=font_value, fill=(255, 255, 255))
    draw.text((col_x[2], row_y), f"{kills_per_hour:.2f} Kills/Hour", font=font_value, fill=(255, 255, 255))
    draw.text((col_x[3], row_y), kdr_estimate, font=font_value, fill=(255, 170, 0))
    
    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 13)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 28), footer_text, 
              font=font_footer, fill=(60, 60, 65))
    
    out = io.BytesIO()
    composite.convert("RGB").save(out, format='PNG')
    out.seek(0)
    return out


class RatiosView(discord.ui.View):
    """View for displaying ratio milestone predictions with period tabs."""
    
    def __init__(self, user_data: dict, ign: str):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.ign = ign
        self.user_data = user_data
        self.meta = user_data.get("meta", {})
        self.current_tab = "all-time"
        self.message = None  # Store message reference for timeout handling
        self._load_color()
        self.update_buttons()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
    
    def _load_color(self):
        """Load color and guild info for this username from database."""
        self.ign_color = None
        self.guild_tag = None
        self.guild_color = None
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                if not self.ign_color:
                    self.ign_color = get_rank_color_hex(meta.get('rank'))
                self.guild_tag = meta.get('guild_tag')
                self.guild_color = meta.get('guild_hex')
        except Exception as e:
            print(f"[WARNING] Failed to load color for {self.ign}: {e}")
    
    def _get_tab_stats(self, tab_name: str) -> dict:
        """Extract stats for a specific tab."""
        # Map tab names to cache keys
        key_map = {"all-time": "lifetime"}
        cache_key = key_map.get(tab_name, tab_name)
        
        stats = self.user_data.get("stats", {})
        wins = stats.get("wins", {}).get(cache_key, 0)
        losses = stats.get("losses", {}).get(cache_key, 0)
        kills = stats.get("kills", {}).get(cache_key, 0)
        deaths = stats.get("deaths", {}).get(cache_key, 0)
        playtime = stats.get("playtime", {}).get(cache_key, 0)
        
        # Calculate ratios
        wlr = wins / losses if losses > 0 else wins
        kdr = kills / deaths if deaths > 0 else kills
        
        # Calculate per-hour rates
        playtime_hours = playtime / 3600 if playtime > 0 else 0
        wins_per_hour = wins / playtime_hours if playtime_hours > 0 else 0
        kills_per_hour = kills / playtime_hours if playtime_hours > 0 else 0
        
        return {
            "wins": int(wins),
            "losses": int(losses),
            "kills": int(kills),
            "deaths": int(deaths),
            "wlr": wlr,
            "kdr": kdr,
            "wins_per_hour": wins_per_hour,
            "kills_per_hour": kills_per_hour
        }
    
    def generate_ratios_image(self, tab_name: str):
        """Generate ratios prediction image for the specified tab."""
        stats = self._get_tab_stats(tab_name)
        
        # Calculate WLR milestone
        current_wlr = stats["wlr"]
        milestone_wlr = calculate_next_milestone(current_wlr)
        wins_remaining = calculate_wins_needed_for_wlr(stats["wins"], stats["losses"], milestone_wlr)
        wlr_estimate = format_time_estimate(wins_remaining, stats["wins_per_hour"])
        
        # Calculate KDR milestone
        current_kdr = stats["kdr"]
        milestone_kdr = calculate_next_milestone(current_kdr)
        kills_remaining = calculate_kills_needed_for_kdr(stats["kills"], stats["deaths"], milestone_kdr)
        kdr_estimate = format_time_estimate(kills_remaining, stats["kills_per_hour"])
        
        # Get level and prestige icon
        level = self.meta.get("level", 0)
        icon = self.meta.get("icon", "")
        
        # Create image
        img_io = create_ratios_image(
            self.ign, level, icon, tab_name,
            milestone_wlr, wins_remaining, stats["wins_per_hour"], wlr_estimate,
            milestone_kdr, kills_remaining, stats["kills_per_hour"], kdr_estimate,
            self.ign_color, self.guild_tag, self.guild_color
        )
        
        return discord.File(img_io, filename=f"{self.ign}_ratios_{tab_name}.png")
    
    def update_buttons(self):
        """Update button styles based on current tab."""
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_tab:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary
    
    async def handle_tab_click(self, interaction: discord.Interaction, tab_name: str):
        """Handle tab button clicks."""
        self.current_tab = tab_name
        self.update_buttons()
        file = self.generate_ratios_image(tab_name)
        await interaction.response.edit_message(attachments=[file], view=self)
    
    @discord.ui.button(label="All-time", custom_id="all-time", style=discord.ButtonStyle.primary)
    async def all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "all-time")
    
    @discord.ui.button(label="Session", custom_id="session", style=discord.ButtonStyle.secondary)
    async def session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "session")
    
    @discord.ui.button(label="Daily", custom_id="daily", style=discord.ButtonStyle.secondary)
    async def daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "daily")
    
    @discord.ui.button(label="Yesterday", custom_id="yesterday", style=discord.ButtonStyle.secondary)
    async def yesterday_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "yesterday")
    
    @discord.ui.button(label="Weekly", custom_id="weekly", style=discord.ButtonStyle.secondary)
    async def weekly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "weekly")
    
    @discord.ui.button(label="Monthly", custom_id="monthly", style=discord.ButtonStyle.secondary)
    async def monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "monthly")


# LEVEL PROGRESS & ETA PREDICTION
# ================================

def create_level_progress_image(ign: str, level: int, icon: str, tab_name: str,
                                 current_xp_progress: int, xp_for_next_level: int,
                                 target_next_level: int, xp_to_next_level: int, exp_per_hour: float, eta_next_level: str,
                                 target_next_prestige: int, xp_to_next_prestige: int, eta_next_prestige: str,
                                 ign_color: str = None, guild_tag: str = None, guild_hex: str = None,
                                 custom_level_data: tuple = None) -> io.BytesIO:
    """Create an image showing level progress and ETA predictions."""
    
    if Image is None:
        raise RuntimeError("Pillow not available")
    
    # Build ordered list of targets (before image setup to know final row count)
    targets = [
        (target_next_level, xp_to_next_level, eta_next_level, "next_level"),
        (target_next_prestige, xp_to_next_prestige, eta_next_prestige, "next_prestige")
    ]
    
    # Add custom level if provided
    if custom_level_data:
        custom_lvl, custom_xp, custom_eta = custom_level_data
        targets.append((custom_lvl, custom_xp, custom_eta, "custom"))
    
    # Remove duplicates based on target level (keep first occurrence)
    seen_levels = set()
    unique_targets = []
    for target in targets:
        if target[0] not in seen_levels:
            seen_levels.add(target[0])
            unique_targets.append(target)
    targets = unique_targets
    
    # Sort by level (ascending)
    targets.sort(key=lambda x: x[0])
    
    # Determine number of rows and adjust canvas height dynamically
    num_rows = len(targets)
    
    # Canvas dimensions - adjust height based on actual number of rows
    canvas_w = 950
    margin = 35
    header_section_height = 180  # Space for title, subtitle, separator, and headers
    row_height = 52
    footer_height = 50  # Space for footer
    canvas_h = header_section_height + (num_rows * row_height) + footer_height
    
    composite = Image.new('RGBA', (canvas_w, canvas_h), (18, 18, 20, 255))
    draw = ImageDraw.Draw(composite)
    
    # Render prestige and username - CENTERED
    ign_rgb = (85, 255, 255)
    if ign_color:
        try:
            ign_rgb = tuple(int(str(ign_color).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass
    
    # Get guild color
    g_rgb = (170, 170, 170)
    if guild_tag and guild_hex:
        if str(guild_hex).upper() in MINECRAFT_NAME_TO_HEX:
            guild_hex = MINECRAFT_NAME_TO_HEX[str(guild_hex).upper()]
        try:
            g_rgb = tuple(int(str(guild_hex).lstrip('#')[j:j+2], 16) for j in (0, 2, 4))
        except:
            pass
    
    # Get prestige segments and render
    segs = get_prestige_segments(level, icon)
    font_title = _load_font("DejaVuSans-Bold.ttf", 24)
    
    # Render prestige
    txt_io = _render_text_segments_to_image(segs, font=font_title, padding=(0, 0))
    prestige_img = Image.open(txt_io).convert("RGBA")
    
    # Calculate total width for centering
    temp_draw = ImageDraw.Draw(Image.new('RGBA', (1, 1)))
    username_text = f" {ign}"
    username_bbox = temp_draw.textbbox((0, 0), username_text, font=font_title)
    username_width = username_bbox[2] - username_bbox[0]
    
    guild_width = 0
    if guild_tag:
        safe_tag = _safe_guild_tag(guild_tag)
        guild_text = f" [{safe_tag}]"
        guild_bbox = temp_draw.textbbox((0, 0), guild_text, font=font_title)
        guild_width = guild_bbox[2] - guild_bbox[0]
    
    total_width = prestige_img.width + username_width + guild_width
    
    # Start position for centered title
    title_y = 25
    x_pos = (canvas_w - total_width) // 2
    
    # Paste prestige
    composite.paste(prestige_img, (x_pos, title_y), prestige_img)
    x_pos += prestige_img.width
    
    # Draw username
    draw.text((x_pos, title_y + prestige_img.height // 2), username_text, 
              font=font_title, fill=ign_rgb, anchor="lm")
    x_pos += username_width
    
    # Draw guild tag if present
    if guild_tag:
        draw.text((x_pos, title_y + prestige_img.height // 2), guild_text, 
                  font=font_title, fill=g_rgb, anchor="lm")
    
    # Subtitle with progress bar
    font_subtitle = _load_font("DejaVuSans.ttf", 16)
    progress_text = f"Progress: {current_xp_progress:,}/{xp_for_next_level:,} XP ({100 * current_xp_progress / xp_for_next_level:.1f}%) - {tab_name.title()}"
    subtitle_bbox = draw.textbbox((0, 0), progress_text, font=font_subtitle)
    subtitle_w = subtitle_bbox[2] - subtitle_bbox[0]
    draw.text(((canvas_w - subtitle_w) // 2, title_y + 38), progress_text, 
              font=font_subtitle, fill=(180, 180, 200))
    
    # Separator line
    line_y = title_y + 70
    draw.line([margin, line_y, canvas_w - margin, line_y], fill=(60, 60, 80), width=2)
    
    # Table headers
    table_y = line_y + 25
    col_widths = [190, 165, 270, 200]
    col_spacing = 20
    col_x = [
        margin,
        margin + col_widths[0] + col_spacing,
        margin + col_widths[0] + col_widths[1] + col_spacing * 2,
        margin + col_widths[0] + col_widths[1] + col_widths[2] + col_spacing * 3
    ]
    
    font_header = _load_font("DejaVuSans-Bold.ttf", 15)
    headers = ["Target Level", "Missing", "Current Rate", "Estimation"]
    
    for i, (x, header) in enumerate(zip(col_x, headers)):
        draw.text((x, table_y), header, font=font_header, fill=(255, 255, 255))
    
    # Header underline
    draw.line([margin, table_y + 26, canvas_w - margin, table_y + 26], fill=(80, 80, 100), width=1)
    
    # Table rows
    font_value = _load_font("DejaVuSans.ttf", 16)
    font_label = _load_font("DejaVuSans-Bold.ttf", 16)
    
    # Render rows (targets already built and sorted earlier)
    row_y = table_y + 42
    for idx, (target_lvl, xp_needed, eta, row_type) in enumerate(targets):
        target_color = get_prestige_color(target_lvl)
        target_icon = get_prestige_icon(target_lvl)
        
        label_text = f"Level {target_lvl} {target_icon}"
        
        draw.text((col_x[0], row_y), label_text, font=font_label, fill=target_color)
        draw.text((col_x[1], row_y), f"{xp_needed:,} XP", font=font_value, fill=(255, 255, 255))
        draw.text((col_x[2], row_y), f"{exp_per_hour:.2f} XP/Hour", font=font_value, fill=(255, 255, 255))
        draw.text((col_x[3], row_y), eta, font=font_value, fill=(255, 170, 0))
        
        # Row separator (except after last row)
        if idx < len(targets) - 1:
            draw.line([margin, row_y + 28, canvas_w - margin, row_y + 28], fill=(50, 50, 60), width=1)
        
        row_y += row_height
    
    # Footer
    footer_text = "Made with ❤ by chuckegg & felix"
    font_footer = _load_font("DejaVuSans.ttf", 13)
    bbox = draw.textbbox((0, 0), footer_text, font=font_footer)
    text_w = bbox[2] - bbox[0]
    draw.text(((canvas_w - text_w) // 2, canvas_h - 28), footer_text, 
              font=font_footer, fill=(60, 60, 65))
    
    out = io.BytesIO()
    composite.convert("RGB").save(out, format='PNG')
    out.seek(0)
    return out


class LevelProgressView(discord.ui.View):
    """View for displaying level progress predictions with period tabs."""
    
    def __init__(self, user_data: dict, ign: str, custom_level: int = None):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.ign = ign
        self.user_data = user_data
        self.meta = user_data.get("meta", {})
        self.current_tab = "all-time"
        self.custom_level = custom_level
        self.message = None  # Store message reference for timeout handling
        self._load_color()
        self.update_buttons()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
    
    def _load_color(self):
        """Load color and guild info for this username from database."""
        self.ign_color = None
        self.guild_tag = None
        self.guild_color = None
        try:
            meta = get_user_meta(self.ign)
            if meta:
                self.ign_color = meta.get('ign_color')
                if not self.ign_color:
                    self.ign_color = get_rank_color_hex(meta.get('rank'))
                self.guild_tag = meta.get('guild_tag')
                self.guild_color = meta.get('guild_hex', '#AAAAAA')
        except Exception as e:
            print(f"[WARNING] Failed to load metadata for {self.ign}: {e}")
    
    def update_buttons(self):
        """Update button styles based on current tab."""
        for item in self.children:
            if isinstance(item, discord.ui.Button):
                if item.custom_id == self.current_tab:
                    item.style = discord.ButtonStyle.primary
                else:
                    item.style = discord.ButtonStyle.secondary
    
    def generate_level_progress_image(self, tab_name: str) -> discord.File:
        """Generate level progress prediction image for a specific period."""
        stats = self.user_data.get("stats", {})
        
        # Map tab names to stats keys
        tab_key = {
            "all-time": "lifetime",
            "session": "session",
            "daily": "daily",
            "yesterday": "yesterday",
            "weekly": "weekly",
            "monthly": "monthly"
        }.get(tab_name, "lifetime")
        
        # Get level and experience from meta or stats
        level = self.meta.get("level", 0)
        current_total_exp = int(stats.get('experience', {}).get('lifetime', 0))
        
        # Calculate progress within current level
        xp_for_current_level = level_to_experience_required(level)
        current_xp_progress = current_total_exp - xp_for_current_level
        xp_for_next_level = get_xp_for_next_level(level)
        
        # XP needed to reach next level
        xp_to_next_level = xp_for_next_level - current_xp_progress
        
        # Calculate next prestige
        next_prestige = ((level // 100) + 1) * 100
        if next_prestige == level:
            next_prestige += 100
        
        # XP needed to reach next prestige
        xp_for_prestige = level_to_experience_required(next_prestige)
        xp_to_next_prestige = xp_for_prestige - current_total_exp
        
        # Get exp/hour for the selected period
        exp_per_hour = 0.0
        period_exp = stats.get('experience', {}).get(tab_key, 0)
        period_playtime = stats.get('playtime', {}).get(tab_key, 0)  # in seconds
        
        if period_playtime > 0:
            exp_per_hour = (period_exp / period_playtime) * 3600
        
        # Calculate ETAs
        eta_next_level = format_time_estimate(xp_to_next_level, exp_per_hour)
        eta_next_prestige = format_time_estimate(xp_to_next_prestige, exp_per_hour)
        
        # Get prestige info
        icon = self.meta.get("icon", get_prestige_icon(level))
        
        # Calculate custom level progress if provided
        custom_level_data = None
        if self.custom_level and self.custom_level > level:
            xp_for_custom = level_to_experience_required(self.custom_level)
            xp_to_custom = xp_for_custom - current_total_exp
            eta_custom = format_time_estimate(xp_to_custom, exp_per_hour)
            custom_level_data = (self.custom_level, xp_to_custom, eta_custom)
        
        # Create image
        img_io = create_level_progress_image(
            ign=self.ign,
            level=level,
            icon=icon,
            tab_name=tab_name,
            current_xp_progress=current_xp_progress,
            xp_for_next_level=xp_for_next_level,
            target_next_level=level + 1,
            xp_to_next_level=xp_to_next_level,
            exp_per_hour=exp_per_hour,
            eta_next_level=eta_next_level,
            target_next_prestige=next_prestige,
            xp_to_next_prestige=xp_to_next_prestige,
            eta_next_prestige=eta_next_prestige,
            ign_color=self.ign_color,
            guild_tag=self.guild_tag,
            guild_hex=self.guild_color,
            custom_level_data=custom_level_data
        )
        
        filename = f"level_progress_{self.ign}_{tab_name}.png"
        return discord.File(img_io, filename=filename)
    
    async def handle_tab_click(self, interaction: discord.Interaction, tab_name: str):
        """Handle tab button clicks."""
        self.current_tab = tab_name
        self.update_buttons()
        file = self.generate_level_progress_image(tab_name)
        await interaction.response.edit_message(attachments=[file], view=self)
    
    @discord.ui.button(label="All-time", custom_id="all-time", style=discord.ButtonStyle.primary)
    async def all_time_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "all-time")
    
    @discord.ui.button(label="Session", custom_id="session", style=discord.ButtonStyle.secondary)
    async def session_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "session")
    
    @discord.ui.button(label="Daily", custom_id="daily", style=discord.ButtonStyle.secondary)
    async def daily_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "daily")
    
    @discord.ui.button(label="Yesterday", custom_id="yesterday", style=discord.ButtonStyle.secondary)
    async def yesterday_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "yesterday")
    
    @discord.ui.button(label="Weekly", custom_id="weekly", style=discord.ButtonStyle.secondary)
    async def weekly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "weekly")
    
    @discord.ui.button(label="Monthly", custom_id="monthly", style=discord.ButtonStyle.secondary)
    async def monthly_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self.handle_tab_click(interaction, "monthly")


class LeaderboardView(discord.ui.View):
    def __init__(self, metric: str, data_cache: dict, category: str = "sheepwars"):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.metric = metric
        self.data_cache = data_cache
        self.category = category
        self.current_period = "lifetime"
        self.page = 0
        self.page_size = 10
        self.message = None  # Store message reference for timeout handling
        
        # Category display names
        self.category_names = {
            "general": "Wool Games",
            "sheepwars": "Sheep Wars",
            "ctw": "CTW",
            "ww": "Wool Wars"
        }
        self.category_display = self.category_names.get(category, category.title())
        
        # Column mappings for each period
        self.column_map = {
            "lifetime": "B",      # All-time values
            # Use the DELTA columns for period comparisons
            "session": "C",       # Session Delta
            "daily": "E",         # Daily Delta
            "yesterday": "G",     # Yesterday Delta
            "monthly": "I",       # Monthly Delta
        }
        
        # Merge all metric labels from CATEGORY_METRICS
        self.metric_labels = {}
        for category_metrics in CATEGORY_METRICS.values():
            self.metric_labels.update(category_metrics)
        
        # Add legacy metric labels for backwards compatibility
        legacy_labels = {
            "kills": "Kills",
            "kills_void": "Void Kills",
            "kills_explosive": "Explosive Kills",
            "kills_melee": "Melee Kills",
            "kills_bow": "Bow Kills",
            "deaths": "Deaths",
            "deaths_void": "Void Deaths",
            "deaths_explosive": "Explosive Deaths",
            "deaths_melee": "Melee Deaths",
            "deaths_bow": "Bow Deaths",
            "kdr": "K/D Ratio",
            "wins": "Wins",
            "losses": "Losses",
            "wlr": "W/L Ratio",
            "experience": "Experience",
            "level": "Level",
            "coins": "Coins",
            "damage_dealt": "Damage Dealt",
            "games_played": "Games Played",
            "sheep_thrown": "Sheep Thrown",
            "magic_wool_hit": "Magic Wool Hit",
            "playtime": "Playtime",
        }
        self.metric_labels.update(legacy_labels)
        
        # Period selector dropdown
        self.period_select = LeaderboardPeriodSelect(self)
        self.add_item(self.period_select)
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
        
    def _get_leaderboard(self, period: str):
        return self.metric_labels[self.metric], self.data_cache.get(period, [])

    def _paginate(self, leaderboard: list, page: int):
        total_pages = max(1, (len(leaderboard) + self.page_size - 1) // self.page_size)
        clamped_page = max(0, min(page, total_pages - 1))
        start_index = clamped_page * self.page_size
        return leaderboard[start_index:start_index + self.page_size], total_pages, clamped_page, start_index

    def generate_leaderboard_image(self, period: str, page: int):
        metric_label, leaderboard = self._get_leaderboard(period)

        if not leaderboard:
            empty_embed = self.get_leaderboard_embed(period, page=0, total_pages=1, leaderboard=leaderboard)
            return empty_embed, None, 1

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard, page)
        self.page = clamped_page

        image_data = []
        for idx, entry in enumerate(sliced):
            player, _, value, is_playtime, level, icon, p_hex, g_tag, g_hex = entry
            rank = start_index + idx + 1
            image_data.append((rank, player, level, icon, p_hex, g_tag, g_hex, value, is_playtime))

        if Image is not None:
            try:
                title_with_category = f"{period.title()} {self.category_display}"
                img_io = create_leaderboard_image(title_with_category, metric_label, image_data, page=clamped_page, total_pages=total_pages)
                filename = f"leaderboard_{self.metric}_{period}_p{clamped_page + 1}.png"
                return None, discord.File(img_io, filename=filename), total_pages
            except Exception as e:
                print(f"[WARNING] Leaderboard image generation failed: {e}")
                return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages
        else:
            return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages

    def get_leaderboard_embed(self, period: str, page: int = 0, total_pages: int = 1, leaderboard: Optional[list] = None):
        metric_label, leaderboard_data = self._get_leaderboard(period) if leaderboard is None else (self.metric_labels[self.metric], leaderboard)

        if not leaderboard_data:
            embed = discord.Embed(
                title=f"{period.title()} {self.category_display} {metric_label} Leaderboard",
                description="No data available",
                color=discord.Color.from_rgb(54, 57, 63)
            )
            return embed

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard_data, page)
        self.page = clamped_page

        embed = discord.Embed(
            title=f"{period.title()} {self.category_display} {metric_label} Leaderboard",
            color=discord.Color.from_rgb(54, 57, 63)
        )

        description_lines = []
        for idx, entry in enumerate(sliced):
            player = entry[0]
            value = entry[2]
            is_playtime = entry[3]
            level_value = entry[4]
            icon = entry[5]

            medal = {1: "1.", 2: "2.", 3: "3."}.get(start_index + idx + 1, f"{start_index + idx + 1}.")
            prestige_display = format_prestige_ansi(level_value, icon)

            if is_playtime:
                formatted_value = format_playtime(int(value))
            elif metric_label == "Level":
                formatted_value = f"{float(value):,.2f}"
            elif "Ratio" in metric_label or "/" in metric_label or "Per" in metric_label:
                formatted_value = f"{float(value):,.2f}"
            else:
                formatted_value = f"{int(value):,}"

            description_lines.append(f"{medal} {prestige_display} {player}: {formatted_value}")

        embed.description = f"```ansi\n" + "\n".join(description_lines) + "\n```"
        embed.set_footer(text=f"Page {clamped_page + 1} of {total_pages}")
        return embed

    async def _refresh(self, interaction: discord.Interaction, *, new_period: Optional[str] = None, page_delta: int = 0):
        if new_period is not None:
            self.current_period = new_period
            self.page = 0
            # sync dropdown defaults
            for option in self.period_select.options:
                option.default = option.value == new_period
        else:
            self.page += page_delta

        embed, file, _ = self.generate_leaderboard_image(self.current_period, self.page)

        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Prev Page", custom_id="page_prev", style=discord.ButtonStyle.secondary)
    async def prev_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=-1)

    @discord.ui.button(label="Next Page", custom_id="page_next", style=discord.ButtonStyle.secondary)
    async def next_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=1)

    @discord.ui.button(label="🔍 Search", custom_id="page_search", style=discord.ButtonStyle.primary)
    async def search(self, interaction: discord.Interaction, button: discord.ui.Button):
        modal = LeaderboardSearchModal(self)
        await interaction.response.send_modal(modal)


class LeaderboardSearchModal(discord.ui.Modal, title="Search Leaderboard"):
    def __init__(self, view: LeaderboardView):
        super().__init__()
        self.view_ref = view
    
    search_input = discord.ui.TextInput(
        label="Search by Rank or Player Name",
        placeholder="Enter rank number (e.g. 500) or player name (e.g. Chuckegg)",
        required=True,
        max_length=100
    )
    
    async def on_submit(self, interaction: discord.Interaction):
        search_term = self.search_input.value.strip()
        
        # Get current leaderboard data
        _, leaderboard = self.view_ref._get_leaderboard(self.view_ref.current_period)
        
        if not leaderboard:
            await interaction.response.send_message("❌ No leaderboard data available", ephemeral=True)
            return
        
        # Try to parse as rank number
        target_page = None
        try:
            rank = int(search_term)
            if 1 <= rank <= len(leaderboard):
                # Calculate which page this rank is on (0-indexed)
                target_page = (rank - 1) // self.view_ref.page_size
            else:
                await interaction.response.send_message(
                    f"❌ Rank {rank} is out of range (1-{len(leaderboard)})",
                    ephemeral=True
                )
                return
        except ValueError:
            # Not a number, search for player name (case-insensitive)
            search_lower = search_term.lower()
            for idx, entry in enumerate(leaderboard):
                player_name = entry[0]
                if player_name.lower() == search_lower or search_lower in player_name.lower():
                    target_page = idx // self.view_ref.page_size
                    break
            
            if target_page is None:
                await interaction.response.send_message(
                    f"❌ Player '{search_term}' not found in this leaderboard",
                    ephemeral=True
                )
                return
        
        # Jump to the target page
        self.view_ref.page = target_page
        embed, file, _ = self.view_ref.generate_leaderboard_image(self.view_ref.current_period, target_page)
        
        if file:
            await interaction.response.edit_message(view=self.view_ref, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self.view_ref)


class LeaderboardPeriodSelect(discord.ui.Select):
    def __init__(self, view: LeaderboardView):
        options = [
            discord.SelectOption(label="Lifetime", value="lifetime", default=True),
            discord.SelectOption(label="Session", value="session"),
            discord.SelectOption(label="Daily", value="daily"),
            discord.SelectOption(label="Yesterday", value="yesterday"),
            discord.SelectOption(label="Weekly", value="weekly"),
            discord.SelectOption(label="Monthly", value="monthly"),
        ]
        super().__init__(
            placeholder="Select leaderboard period",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="leaderboard_period_select",
        )
        self.view_ref = view

    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref._refresh(interaction, new_period=selected)


def _load_leaderboard_data_from_excel(metric: str, category: str = "sheepwars"):
    """Load leaderboard data from database.
    
    Args:
        metric: The stat to rank by
        category: Which stat category (general, sheepwars, ctw, ww)
    
    Returns dict with period -> list of (username, value, display_value, is_playtime, level, icon, color, guild_tag, guild_color)
    """
    periods = ["lifetime", "session", "daily", "yesterday", "weekly", "monthly"]
    result = {p: [] for p in periods}
    
    try:
        # Get all users from database
        usernames = get_all_usernames()
        
        # Load user colors
        user_colors = load_user_colors()
        if not user_colors:
            print("[WARN] No user colors loaded in leaderboard generation")
        # else:
            # print(f"[DEBUG] Loaded {len(user_colors)} user colors for leaderboard")
        
        for username in usernames:
            try:
                # Check if user is tracked (for filtering non-lifetime periods)
                user_is_tracked = is_tracked_user(username)
                
                # Get stats with deltas
                stats_dict = get_user_stats_with_deltas(username)
                
                if not stats_dict:
                    continue
                
                # Get metadata
                user_meta = user_colors.get(username.lower(), {})
                
                # Try to get level from stats_dict first
                level = stats_dict.get("level", {}).get("lifetime", 0)
                if not level:
                    level = stats_dict.get("prestige level", {}).get("lifetime", 0)
                if not level:
                    level = 0
                else:
                    level = int(level) if level else 0
                
                icon = user_meta.get("icon") or ""
                ign_color = user_meta.get("color")
                if not ign_color:
                    ign_color = get_rank_color_hex(user_meta.get("rank"))
                
                guild_tag = user_meta.get("guild_tag")
                # Fix: Handle None correctly for guild_color
                raw_g = str(user_meta.get('guild_color') or 'GRAY').upper()
                guild_color = raw_g if raw_g.startswith('#') else MINECRAFT_NAME_TO_HEX.get(raw_g, "#AAAAAA")
                
                # Process each period
                for period in periods:
                    # Skip untracked users for non-lifetime periods
                    # (they don't get periodic updates so their deltas are always 0)
                    if period != "lifetime" and not user_is_tracked:
                        continue
                    
                    val = 0
                    
                    # Special handling for level metric - calculate from experience delta
                    if metric == "level":
                        if period == "lifetime":
                            # For lifetime, use the absolute level value
                            val = stats_dict.get("level", {}).get(period, 0)
                        else:
                            # For deltas, calculate level gain from experience gain
                            exp_delta = stats_dict.get("experience", {}).get(period, 0)
                            val = experience_delta_to_level_delta(exp_delta)
                    # Handle combined general stats (Sheep Wars + Wool Wars + CTW)
                    elif metric == "total_kills":
                        sw_kills = stats_dict.get("kills", {}).get(period, 0)
                        ctw_kills = stats_dict.get("ctw_kills", {}).get(period, 0)
                        ww_kills = stats_dict.get("ww_kills", {}).get(period, 0)
                        val = sw_kills + ctw_kills + ww_kills
                    elif metric == "total_deaths":
                        sw_deaths = stats_dict.get("deaths", {}).get(period, 0)
                        ctw_deaths = stats_dict.get("ctw_deaths", {}).get(period, 0)
                        ww_deaths = stats_dict.get("ww_deaths", {}).get(period, 0)
                        val = sw_deaths + ctw_deaths + ww_deaths
                    elif metric == "total_wins":
                        sw_wins = stats_dict.get("wins", {}).get(period, 0)
                        ctw_wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
                        ww_wins = stats_dict.get("ww_wins", {}).get(period, 0)
                        val = sw_wins + ctw_wins + ww_wins
                    elif metric == "total_losses":
                        sw_losses = stats_dict.get("losses", {}).get(period, 0)
                        ctw_losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
                        ww_games = stats_dict.get("ww_games_played", {}).get(period, 0)
                        ww_wins = stats_dict.get("ww_wins", {}).get(period, 0)
                        ww_losses = ww_games - ww_wins
                        val = sw_losses + ctw_losses + ww_losses
                    elif metric == "total_games_played":
                        sw_games = stats_dict.get("games_played", {}).get(period, 0)
                        ctw_wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
                        ctw_losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
                        ctw_games = ctw_wins + ctw_losses
                        ww_games = stats_dict.get("ww_games_played", {}).get(period, 0)
                        val = sw_games + ctw_games + ww_games
                    # Map metric names to database columns
                    # Try to get the metric directly first (works for all stats)
                    elif metric in stats_dict:
                        val = stats_dict.get(metric, {}).get(period, 0)
                        # Make gold_spent positive
                        if metric == "ctw_gold_spent":
                            val = abs(val)
                    # Handle special ratio calculations
                    elif metric == "kdr":
                        k = stats_dict.get("kills", {}).get(period, 0)
                        d = stats_dict.get("deaths", {}).get(period, 0)
                        val = k / d if d > 0 else k
                    elif metric == "wlr":
                        w = stats_dict.get("wins", {}).get(period, 0)
                        l = stats_dict.get("losses", {}).get(period, 0)
                        val = w / l if l > 0 else w
                    else:
                        # Default fallback - return 0 if stat doesn't exist
                        val = 0
                    
                    is_playtime = (metric == "playtime")
                    result[period].append((
                        username, float(val), val, is_playtime, level, icon, ign_color, guild_tag, guild_color
                    ))
            except Exception as e:
                print(f"[LEADERBOARD] Error processing {username}: {e}")
                continue
            
        # Sort each period by value descending
        for p in result:
            result[p].sort(key=lambda x: x[1], reverse=True)
        
        return result
    except Exception as e:
        print(f"[LEADERBOARD] Error loading from Excel: {e}")
        return result

def _process_leaderboard_data(cache_data, metric):
    periods = ["lifetime", "session", "daily", "yesterday", "weekly", "monthly"]
    result = {p: [] for p in periods}

    for username, data in cache_data.items():
        stats = data.get("stats", {})
        meta = data.get("meta", {})
        
        for period in periods:
            val = 0
            if metric == "kdr":
                k = stats.get("kills", {}).get(period, 0)
                d = stats.get("deaths", {}).get(period, 0)
                val = k / d if d > 0 else k
            elif metric == "wlr":
                w = stats.get("wins", {}).get(period, 0)
                l = stats.get("losses", {}).get(period, 0)
                val = w / l if l > 0 else w
            else:
                val = stats.get(metric, {}).get(period, 0)
            
            is_playtime = (metric == "playtime")
            result[period].append((
                meta.get("username", username), float(val), val, is_playtime,
                meta.get("level", 0), meta.get("icon", ""), meta.get("ign_color"), meta.get("guild_tag"), meta.get("guild_hex")
            ))

    for p in result:
        result[p].sort(key=lambda x: x[1], reverse=True)
    return result

def _calculate_ratio_value_from_excel(stats_dict: dict, period: str, metric: str):
    """Calculate ratio values from Excel data."""
    try:
        # General combined stats (Sheep Wars + Wool Wars + CTW)
        if metric == "total_kdr":
            # Sum kills from all game modes
            sw_kills = stats_dict.get("kills", {}).get(period, 0)
            ctw_kills = stats_dict.get("ctw_kills", {}).get(period, 0)
            ww_kills = stats_dict.get("ww_kills", {}).get(period, 0)
            total_kills = sw_kills + ctw_kills + ww_kills
            
            # Sum deaths from all game modes
            sw_deaths = stats_dict.get("deaths", {}).get(period, 0)
            ctw_deaths = stats_dict.get("ctw_deaths", {}).get(period, 0)
            ww_deaths = stats_dict.get("ww_deaths", {}).get(period, 0)
            total_deaths = sw_deaths + ctw_deaths + ww_deaths
            
            return round(total_kills / total_deaths, 2) if total_deaths > 0 else total_kills
        elif metric == "total_wlr":
            # Sum wins from all game modes
            sw_wins = stats_dict.get("wins", {}).get(period, 0)
            ctw_wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            ww_wins = stats_dict.get("ww_wins", {}).get(period, 0)
            total_wins = sw_wins + ctw_wins + ww_wins
            
            # Sum losses from all game modes
            sw_losses = stats_dict.get("losses", {}).get(period, 0)
            ctw_losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            ww_games = stats_dict.get("ww_games_played", {}).get(period, 0)
            ww_losses = ww_games - ww_wins  # Wool Wars doesn't store losses directly
            total_losses = sw_losses + ctw_losses + ww_losses
            
            return round(total_wins / total_losses, 2) if total_losses > 0 else total_wins
        
        # Sheep Wars ratios
        if metric == "wl_ratio" or metric == "wlr":
            wins = stats_dict.get("wins", {}).get(period, 0)
            losses = stats_dict.get("losses", {}).get(period, 0)
            return round(wins / losses, 2) if losses > 0 else wins
        elif metric == "kd_ratio" or metric == "kdr":
            kills = stats_dict.get("kills", {}).get(period, 0)
            deaths = stats_dict.get("deaths", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "kills_per_game":
            kills = stats_dict.get("kills", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(kills / games, 2) if games > 0 else 0
        elif metric == "kills_per_win":
            kills = stats_dict.get("kills", {}).get(period, 0)
            wins = stats_dict.get("wins", {}).get(period, 0)
            return round(kills / wins, 2) if wins > 0 else 0
        elif metric == "damage_per_game":
            damage = stats_dict.get("damage_dealt", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(damage / games, 2) if games > 0 else 0
        elif metric == "damage_per_sheep":
            damage = stats_dict.get("damage_dealt", {}).get(period, 0)
            sheep = stats_dict.get("sheep_thrown", {}).get(period, 0)
            return round(damage / sheep, 2) if sheep > 0 else 0
        elif metric == "wools_per_game":
            wools = stats_dict.get("magic_wool_hit", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(wools / games, 2) if games > 0 else 0
        elif metric == "void_kd_ratio":
            void_kills = stats_dict.get("kills_void", {}).get(period, 0)
            void_deaths = stats_dict.get("deaths_void", {}).get(period, 0)
            return round(void_kills / void_deaths, 2) if void_deaths > 0 else void_kills
        elif metric == "explosive_kd_ratio":
            exp_kills = stats_dict.get("kills_explosive", {}).get(period, 0)
            exp_deaths = stats_dict.get("deaths_explosive", {}).get(period, 0)
            return round(exp_kills / exp_deaths, 2) if exp_deaths > 0 else exp_kills
        elif metric == "bow_kd_ratio":
            bow_kills = stats_dict.get("kills_bow", {}).get(period, 0)
            bow_deaths = stats_dict.get("deaths_bow", {}).get(period, 0)
            return round(bow_kills / bow_deaths, 2) if bow_deaths > 0 else bow_kills
        elif metric == "melee_kd_ratio":
            melee_kills = stats_dict.get("kills_melee", {}).get(period, 0)
            melee_deaths = stats_dict.get("deaths_melee", {}).get(period, 0)
            return round(melee_kills / melee_deaths, 2) if melee_deaths > 0 else melee_kills
        elif metric == "exp_per_hour":
            exp = stats_dict.get("experience", {}).get(period, 0)
            playtime = stats_dict.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(exp / hours, 2) if hours > 0 else 0
        elif metric == "exp_per_game":
            exp = stats_dict.get("experience", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(exp / games, 2) if games > 0 else 0
        elif metric == "wins_per_hour":
            wins = stats_dict.get("wins", {}).get(period, 0)
            playtime = stats_dict.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(wins / hours, 2) if hours > 0 else 0
        elif metric == "kills_per_hour":
            kills = stats_dict.get("kills", {}).get(period, 0)
            playtime = stats_dict.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(kills / hours, 2) if hours > 0 else 0
        elif metric == "sheeps_per_game":
            sheep = stats_dict.get("sheep_thrown", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            return round(sheep / games, 2) if games > 0 else 0
        elif metric == "survival_rate":
            games = stats_dict.get("games_played", {}).get(period, 0)
            deaths = stats_dict.get("deaths", {}).get(period, 0)
            return round((games - deaths) / games, 2) if games > 0 else 0
        elif metric == "carried_score":
            wins = stats_dict.get("wins", {}).get(period, 0)
            losses = stats_dict.get("losses", {}).get(period, 0)
            kills = stats_dict.get("kills", {}).get(period, 0)
            deaths = stats_dict.get("deaths", {}).get(period, 0)
            games = stats_dict.get("games_played", {}).get(period, 0)
            if games > 0 and deaths > 0:
                return calculate_carried_score_average(wins, losses, kills, deaths, games)
            return 0
        
        # CTW Ratios
        elif metric == "ctw_wl_ratio":
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            return round(wins / losses, 2) if losses > 0 else wins
        elif metric == "ctw_kd_ratio":
            kills = stats_dict.get("ctw_kills", {}).get(period, 0)
            deaths = stats_dict.get("ctw_deaths", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "ctw_kills_per_game":
            kills = stats_dict.get("ctw_kills", {}).get(period, 0)
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            games = wins + losses
            return round(kills / games, 2) if games > 0 else 0
        elif metric == "ctw_deaths_per_game":
            deaths = stats_dict.get("ctw_deaths", {}).get(period, 0)
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            games = wins + losses
            return round(deaths / games, 2) if games > 0 else 0
        elif metric == "ctw_kd_on_woolholder":
            kills = stats_dict.get("ctw_kills_on_woolholder", {}).get(period, 0)
            deaths = stats_dict.get("ctw_deaths_to_woolholder", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "ctw_kd_as_woolholder":
            kills = stats_dict.get("ctw_kills_with_wool", {}).get(period, 0)
            deaths = stats_dict.get("ctw_deaths_with_wool", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "ctw_woolholder_kills_per_game":
            kills = stats_dict.get("ctw_kills_on_woolholder", {}).get(period, 0)
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            games = wins + losses
            return round(kills / games, 2) if games > 0 else 0
        elif metric == "ctw_woolholder_kills_per_kill":
            wh_kills = stats_dict.get("ctw_kills_on_woolholder", {}).get(period, 0)
            kills = stats_dict.get("ctw_kills", {}).get(period, 0)
            return round(wh_kills / kills, 2) if kills > 0 else 0
        elif metric == "ctw_wools_captured_per_game":
            wools = stats_dict.get("ctw_wools_captured", {}).get(period, 0)
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            games = wins + losses
            return round(wools / games, 2) if games > 0 else 0
        elif metric == "ctw_wools_captured_per_death":
            wools = stats_dict.get("ctw_wools_captured", {}).get(period, 0)
            deaths = stats_dict.get("ctw_deaths", {}).get(period, 0)
            return round(wools / deaths, 2) if deaths > 0 else 0
        elif metric == "ctw_gold_earned_per_game":
            gold = stats_dict.get("ctw_gold_earned", {}).get(period, 0)
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            games = wins + losses
            return round(gold / games, 2) if games > 0 else 0
        elif metric == "ctw_gold_spent_per_game":
            gold = abs(stats_dict.get("ctw_gold_spent", {}).get(period, 0))
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            games = wins + losses
            return round(gold / games, 2) if games > 0 else 0
        elif metric == "ctw_wools_stolen_per_game":
            wools = stats_dict.get("ctw_wools_stolen", {}).get(period, 0)
            wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            games = wins + losses
            return round(wools / games, 2) if games > 0 else 0
        
        # WW Ratios
        elif metric == "ww_wl_ratio":
            wins = stats_dict.get("ww_wins", {}).get(period, 0)
            games = stats_dict.get("ww_games_played", {}).get(period, 0)
            losses = games - wins
            return round(wins / losses, 2) if losses > 0 else wins
        elif metric == "ww_kd_ratio":
            kills = stats_dict.get("ww_kills", {}).get(period, 0)
            deaths = stats_dict.get("ww_deaths", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "ww_kills_per_game":
            kills = stats_dict.get("ww_kills", {}).get(period, 0)
            games = stats_dict.get("ww_games_played", {}).get(period, 0)
            return round(kills / games, 2) if games > 0 else 0
        elif metric == "ww_assists_per_game":
            assists = stats_dict.get("ww_assists", {}).get(period, 0)
            games = stats_dict.get("ww_games_played", {}).get(period, 0)
            return round(assists / games, 2) if games > 0 else 0
        elif metric == "ww_kill_assist_ratio":
            kills = stats_dict.get("ww_kills", {}).get(period, 0)
            assists = stats_dict.get("ww_assists", {}).get(period, 0)
            return round(kills / assists, 2) if assists > 0 else kills
        elif metric == "ww_assists_per_death":
            assists = stats_dict.get("ww_assists", {}).get(period, 0)
            deaths = stats_dict.get("ww_deaths", {}).get(period, 0)
            return round(assists / deaths, 2) if deaths > 0 else 0
        
        # WW Class-specific ratios (tank, assault, golem, swordsman, archer, engineer)
        elif metric.startswith("ww_") and ("_kd_ratio" in metric or "_assists_per_death" in metric or "_kill_assist_ratio" in metric):
            # Extract class name
            for class_name in ["tank", "assault", "golem", "swordsman", "archer", "engineer"]:
                if f"ww_{class_name}_" in metric:
                    if metric.endswith("_kd_ratio"):
                        kills = stats_dict.get(f"ww_{class_name}_kills", {}).get(period, 0)
                        deaths = stats_dict.get(f"ww_{class_name}_deaths", {}).get(period, 0)
                        return round(kills / deaths, 2) if deaths > 0 else kills
                    elif metric.endswith("_assists_per_death"):
                        assists = stats_dict.get(f"ww_{class_name}_assists", {}).get(period, 0)
                        deaths = stats_dict.get(f"ww_{class_name}_deaths", {}).get(period, 0)
                        return round(assists / deaths, 2) if deaths > 0 else 0
                    elif metric.endswith("_kill_assist_ratio"):
                        kills = stats_dict.get(f"ww_{class_name}_kills", {}).get(period, 0)
                        assists = stats_dict.get(f"ww_{class_name}_assists", {}).get(period, 0)
                        return round(kills / assists, 2) if assists > 0 else kills
                    break
    except Exception as e:
        print(f"[ERROR] Exception in _calculate_ratio_value_from_excel for metric {metric}: {e}")
        return None
    return None

def _load_ratio_leaderboard_data_from_excel(metric: str, category: str = "sheepwars"):
    """Load ratio leaderboard data from database.
    
    Args:
        metric: The ratio stat to rank by
        category: Which stat category (general, sheepwars, ctw, ww)
    """
    periods = ["lifetime", "session", "daily", "yesterday", "weekly", "monthly"]
    result = {p: [] for p in periods}
    
    try:
        # Get all users from database
        usernames = get_all_usernames()
        
        # Load user colors
        user_colors = load_user_colors()
        if not user_colors:
            print("[WARN] No user colors loaded in ratio leaderboard generation")
        # else:
            # print(f"[DEBUG] Loaded {len(user_colors)} user colors for ratio leaderboard")
        
        for username in usernames:
            try:
                # Check if user is tracked (for filtering non-lifetime periods)
                user_is_tracked = is_tracked_user(username)
                
                # Get stats with deltas
                stats_dict = get_user_stats_with_deltas(username)
                
                if not stats_dict:
                    continue
                
                # Get metadata
                user_meta = user_colors.get(username.lower(), {})
                
                # Try to get level from stats_dict first
                level = stats_dict.get("level", {}).get("lifetime", 0)
                if not level:
                    level = stats_dict.get("prestige level", {}).get("lifetime", 0)
                if not level:
                    level = 0
                else:
                    level = int(level) if level else 0
                
                icon = user_meta.get("icon") or ""
                ign_color = user_meta.get("color")
                if not ign_color:
                    ign_color = get_rank_color_hex(user_meta.get("rank"))
                
                guild_tag = user_meta.get("guild_tag")
                # Fix: Handle None correctly for guild_color
                raw_g = str(user_meta.get('guild_color') or 'GRAY').upper()
                guild_color = raw_g if raw_g.startswith('#') else MINECRAFT_NAME_TO_HEX.get(raw_g, "#AAAAAA")
                
                # Process each period
                for period in periods:
                    # Skip untracked users for non-lifetime periods
                    # (they don't get periodic updates so their deltas are always 0)
                    if period != "lifetime" and not user_is_tracked:
                        continue
                    
                    val = _calculate_ratio_value_from_excel(stats_dict, period, metric)
                    # Filter out invalid values
                    # - survival_rate: 0 means no games played
                    # - carried_score: 0 means insufficient data (not a valid score)
                    should_add = val is not None and val != 0 if metric in ["survival_rate", "carried_score"] else val is not None
                    if should_add:
                        result[period].append((
                            username, float(val), val, level, icon, ign_color, guild_tag, guild_color
                        ))
            except Exception as e:
                print(f"[LEADERBOARD] Error processing {username}: {e}")
                continue
            
        # Sort each period by value descending (except carried_score which is ascending)
        for p in result:
            if metric == "carried_score":
                result[p].sort(key=lambda x: x[1], reverse=False)  # Ascending: lower is better
            else:
                result[p].sort(key=lambda x: x[1], reverse=True)  # Descending: higher is better
        
        return result
    except Exception as e:
        print(f"[LEADERBOARD] Error loading ratio data from Excel: {e}")
        return result

def _calculate_ratio_value_from_cache(stats, period, metric):
    try:
        if metric == "wl_ratio" or metric == "wlr":
            wins = stats.get("wins", {}).get(period, 0)
            losses = stats.get("losses", {}).get(period, 0)
            return round(wins / losses, 2) if losses > 0 else wins
        elif metric == "kd_ratio" or metric == "kdr":
            kills = stats.get("kills", {}).get(period, 0)
            deaths = stats.get("deaths", {}).get(period, 0)
            return round(kills / deaths, 2) if deaths > 0 else kills
        elif metric == "kills_per_game":
            kills = stats.get("kills", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(kills / games, 2) if games > 0 else 0
        elif metric == "kills_per_win":
            kills = stats.get("kills", {}).get(period, 0)
            wins = stats.get("wins", {}).get(period, 0)
            return round(kills / wins, 2) if wins > 0 else 0
        elif metric == "damage_per_game":
            damage = stats.get("damage_dealt", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(damage / games, 2) if games > 0 else 0
        elif metric == "damage_per_sheep":
            damage = stats.get("damage_dealt", {}).get(period, 0)
            sheep = stats.get("sheep_thrown", {}).get(period, 0)
            return round(damage / sheep, 2) if sheep > 0 else 0
        elif metric == "wools_per_game":
            wools = stats.get("magic_wool_hit", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(wools / games, 2) if games > 0 else 0
        elif metric == "void_kd_ratio":
            void_kills = stats.get("kills_void", {}).get(period, 0)
            void_deaths = stats.get("deaths_void", {}).get(period, 0)
            return round(void_kills / void_deaths, 2) if void_deaths > 0 else void_kills
        elif metric == "explosive_kd_ratio":
            exp_kills = stats.get("kills_explosive", {}).get(period, 0)
            exp_deaths = stats.get("deaths_explosive", {}).get(period, 0)
            return round(exp_kills / exp_deaths, 2) if exp_deaths > 0 else exp_kills
        elif metric == "bow_kd_ratio":
            bow_kills = stats.get("kills_bow", {}).get(period, 0)
            bow_deaths = stats.get("deaths_bow", {}).get(period, 0)
            return round(bow_kills / bow_deaths, 2) if bow_deaths > 0 else bow_kills
        elif metric == "melee_kd_ratio":
            melee_kills = stats.get("kills_melee", {}).get(period, 0)
            melee_deaths = stats.get("deaths_melee", {}).get(period, 0)
            return round(melee_kills / melee_deaths, 2) if melee_deaths > 0 else melee_kills
        elif metric == "exp_per_hour":
            exp = stats.get("experience", {}).get(period, 0)
            playtime = stats.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(exp / hours, 2) if hours > 0 else 0
        elif metric == "exp_per_game":
            exp = stats.get("experience", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(exp / games, 2) if games > 0 else 0
        elif metric == "wins_per_hour":
            wins = stats.get("wins", {}).get(period, 0)
            playtime = stats.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(wins / hours, 2) if hours > 0 else 0
        elif metric == "kills_per_hour":
            kills = stats.get("kills", {}).get(period, 0)
            playtime = stats.get("playtime", {}).get(period, 0)
            hours = playtime / 3600
            return round(kills / hours, 2) if hours > 0 else 0
        elif metric == "sheeps_per_game":
            sheep = stats.get("sheep_thrown", {}).get(period, 0)
            games = stats.get("games_played", {}).get(period, 0)
            return round(sheep / games, 2) if games > 0 else 0
        elif metric == "survival_rate":
            games = stats.get("games_played", {}).get(period, 0)
            deaths = stats.get("deaths", {}).get(period, 0)
            return round((games - deaths) / games, 2) if games > 0 else 0
    except:
        return None
    return None

def _process_ratio_data(cache_data, metric):
    periods = ["lifetime", "session", "daily", "yesterday", "weekly", "monthly"]
    result = {p: [] for p in periods}

    for username, data in cache_data.items():
        stats = data.get("stats", {})
        meta = data.get("meta", {})
        
        for period in periods:
            val = _calculate_ratio_value_from_cache(stats, period, metric)
            if val is not None:
                result[period].append((
                    meta.get("username", username), float(val), val, 
                    meta.get("level", 0), meta.get("icon", ""), meta.get("ign_color"), meta.get("guild_tag"), meta.get("guild_hex")
                ))

    for p in result:
        result[p].sort(key=lambda x: x[1], reverse=True)
    return result

class RatioLeaderboardView(discord.ui.View):
    def __init__(self, metric: str, data_cache: dict, category: str = "sheepwars"):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.metric = metric
        self.data_cache = data_cache
        self.category = category
        self.current_period = "lifetime"
        self.message = None  # Store message reference for timeout handling
        self.page = 0
        self.page_size = 10
        
        # Category display names
        self.category_names = {
            "general": "Wool Games",
            "sheepwars": "Sheep Wars",
            "ctw": "CTW",
            "ww": "Wool Wars"
        }
        self.category_display = self.category_names.get(category, category.title())
        
        # Column mappings for each period
        self.column_map = {
            "lifetime": "B",      # All-time values
            # Use the DELTA columns for period comparisons
            "session": "C",       # Session Delta
            "daily": "E",         # Daily Delta
            "yesterday": "G",     # Yesterday Delta
            "weekly": "H",        # Weekly Delta
            "monthly": "I",       # Monthly Delta
        }
        
        # Merge all metric labels from CATEGORY_METRICS
        self.metric_labels = {}
        for category_metrics in CATEGORY_METRICS.values():
            self.metric_labels.update(category_metrics)
        
        # Add legacy ratio metric labels for backwards compatibility
        legacy_ratio_labels = {
            "wl_ratio": "W/L Ratio",
            "kd_ratio": "K/D Ratio",
            "kills_per_game": "Kills/Game",
            "kills_per_win": "Kills/Win",
            "kills_per_hour": "Kills/Hour",
            "damage_per_game": "Damage/Game",
            "damage_per_sheep": "Damage/Sheep",
            "wools_per_game": "Wools/Game",
            "sheeps_per_game": "Sheeps/Game",
            "void_kd_ratio": "Void K/D Ratio",
            "explosive_kd_ratio": "Explosive K/D Ratio",
            "bow_kd_ratio": "Bow K/D Ratio",
            "melee_kd_ratio": "Melee K/D Ratio",
            "wins_per_hour": "Wins/Hour",
            "exp_per_hour": "EXP/Hour",
            "exp_per_game": "EXP/Game",
            "survival_rate": "Survival Rate",
            "carried_score": "Carried Score",
        }
        self.metric_labels.update(legacy_ratio_labels)
        
        # Period selector dropdown
        self.period_select = RatioPeriodSelect(self)
        self.add_item(self.period_select)
        
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
        
    def _get_leaderboard(self, period: str):
        return self.metric_labels[self.metric], self.data_cache.get(period, [])

    def _paginate(self, leaderboard: list, page: int):
        total_pages = max(1, (len(leaderboard) + self.page_size - 1) // self.page_size)
        clamped_page = max(0, min(page, total_pages - 1))
        start_index = clamped_page * self.page_size
        return leaderboard[start_index:start_index + self.page_size], total_pages, clamped_page, start_index

    def generate_leaderboard_image(self, period: str, page: int):
        metric_label, leaderboard = self._get_leaderboard(period)

        if not leaderboard:
            empty_embed = self.get_leaderboard_embed(period, page=0, total_pages=1, leaderboard=leaderboard)
            return empty_embed, None, 1

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard, page)
        self.page = clamped_page

        image_data = []
        for idx, entry in enumerate(sliced):
            player, _, value, level_value, icon, ign_color, g_tag, g_hex = entry
            rank = start_index + idx + 1
            image_data.append((rank, player, level_value, icon, ign_color, g_tag, g_hex, value, False))

        if Image is not None:
            try:
                title_with_category = f"{period.title()} {self.category_display}"
                img_io = create_leaderboard_image(title_with_category, metric_label, image_data, page=clamped_page, total_pages=total_pages)
                filename = f"ratio_leaderboard_{self.metric}_{period}_p{clamped_page + 1}.png"
                return None, discord.File(img_io, filename=filename), total_pages
            except Exception as e:
                print(f"[WARNING] Ratio leaderboard image generation failed: {e}")
                return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages
        else:
            return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages

    def get_leaderboard_embed(self, period: str, page: int = 0, total_pages: int = 1, leaderboard: Optional[list] = None):
        metric_label, leaderboard_data = self._get_leaderboard(period) if leaderboard is None else (self.metric_labels[self.metric], leaderboard)

        if not leaderboard_data:
            embed = discord.Embed(
                title=f"{period.title()} {self.category_display} {metric_label} Leaderboard",
                description="No data available",
                color=discord.Color.from_rgb(54, 57, 63)
            )
            return embed

        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard_data, page)
        self.page = clamped_page

        embed = discord.Embed(
            title=f"{period.title()} {self.category_display} {metric_label} Leaderboard",
            color=discord.Color.from_rgb(54, 57, 63)
        )

        description_lines = []
        for idx, entry in enumerate(sliced):
            player = entry[0]
            value = entry[2]
            level_value = entry[3]
            icon = entry[4]

            medal = {1: "1.", 2: "2.", 3: "3."}.get(start_index + idx + 1, f"{start_index + idx + 1}.")
            prestige_display = format_prestige_ansi(level_value, icon)

            if "Ratio" in metric_label or "/" in metric_label or "Per" in metric_label:
                formatted_value = f"{float(value):,.2f}"
            else:
                formatted_value = f"{value}"

            description_lines.append(f"{medal} {prestige_display} {player}: {formatted_value}")

        embed.description = f"```ansi\n" + "\n".join(description_lines) + "\n```"
        embed.set_footer(text=f"Page {clamped_page + 1} of {total_pages}")
        return embed

    async def _refresh(self, interaction: discord.Interaction, *, new_period: Optional[str] = None, page_delta: int = 0):
        if new_period is not None:
            self.current_period = new_period
            self.page = 0
            # sync dropdown defaults
            for option in self.period_select.options:
                option.default = option.value == new_period
        else:
            self.page += page_delta

        embed, file, _ = self.generate_leaderboard_image(self.current_period, self.page)

        if file:
            await interaction.response.edit_message(view=self, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Prev Page", custom_id="page_prev_ratio", style=discord.ButtonStyle.secondary)
    async def prev_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=-1)

    @discord.ui.button(label="Next Page", custom_id="page_next_ratio", style=discord.ButtonStyle.secondary)
    async def next_page(self, interaction: discord.Interaction, button: discord.ui.Button):
        await self._refresh(interaction, page_delta=1)

    @discord.ui.button(label="🔍 Search", custom_id="page_search_ratio", style=discord.ButtonStyle.primary)
    async def search(self, interaction: discord.Interaction, button: discord.ui.Button):
        modal = RatioLeaderboardSearchModal(self)
        await interaction.response.send_modal(modal)


class RatioLeaderboardSearchModal(discord.ui.Modal, title="Search Leaderboard"):
    def __init__(self, view: RatioLeaderboardView):
        super().__init__()
        self.view_ref = view
    
    search_input = discord.ui.TextInput(
        label="Search by Rank or Player Name",
        placeholder="Enter rank number (e.g. 500) or player name (e.g. Chuckegg)",
        required=True,
        max_length=100
    )
    
    async def on_submit(self, interaction: discord.Interaction):
        search_term = self.search_input.value.strip()
        
        # Get current leaderboard data
        _, leaderboard = self.view_ref._get_leaderboard(self.view_ref.current_period)
        
        if not leaderboard:
            await interaction.response.send_message("❌ No leaderboard data available", ephemeral=True)
            return
        
        # Try to parse as rank number
        target_page = None
        try:
            rank = int(search_term)
            if 1 <= rank <= len(leaderboard):
                # Calculate which page this rank is on (0-indexed)
                target_page = (rank - 1) // self.view_ref.page_size
            else:
                await interaction.response.send_message(
                    f"❌ Rank {rank} is out of range (1-{len(leaderboard)})",
                    ephemeral=True
                )
                return
        except ValueError:
            # Not a number, search for player name (case-insensitive)
            search_lower = search_term.lower()
            for idx, entry in enumerate(leaderboard):
                player_name = entry[0]
                if player_name.lower() == search_lower or search_lower in player_name.lower():
                    target_page = idx // self.view_ref.page_size
                    break
            
            if target_page is None:
                await interaction.response.send_message(
                    f"❌ Player '{search_term}' not found in this leaderboard",
                    ephemeral=True
                )
                return
        
        # Jump to the target page
        self.view_ref.page = target_page
        embed, file, _ = self.view_ref.generate_leaderboard_image(self.view_ref.current_period, target_page)
        
        if file:
            await interaction.response.edit_message(view=self.view_ref, attachments=[file])
        else:
            await interaction.response.edit_message(embed=embed, view=self.view_ref)


class RatioPeriodSelect(discord.ui.Select):
    def __init__(self, view: RatioLeaderboardView):
        options = [
            discord.SelectOption(label="Lifetime", value="lifetime", default=True),
            discord.SelectOption(label="Session", value="session"),
            discord.SelectOption(label="Daily", value="daily"),
            discord.SelectOption(label="Yesterday", value="yesterday"),
            discord.SelectOption(label="Weekly", value="weekly"),
            discord.SelectOption(label="Monthly", value="monthly"),
        ]
        super().__init__(
            placeholder="Select leaderboard period",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="ratio_leaderboard_period_select",
        )
        self.view_ref = view

    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref._refresh(interaction, new_period=selected)


# Create bot with command tree for slash commands
intents = discord.Intents.default()
# Enabled intents: members and presences required for member/presence features;
# message_content allows reading message content if needed (user enabled in Dev Portal).
intents.members = True
intents.presences = True
intents.message_content = True
bot = commands.Bot(command_prefix="!", intents=intents)

# In-memory pending claim registry so approvals can be handled via slash command even after buttons expire
PENDING_CLAIMS: dict[int, dict] = {}
PENDING_STREAKS: dict[int, dict] = {}

# Approval system for claim command
class ApprovalView(discord.ui.View):
    def __init__(self, ign: str, requester: str, requester_id: int, original_interaction: discord.Interaction):
        super().__init__(timeout=None)
        self.ign = ign
        self.requester = requester
        self.requester_id = requester_id
        self.original_interaction = original_interaction
        self.approved = None
        self.done_event = asyncio.Event()
        self.processed_by_admin_command = False
        self.admin_messages = []
    
    async def _update_other_admins(self, interaction: discord.Interaction, action: str):
        admin_name = interaction.user.name
        for msg in self.admin_messages:
            if msg.id == interaction.message.id:
                continue
            try:
                await msg.edit(content=f"{admin_name} {action} claim for {self.ign}.", view=None)
            except Exception:
                pass
    
    @discord.ui.button(label="Accept", style=discord.ButtonStyle.success)
    async def accept_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.approved = True
        PENDING_CLAIMS.pop(self.requester_id, None)
        self.done_event.set()
        await interaction.response.edit_message(content=f"You accepted claim for {self.ign}.", view=None)
        await self._update_other_admins(interaction, "accepted")
    
    @discord.ui.button(label="Deny", style=discord.ButtonStyle.danger)
    async def deny_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.approved = False
        PENDING_CLAIMS.pop(self.requester_id, None)
        self.done_event.set()
        await interaction.response.edit_message(content=f"You denied claim for {self.ign}.", view=None)
        await self._update_other_admins(interaction, "denied")


class StreakApprovalView(discord.ui.View):
    def __init__(self, ign: str, requester: str, requester_id: int, stats_snapshot: dict):
        super().__init__(timeout=None)
        self.ign = ign
        self.requester = requester
        self.requester_id = requester_id
        self.stats_snapshot = stats_snapshot or {}
        self.approved = None
        self.done_event = asyncio.Event()
        self.processed_by_admin_command = False
        self.admin_messages = []

    async def _update_other_admins(self, interaction: discord.Interaction, action: str):
        admin_name = interaction.user.name
        for msg in self.admin_messages:
            if msg.id == interaction.message.id:
                continue
            try:
                await msg.edit(content=f"{admin_name} {action} streak tracking for {self.ign}.", view=None)
            except Exception:
                pass

    @discord.ui.button(label="Accept", style=discord.ButtonStyle.success)
    async def accept_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        try:
            initialize_streak_entry(self.ign, self.stats_snapshot)
            PENDING_STREAKS.pop(self.requester_id, None)
            self.approved = True
            self.done_event.set()
            try:
                requester_user = interaction.client.get_user(self.requester_id) or await interaction.client.fetch_user(self.requester_id)
                if requester_user:
                    await requester_user.send(f"✅ Your streak tracking request for {self.ign} was approved.")
            except Exception:
                pass
            await interaction.response.edit_message(content=f"You approved streak tracking for {self.ign}.", view=None)
            await self._update_other_admins(interaction, "approved")
        except Exception as e:
            await interaction.response.edit_message(content=f"[ERROR] Failed to initialize streaks: {e}", view=None)

    @discord.ui.button(label="Deny", style=discord.ButtonStyle.danger)
    async def deny_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        PENDING_STREAKS.pop(self.requester_id, None)
        self.approved = False
        self.done_event.set()
        try:
            requester_user = interaction.client.get_user(self.requester_id) or await interaction.client.fetch_user(self.requester_id)
            if requester_user:
                await requester_user.send(f"❌ Your streak tracking request for {self.ign} was denied.")
        except Exception:
            pass
        await interaction.response.edit_message(content=f"You denied streak tracking for {self.ign}.", view=None)
        await self._update_other_admins(interaction, "denied")


class StreakRequestView(discord.ui.View):
    def __init__(self, ign: str, requester: discord.User, stats_snapshot: dict):
        super().__init__()
        self.ign = ign
        self.requester = requester
        self.requester_id = requester.id
        self.stats_snapshot = stats_snapshot or {}

    @discord.ui.button(label="Request tracking", style=discord.ButtonStyle.primary)
    async def request_tracking(self, interaction: discord.Interaction, button: discord.ui.Button):
        if interaction.user.id != self.requester_id:
            await interaction.response.send_message("This request button is only for the original requester.", ephemeral=True)
            return

        if self.requester_id in PENDING_STREAKS:
            await interaction.response.send_message("You already have a pending streak tracking request.", ephemeral=True)
            return

        admins = []
        for admin_id in ADMIN_IDS:
            try:
                uid = int(admin_id)
                user = interaction.client.get_user(uid) or await interaction.client.fetch_user(uid)
                if user:
                    admins.append(user)
            except Exception:
                pass

        if not admins:
            await interaction.response.send_message("[ERROR] Cannot reach administrators for approval.", ephemeral=True)
            return

        view = StreakApprovalView(self.ign, self.requester.name, self.requester_id, self.stats_snapshot)
        _register_pending_streak(self.requester_id, self.ign, self.stats_snapshot, view)

        sent_count = 0
        for admin in admins:
            try:
                msg = await admin.send(
                    f"{self.requester.name} ({self.requester_id}) requests streak tracking for {self.ign}.\n"
                    f"Click Accept/Deny below or run /verification-streak user:{self.requester_id} option: accept/deny.",
                    view=view,
                )
                view.admin_messages.append(msg)
                sent_count += 1
            except Exception:
                pass

        if sent_count == 0:
            _pop_pending_streak(self.requester_id)
            await interaction.response.send_message(f"[ERROR] Could not send streak approval request to administrators.", ephemeral=True)
            return

        await interaction.response.send_message("✅ Sent streak tracking request for approval.", ephemeral=True)


def _register_pending_claim(user_id: int, ign: str, view: ApprovalView):
    PENDING_CLAIMS[user_id] = {"ign": ign, "view": view}


def _pop_pending_claim(user_id: int):
    return PENDING_CLAIMS.pop(user_id, None)


def _register_pending_streak(user_id: int, ign: str, stats_snapshot: dict, view):
    PENDING_STREAKS[user_id] = {"ign": ign, "stats": stats_snapshot, "view": view}


def _pop_pending_streak(user_id: int):
    return PENDING_STREAKS.pop(user_id, None)


# Error Reporting System
class ErrorReportView(discord.ui.View):
    """View for users to report errors to administrators."""
    def __init__(self, error_details: str, command_name: str, user_info: str, context: str):
        super().__init__(timeout=300)  # 5 minute timeout for error reporting
        self.error_details = error_details
        self.command_name = command_name
        self.user_info = user_info
        self.context = context
        self.reported = False
    
    @discord.ui.button(label="📩 Report to Admins", style=discord.ButtonStyle.primary)
    async def report_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        if self.reported:
            await interaction.response.send_message("✅ This error has already been reported.", ephemeral=True)
            return
        
        # Disable the button after reporting
        self.reported = True
        button.disabled = True
        button.label = "✅ Reported"
        button.style = discord.ButtonStyle.success
        
        # Send report to admins
        admins = []
        for admin_id in ADMIN_IDS:
            try:
                uid = int(admin_id)
                user = interaction.client.get_user(uid) or await interaction.client.fetch_user(uid)
                if user:
                    admins.append(user)
            except Exception:
                pass
        
        report_embed = discord.Embed(
            title="🚨 Error Report",
            description=f"User {self.user_info} encountered an error",
            color=discord.Color.red(),
            timestamp=discord.utils.utcnow()
        )
        report_embed.add_field(name="Command", value=f"`{self.command_name}`", inline=False)
        report_embed.add_field(name="Context", value=self.context[:1024], inline=False)
        report_embed.add_field(name="Error Details", value=f"```{self.error_details[:1000]}```", inline=False)
        
        sent_count = 0
        for admin in admins:
            try:
                await admin.send(embed=report_embed)
                sent_count += 1
            except Exception:
                pass
        
        if sent_count > 0:
            try:
                if interaction.response.is_done():
                    await interaction.followup.send(
                        content="✅ Error reported to administrators. Thank you for helping improve the bot!",
                        ephemeral=True
                    )
                else:
                    await interaction.response.edit_message(
                        content="✅ Error reported to administrators. Thank you for helping improve the bot!",
                        view=self
                    )
            except (discord.errors.NotFound, discord.errors.HTTPException):
                # Interaction expired, but report was sent successfully
                pass
        else:
            try:
                if interaction.response.is_done():
                    await interaction.followup.send(
                        content="❌ Failed to send report to administrators. Please try again later or contact them directly.",
                        ephemeral=True
                    )
                else:
                    await interaction.response.edit_message(
                        content="❌ Failed to send report to administrators. Please try again later or contact them directly.",
                        view=self
                    )
            except (discord.errors.NotFound, discord.errors.HTTPException):
                pass


async def send_error_with_report(
    interaction: discord.Interaction,
    user_message: str,
    technical_details: str,
    command_name: str,
    context: str = "",
    ephemeral: bool = False
) -> None:
    """
    Send a user-friendly error message with a button to report the error to admins.
    
    Args:
        interaction: The Discord interaction
        user_message: User-friendly error message
        technical_details: Technical error details for admin report
        command_name: Name of the command that failed
        context: Additional context (e.g., "IGN: player123")
        ephemeral: Whether the message should be ephemeral
    """
    user_info = f"{interaction.user.name} ({interaction.user.id})"
    
    # Create error reporting view
    view = ErrorReportView(
        error_details=technical_details,
        command_name=command_name,
        user_info=user_info,
        context=context
    )
    
    # Send user-friendly message with report button
    full_message = f"❌ {user_message}\n\n*If this problem persists, you can report it to the administrators using the button below.*"
    
    try:
        if interaction.response.is_done():
            await interaction.followup.send(content=full_message, view=view, ephemeral=ephemeral)
        else:
            await interaction.response.send_message(content=full_message, view=view, ephemeral=ephemeral)
    except Exception as e:
        print(f"[ERROR] Failed to send error message: {e}")
        # Fallback to simple message if view fails
        try:
            if interaction.response.is_done():
                await interaction.followup.send(content=user_message, ephemeral=ephemeral)
            else:
                await interaction.response.send_message(content=user_message, ephemeral=ephemeral)
        except Exception:
            pass


# Database Repair Confirmation View
class DatabaseRepairConfirmView(discord.ui.View):
    """Confirmation view for database repair operations."""
    def __init__(self, backup_file: str):
        super().__init__(timeout=60)  # 1 minute timeout
        self.backup_file = backup_file
        self.confirmed = None
        self.done_event = asyncio.Event()
    
    @discord.ui.button(label="✅ Confirm Repair", style=discord.ButtonStyle.danger)
    async def confirm_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_admin(interaction.user):
            await interaction.response.send_message("❌ Only administrators can confirm database repairs.", ephemeral=True)
            return
        
        self.confirmed = True
        self.done_event.set()
        
        # Disable all buttons
        for item in self.children:
            item.disabled = True
        
        await interaction.response.edit_message(
            content=f"🔧 Repair confirmed. Processing...",
            view=self
        )
    
    @discord.ui.button(label="❌ Cancel", style=discord.ButtonStyle.secondary)
    async def cancel_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        if not is_admin(interaction.user):
            await interaction.response.send_message("❌ Only administrators can cancel database repairs.", ephemeral=True)
            return
        
        self.confirmed = False
        self.done_event.set()
        
        # Disable all buttons
        for item in self.children:
            item.disabled = True
        
        await interaction.response.edit_message(
            content="❌ Database repair cancelled.",
            view=self
        )
    
    async def on_timeout(self):
        """Handle timeout - auto-cancel if no response."""
        if self.confirmed is None:
            self.confirmed = False
            self.done_event.set()


# Bot token
# Read from BOT_TOKEN.txt in the same directory
TOKEN_FILE = os.path.join(os.path.dirname(__file__), "BOT_TOKEN.txt")
try:
    with open(TOKEN_FILE, "r", encoding="utf-8") as f:
        DISCORD_TOKEN = f.read().strip()
except Exception as e:
    DISCORD_TOKEN = None
    print(f"[ERROR] Failed to read BOT_TOKEN.txt: {e}")
if not DISCORD_TOKEN:
    raise ValueError("BOT_TOKEN.txt is missing or empty")

@bot.event
async def on_ready():
    import time
    bot_instance_id = int(time.time() * 1000) % 100000
    print(f"[OK] Bot logged in as {bot.user} - Instance ID: {bot_instance_id}")
    
    # Initialize database schema to ensure all tables exist
    await asyncio.to_thread(init_database)
    
    # Check for legacy data migration
    await check_legacy_migration()
    
    # Verify API key
    await asyncio.to_thread(verify_api_key)
    
    try:
        synced = await bot.tree.sync()
        print(f"[OK] Synced {len(synced)} command(s) - Instance ID: {bot_instance_id}")

    except Exception as e:
        print(f"[ERROR] Failed to sync commands: {e}")
    # start background tasks once
    if not getattr(bot, "background_tasks_started", False):
        # Store task references for graceful shutdown
        bot.background_tasks = [
            bot.loop.create_task(scheduler_loop()),
            bot.loop.create_task(staggered_stats_refresher(interval_minutes=5)),
            bot.loop.create_task(streak_stats_refresher(interval_seconds=60)),
            bot.loop.create_task(presence_updater_loop(interval_seconds=5)),
            bot.loop.create_task(guild_updater_hourly())
        ]
        bot.background_tasks_started = True
        print(f"[OK] Background tasks started (including hourly guild updater) - Instance ID: {bot_instance_id}")

@bot.tree.command(name="track", description="Create a stats sheet for a player (no authorization required)")
@discord.app_commands.describe(ign="Minecraft IGN")
async def track(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Check if user is already tracked (UUID-aware)
        if is_tracked_user(ign):
            # Get the current username for this user (in case ign is old username)
            resolved = resolve_username_to_uuid(ign)
            if resolved:
                current_name = resolved[1]
                await interaction.followup.send(f"{current_name} is already being tracked.")
            else:
                await interaction.followup.send(f"{ign} is already being tracked.")
            return
        
        # Create sheet using api_get.py
        # Initialize session, daily, and monthly snapshots (yesterday will be populated from daily rotation)
        result = run_script("api_get.py", ["-ign", ign, "-session", "-yesterday", "-daily", "-monthly"])

        if result.returncode == 0:
            print(f"[OK] api_get.py succeeded for {ign}")
            
            # Parse output to get proper username
            actual_ign = ign
            if result.stdout:
                try:
                    for line in reversed(result.stdout.splitlines()):
                        if line.strip().startswith('{'):
                            data = json.loads(line.strip())
                            if 'username' in data:
                                actual_ign = data['username']
                                break
                except:
                    pass
            
            # Verify user exists in DB
            if not user_exists(actual_ign):
                await interaction.followup.send(f"[ERROR] Database entry for {ign} was not created.")
                return
            
            # Add to tracked users list using the properly-cased username
            added = add_tracked_user(actual_ign)
            
            if added:
                await interaction.followup.send(f"{actual_ign} is now being tracked. Use `/claim ign:{actual_ign}` to link this username to your Discord account.")
            else:
                await interaction.followup.send(f"{actual_ign} is already being tracked.")
        else:
            err = (result.stderr or result.stdout) or "Unknown error"
            print(f"[ERROR] api_get.py failed for {ign}:")
            print(f"  stdout: {result.stdout}")
            print(f"  stderr: {result.stderr}")
            await interaction.followup.send(f"Error creating sheet for {ign}:\n```{sanitize_output(err[:500])}```")
            
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="add", description="[Admin] Add user(s) to database without tracking")
@discord.app_commands.describe(
    ign="Single Minecraft IGN to add",
    ignswithcommas="Multiple IGNs separated by commas (e.g., user1,user2,user3)"
)
async def add_users(interaction: discord.Interaction, ign: str = None, ignswithcommas: str = None):
    """Add one or multiple users to the database without tracking them."""
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Admin only
    if not is_admin(interaction.user):
        await interaction.edit_original_response(content="❌ [ERROR] This command is admin-only.")
        return
    
    # Check that exactly one option is provided
    if (ign is None and ignswithcommas is None) or (ign is not None and ignswithcommas is not None):
        await interaction.edit_original_response(
            content="❌ Error: Please provide exactly one of the following options:\n"
            "• `ign:` for a single username\n"
            "• `ignswithcommas:` for multiple usernames separated by commas"
        )
        return
    
    try:
        if ign is not None:
            # Single user - use api_get.py directly
            # Validate username
            ok, proper_ign = validate_and_normalize_ign(ign)
            if not ok:
                await interaction.edit_original_response(content=f"❌ The username `{ign}` is invalid.")
                return
            ign = proper_ign
            
            # Check if already tracked
            if is_tracked_user(ign):
                await interaction.edit_original_response(content=f"⚠️ `{ign}` is already tracked.")
                return
            
            # Check if already registered (but not tracked)
            if is_registered_user(ign):
                await interaction.edit_original_response(content=f"⚠️ `{ign}` is already registered in the database.")
                return
            
            # Fetch stats using api_get.py
            await interaction.edit_original_response(content=f"🔄 Fetching stats for `{ign}`...")
            result = run_script("api_get.py", ["-ign", ign])
            
            if result.returncode == 0:
                # Parse output to get proper username
                actual_ign = ign
                if result.stdout:
                    try:
                        for line in reversed(result.stdout.splitlines()):
                            if line.strip().startswith('{'):
                                data = json.loads(line.strip())
                                if 'username' in data:
                                    actual_ign = data['username']
                                    break
                    except:
                        pass
                
                # Register user (but don't track)
                if register_user(actual_ign):
                    await interaction.edit_original_response(
                        content=f"✅ Successfully added `{actual_ign}` to the database!\n"
                        f"💡 This user will appear in leaderboards but won't receive periodic updates.\n"
                        f"Use `/track ign:{actual_ign}` to start tracking."
                    )
                else:
                    await interaction.edit_original_response(content=f"⚠️ `{actual_ign}` may already be in the database.")
            else:
                err = (result.stderr or result.stdout) or "Unknown error"
                await interaction.edit_original_response(content=f"❌ Error fetching stats for `{ign}`:\n```{sanitize_output(err[:500])}```")
                
        else:
            # Multiple users - use add_to_db.py with temporary file
            # Remove spaces and split by comma
            cleaned = ignswithcommas.replace(" ", "")
            usernames = [u.strip() for u in cleaned.split(",") if u.strip()]
            
            if not usernames:
                await interaction.edit_original_response(content="❌ No valid usernames found in the comma-separated list.")
                return
            
            # Validate all usernames first
            invalid_users = []
            valid_users = []
            for username in usernames:
                ok, proper_ign = validate_and_normalize_ign(username)
                if ok:
                    valid_users.append(proper_ign)
                else:
                    invalid_users.append(username)
            
            if invalid_users and not valid_users:
                await interaction.edit_original_response(
                    content=f"❌ Invalid username(s): {', '.join(f'`{u}`' for u in invalid_users)}\n"
                    f"No valid usernames to process."
                )
                return
            
            if not valid_users:
                await interaction.edit_original_response(content="❌ No valid usernames to process.")
                return
            
            # Create temporary file with usernames
            import tempfile
            with tempfile.NamedTemporaryFile(mode='w', suffix='.txt', delete=False, dir=BOT_DIR) as temp_file:
                temp_filename = temp_file.name
                for username in valid_users:
                    temp_file.write(f"{username}\n")
            
            try:
                # Show initial status message
                status_msg = f"🔄 Processing {len(valid_users)} user(s)...\n"
                if invalid_users:
                    status_msg += f"⚠️ Skipping invalid username(s): {', '.join(f'`{u}`' for u in invalid_users)}\n"
                status_msg += "This may take a few moments."
                await interaction.edit_original_response(content=status_msg)
                
                # Run add_to_db.py with the temporary file
                result = run_script_batch("add_to_db.py", [temp_filename])
                
                if result.returncode in [0, 2]:  # 0 = success, 2 = some failed
                    # Parse the output to extract statistics
                    output = result.stdout or ""
                    
                    # Send the summary
                    summary_lines = []
                    in_summary = False
                    for line in output.splitlines():
                        if "SUMMARY" in line or "="*20 in line:
                            in_summary = True
                        if in_summary and (line.strip().startswith("Total") or 
                                          line.strip().startswith("✅") or 
                                          line.strip().startswith("⏭️") or
                                          line.strip().startswith("🔒") or
                                          line.strip().startswith("❌")):
                            summary_lines.append(line)
                    
                    if summary_lines:
                        summary = "\n".join(summary_lines)
                        await interaction.edit_original_response(
                            content=f"✅ Batch processing complete!\n```\n{summary}\n```\n"
                            f"💡 Added users will appear in leaderboards but won't receive periodic updates."
                        )
                    else:
                        await interaction.edit_original_response(
                            content=f"✅ Batch processing complete!\n"
                            f"Processed {len(valid_users)} user(s)."
                        )
                else:
                    err = (result.stderr or result.stdout) or "Unknown error"
                    await interaction.edit_original_response(
                        content=f"❌ Error processing users:\n```{sanitize_output(err[:1000])}```"
                    )
            finally:
                # Clean up temporary file
                try:
                    os.unlink(temp_filename)
                except:
                    pass
                    
    except subprocess.TimeoutExpired:
        await interaction.edit_original_response(content="❌ [ERROR] Command timed out. Too many users or API is slow.")
    except Exception as e:
        await interaction.edit_original_response(content=f"❌ [ERROR] {str(e)}")

@bot.tree.command(name="trackguild", description="Track guild experience stats for a guild")
@discord.app_commands.describe(guild="Guild name")
async def trackguild(interaction: discord.Interaction, guild: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Check if guild is already tracked
        already_tracked = is_tracked_guild(guild)
        
        # Fetch guild data using api_get.py guild function
        # We need to import and call api_update_guild_database
        from api_get import api_update_guild_database, read_api_key_file
        
        api_key = read_api_key_file()
        if not api_key:
            await interaction.followup.send("❌ [ERROR] API key not configured.")
            return
        
        # Initialize session, daily, and monthly snapshots
        if already_tracked:
            await interaction.edit_original_response(content=f"🔄 Updating guild data for `{guild}`...")
        else:
            await interaction.edit_original_response(content=f"🔄 Fetching guild data for `{guild}`...")
        
        result = api_update_guild_database(
            guild, 
            api_key, 
            snapshot_sections={'session', 'daily', 'yesterday', 'weekly', 'monthly'} if not already_tracked else None
        )
        
        if 'error' in result:
            await interaction.edit_original_response(
                content=f"❌ Error fetching guild data for `{guild}`:\n```{result['error'][:500]}```"
            )
            return
        
        # Get the proper guild name from the result
        proper_name = result.get('guild_name', guild)
        guild_tag = result.get('guild_tag', '')
        guild_tag_color = result.get('guild_tag_color', '')
        
        # Add to tracked guilds list with tag info
        added = add_tracked_guild(proper_name, guild_tag, guild_tag_color)
        
        games_count = len(result.get('games', []))
        total_exp = result.get('total_exp', 0)
        
        if already_tracked:
            await interaction.edit_original_response(
                content=f"✅ Guild `{proper_name}` data has been updated!\n"
                       f"📊 Tag: [{guild_tag}] • {games_count} game types • {total_exp:,} total experience."
            )
        else:
            await interaction.edit_original_response(
                content=f"✅ Guild `{proper_name}` is now being tracked!\n"
                       f"📊 Tag: [{guild_tag}] • {games_count} game types • {total_exp:,} total experience."
            )
    except Exception as e:
        await interaction.edit_original_response(content=f"❌ [ERROR] {str(e)}")

@bot.tree.command(name="untrackguild", description="[Admin] Stop tracking a guild")
@discord.app_commands.describe(guild="Guild name")
async def untrackguild(interaction: discord.Interaction, guild: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Admin only
    if not is_admin(interaction.user):
        await interaction.edit_original_response(content="❌ [ERROR] This command is admin-only.")
        return
    
    try:
        # Check if guild is tracked
        if not is_tracked_guild(guild):
            await interaction.followup.send(f"Guild `{guild}` is not being tracked.")
            return
        
        # Remove from tracked guilds
        removed = remove_tracked_guild(guild)
        
        if removed:
            await interaction.followup.send(f"✅ Guild `{guild}` has been removed from tracking.")
        else:
            await interaction.followup.send(f"❌ Failed to remove guild `{guild}` from tracking.")
    except Exception as e:
        await interaction.followup.send(f"❌ [ERROR] {str(e)}")

@bot.tree.command(name="updateguilds", description="[Admin] Update tag data for all tracked guilds")
async def updateguilds(interaction: discord.Interaction):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Admin only
    if not is_admin(interaction.user):
        await interaction.edit_original_response(content="❌ [ERROR] This command is admin-only.")
        return
    
    try:
        from api_get import api_update_guild_database, read_api_key_file
        
        api_key = read_api_key_file()
        if not api_key:
            await interaction.edit_original_response(content="❌ [ERROR] API key not configured.")
            return
        
        # Get guilds that need periodic updates (guilds without tracked members)
        # Guilds with tracked members get updated automatically when those members are updated
        guilds_to_update = get_guilds_for_periodic_updates()
        all_tracked = get_tracked_guilds()
        skipped_count = len(all_tracked) - len(guilds_to_update)
        
        # Print detailed breakdown to console
        print("\n" + "="*80)
        print(f"[GUILD UPDATE] Tracked Guilds Analysis ({len(all_tracked)} total)")
        print("="*80)
        
        guilds_to_update_set = set(guilds_to_update)
        
        for guild_name in all_tracked:
            will_update = guild_name in guilds_to_update_set
            has_members = guild_has_tracked_members(guild_name)
            
            if will_update:
                print(f"✅ {guild_name}")
                print(f"   → Will receive periodic updates (no tracked members)")
            else:
                print(f"⏭️  {guild_name}")
                print(f"   → Skipping periodic updates (has tracked members, updates automatically)")
        
        print("="*80)
        print(f"[SUMMARY] Updates needed: {len(guilds_to_update)} | Skipped: {skipped_count}")
        print("="*80 + "\n")
        
        if not guilds_to_update:
            if skipped_count > 0:
                await interaction.edit_original_response(
                    content=f"✅ All {len(all_tracked)} tracked guild(s) have tracked members and will be updated automatically.\nNo periodic updates needed."
                )
            else:
                await interaction.edit_original_response(content="❌ No guilds are being tracked.")
            return
        
        status_msg = f"🔄 Updating {len(guilds_to_update)} guild(s)..."
        if skipped_count > 0:
            status_msg += f"\n(Skipping {skipped_count} guild(s) with tracked members)"
        status_msg += "\nThis may take a moment."
        await interaction.edit_original_response(content=status_msg)
        
        success_count = 0
        failed = []
        
        for guild_name in guilds_to_update:
            try:
                result = api_update_guild_database(guild_name, api_key)
                if 'error' not in result:
                    success_count += 1
                else:
                    failed.append(f"{guild_name}: {result['error'][:50]}")
            except Exception as e:
                failed.append(f"{guild_name}: {str(e)[:50]}")
        
        # Build response message
        msg = f"✅ Successfully updated {success_count}/{len(guilds_to_update)} guild(s)."
        if skipped_count > 0:
            msg += f"\n(Skipped {skipped_count} guild(s) with tracked members)"
        if failed:
            msg += f"\n\n❌ Failed ({len(failed)}):\n" + "\n".join(f"• {f}" for f in failed[:5])
            if len(failed) > 5:
                msg += f"\n... and {len(failed) - 5} more"
        
        await interaction.edit_original_response(content=msg)
    except Exception as e:
        await interaction.edit_original_response(content=f"❌ [ERROR] {str(e)}")

@bot.tree.command(name="claim", description="Link a Minecraft username to your Discord account")
@discord.app_commands.describe(ign="Minecraft IGN")
async def claim(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Check if username is tracked (UUID-aware, handles old usernames)
        from db_helper import is_tracked_user
        if not is_tracked_user(ign):
            await interaction.followup.send(f"[ERROR] `{ign}` is not being tracked. Use `/track ign:{ign}` first.")
            return
        
        # Check if already claimed
        links = load_user_links()
        if ign.casefold() in links:
            claimed_by = links[ign.casefold()]
            if claimed_by == str(interaction.user.id):
                await interaction.followup.send(f"[ERROR] You have already claimed {ign}.")
            else:
                await interaction.followup.send(f"[ERROR] {ign} is already claimed by another user.")
            return
        
        # Get admin users
        admins = []
        for admin_id in ADMIN_IDS:
            try:
                uid = int(admin_id)
                user = bot.get_user(uid) or await bot.fetch_user(uid)
                if user:
                    admins.append(user)
            except Exception:
                pass
        
        if not admins:
            await interaction.followup.send("[ERROR] Cannot reach administrators for approval.")
            return
        
        # Send waiting message to requester
        requester_name = interaction.user.name
        await interaction.followup.send(f"Asked administrators for approval to claim {ign}. Please wait for confirmation.")
        
        # Create approval view and send to admins
        view = ApprovalView(ign, requester_name, interaction.user.id, interaction)
        _register_pending_claim(interaction.user.id, ign, view)
        
        sent_count = 0
        for admin in admins:
            try:
                msg = await admin.send(
                    f"{requester_name} ({interaction.user.id}) wants to claim {ign}.\n"
                    f"Click Accept/Deny below or run /verification user:{interaction.user.id} option: accept/deny.",
                    view=view,
                )
                view.admin_messages.append(msg)
                sent_count += 1
            except Exception:
                pass
        
        if sent_count == 0:
            _pop_pending_claim(interaction.user.id)
            await interaction.followup.send(f"[ERROR] Could not send approval request to administrators.")
            return
        
        # Wait for approval (no timeout)
        await view.done_event.wait()
        
        # Process based on approval
        if getattr(view, "processed_by_admin_command", False):
            # Manual /verification already handled linking/notification
            if view.approved:
                await interaction.followup.send(f"Chuckegg has approved your claim. {ign} is now linked to your Discord account.")
            else:
                await interaction.followup.send(f"Chuckegg has denied your claim for {ign}.")
            return

        _pop_pending_claim(interaction.user.id)

        if view.approved:
            link_user_to_ign(interaction.user.id, ign)
            await interaction.followup.send(f"An administrator has approved your claim. {ign} is now linked to your Discord account.")
        else:
            await interaction.followup.send(f"An administrator has denied your claim for {ign}.")
            
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="unclaim", description="Unlink a Minecraft username from your Discord account")
@discord.app_commands.describe(ign="Minecraft IGN to unclaim")
async def unclaim(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Check if user is authorized to unclaim this username
    if not (is_admin(interaction.user) or is_user_authorized(interaction.user.id, ign)):
        await interaction.followup.send(f"[ERROR] You are not authorized to unclaim {ign}. Only the user who claimed this username or an admin can unclaim it.")
        return
    
    try:
        # Remove from user links
        removed_link = unlink_user_from_ign(ign)
        
        if removed_link:
            await interaction.followup.send(f"Successfully unclaimed {ign}. You are no longer linked to this username.")
        else:
            await interaction.followup.send(f"[WARNING] No claim found for {ign}.")
            
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to unclaim: {str(e)}")


@bot.tree.command(name="verification", description="Manually approve or deny a claim (admin only)")
@discord.app_commands.describe(option="Accept or deny the claim", user="Discord user ID of the requester")
@discord.app_commands.choices(option=[
    discord.app_commands.Choice(name="Accept", value="accept"),
    discord.app_commands.Choice(name="Deny", value="deny"),
])
async def verification(interaction: discord.Interaction, option: discord.app_commands.Choice[str], user: str):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    try:
        user_id = int(user)
    except ValueError:
        await interaction.followup.send("[ERROR] Invalid user ID.", ephemeral=True)
        return

    pending = _pop_pending_claim(user_id)
    if not pending:
        await interaction.followup.send("[ERROR] No pending claim found for that user.", ephemeral=True)
        return

    ign = pending.get("ign")
    view = pending.get("view")
    approved = option.value == "accept"

    # If we have the original view, mark it and unblock the waiting task
    if view:
        view.approved = approved
        view.processed_by_admin_command = True
        view.done_event.set()
        
        # Update admin messages
        admin_name = interaction.user.name
        action_text = "accepted" if approved else "denied"
        for msg in getattr(view, 'admin_messages', []):
            try:
                await msg.edit(content=f"{admin_name} {action_text} claim for {ign} (via command).", view=None)
            except Exception:
                pass

    requester_user = None
    try:
        requester_user = bot.get_user(user_id) or await bot.fetch_user(user_id)
    except Exception:
        requester_user = None

    if approved:
        link_user_to_ign(user_id, ign)
        if requester_user:
            try:
                await requester_user.send(f"✅ Your claim for {ign} was approved by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Approved claim: {ign} linked to <@{user_id}>.", ephemeral=True)
    else:
        if requester_user:
            try:
                await requester_user.send(f"❌ Your claim for {ign} was denied by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Denied claim for {ign} (requester <@{user_id}>).", ephemeral=True)


@bot.tree.command(name="verification-streak", description="Approve or deny a streak tracking request (admin only)")
@discord.app_commands.describe(option="Accept or deny the streak request", user="Discord user ID of the requester")
@discord.app_commands.choices(option=[
    discord.app_commands.Choice(name="Accept", value="accept"),
    discord.app_commands.Choice(name="Deny", value="deny"),
])
async def verification_streak(interaction: discord.Interaction, option: discord.app_commands.Choice[str], user: str):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    try:
        user_id = int(user)
    except ValueError:
        await interaction.followup.send("[ERROR] Invalid user ID.", ephemeral=True)
        return

    pending = _pop_pending_streak(user_id)
    if not pending:
        await interaction.followup.send("[ERROR] No pending streak request found for that user.", ephemeral=True)
        return

    ign = pending.get("ign")
    stats_snapshot = pending.get("stats", {})
    view = pending.get("view")
    approved = option.value == "accept"

    requester_user = None
    try:
        requester_user = bot.get_user(user_id) or await bot.fetch_user(user_id)
    except Exception:
        requester_user = None

    if approved:
        initialize_streak_entry(ign, stats_snapshot)
        if view:
            view.approved = True
            view.processed_by_admin_command = True
            view.done_event.set()
            
            # Update admin messages
            admin_name = interaction.user.name
            for msg in getattr(view, 'admin_messages', []):
                try:
                    await msg.edit(content=f"{admin_name} approved streak tracking for {ign} (via command).", view=None)
                except Exception:
                    pass
        if requester_user:
            try:
                await requester_user.send(f"✅ Your streak tracking request for {ign} was approved by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Approved streak tracking for {ign} (requester <@{user_id}>).", ephemeral=True)
    else:
        if view:
            view.approved = False
            view.processed_by_admin_command = True
            view.done_event.set()
            
            # Update admin messages
            admin_name = interaction.user.name
            for msg in getattr(view, 'admin_messages', []):
                try:
                    await msg.edit(content=f"{admin_name} denied streak tracking for {ign} (via command).", view=None)
                except Exception:
                    pass
        if requester_user:
            try:
                await requester_user.send(f"❌ Your streak tracking request for {ign} was denied by an admin.")
            except Exception:
                pass
        await interaction.followup.send(f"Denied streak tracking for {ign} (requester <@{user_id}>).", ephemeral=True)

@bot.tree.command(name="untrack", description="Remove all tracking data for a Minecraft username")
@discord.app_commands.describe(ign="Minecraft IGN to untrack")
async def untrack(interaction: discord.Interaction, ign: str):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Allow creator override; otherwise require claim authorization
    if not (is_admin(interaction.user) or is_user_authorized(interaction.user.id, ign)):
        await interaction.followup.send(f"[ERROR] You are not authorized to untrack {ign}. Only the user who claimed this username or the creator can untrack it.")
        return
    
    try:
        actual_ign = ign
        
        # Remove from all locations
        removed_tracked = remove_tracked_user(actual_ign)
        removed_link = unlink_user_from_ign(actual_ign)
        removed_color = remove_user_color(actual_ign)
        removed_sheet = delete_user_sheet(actual_ign)
        
        if removed_tracked or removed_link or removed_color or removed_sheet:
            results = []
            if removed_tracked:
                results.append("tracked users list")
            if removed_link:
                results.append("user links")
            if removed_color:
                results.append("user colors")
            if removed_sheet:
                results.append("stats sheet")
            
            await interaction.followup.send(f"Successfully untracked {actual_ign}. Removed from: {', '.join(results)}.")
        else:
            await interaction.followup.send(f"[WARNING] {ign} was not found in any tracking data.")
            
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to untrack: {str(e)}")

@bot.tree.command(name="remove", description="[Admin] Force remove a user from tracking (case-sensitive)")
@discord.app_commands.describe(ign="Exact Minecraft IGN to remove (case-sensitive)")
async def remove_tracked(interaction: discord.Interaction, ign: str):
    """Admin-only command to force remove a tracked user with exact casing."""
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    # Admin only
    if not is_admin(interaction.user):
        await interaction.followup.send("❌ [ERROR] This command is admin-only.")
        return
    
    try:
        # Use exact casing - query database with case-sensitive match
        with get_db_connection() as conn:
            cursor = conn.cursor()
            
            # Remove from tracked_users (exact match)
            cursor.execute('DELETE FROM tracked_users WHERE username = ?', (ign,))
            removed_tracked = cursor.rowcount > 0
            
            # Remove from user_links (exact match)
            cursor.execute('DELETE FROM user_links WHERE username = ?', (ign,))
            removed_link = cursor.rowcount > 0
            
            # Remove from user_meta (case-insensitive for color/metadata)
            cursor.execute('DELETE FROM user_meta WHERE LOWER(username) = LOWER(?)', (ign,))
            removed_meta = cursor.rowcount > 0
            
            conn.commit()
        
        # Also try to delete the sheet
        removed_sheet = delete_user_sheet(ign)
        
        if removed_tracked or removed_link or removed_meta or removed_sheet:
            results = []
            if removed_tracked:
                results.append(f"tracked users ({ign})")
            if removed_link:
                results.append("user links")
            if removed_meta:
                results.append("user metadata")
            if removed_sheet:
                results.append("stats sheet")
            
            await interaction.followup.send(
                f"✅ Successfully force-removed **{ign}** (exact casing).\n"
                f"Removed from: {', '.join(results)}.\n\n"
                f"⚠️ Note: This was a case-sensitive removal. If duplicate casings exist, run this command again with the other casing."
            )
        else:
            await interaction.followup.send(
                f"⚠️ [WARNING] **{ign}** (exact casing) was not found in tracking data.\n\n"
                f"💡 Tip: This command is case-sensitive. Make sure the casing matches exactly."
            )
            
    except Exception as e:
        await interaction.followup.send(f"❌ [ERROR] Failed to remove: {str(e)}")

# Create color choices from MINECRAFT_CODE_TO_HEX
COLOR_CHOICES = [
    discord.app_commands.Choice(name="Black", value="0"),
    discord.app_commands.Choice(name="Dark Blue", value="1"),
    discord.app_commands.Choice(name="Dark Green", value="2"),
    discord.app_commands.Choice(name="Dark Aqua", value="3"),
    discord.app_commands.Choice(name="Dark Red", value="4"),
    discord.app_commands.Choice(name="Dark Purple", value="5"),
    discord.app_commands.Choice(name="Gold", value="6"),
    discord.app_commands.Choice(name="Gray", value="7"),
    discord.app_commands.Choice(name="Dark Gray", value="8"),
    discord.app_commands.Choice(name="Blue", value="9"),
    discord.app_commands.Choice(name="Green", value="a"),
    discord.app_commands.Choice(name="Aqua", value="b"),
    discord.app_commands.Choice(name="Red", value="c"),
    discord.app_commands.Choice(name="Light Purple/Pink", value="d"),
    discord.app_commands.Choice(name="Yellow", value="e"),
    discord.app_commands.Choice(name="White", value="f"),
]

@bot.tree.command(name="color", description="Set a custom color for your username in stats displays")
@discord.app_commands.describe(
    ign="Minecraft IGN (optional if you set /default)",
    color="Color for your username"
)
@discord.app_commands.choices(color=COLOR_CHOICES)
async def color(interaction: discord.Interaction, ign: str = None, color: discord.app_commands.Choice[str] = None):
    # Resolve default IGN if not provided, and validate before any heavy work
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username via Mojang API and simple format
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            pass
    
    # Check if user is authorized to change color for this username
    if not (is_admin(interaction.user) or is_user_authorized(interaction.user.id, ign)):
        await interaction.followup.send(f"[ERROR] You are not authorized to change the color for {ign}. Only the user who claimed this username or an admin can change its color.", ephemeral=True)
        return
    
    try:
        # Get hex color from code
        color_code = color.value
        hex_color = MINECRAFT_CODE_TO_HEX.get(color_code, '#FFFFFF')
        
        # Update color in database
        update_user_meta(ign, ign_color=hex_color)
        
        # Invalidate cache to force reload with new color
        await STATS_CACHE.invalidate()
        
        await interaction.followup.send(f"Successfully set {ign}'s username color to {color.name}!", ephemeral=True)
        
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to set color: {str(e)}", ephemeral=True)

@bot.tree.command(name="reset", description="Reset session snapshot for a player")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def reset(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN and validate before any heavy work
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            pass
    
    # Check if user is authorized to reset session for this username
    if not (is_admin(interaction.user) or is_user_authorized(interaction.user.id, ign)):
        await interaction.followup.send(f"[ERROR] You are not authorized to reset session for {ign}. Only the user who claimed this username or an admin can reset its session.", ephemeral=True)
        return
    
    try:
        result = run_script("api_get.py", ["-ign", ign, "-session"])

        if result.returncode == 0:
            await interaction.followup.send(f"Session snapshot reset for {ign}.", ephemeral=True)
        else:
            err = (result.stderr or result.stdout) or "Unknown error"
            await interaction.followup.send(f"[ERROR] {sanitize_output(err)}", ephemeral=True)
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)", ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}", ephemeral=True)

@bot.tree.command(name="dmme", description="Send yourself a test DM from the bot")
async def dmme(interaction: discord.Interaction):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    try:
        await interaction.user.send("Hello! This is a private message from the bot.")
        await interaction.followup.send("Sent you a DM.", ephemeral=True)
    except Exception as e:
        await interaction.followup.send("Couldn't DM you. Check your privacy settings (Allow DMs from server members).", ephemeral=True)


@bot.tree.command(name="default", description="Set your default Minecraft username")
@discord.app_commands.describe(ign="Minecraft IGN to use by default")
async def default(interaction: discord.Interaction, ign: str):
    # Validate username before persisting
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return

    # Quick response, no heavy work
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            pass
    try:
        # Optionally validate tracked status to help users
        excel_file = BOT_DIR / "stats.xlsx"
        is_known = False
        if excel_file.exists():
            wb = None
            try:
                wb = load_workbook(str(excel_file), read_only=True, data_only=True)
                key = proper_ign.casefold()
                for sheet_name in wb.sheetnames:
                    if sheet_name.casefold() == key:
                        is_known = True
                        break
            except Exception:
                pass
            finally:
                if wb is not None:
                    try:
                        wb.close()
                    except Exception:
                        pass

        set_default_user(interaction.user.id, proper_ign)
        if is_known:
            await interaction.followup.send(f"Default username set to {proper_ign}.", ephemeral=True)
        else:
            await interaction.followup.send(f"Default username set to {proper_ign}. Note: {proper_ign} is not tracked yet; some commands may fail until you run /track.", ephemeral=True)
    except Exception as e:
        await interaction.followup.send(f"[ERROR] Failed to set default: {str(e)}", ephemeral=True)

@bot.tree.command(name="removedefault", description="Remove your default Minecraft username")
async def removedefault(interaction: discord.Interaction):
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    removed = remove_default_user(interaction.user.id)
    if removed:
        await interaction.followup.send("Your default username has been removed.", ephemeral=True)
    else:
        await interaction.followup.send("You don't have a default username set.", ephemeral=True)

@bot.tree.command(name="prestige", description="Display a prestige prefix for any level")
@discord.app_commands.describe(
    level="The prestige level (e.g., 1964)",
    ign="Optional: Username to display after the prefix"
)
async def prestige(interaction: discord.Interaction, level: int, ign: str = None):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Validate level range
        if level < 0 or level > 10000:
            await interaction.followup.send("[ERROR] Level must be between 0 and 10000")
            return
        
        # Get prestige icon for this level
        icon = get_prestige_icon(level)
        
        # Build the colored prefix
        colored_prefix = format_prestige_ansi(level, icon)
        
        # Add IGN if provided
        if ign:
            display_text = f"{colored_prefix} {ign}"
        else:
            display_text = colored_prefix
        
        # Try to render as image if Pillow is available
        if Image is not None:
            try:
                base = (level // 100) * 100
                raw = PRESTIGE_RAW_PATTERNS.get(base)
                
                if raw:
                    # Parse the pattern and replace the level number
                    parts = _parse_raw_pattern(raw)
                    
                    # Build segments with the actual level
                    concat = ''.join(t for (_, t) in parts)
                    m = re.search(r"\d+", concat)
                    
                    segments = []
                    if m:
                        num_start, num_end = m.start(), m.end()
                        pos = 0
                        replaced = False
                        
                        for code, text in parts:
                            part_start = pos
                            part_end = pos + len(text)
                            pos = part_end
                            hexcol = MINECRAFT_CODE_TO_HEX.get(code.lower(), '#FFFFFF')
                            
                            if part_end <= num_start or part_start >= num_end:
                                segments.append((hexcol, text))
                                continue
                            
                            # Prefix before number
                            prefix_len = max(0, num_start - part_start)
                            if prefix_len > 0:
                                segments.append((hexcol, text[:prefix_len]))
                            
                            # Replace with actual level
                            if not replaced:
                                # Check if this is a rainbow prestige
                                rainbow_bases = {k for k, v in PRESTIGE_COLORS.items() if v is None}
                                if base in rainbow_bases:
                                    # Build rainbow colors
                                    colors_in_span = []
                                    pos2 = 0
                                    for c_code, c_text in parts:
                                        part_s = pos2
                                        part_e = pos2 + len(c_text)
                                        pos2 = part_e
                                        overlap_s = max(part_s, num_start)
                                        overlap_e = min(part_e, num_end)
                                        if overlap_e > overlap_s:
                                            hexcol_span = MINECRAFT_CODE_TO_HEX.get(c_code.lower(), '#FFFFFF')
                                            for _ in range(overlap_e - overlap_s):
                                                colors_in_span.append(hexcol_span)
                                    
                                    if not colors_in_span:
                                        RAINBOW_CODES = ['c', '6', 'e', 'a', 'b', 'd', '9', '3']
                                        colors_in_span = [MINECRAFT_CODE_TO_HEX.get(c, '#FFFFFF') for c in RAINBOW_CODES]
                                    
                                    # Apply colors to level digits
                                    for i, ch in enumerate(str(level)):
                                        col = colors_in_span[i % len(colors_in_span)]
                                        segments.append((col, ch))
                                else:
                                    segments.append((hexcol, str(level)))
                                replaced = True
                            
                            # Suffix after number
                            suffix_start_in_part = max(0, num_end - part_start)
                            if suffix_start_in_part < len(text):
                                segments.append((hexcol, text[suffix_start_in_part:]))
                    else:
                        # No number found, just use the pattern as-is with level prepended
                        segments = [(MINECRAFT_CODE_TO_HEX.get(parts[0][0], '#FFFFFF'), f"[{level}")]
                        segments.extend([(MINECRAFT_CODE_TO_HEX.get(code, '#FFFFFF'), text) for code, text in parts[1:]])
                    
                    # Add IGN if provided
                    if ign:
                        segments.append((MINECRAFT_CODE_TO_HEX.get('f', '#FFFFFF'), f" {ign}"))
                    
                    # Render to image
                    img_io = _render_text_segments_to_image(segments)
                    filename = f"prestige_{level}" + (f"_{ign}" if ign else "") + ".png"
                    await interaction.followup.send(file=discord.File(img_io, filename=filename))
                    return
            except Exception as e:
                # Fall back to ANSI if image rendering fails
                print(f"[WARNING] Image rendering failed: {e}")
        
        # Fallback: send as ANSI text
        await interaction.followup.send(f"```ansi\n{display_text}\n```")
        
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")



@bot.tree.command(name="levelprogress", description="View progress to next level/prestige with playtime estimates")
@discord.app_commands.describe(
    ign="Minecraft IGN (optional if you set /default)",
    level="Optional target level to include in progress display"
)
async def levelprogress(interaction: discord.Interaction, ign: str = None, level: int = None):
    """Display level progress and ETA predictions."""
    print(f"[DEBUG] /levelprogress triggered for IGN: {ign} by user: {interaction.user.name} in guild: {interaction.guild.name if interaction.guild else 'DM'}")
    
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign
    
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[DEBUG] Defer failed for {ign} in /levelprogress: {e}")
            return
    
    try:
        # Fetch fresh stats
        print(f"[DEBUG] Running api_get.py for IGN: {ign} (/levelprogress)")
        result = run_script("api_get.py", ["-ign", ign])
        
        if result.returncode != 0:
            if result.stderr and "429" in result.stderr:
                print(f"[DEBUG] Rate limited for {ign} (/levelprogress), attempting to use existing data")
            else:
                error_msg = result.stderr or result.stdout or "Unknown error"
                await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
                return
        
        # Optimistically update cache if we have JSON output
        try:
            if result.stdout:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                    ign = json_data["username"]
        except Exception as e:
            print(f"[WARNING] Failed to update cache from output: {e}")
        
        # Get user data from cache
        cache_data = await STATS_CACHE.get_data()
        
        # Find user in cache case-insensitively
        key = ign.casefold()
        user_data = None
        actual_ign = ign
        for name, data in cache_data.items():
            if name.casefold() == key:
                user_data = data
                actual_ign = name
                break
        
        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return
        
        # Check if user is tracked (UUID-aware)
        is_tracked = is_tracked_user(actual_ign)
        
        # Validate custom level if provided
        custom_level = None
        if level is not None:
            if level < 1 or level > 10000:
                await interaction.followup.send("[ERROR] Custom level must be between 1 and 10,000")
                return
            # Make sure it's greater than current level
            current_level = user_data.get("meta", {}).get("level", 0)
            if level <= current_level:
                await interaction.followup.send(f"[ERROR] Custom level {level} must be greater than current level {current_level}")
                return
            custom_level = level
        
        # Create view and generate image
        view = LevelProgressView(user_data, actual_ign, custom_level=custom_level)
        file = view.generate_level_progress_image("all-time")
        
        if is_tracked:
            message = await interaction.followup.send(file=file, view=view)
            view.message = message  # Store message reference for timeout handling
        else:
            msg = f"`{actual_ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{actual_ign}` to start tracking and enable session/daily/monthly stats."
            await interaction.followup.send(content=msg, file=file)
            # Register user in database for leaderboard accuracy (but don't actively track)
            register_user(actual_ign)
    
    except subprocess.TimeoutExpired:
        await send_error_with_report(
            interaction,
            "The command took too long to complete.",
            "Command timed out (30s limit)",
            "/levelprogress",
            f"IGN: {ign}, Level: {level}"
        )
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        await send_error_with_report(
            interaction,
            "An unexpected error occurred while generating level progress predictions.",
            f"{str(e)}\n\n{error_traceback}",
            "/levelprogress",
            f"IGN: {ign}, Level: {level}"
        )


class InstructionsView(discord.ui.View):
    def __init__(self, instructions_data):
        super().__init__(timeout=300)  # 5 minutes timeout
        self.data = instructions_data
        self.current_section = "home"
        self.message = None
        self.update_button_styles()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass
    
    def update_button_styles(self):
        """Update button styles based on current section."""
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_section:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary
    
    def get_embed(self):
        """Generate embed for current section."""
        section_data = self.data.get(self.current_section, {})
        
        embed = discord.Embed(
            title=section_data.get("title", "🐑 Sheep Wars Bot"),
            description=section_data.get("description", "Select a section below."),
            color=discord.Color.blue()
        )
        
        # Add fields if any
        for field in section_data.get("fields", []):
            embed.add_field(
                name=field["name"],
                value=field["value"],
                inline=field.get("inline", False)
            )
        
        embed.set_footer(text=section_data.get("footer", "Use the buttons below to navigate"))
        return embed
    
    @discord.ui.button(label="🏠 Home", custom_id="home", row=0)
    async def home_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_section = "home"
        self.update_button_styles()
        await interaction.response.edit_message(embed=self.get_embed(), view=self)
    
    @discord.ui.button(label="📊 Stats", custom_id="stats", row=0)
    async def stats_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_section = "stats"
        self.update_button_styles()
        await interaction.response.edit_message(embed=self.get_embed(), view=self)
    
    @discord.ui.button(label="🏆 Leaderboards", custom_id="leaderboards", row=0)
    async def leaderboards_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_section = "leaderboards"
        self.update_button_styles()
        await interaction.response.edit_message(embed=self.get_embed(), view=self)
    
    @discord.ui.button(label="⚙️ Utility", custom_id="utility", row=1)
    async def utility_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_section = "utility"
        self.update_button_styles()
        await interaction.response.edit_message(embed=self.get_embed(), view=self)
    
    @discord.ui.button(label="👑 Admin", custom_id="admin", row=1)
    async def admin_button(self, interaction: discord.Interaction, button: discord.ui.Button):
        self.current_section = "admin"
        self.update_button_styles()
        await interaction.response.edit_message(embed=self.get_embed(), view=self)


@bot.tree.command(name="instructions", description="Display bot usage instructions")
async def instructions(interaction: discord.Interaction):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Define sections with organized content from instructions.txt
        instructions_data = {
            "home": {
                "title": "🐑 Sheep Wars Bot — Quick Start Guide",
                "description": "Welcome! This bot tracks Sheep Wars statistics, leaderboards, and player data.\n\n**Getting Started:**",
                "fields": [
                    {"name": "1️⃣ Track a Player", "value": "`/track ign:<username>`\nCreates database entry and snapshots. Does not recover historic data.\nTracked players appear in non-lifetime leaderboards.", "inline": False},
                    {"name": "2️⃣ Claim Your IGN (Optional)", "value": "`/claim ign:<username>`\nLinks Minecraft username to Discord. Requires admin approval.\nAllows you to manage your profile (reset session, set color, etc.).", "inline": False},
                    {"name": "3️⃣ Set Default IGN", "value": "`/default ign:<username>`\nSaves default IGN so commands work without typing it each time.\nRemove with `/removedefault`", "inline": False},
                    {"name": "📚 Navigation", "value": "**📊 Stats** - View player statistics\n**🏆 Leaderboards** - Rankings and comparisons\n**⚙️ Utility** - Setup and personalization\n**👑 Admin** - Administrator commands", "inline": False},
                ],
                "footer": "Note: This is a vibe coded bot, bugs are normal • Buttons expire after 5 minutes"
            },
            "stats": {
                "title": "📊 Statistics Commands",
                "description": "View detailed player statistics across different game modes and time periods.",
                "fields": [
                    {"name": "/sheepwars ign:<username>", "value": "Sheep Wars stats rendered as image. Periods: all-time, session, daily, yesterday, weekly, monthly.", "inline": False},
                    {"name": "/ww ign:<username>", "value": "Wool Wars stats with class breakdown and interactive class selector.", "inline": False},
                    {"name": "/ctw ign:<username>", "value": "Capture The Wool stats (wools captured, kills on wool holder, gold earned/spent).", "inline": False},
                    {"name": "/stats ign:<username>", "value": "Full sheep wars layout showing many boxes and derived metrics.", "inline": False},
                    {"name": "/compare ign1:<u1> ign2:<u2> [stat:<metric>]", "value": "Compare two players. Without stat shows full comparison; with stat focuses on that metric.", "inline": False},
                    {"name": "/levelprogress ign:<username> [level:<target>]", "value": "Progress to next level/prestige with ETA based on recent exp/hour.", "inline": False},
                    {"name": "/ratios ign:<username>", "value": "Predicts wins/kills and time to reach next WLR (W/L) and KDR (K/D) milestones.", "inline": False},
                    {"name": "/streak ign:<username>", "value": "View current win/kill streaks (requires streak tracking approval).", "inline": False},
                    {"name": "/layout ign:<username>", "value": "View hotbar/hotkey layouts for Sheep Wars, Wool Wars, and CTW.", "inline": False},
                    {"name": "/killdistribution & /deathdistribution", "value": "Render kill-type or death-type distribution as pie chart.", "inline": False},
                ],
                "footer": "💡 Uses cached data for tracked users, fetches live for untracked when possible"
            },
            "leaderboards": {
                "title": "🏆 Leaderboards & Rankings",
                "description": "View rankings, compare positions, and explore top players.",
                "fields": [
                    {"name": "/leaderboard <category> metric:<metric>", "value": "Categories: **general**, **sheepwars**, **ctw**, **ww**, **guild**\nPeriods: lifetime, session, daily, yesterday, weekly, monthly\nResults are paged and searchable with interactive UI.", "inline": False},
                    {"name": "/rankings <category> ign:<username>", "value": "Show player's rank for every metric in that category across all periods.", "inline": False},
                    {"name": "/rankings guild guild_name:<guild>", "value": "Show guild's rank across game types and periods.", "inline": False},
                ],
                "footer": "Tracked players receive periodic snapshot updates for time-based leaderboards"
            },
            "utility": {
                "title": "⚙️ Utility & Personalization",
                "description": "Setup, customization, and account management commands.",
                "fields": [
                    {"name": "Account Management", "value": "`/track ign:<username>` - Start tracking\n`/claim ign:<username>` - Link to Discord (needs approval)\n`/unclaim` - Unlink from Discord\n`/default ign:<username>` - Set default IGN\n`/removedefault` - Remove default", "inline": False},
                    {"name": "Customization", "value": "`/color ign:<username> color:<choice>` - Set display color\n`/reset ign:<username>` - Reset session snapshot\n`/streak-remove ign:<username>` - Stop streak tracking", "inline": False},
                    {"name": "Other Utilities", "value": "`/dmme` - Test DM from bot\n`/prestige level:<number>` - Preview prestige bracket\n`/version` - Bot uptime and path\n`/api-stats` - Hypixel API usage statistics", "inline": False},
                ],
                "footer": "💡 Only claiming user or admin can change colors/settings for an IGN"
            },
            "admin": {
                "title": "👑 Administrator Commands",
                "description": "Admin-only commands for managing users, guilds, and database.",
                "fields": [
                    {"name": "User Management", "value": "`/add ign:<username>` or `ignswithcommas:<u1,u2,...>`\nAdds users without periodic tracking.\n\n`/remove ign:<Exact_Case_Username>` - Force-remove tracked user (case-sensitive)", "inline": False},
                    {"name": "Guild Management", "value": "`/trackguild` - Track a guild\n`/untrackguild` - Stop tracking guild\n`/updateguilds` - Refresh guild metadata\n`/fixguildtracking`, `/fixguilds` - Repair guild data", "inline": False},
                    {"name": "Snapshots & Updates", "value": "`/refresh mode:<...> [ign:<username>]` - Manual snapshot trigger\nModes: batch or single-user (requires permission)", "inline": False},
                    {"name": "Verification", "value": "`/verification user:<id> option:<accept|deny>` - Approve/deny claims\n`/verification-streak` - Handle streak tracking requests", "inline": False},
                    {"name": "Database", "value": "`/repairdatabase` - Check integrity, attempt repair/restore from backups (requires confirmation)", "inline": False},
                    {"name": "Monitoring", "value": "`/whatamirunningon` - Creator-only: IP and bot path", "inline": False},
                ],
                "footer": "⚠️ Admin commands require proper permissions"
            }
        }
        
        view = InstructionsView(instructions_data)
        embed = view.get_embed()
        message = await interaction.followup.send(embed=embed, view=view)
        view.message = message
        
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="whatamirunningon", description="Creator-only: show public IP and bot file path")
async def whatamirunningon(interaction: discord.Interaction):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    # Only allow the creator (by ID) to run this
    allowed = is_admin(interaction.user)

    if not allowed:
        await interaction.followup.send("[ERROR] You are not authorized to run this command.", ephemeral=True)
        return

    # Try to get public IP, fall back to local hostname IP
    ip = None
    try:
        import urllib.request, json, ssl
        ctx = ssl.create_default_context()
        with urllib.request.urlopen("https://api.ipify.org?format=json", timeout=5, context=ctx) as resp:
            data = json.load(resp)
            ip = data.get("ip")
    except Exception:
        try:
            import socket
            ip = socket.gethostbyname(socket.gethostname())
        except Exception:
            ip = "unknown"

    path = str(BOT_DIR / "discord_bot.py")
    await interaction.followup.send(f"IP: {ip}\nPath: {path}", ephemeral=True)

@bot.tree.command(name="version", description="Show bot uptime and verify code version")
async def version(interaction: discord.Interaction):
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    uptime = int(time.time() - START_TIME)
    hours, rem = divmod(uptime, 3600)
    minutes, seconds = divmod(rem, 60)
    
    await interaction.followup.send(f"Bot Uptime: {hours}h {minutes}m {seconds}s\nRunning from: {BOT_DIR}", ephemeral=True)


# ==============================
# TIMELINE COMMAND
# ==============================

async def guild_autocomplete(interaction: discord.Interaction, current: str) -> list[discord.app_commands.Choice[str]]:
    """Autocomplete for guilds that have historical data."""
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute('SELECT DISTINCT name FROM guild_historical ORDER BY name')
        guilds = [row['name'] for row in cursor.fetchall()]
    
    current_lower = current.lower()
    matches = [
        discord.app_commands.Choice(name=guild, value=guild)
        for guild in guilds
        if current_lower in guild.lower()
    ]
    return matches[:25]


def generate_timeline_graph(guild_name: str, game: str, show_delta: bool = False, width: int = 920, height: int = 520):
    """Generate a timeline graph similar to the reference image.
    
    Args:
        guild_name: Name of the guild
        game: Game type (e.g., 'GENERAL', 'WOOL_WARS')
        show_delta: If True, show daily change instead of absolute values
        width: Image width in pixels
        height: Image height in pixels
    
    Returns:
        BytesIO: PNG image data
    """
    from datetime import datetime
    import io
    
    # Get historical data
    history = get_guild_historical_data(guild_name, game=game)
    
    # Filter out zero values
    history = [record for record in history if record['lifetime_exp'] > 0]
    
    if not history or len(history) < 2:
        # Create error image
        img = Image.new('RGB', (width, height), (18, 18, 20))
        draw = ImageDraw.Draw(img)
        try:
            error_font = ImageFont.truetype("/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf", 20)
        except:
            error_font = ImageFont.load_default()
        text = "No historical data available" if not history else "Insufficient data (need at least 2 data points)"
        bbox = draw.textbbox((0, 0), text, font=error_font)
        text_width = bbox[2] - bbox[0]
        text_height = bbox[3] - bbox[1]
        draw.text(((width - text_width) // 2, (height - text_height) // 2), text, fill=(200, 100, 100), font=error_font)
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return buf
    
    # Extract data points
    timestamps = [record['timestamp'] for record in history]
    exp_values = [record['lifetime_exp'] for record in history]
    
    # Calculate values to display (either absolute or delta)
    if show_delta:
        # Calculate daily changes
        display_values = [0]  # First value has no previous, so delta is 0
        for i in range(1, len(exp_values)):
            delta = exp_values[i] - exp_values[i-1]
            display_values.append(delta)
    else:
        display_values = exp_values
    
    # Calculate statistics
    start_exp = exp_values[0]
    end_exp = exp_values[-1]
    total_gain = end_exp - start_exp
    avg_per_day = total_gain / (len(history) - 1) if len(history) > 1 else 0
    
    # Create image
    bg_color = (24, 28, 38)
    img = Image.new('RGB', (width, height), bg_color)
    draw = ImageDraw.Draw(img)
    
    # Margins
    margin_left = 80
    margin_right = 40
    margin_top = 100
    margin_bottom = 80
    
    plot_width = width - margin_left - margin_right
    plot_height = height - margin_top - margin_bottom
    
    # Calculate scales with padding
    min_val = min(display_values)
    max_val = max(display_values)
    
    # For delta view, include 0 in the range if there are negative values
    if show_delta and min_val < 0:
        min_val = min(min_val, 0)
        max_val = max(max_val, 0)
    
    # Add padding above and below (10% on each side)
    val_range = max_val - min_val if max_val != min_val else 1
    padding = val_range * 0.1
    min_val -= padding
    max_val += padding
    val_range = max_val - min_val
    
    min_time = min(timestamps)
    max_time = max(timestamps)
    time_range = max_time - min_time if max_time != min_time else 1
    
    # Transform data to pixel coordinates
    points = []
    for ts, val in zip(timestamps, display_values):
        x = margin_left + int((ts - min_time) / time_range * plot_width)
        y = margin_top + plot_height - int((val - min_val) / val_range * plot_height)
        points.append((x, y))
    
    # Draw filled area under the line with gradient effect
    # Create vertical gradient fill effect (darker at bottom, brighter at top)
    for i in range(len(points) - 1):
            x1, y1 = points[i]
            x2, y2 = points[i + 1]
            
            # Draw gradient segments vertically from bottom to line
            bottom_y = height - margin_bottom
            
            # Use the full height from bottom to the line
            num_segments = 50
            
            for seg in range(num_segments):
                # Calculate segment boundaries (from bottom going up)
                seg_y_start = bottom_y - (bottom_y - bottom_y) * seg / num_segments
                seg_y_end = bottom_y - (bottom_y - bottom_y) * (seg + 1) / num_segments
                
                # Calculate the line position at this height
                # Linear interpolation between y1 and y2
                line_y1 = y1 + (bottom_y - y1) * (num_segments - seg) / num_segments
                line_y2 = y2 + (bottom_y - y2) * (num_segments - seg) / num_segments
                line_y1_next = y1 + (bottom_y - y1) * (num_segments - seg - 1) / num_segments
                line_y2_next = y2 + (bottom_y - y2) * (num_segments - seg - 1) / num_segments
                
                # Calculate color gradient (darker at bottom, brighter at top)
                # Factor goes from 0 (bottom) to 1 (top)
                gradient_factor = seg / num_segments
                
                # Darker, more subtle blue gradient
                r = int(25 + 35 * gradient_factor)
                g = int(50 + 80 * gradient_factor)  
                b = int(75 + 105 * gradient_factor)
                
                fill_color = (r, g, b)
                
                # Draw trapezoid segment that follows the line
                polygon = [
                    (x1, line_y1),
                    (x2, line_y2),
                    (x2, line_y2_next),
                    (x1, line_y1_next)
                ]
                draw.polygon(polygon, fill=fill_color)
    
    # Draw the line connecting points
    line_color = (100, 180, 255)
    for i in range(len(points) - 1):
        if show_delta:
            # Color based on value
            if display_values[i] < 0:
                line_color = (255, 100, 100)
            else:
                line_color = (100, 180, 255)
        draw.line([points[i], points[i + 1]], fill=line_color, width=3)
    
    # Draw points (circles) on the line
    point_radius = 4
    for i, (x, y) in enumerate(points):
        if show_delta:
            point_color = (255, 100, 100) if display_values[i] < 0 else (100, 180, 255)
        else:
            point_color = (100, 180, 255)
        
        # Draw filled circle
        draw.ellipse([x - point_radius, y - point_radius, x + point_radius, y + point_radius], 
                    fill=point_color, outline=point_color)
    
    # Draw zero line for delta view
    if show_delta and min_val < 0:
        zero_y = margin_top + plot_height - int((0 - min_val) / val_range * plot_height)
        draw.line([(margin_left, zero_y), (width - margin_right, zero_y)], fill=(100, 100, 100), width=2)
    
    # Load font (try to use a better font if available)
    try:
        title_font = ImageFont.truetype("/usr/share/fonts/dejavu-sans-fonts/DejaVuSans-Bold.ttf", 26)
        label_font = ImageFont.truetype("/usr/share/fonts/dejavu-sans-fonts/DejaVuSans.ttf", 16)
        subtitle_font = ImageFont.truetype("/usr/share/fonts/dejavu-sans-fonts/DejaVuSans.ttf", 16)
    except Exception as e:
        print(f"Font loading error: {e}")
        title_font = ImageFont.load_default()
        label_font = ImageFont.load_default()
        subtitle_font = ImageFont.load_default()
    
    # Format game name
    game_names = {
        'GENERAL': 'General',
        'CTW': 'CTW',
        'WOOL_WARS': 'Wool Wars',
        'CTF': 'CTF',
        'BRIDGE': 'Bridge',
        'DESTROY': 'Destroy'
    }
    game_display = game_names.get(game, game.replace('_', ' ').title())
    
    # Draw title (same format for both delta and absolute modes)
    title = f"{game_display} Experience · {guild_name}"
    
    subtitle = f"{start_exp:,.0f} → {end_exp:,.0f} (+{total_gain:,.0f}) · Avg/day +{avg_per_day:,.1f}"
    
    draw.text((margin_left, 20), title, fill=(255, 255, 255), font=title_font)
    draw.text((margin_left, 78), subtitle, fill=(160, 160, 165), font=subtitle_font)
    
    # Draw Y-axis labels
    num_y_labels = 5
    for i in range(num_y_labels):
        value = min_val + (val_range * i / (num_y_labels - 1))
        y = margin_top + plot_height - int((value - min_val) / val_range * plot_height)
        
        # Format large numbers
        if abs(value) >= 1_000_000:
            label = f"{value / 1_000_000:+.1f}M" if show_delta else f"{value / 1_000_000:.1f}M"
        elif abs(value) >= 1_000:
            label = f"{value / 1_000:+.1f}K" if show_delta else f"{value / 1_000:.1f}K"
        else:
            label = f"{value:+.0f}" if show_delta else f"{value:.0f}"
        
        draw.text((10, y - 8), label, fill=(150, 150, 150), font=label_font)
        draw.line([(margin_left, y), (width - margin_right, y)], fill=(40, 40, 50), width=1)
    
    # Draw X-axis labels (dates)
    num_x_labels = min(5, len(points))
    for i in range(num_x_labels):
        idx = int(i * (len(history) - 1) / (num_x_labels - 1)) if num_x_labels > 1 else 0
        ts = timestamps[idx]
        date_str = datetime.fromtimestamp(ts).strftime('%m/%d')
        x = points[idx][0]
        
        bbox = draw.textbbox((0, 0), date_str, font=label_font)
        text_width = bbox[2] - bbox[0]
        draw.text((x - text_width // 2, height - margin_bottom + 15), date_str, fill=(150, 150, 150), font=label_font)
    
    # Draw start and end dates at bottom
    start_date = datetime.fromtimestamp(timestamps[0]).strftime('%Y-%m-%d')
    end_date = datetime.fromtimestamp(timestamps[-1]).strftime('%Y-%m-%d')
    draw.text((margin_left, height - 30), f"Start: {start_date}", fill=(120, 120, 120), font=subtitle_font)
    draw.text((width - margin_right - 180, height - 30), f"Latest: {end_date}", fill=(120, 120, 120), font=subtitle_font)
    
    # Save to buffer
    buf = io.BytesIO()
    img.save(buf, format='PNG')
    buf.seek(0)
    return buf


class TimelineGameSelect(discord.ui.View):
    """Dropdown menu and buttons for timeline controls."""
    
    def __init__(self, guild_name: str, current_game: str = 'GENERAL', show_delta: bool = False):
        super().__init__(timeout=180)
        self.guild_name = guild_name
        self.current_game = current_game
        self.show_delta = show_delta
        
        # Get available games for this guild
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('''
                SELECT DISTINCT game FROM guild_historical 
                WHERE name = ? AND lifetime_exp > 0
                ORDER BY game
            ''', (guild_name,))
            games = [row['game'] for row in cursor.fetchall()]
        
        # Create select menu
        options = []
        for game in games:
            display_name = game.replace('_', ' ').title()
            options.append(discord.SelectOption(
                label=display_name,
                value=game,
                default=(game == current_game)
            ))
        
        # Add select menu to view
        select = discord.ui.Select(
            placeholder="Choose a game type...",
            options=options[:25],  # Discord limit
            custom_id="game_select"
        )
        select.callback = self.game_callback
        self.add_item(select)
        
        # Add toggle button for absolute/delta view
        toggle_button = discord.ui.Button(
            label="Show Delta" if not show_delta else "Show Absolute",
            style=discord.ButtonStyle.primary if not show_delta else discord.ButtonStyle.secondary,
            custom_id="toggle_delta"
        )
        toggle_button.callback = self.toggle_callback
        self.add_item(toggle_button)
    
    async def game_callback(self, interaction: discord.Interaction):
        """Handle game selection."""
        game = interaction.data['values'][0]
        self.current_game = game
        
        await interaction.response.defer()
        
        try:
            # Generate new graph
            graph_buf = generate_timeline_graph(self.guild_name, game, show_delta=self.show_delta)
            file = discord.File(graph_buf, filename=f"timeline_{self.guild_name}_{game}.png")
            
            # Create new view with updated state
            new_view = TimelineGameSelect(self.guild_name, current_game=game, show_delta=self.show_delta)
            
            # Update message
            game_display = game.replace('_', ' ').title()
            mode_text = " (Daily Change)" if self.show_delta else ""
            await interaction.followup.edit_message(
                interaction.message.id,
                content=f"**Timeline for {self.guild_name} - {game_display}{mode_text}**",
                attachments=[file],
                view=new_view
            )
        except Exception as e:
            await interaction.followup.send(f"Error generating timeline: {str(e)}", ephemeral=True)
    
    async def toggle_callback(self, interaction: discord.Interaction):
        """Handle delta/absolute toggle."""
        self.show_delta = not self.show_delta
        
        await interaction.response.defer()
        
        try:
            # Generate new graph with toggled mode
            graph_buf = generate_timeline_graph(self.guild_name, self.current_game, show_delta=self.show_delta)
            file = discord.File(graph_buf, filename=f"timeline_{self.guild_name}_{self.current_game}.png")
            
            # Create new view with updated state
            new_view = TimelineGameSelect(self.guild_name, current_game=self.current_game, show_delta=self.show_delta)
            
            # Update message
            game_display = self.current_game.replace('_', ' ').title()
            mode_text = " (Daily Change)" if self.show_delta else ""
            await interaction.followup.edit_message(
                interaction.message.id,
                content=f"**Timeline for {self.guild_name} - {game_display}{mode_text}**",
                attachments=[file],
                view=new_view
            )
        except Exception as e:
            await interaction.followup.send(f"Error toggling view: {str(e)}", ephemeral=True)


@bot.tree.command(name="timeline", description="View guild experience timeline graph")
@discord.app_commands.describe(guild="Guild name")
@discord.app_commands.autocomplete(guild=guild_autocomplete)
async def timeline(interaction: discord.Interaction, guild: str):
    """Display a timeline graph of guild experience over time."""
    if not interaction.response.is_done():
        await interaction.response.defer()
    
    try:
        # Validate guild has historical data
        history = get_guild_historical_data(guild, game='GENERAL')
        if not history:
            await interaction.followup.send(f"❌ No historical data found for guild '{guild}'.")
            return
        
        # Generate initial graph (GENERAL by default)
        graph_buf = generate_timeline_graph(guild, 'GENERAL')
        file = discord.File(graph_buf, filename=f"timeline_{guild}_GENERAL.png")
        
        # Create view with game selector
        view = TimelineGameSelect(guild)
        
        await interaction.followup.send(
            content=f"**Timeline for {guild} - General**",
            file=file,
            view=view
        )
    
    except Exception as e:
        print(f"[ERROR] Timeline command failed: {e}")
        import traceback
        traceback.print_exc()
        await interaction.followup.send(f"❌ Error generating timeline: {str(e)}")


@bot.tree.command(name="api-stats", description="View Hypixel API usage statistics for the current 5-minute window")
async def api_stats(interaction: discord.Interaction):
    """Show API request statistics for the current 5-minute window."""
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    # Get stats from tracker
    stats = await API_TRACKER.get_stats()
    
    # Calculate time until next reset
    current_time = time.time()
    window_start = stats['window_start']
    next_reset = window_start + 300  # 5 minutes in seconds
    seconds_until_reset = int(next_reset - current_time)
    
    minutes, seconds = divmod(seconds_until_reset, 60)
    
    # Format timestamp for window start
    from datetime import datetime
    window_time = datetime.fromtimestamp(window_start, tz=CREATOR_TZ).strftime('%I:%M %p')
    
    # Build embed
    embed = discord.Embed(
        title="📊 Hypixel API Usage Statistics",
        description=f"Current 5-minute window: **{window_time}**\nResets in: **{minutes}m {seconds}s**",
        color=discord.Color.blue()
    )
    
    # Total requests
    total = stats['total']
    embed.add_field(
        name="Total Requests",
        value=f"```{total}```",
        inline=False
    )
    
    # Breakdown
    breakdown = stats['breakdown']
    breakdown_text = "\n".join([
        f"**{category.capitalize()}**: {count}"
        for category, count in sorted(breakdown.items(), key=lambda x: x[1], reverse=True)
        if count > 0
    ])
    
    if not breakdown_text:
        breakdown_text = "*No requests recorded in this window*"
    
    embed.add_field(
        name="Breakdown by Type",
        value=breakdown_text,
        inline=False
    )
    
    # Rate limit info
    embed.add_field(
        name="ℹ️ Hypixel API Limits",
        value="• 120 requests/minute per API key\n• 600 requests per 5 minutes\n• Rate limits reset every 5 minutes",
        inline=False
    )
    
    # Usage percentage (assuming 600 limit per 5 minutes)
    usage_pct = (total / 600) * 100
    if usage_pct < 50:
        usage_emoji = "🟢"
        usage_status = "Low"
    elif usage_pct < 80:
        usage_emoji = "🟡"
        usage_status = "Moderate"
    else:
        usage_emoji = "🔴"
        usage_status = "High"
    
    embed.set_footer(text=f"{usage_emoji} API Usage: {usage_status} ({usage_pct:.1f}% of 5-min limit)")
    
    await interaction.followup.send(embed=embed, ephemeral=True)


@bot.tree.command(name="repairdatabase", description="[ADMIN] Check and repair corrupted database, restoring from backup if needed")
async def repairdatabase(interaction: discord.Interaction):
    """Admin command to repair corrupted database."""
    # Admin check
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to administrators.", ephemeral=True)
        return
    
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    try:
        # Check database integrity
        is_valid, error_msg = check_database_integrity(DB_PATH)
        
        if is_valid:
            embed = discord.Embed(
                title="✅ Database Health Check",
                description="The database is healthy and does not need repair.",
                color=discord.Color.green()
            )
            await interaction.followup.send(embed=embed, ephemeral=True)
            return
        
        # Database is corrupted - show confirmation dialog
        embed = discord.Embed(
            title="⚠️ Database Corruption Detected",
            description="The database has corruption issues and needs repair.",
            color=discord.Color.orange()
        )
        embed.add_field(
            name="Corruption Details",
            value=f"```{error_msg[:500]}```",
            inline=False
        )
        
        # Find latest valid backup
        backup_path, timestamp_str = find_latest_valid_backup()
        
        if backup_path:
            backup_display = timestamp_str.replace("_", " ").replace("-", "/")
            embed.add_field(
                name="📦 Backup Available",
                value=f"Latest valid backup: `{timestamp_str}.db`\nTimestamp: {backup_display}",
                inline=False
            )
        else:
            embed.add_field(
                name="⚠️ No Valid Backup Found",
                value="No non-corrupted backups were found. Only repair attempt will be made.",
                inline=False
            )
        
        embed.add_field(
            name="🔧 Repair Strategy",
            value="1. **First**: Attempt to repair using SQLite dump/restore\n2. **If repair fails**: Restore from latest valid backup\n\nThe corrupted database will be saved with timestamp for investigation.",
            inline=False
        )
        
        # Create confirmation view
        view = DatabaseRepairConfirmView(timestamp_str if backup_path else "none")
        
        await interaction.followup.send(
            content=f"**⚠️ Database repair required. Confirm to proceed:**",
            embed=embed,
            view=view,
            ephemeral=True
        )
        
        # Wait for confirmation (60 second timeout)
        try:
            await asyncio.wait_for(view.done_event.wait(), timeout=60)
        except asyncio.TimeoutError:
            await interaction.followup.send(
                "❌ Confirmation timeout - database repair cancelled.",
                ephemeral=True
            )
            return
        
        if not view.confirmed:
            # User cancelled
            return
        
        # Perform repair
        status_msg = await interaction.followup.send(
            "🔧 Starting database repair process...",
            ephemeral=True
        )
        
        success, message, backup_used = await perform_database_repair()
        
        if success:
            result_embed = discord.Embed(
                title="✅ Database Repair Successful",
                description=message,
                color=discord.Color.green()
            )
            if backup_used:
                result_embed.add_field(
                    name="📦 Restored From Backup",
                    value=f"`{backup_used}.db`",
                    inline=False
                )
            result_embed.add_field(
                name="ℹ️ Next Steps",
                value="• The bot will continue running with the repaired database\n• Corrupted database was saved for analysis\n• Monitor for any issues",
                inline=False
            )
            
            # Reload cache after repair
            await STATS_CACHE.reload_cache()
            result_embed.add_field(
                name="🔄 Cache Status",
                value="Stats cache reloaded successfully",
                inline=False
            )
        else:
            result_embed = discord.Embed(
                title="❌ Database Repair Failed",
                description=message,
                color=discord.Color.red()
            )
            result_embed.add_field(
                name="⚠️ Manual Intervention Required",
                value="Please check the backups directory for valid backups and manually restore if needed.",
                inline=False
            )
        
        await interaction.followup.send(embed=result_embed, ephemeral=True)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        error_embed = discord.Embed(
            title="❌ Error During Repair Process",
            description=f"An unexpected error occurred:\n```{str(e)[:500]}```",
            color=discord.Color.red()
        )
        await interaction.followup.send(embed=error_embed, ephemeral=True)
        print(f"[ERROR] Database repair command failed: {error_details}")


@bot.tree.command(name="fixguildtracking", description="[ADMIN] Auto-track all guilds with tracked members")
async def fixguildtracking(interaction: discord.Interaction):
    """Admin command to automatically set is_tracked=1 for all guilds with tracked members."""
    # Admin check
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to administrators.", ephemeral=True)
        return
    
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    try:
        def fix_guild_tracking():
            """Find all guilds with tracked members and mark them as tracked."""
            with get_db_connection() as conn:
                cursor = conn.cursor()
                
                # Find all guilds that have tracked members but aren't marked as tracked
                cursor.execute('''
                    SELECT DISTINCT um.guild_name, tg.is_tracked
                    FROM user_meta um
                    JOIN tracked_users tu ON LOWER(um.username) = LOWER(tu.username)
                    LEFT JOIN tracked_guilds tg ON um.guild_name = tg.name
                    WHERE tu.is_tracked = 1 AND um.guild_name IS NOT NULL AND um.guild_name != ''
                    ORDER BY um.guild_name
                ''')
                
                guilds_info = cursor.fetchall()
                needs_fix = []
                already_tracked = []
                
                for row in guilds_info:
                    guild_name = row['guild_name']
                    is_tracked = row['is_tracked']
                    
                    if is_tracked != 1:
                        needs_fix.append(guild_name)
                    else:
                        already_tracked.append(guild_name)
                
                # Fix guilds that need tracking
                fixed = []
                for guild_name in needs_fix:
                    # Check if guild exists in tracked_guilds
                    cursor.execute('SELECT name FROM tracked_guilds WHERE name = ?', (guild_name,))
                    exists = cursor.fetchone()
                    
                    if exists:
                        # Update existing entry
                        cursor.execute('UPDATE tracked_guilds SET is_tracked = 1 WHERE name = ?', (guild_name,))
                    else:
                        # Insert new entry
                        cursor.execute('''
                            INSERT INTO tracked_guilds (name, added_at, is_tracked)
                            VALUES (?, strftime('%s', 'now'), 1)
                        ''', (guild_name,))
                    
                    fixed.append(guild_name)
                
                conn.commit()
                return fixed, already_tracked
        
        fixed, already_tracked = await asyncio.to_thread(fix_guild_tracking)
        
        embed = discord.Embed(
            title="🔧 Guild Tracking Fix Results",
            color=discord.Color.blue()
        )
        
        if fixed:
            embed.add_field(
                name=f"✅ Fixed ({len(fixed)} guilds)",
                value="\n".join([f"• {g}" for g in fixed[:10]]) + (f"\n... and {len(fixed) - 10} more" if len(fixed) > 10 else ""),
                inline=False
            )
        
        if already_tracked:
            embed.add_field(
                name=f"ℹ️ Already Tracked ({len(already_tracked)} guilds)",
                value="\n".join([f"• {g}" for g in already_tracked[:10]]) + (f"\n... and {len(already_tracked) - 10} more" if len(already_tracked) > 10 else ""),
                inline=False
            )
        
        if not fixed and not already_tracked:
            embed.description = "No guilds found with tracked members."
        
        embed.set_footer(text="Tracked guilds will now appear in leaderboards and receive snapshot updates.")
        
        await interaction.followup.send(embed=embed, ephemeral=True)
        
    except Exception as e:
        import traceback
        error_details = traceback.format_exc()
        await interaction.followup.send(f"❌ Error: {str(e)}\n```{error_details[:1000]}```", ephemeral=True)
        print(f"[ERROR] fixguildtracking command failed: {error_details}")


@bot.tree.command(name="refresh", description="Manually run batch snapshot update for all tracked users")
@discord.app_commands.describe(mode="One of: session, daily, yesterday, monthly, all, or all+session", ign="Optional: Minecraft IGN to refresh")
@discord.app_commands.choices(mode=[
    discord.app_commands.Choice(name="session", value="session"),
    discord.app_commands.Choice(name="daily", value="daily"),
    discord.app_commands.Choice(name="yesterday", value="yesterday"),
    discord.app_commands.Choice(name="monthly", value="monthly"),
    discord.app_commands.Choice(name="all (daily + yesterday + monthly)", value="all"),
    discord.app_commands.Choice(name="all+session (session + daily + yesterday + monthly)", value="all-session"),
])
async def refresh(interaction: discord.Interaction, mode: discord.app_commands.Choice[str], ign: str = None):
    if not interaction.response.is_done():
        try:
            await interaction.response.defer(ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[REFRESH] Failed to defer interaction: {e}")
            return
    try:
        # If an IGN was supplied, run per-user api_get.py with appropriate flags
        if ign:
            # Only allow creator or the Discord user who claimed the IGN
            allowed = is_admin(interaction.user)
            if not allowed and not is_user_authorized(interaction.user.id, ign):
                await interaction.followup.send(f"[ERROR] You are not authorized to refresh {ign}.", ephemeral=True)
                return

            # Map mode to api_get flags
            mode_map = {
                'session': ['-session'],
                'daily': ['-daily'],
                'yesterday': ['-yesterday'],
                'monthly': ['-monthly'],
                'all': ['-daily', '-yesterday', '-monthly'],
                'all-session': ['-session', '-daily', '-yesterday', '-monthly'],
            }
            flags = mode_map.get(mode.value, [])

            args = ['-ign', ign, *flags]

            # Use batch runner for potentially longer single-user operations
            result = await asyncio.to_thread(run_script_batch, "api_get.py", args)

            if result.returncode == 0:
                msg = f"Refresh completed for {ign} (schedule: {mode.name})"
            else:
                error_msg = result.stderr or result.stdout or "Unknown error"
                print(f"[REFRESH] Single-user refresh FAILED for {ign} - returncode: {result.returncode}")
                print(f"[REFRESH] stdout: {result.stdout[:500] if result.stdout else '(empty)'}")
                print(f"[REFRESH] stderr: {result.stderr[:500] if result.stderr else '(empty)'}")
                msg = f"Refresh failed for {ign}: {error_msg[:300]}"
        else:
            # Run batch_update.py with selected schedule (use extended timeout)
            def run_batch():
                return run_script_batch("batch_update.py", ["-schedule", mode.value])

            result = await asyncio.to_thread(run_batch)

            if result.returncode == 0:
                msg = f"Batch snapshot update completed for schedule: {mode.name}"
            else:
                error_msg = result.stderr or result.stdout or "Unknown error"
                print(f"[REFRESH] Batch update FAILED for schedule {mode.value} - returncode: {result.returncode}")
                print(f"[REFRESH] Full stdout:")
                print(result.stdout if result.stdout else "(empty)")
                print(f"[REFRESH] Full stderr:")
                print(result.stderr if result.stderr else "(empty)")
                msg = f"Batch update failed: {error_msg[:800]}"
        
        # Try to DM the invoking user directly
        try:
            await interaction.user.send(msg)
            try:
                await interaction.followup.send("Sent you a DM with the results.", ephemeral=True)
            except (discord.errors.NotFound, discord.errors.HTTPException):
                print(f"[REFRESH] Interaction expired, but DM was sent to {interaction.user.name}")
        except Exception as dm_error:
            # Fallback to ephemeral if DMs are closed
            print(f"[REFRESH] Failed to send DM: {dm_error}")
            try:
                await interaction.followup.send(msg, ephemeral=True)
            except (discord.errors.NotFound, discord.errors.HTTPException):
                print(f"[REFRESH] Interaction expired, couldn't send results to {interaction.user.name}")
    except subprocess.TimeoutExpired:
        try:
            await interaction.followup.send(f"[ERROR] Batch update timed out after 5 minutes. Try a smaller schedule (e.g., just 'daily' or 'session').", ephemeral=True)
        except (discord.errors.NotFound, discord.errors.HTTPException):
            print(f"[REFRESH] Timeout error but interaction expired")
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        await send_error_with_report(
            interaction,
            "An unexpected error occurred while refreshing the cache.",
            f"{str(e)}\n\n{error_traceback}",
            "/refresh",
            "Cache refresh attempt",
            ephemeral=True
        )

@bot.tree.command(name="fixguilds", description="Admin: Force refresh all guild tags and colors")
async def fixguilds(interaction: discord.Interaction):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return

    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)

    await interaction.followup.send("Starting guild data repair... This may take a while.", ephemeral=True)
    
    try:
        # Run the fix script
        result = await asyncio.to_thread(run_script, "fix_guilds.py", [], timeout=600)
        
        if result.returncode == 0:
            # Force cache refresh so the bot sees the new tags immediately
            await STATS_CACHE.refresh()
            await interaction.followup.send("✅ Guild data repair completed successfully!", ephemeral=True)
        else:
            err = result.stderr or result.stdout or "Unknown error"
            await interaction.followup.send(f"❌ Guild repair failed:\n```{sanitize_output(err[:1000])}```", ephemeral=True)
            
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        await send_error_with_report(
            interaction,
            "An unexpected error occurred during guild data repair.",
            f"{str(e)}\n\n{error_traceback}",
            "/fixguilds",
            "Admin command execution",
            ephemeral=True
        )

# ==============================
# STATS COMMAND HELPERS
# ==============================

async def _send_stats_response(interaction: discord.Interaction, ign: str, user_data: dict, is_tracked: bool, is_initial: bool = False):
    """Helper function to send stats response (used for both cached and fresh data)."""
    EXCEL_FILE = BOT_DIR / "stats.xlsx"
    if not EXCEL_FILE.exists():
        await interaction.followup.send("[ERROR] Excel file not found")
        return

    view = StatsFullView(user_data, ign)
    embed, file = view.generate_full_image("all-time")
    
    # Add note if showing cached data
    cached_note = ""
    if is_initial and not is_tracked:
        cached_note = "⚡ Showing cached data (updating in background)...\n"

    if file:
        if is_tracked:
            message = await interaction.followup.send(view=view, file=file)
            view.message = message  # Store message reference for timeout handling
        else:
            msg = f"{cached_note}`{ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{ign}` to start tracking and enable session/daily/monthly stats."
            await interaction.followup.send(content=msg, file=file)
            # Register user in database for leaderboard accuracy (but don't actively track)
            register_user(ign)
    else:
        if is_tracked:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
        else:
            msg = f"{cached_note}`{ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{ign}` to start tracking and enable session/daily/monthly stats."
            await interaction.followup.send(content=msg, embed=embed)
            # Register user in database for leaderboard accuracy (but don't actively track)
            register_user(ign)


async def _background_update_stats(ign: str, interaction: discord.Interaction):
    """Background task to update stats for untracked users after showing cached data.
    
    After the update completes, edits the original Discord message to show fresh data.
    """
    try:
        print(f"[OPTIMIZE] Background update started for {ign}")
        result = run_script("api_get.py", ["-ign", ign])
        
        # Track API calls
        if result.returncode == 0 and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                if json_data:
                    # Log API calls
                    if 'api_calls' in json_data:
                        api_calls = json_data['api_calls']
                        if api_calls.get('player', 0) > 0:
                            await API_TRACKER.log_request('player')
                        if api_calls.get('guild', 0) > 0:
                            await API_TRACKER.log_request('guild')
                    
                    # Update cache
                    if "processed_stats" in json_data and "username" in json_data:
                        await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                        print(f"[OPTIMIZE] Background update completed for {ign}, updating Discord message")
                        
                        # Get the fresh data and edit the original message
                        cache_data = await STATS_CACHE.get_data()
                        updated_ign = json_data["username"]
                        key = updated_ign.casefold()
                        user_data = None
                        for name, data in cache_data.items():
                            if name.casefold() == key:
                                user_data = data
                                updated_ign = name
                                break
                        
                        if user_data:
                            # Generate fresh stats image/embed
                            EXCEL_FILE = BOT_DIR / "stats.xlsx"
                            if EXCEL_FILE.exists():
                                view = StatsFullView(user_data, updated_ign)
                                embed, file = view.generate_full_image("all-time")
                                
                                # Try to edit the original message
                                try:
                                    # Get the original message
                                    original_messages = [msg async for msg in interaction.channel.history(limit=10)]
                                    for msg in original_messages:
                                        if msg.interaction and msg.interaction.id == interaction.id:
                                            # Found the original response, edit it
                                            if file:
                                                msg_content = f"`{updated_ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{updated_ign}` to start tracking and enable session/daily/monthly stats.\n\n✅ **Updated with fresh data**"
                                                await msg.edit(content=msg_content, attachments=[file])
                                            else:
                                                msg_content = f"`{updated_ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{updated_ign}` to start tracking and enable session/daily/monthly stats.\n\n✅ **Updated with fresh data**"
                                                await msg.edit(content=msg_content, embed=embed)
                                            print(f"[OPTIMIZE] Successfully edited Discord message with fresh data for {ign}")
                                            break
                                except Exception as e:
                                    print(f"[WARNING] Failed to edit Discord message: {e}")
            except Exception as e:
                print(f"[WARNING] Failed to process background update result: {e}")
        else:
            print(f"[WARNING] Background update failed for {ign}: {result.stderr if result.stderr else 'Unknown error'}")
    except Exception as e:
        print(f"[ERROR] Background update exception for {ign}: {e}")


@bot.tree.command(name="stats", description="Get full player stats (Template.xlsx layout) with deltas")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def stats(interaction: discord.Interaction, ign: str = None):
    print(f"[DEBUG] /stats triggered for IGN: {ign} by user: {interaction.user.name} in guild: {interaction.guild.name if interaction.guild else 'DM'}")
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign
    
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[DEBUG] Defer failed for {ign} in /stats: {e}")
            return

    # Try to find in cache first
    cache_data = await STATS_CACHE.get_data()
    user_data = None
    key = ign.casefold()
    for name, data in cache_data.items():
        if name.casefold() == key:
            user_data = data
            ign = name
            break

    # Check if user is tracked
    user_is_tracked = is_tracked_user(ign)
    
    # OPTIMIZATION: Tracked users ALWAYS use cached data (no API call)
    if user_is_tracked and user_data:
        print(f"[OPTIMIZE] User {ign} is tracked, using cached data (no API call)")
        await _send_stats_response(interaction, ign, user_data, user_is_tracked, is_initial=False)
        return
    
    # For untracked users: try to fetch fresh data, fall back to cache if rate limited
    if not user_data:
        try:
            # Fetch fresh stats
            print(f"[DEBUG] Running api_get.py for IGN: {ign} (/stats)")
            result = run_script("api_get.py", ["-ign", ign])
            print(f"[DEBUG] api_get.py returncode (/stats): {result.returncode}")
            print(f"[DEBUG] api_get.py stdout (/stats): {result.stdout if result.stdout else 'None'}")
            print(f"[DEBUG] api_get.py stderr (/stats): {result.stderr if result.stderr else 'None'}")

            if result.returncode != 0:
                # Check if it's a rate limit in stderr or stdout
                stdout_msg = result.stdout or ""
                stderr_msg = result.stderr or ""
                is_rate_limited = False
                
                if "429" in stderr_msg:
                    is_rate_limited = True
                
                # Also check JSON in stdout for rate limit
                if not is_rate_limited:
                    try:
                        for line in reversed(stdout_msg.splitlines()):
                            line = line.strip()
                            if line.startswith('{') and line.endswith('}'):
                                try:
                                    json_data = json.loads(line)
                                    if json_data.get("skipped") and json_data.get("reason") == "rate_limited":
                                        is_rate_limited = True
                                        break
                                except json.JSONDecodeError:
                                    continue
                    except Exception:
                        pass
                
                if is_rate_limited:
                    print(f"[DEBUG] Rate limited for {ign} (/stats), attempting to use existing data")
                    # Don't return - continue to try to use cached data below
                else:
                    stdout_msg = result.stdout or ""
                    stderr_msg = result.stderr or ""
                    full_output = f"STDOUT:\n{stdout_msg}\n\nSTDERR:\n{stderr_msg}"
                    
                    # Try to parse JSON error from stdout
                    error_details = None
                    try:
                        for line in reversed(stdout_msg.splitlines()):
                            line = line.strip()
                            if line.startswith('{') and line.endswith('}'):
                                try:
                                    json_data = json.loads(line)
                                    if json_data.get("skipped") and json_data.get("error"):
                                        error_details = json_data.get("error")
                                        break
                                except json.JSONDecodeError:
                                    continue
                    except Exception:
                        pass
                    
                    # Check for specific error types
                    user_msg = None
                    if "never played" in full_output.lower() or "no wool games data" in full_output.lower():
                        await interaction.followup.send(
                            f"`{ign}` has no Wool Games data. They have either never played the game (they are missing out) or on stat freeze. Alternatively, chuck managed to corrupt the database again."
                        )
                        return
                    elif error_details:
                        # Parse the error details for user-friendly message
                        if "403" in error_details or "Forbidden" in error_details:
                            await interaction.followup.send("Invalid API key. Chuck is slow and probably broke something.")
                            return
                        elif "429" in error_details or "Rate" in error_details:
                            user_msg = f"Too many requests to the Hypixel API. Please wait a moment and try again.\n\n**Technical details:** {error_details}"
                        elif "404" in error_details:
                            user_msg = f"Player `{ign}` not found on Hypixel.\n\n**Technical details:** {error_details}"
                        else:
                            user_msg = f"API Error: {error_details}"
                    else:
                        user_msg = "Failed to fetch player statistics. The Hypixel API may be unavailable or the player might not exist."
                    
                    await send_error_with_report(
                        interaction,
                        user_msg,
                        full_output,
                        "/stats",
                        f"IGN: {ign}"
                    )
                    return

            # Optimistically update cache if we have JSON output
            try:
                if result.stdout:
                    # Try to find the JSON object in the output (ignoring debug logs)
                    json_data = None
                    for line in reversed(result.stdout.splitlines()):
                        line = line.strip()
                        if line.startswith('{') and line.endswith('}'):
                            try:
                                json_data = json.loads(line)
                                break
                            except json.JSONDecodeError:
                                continue
                    
                    # Check if the API request was skipped due to an error
                    if json_data and json_data.get("skipped"):
                        error_msg = json_data.get("error", "Unknown error")
                        reason = json_data.get("reason", "unknown")
                        
                        # RATE LIMITED - Use cached data instead of showing error
                        if reason == "rate_limited":
                            print(f"[DEBUG] Rate limited for {ign}, using cached data")
                            # Refresh cache to get latest data (snapshots were written)
                            cache_data = await STATS_CACHE.get_data()
                            user_data = None
                            key = ign.casefold()
                            for name, data in cache_data.items():
                                if name.casefold() == key:
                                    user_data = data
                                    ign = name
                                    break
                            # Continue to show cached data (don't return here)
                        elif "403" in error_msg or "Forbidden" in error_msg:
                            await interaction.followup.send("Invalid API key. Chuck is slow and probably broke something.")
                            return
                        elif reason == "api_error":
                            user_msg = f"API Error occurred while fetching data.\n\n**Technical details:** {error_msg}"
                            await send_error_with_report(
                                interaction,
                                user_msg,
                                f"STDOUT:\n{result.stdout}\n\nSTDERR:\n{result.stderr or 'None'}",
                                "/stats",
                                f"IGN: {ign}"
                            )
                            return
                        else:
                            user_msg = f"Unable to fetch data for `{ign}`.\n\n**Technical details:** {error_msg}"
                            await send_error_with_report(
                                interaction,
                                user_msg,
                                f"STDOUT:\n{result.stdout}\n\nSTDERR:\n{result.stderr or 'None'}",
                                "/stats",
                                f"IGN: {ign}"
                            )
                            return
                    
                    if json_data and "processed_stats" in json_data and "username" in json_data:
                        await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                        # Update ign to the proper case returned by API
                        ign = json_data["username"]
            except Exception as e:
                print(f"[WARNING] Failed to update cache from output: {e}")

            # Refresh cache data after api_get
            cache_data = await STATS_CACHE.get_data()
            user_data = None
            key = ign.casefold()
            for name, data in cache_data.items():
                if name.casefold() == key:
                    user_data = data
                    ign = name
                    break
        except subprocess.TimeoutExpired:
            await send_error_with_report(
                interaction,
                "The command took too long to complete. The Hypixel API might be slow or unresponsive.",
                "Command timed out (30s limit)",
                "/stats",
                f"IGN: {ign}"
            )
            return
        except Exception as e:
            import traceback
            error_traceback = traceback.format_exc()
            await send_error_with_report(
                interaction,
                "An unexpected error occurred while fetching stats. Please try again later.",
                f"{str(e)}\n\n{error_traceback}",
                "/stats",
                f"IGN: {ign}"
            )
            return
    
    if not user_data:
        # Check if this is because they have no Wool Games data
        # This can happen if api_get ran but the player has no stats
        await interaction.followup.send(
            f"`{ign}` has no Wool Games data. They have either never played the game (they are missing out) or on stat freeze. Alternatively, chuck managed to corrupt the database again."
        )
        return

    # Send stats using helper function
    await _send_stats_response(interaction, ign, user_data, is_tracked_user(ign), is_initial=False)


@bot.tree.command(name="streak", description="View current win/kill streaks (approved users)")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def streak(interaction: discord.Interaction, ign: str = None):
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign

    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    try:
        # Ensure user is cached (fetches if needed)
        cached, actual_ign = await ensure_user_cached(ign)
        if not cached:
            await interaction.followup.send(f"[ERROR] Could not find or fetch data for {ign}")
            return

        cache_data = await STATS_CACHE.get_data()
        key = actual_ign.casefold()
        user_data = None
        for name, data in cache_data.items():
            if name.casefold() == key:
                user_data = data
                actual_ign = name
                break

        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{actual_ign}' not found")
            return

        # Require user to be tracked (UUID-aware)
        if not is_tracked_user(actual_ign):
            await interaction.followup.send(f"`{actual_ign}` is not currently tracked. Use `/track ign:{actual_ign}` first.")
            return

        streaks = load_tracked_streaks()
        entry_key = None
        for k in streaks.keys():
            if k.casefold() == actual_ign.casefold():
                entry_key = k
                break

        if entry_key:
            entry = streaks.get(entry_key, {})
            winstreak = int(entry.get("winstreak", 0))
            killstreak = int(entry.get("killstreak", 0))

            meta = user_data.get("meta", {})
            level = meta.get("level", 0)
            icon = meta.get("icon", "")
            ign_color = meta.get("ign_color")
            guild_tag = meta.get("guild_tag")
            guild_color = meta.get("guild_hex")

            if Image is not None:
                try:
                    img_io = create_streaks_image(actual_ign, level, icon, ign_color, guild_tag, guild_color, winstreak, killstreak)
                    filename = f"{actual_ign}_streaks.png"
                    await interaction.followup.send(file=discord.File(img_io, filename=filename))
                    return
                except Exception as e:
                    print(f"[STREAK] Failed to render streaks image: {e}")

            embed = discord.Embed(title=f"{actual_ign} Streaks")
            embed.add_field(name="Current Winstreak", value=f"```{winstreak}```", inline=True)
            embed.add_field(name="Current Killstreak", value=f"```{killstreak}```", inline=True)
            await interaction.followup.send(embed=embed)
        else:
            message = (
                f"{actual_ign} does not have streaks tracked. "
                f"Click the button to request to track {actual_ign}'s streaks."
            )
            view = StreakRequestView(actual_ign, interaction.user, user_data.get("stats", {}))
            await interaction.followup.send(content=message, view=view)

    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="streak-remove", description="Stop tracking streaks for a user (Admin or self)")
@discord.app_commands.describe(ign="Minecraft IGN to stop tracking streaks for")
async def streak_remove(interaction: discord.Interaction, ign: str):
    """Remove a user from streak tracking.
    
    Admins can remove anyone. Regular users can only remove themselves.
    """
    # Validate username
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign
    
    # Check permissions
    is_user_admin = is_admin(interaction.user)
    
    # If not admin, check if user is trying to remove themselves
    if not is_user_admin:
        # Get user's linked IGN or default
        user_linked_ign = None
        user_links = load_user_links()
        for linked_ign, discord_id in user_links.items():
            if str(discord_id) == str(interaction.user.id):
                user_linked_ign = linked_ign
                break
        
        if not user_linked_ign:
            user_linked_ign = get_default_user(interaction.user.id)
        
        # Check if they're trying to remove themselves
        if not user_linked_ign or user_linked_ign.lower() != ign.lower():
            await interaction.response.send_message(
                "❌ You can only remove streak tracking for your own account.\n"
                f"Admins can remove streak tracking for any user.",
                ephemeral=True
            )
            return
    
    if not interaction.response.is_done():
        await interaction.response.defer(ephemeral=True)
    
    try:
        # Check if user has streak tracking
        streaks = load_tracked_streaks()
        entry_key = None
        for k in streaks.keys():
            if k.casefold() == ign.casefold():
                entry_key = k
                break
        
        if not entry_key:
            await interaction.followup.send(
                f"❌ `{ign}` does not have streak tracking enabled.",
                ephemeral=True
            )
            return
        
        # Remove from streak tracking
        del streaks[entry_key]
        with get_db_connection() as conn:
            cursor = conn.cursor()
            cursor.execute('DELETE FROM tracked_streaks WHERE LOWER(username) = LOWER(?)', (ign,))
            conn.commit()
        
        await interaction.followup.send(
            f"✅ Removed streak tracking for `{ign}`.\n\n"
            f"💡 This user will now receive updates every **10 minutes** instead of every **60 seconds**, "
            f"significantly reducing API usage.",
            ephemeral=True
        )
        
        print(f"[STREAK_REMOVE] {interaction.user.name} removed streak tracking for {ign}")
        
    except Exception as e:
        await interaction.followup.send(f"❌ [ERROR] {str(e)}", ephemeral=True)


# Helper classes for /layout command
class LayoutKitSelect(discord.ui.Select):
    """Dropdown for selecting Wool Wars kit."""
    def __init__(self, view, kits):
        options = [discord.SelectOption(label=kit, value=kit, default=(i==0)) for i, kit in enumerate(kits)]
        super().__init__(
            placeholder="Select kit...",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="layout_kit_select",
        )
        self.view_ref = view

    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref.handle_kit_change(interaction, selected)


class LayoutTabView(discord.ui.View):
    """View for hotbar layout display with game tabs."""
    def __init__(self, username, layouts):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.username = username
        self.layouts = layouts
        self.current_game = "sheep_wars"
        self.current_kit = None
        self.kit_selector = None
        self.message = None  # Store message reference for timeout handling
        
        # Organize layouts by game
        self.game_layouts = {}
        for layout in layouts:
            game = layout['game']
            if game not in self.game_layouts:
                self.game_layouts[game] = []
            self.game_layouts[game].append(layout)
        
        # Initialize kit for wool wars if it has multiple kits
        if 'wool_wars' in self.game_layouts:
            ww_layouts = self.game_layouts['wool_wars']
            # Filter out NULL kits
            non_null_kits = [l for l in ww_layouts if l.get('kit') != 'NULL']
            if non_null_kits:
                self.current_kit = non_null_kits[0]['kit']
        
        self.update_button_styles()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible

    def update_button_styles(self):
        """Set active button to primary style, others to secondary."""
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                if child.custom_id == self.current_game:
                    child.style = discord.ButtonStyle.primary
                else:
                    child.style = discord.ButtonStyle.secondary

    def add_kit_selector_if_needed(self):
        """Add kit selector only for wool wars tab with multiple kits."""
        # Remove existing kit selector if present
        if self.kit_selector:
            self.remove_item(self.kit_selector)
            self.kit_selector = None
        
        # Add kit selector only if on wool wars tab with multiple kits
        if self.current_game == 'wool_wars' and 'wool_wars' in self.game_layouts:
            ww_layouts = self.game_layouts['wool_wars']
            non_null_kits = [l for l in ww_layouts if l.get('kit') != 'NULL']
            if len(non_null_kits) > 1:
                kit_names = [l['kit'] for l in non_null_kits]
                self.kit_selector = LayoutKitSelect(self, kit_names)
                self.add_item(self.kit_selector)

    def format_layout_embed(self, game):
        """Format the layout as a Discord embed."""
        game_display = game.replace('_', ' ').title()
        
        if game not in self.game_layouts:
            embed = discord.Embed(
                title=f"{self.username}'s Hotbar Layout",
                description=f"No {game_display} layout data available.",
                color=0x808080
            )
            return embed
        
        layouts = self.game_layouts[game]
        
        # For wool wars with kit selection
        if game == 'wool_wars' and self.current_kit:
            layout = next((l for l in layouts if l.get('kit') == self.current_kit), None)
            if not layout:
                layout = layouts[0]
        else:
            layout = layouts[0]
        
        # Check if layout has any data
        has_data = any(layout.get(f'slot_{i}') for i in range(9))
        if not has_data:
            embed = discord.Embed(
                title=f"{self.username}'s Hotbar Layout",
                description=f"No {game_display} layout data available.",
                color=0x808080
            )
            return embed
        
        # Create embed
        kit_name = layout.get('kit', 'N/A')
        kit_display = f" - {kit_name}" if kit_name and kit_name != 'NULL' else ""
        
        embed = discord.Embed(
            title=f"{self.username}'s Hotbar Layout",
            description=f"**Game:** {game_display}{kit_display}",
            color=0x00AAFF
        )
        
        # Add slots as fields
        for i in range(9):
            slot_value = layout.get(f'slot_{i}')
            if slot_value:
                embed.add_field(name=f"Slot {i+1}", value=slot_value, inline=True)
            else:
                embed.add_field(name=f"Slot {i+1}", value="[Empty]", inline=True)
        
        return embed

    async def handle_tab_click(self, interaction: discord.Interaction, game: str):
        self.current_game = game
        self.update_button_styles()
        
        # Reset kit selection for wool wars
        if game == 'wool_wars' and 'wool_wars' in self.game_layouts:
            ww_layouts = self.game_layouts['wool_wars']
            non_null_kits = [l for l in ww_layouts if l.get('kit') != 'NULL']
            if non_null_kits:
                self.current_kit = non_null_kits[0]['kit']
        
        # Update kit selector visibility
        self.add_kit_selector_if_needed()
        
        embed = self.format_layout_embed(game)
        await interaction.response.edit_message(embed=embed, view=self)

    async def handle_kit_change(self, interaction: discord.Interaction, kit: str):
        self.current_kit = kit
        embed = self.format_layout_embed(self.current_game)
        await interaction.response.edit_message(embed=embed, view=self)

    @discord.ui.button(label="Sheep Wars", custom_id="sheep_wars")
    async def sheep_wars(self, interaction, button):
        await self.handle_tab_click(interaction, "sheep_wars")

    @discord.ui.button(label="Wool Wars", custom_id="wool_wars")
    async def wool_wars(self, interaction, button):
        await self.handle_tab_click(interaction, "wool_wars")

    @discord.ui.button(label="Capture The Wool", custom_id="capture_the_wool")
    async def capture_the_wool(self, interaction, button):
        await self.handle_tab_click(interaction, "capture_the_wool")


@bot.tree.command(name="layout", description="View hotbar layouts for a player")
@discord.app_commands.describe(ign="Minecraft IGN (required)")
async def layout(interaction: discord.Interaction, ign: str):
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    try:
        # Check if user is tracked - if so, get from database
        needs_fetch = False
        if is_tracked_user(proper_ign):
            layouts = get_hotbar_layouts(proper_ign)
            if not layouts:
                # Tracked but no layout data - fetch it
                needs_fetch = True
        else:
            # Not tracked - fetch data
            needs_fetch = True
        
        if needs_fetch:
            # Send initial status message
            status_msg = await interaction.followup.send(f"Fetching layout data for {proper_ign}...", wait=True)
            result = await asyncio.to_thread(run_script, "api_get.py", ["-ign", proper_ign], timeout=30)
            layouts = get_hotbar_layouts(proper_ign)
            
            if not layouts:
                await status_msg.edit(content=f"No layout data found for {proper_ign}.")
                return
            
            # Create view and edit the message with the data
            view = LayoutTabView(proper_ign, layouts)
            embed = view.format_layout_embed("sheep_wars")
            message = await status_msg.edit(content=None, embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
        else:
            # Data already available
            view = LayoutTabView(proper_ign, layouts)
            embed = view.format_layout_embed("sheep_wars")
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling

    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="compare", description="Compare stats between two players")
@discord.app_commands.describe(
    ign1="First Minecraft IGN",
    ign2="Second Minecraft IGN",
    stat="Specific stat to compare (optional - shows all stats if not specified)"
)
@discord.app_commands.choices(stat=[
    discord.app_commands.Choice(name="Kills", value="kills"),
    discord.app_commands.Choice(name="Deaths", value="deaths"),
    discord.app_commands.Choice(name="K/D Ratio", value="kdr"),
    discord.app_commands.Choice(name="Wins", value="wins"),
    discord.app_commands.Choice(name="Losses", value="losses"),
    discord.app_commands.Choice(name="W/L Ratio", value="wlr"),
    discord.app_commands.Choice(name="Damage Dealt", value="damage"),
    discord.app_commands.Choice(name="Games Played", value="games_played"),
    discord.app_commands.Choice(name="Sheep Thrown", value="sheeps_thrown"),
    discord.app_commands.Choice(name="Magic Wool Hit", value="magic_wools"),
    discord.app_commands.Choice(name="Void Kills", value="void_kills"),
    discord.app_commands.Choice(name="Explosive Kills", value="explosive_kills"),
    discord.app_commands.Choice(name="Melee Kills", value="melee_kills"),
    discord.app_commands.Choice(name="Bow Kills", value="bow_kills"),
    discord.app_commands.Choice(name="Void Deaths", value="void_deaths"),
    discord.app_commands.Choice(name="Explosive Deaths", value="explosive_deaths"),
    discord.app_commands.Choice(name="Melee Deaths", value="melee_deaths"),
    discord.app_commands.Choice(name="Bow Deaths", value="bow_deaths"),
    discord.app_commands.Choice(name="Void K/D Ratio", value="void_kdr"),
    discord.app_commands.Choice(name="Explosive K/D Ratio", value="explosive_kdr"),
    discord.app_commands.Choice(name="Bow K/D Ratio", value="bow_kdr"),
    discord.app_commands.Choice(name="Melee K/D Ratio", value="melee_kdr"),
    discord.app_commands.Choice(name="Kills per Game", value="kills_per_game"),
    discord.app_commands.Choice(name="Damage per Game", value="damage_per_game"),
    discord.app_commands.Choice(name="Wools per Game", value="wools_per_game"),
])
async def compare(interaction: discord.Interaction, ign1: str, ign2: str, stat: str = None):
    print(f"[DEBUG] /compare triggered for IGN1: {ign1}, IGN2: {ign2}, stat: {stat} by user: {interaction.user.name} in guild: {interaction.guild.name if interaction.guild else 'DM'}")
    
    # Validate both usernames
    ok1, proper_ign1 = validate_and_normalize_ign(ign1)
    if not ok1:
        await interaction.response.send_message(f"The username {ign1} is invalid.", ephemeral=True)
        return
    ign1 = proper_ign1
    
    ok2, proper_ign2 = validate_and_normalize_ign(ign2)
    if not ok2:
        await interaction.response.send_message(f"The username {ign2} is invalid.", ephemeral=True)
        return
    ign2 = proper_ign2
    
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[DEBUG] Defer failed for {ign1} vs {ign2} in /compare: {e}")
            return

    try:
        # Use cached data for fast comparison
        cache_data = await STATS_CACHE.get_data()
        
        # Find users in cache case-insensitively
        key1 = ign1.casefold()
        user_data1 = None
        actual_ign1 = ign1
        for name, data in cache_data.items():
            if name.casefold() == key1:
                user_data1 = data
                actual_ign1 = name
                break
        
        key2 = ign2.casefold()
        user_data2 = None
        actual_ign2 = ign2
        for name, data in cache_data.items():
            if name.casefold() == key2:
                user_data2 = data
                actual_ign2 = name
                break
        
        # If either player not in cache, try to fetch them
        if not user_data1:
            print(f"[DEBUG] {ign1} not in cache, fetching...")
            cached1, actual_ign1 = await ensure_user_cached(ign1)
            if not cached1:
                await interaction.followup.send(f"[ERROR] Could not find or fetch data for {ign1}")
                return
            # Reload cache after fetch
            cache_data = await STATS_CACHE.get_data()
            for name, data in cache_data.items():
                if name.casefold() == actual_ign1.casefold():
                    user_data1 = data
                    print(f"[DEBUG] Found {actual_ign1} in cache after fetch. Has {len(data.get('stats', {}))} stats.")
                    break
        
        if not user_data2:
            print(f"[DEBUG] {ign2} not in cache, fetching...")
            cached2, actual_ign2 = await ensure_user_cached(ign2)
            if not cached2:
                await interaction.followup.send(f"[ERROR] Could not find or fetch data for {ign2}")
                return
            # Reload cache after fetch
            cache_data = await STATS_CACHE.get_data()
            for name, data in cache_data.items():
                if name.casefold() == actual_ign2.casefold():
                    user_data2 = data
                    print(f"[DEBUG] Found {actual_ign2} in cache after fetch. Has {len(data.get('stats', {}))} stats.")
                    break
        
        if not user_data1:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign1}' not found")
            return
        
        if not user_data2:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign2}' not found")
            return

        # Check if users are tracked (UUID-aware)
        is_tracked1 = is_tracked_user(actual_ign1)
        is_tracked2 = is_tracked_user(actual_ign2)

        view = CompareView(user_data1, user_data2, actual_ign1, actual_ign2, stat=stat)
        embed, file = view.generate_compare_image("all-time")

        warning_msg = ""
        if not is_tracked1 and not is_tracked2:
            warning_msg = f"`{actual_ign1}` and `{actual_ign2}` are not currently tracked. Only all-time stats are available.\nUse `/track` to start tracking and enable session/daily/monthly stats."
        elif not is_tracked1:
            warning_msg = f"`{actual_ign1}` is not currently tracked. Only all-time stats are available for them.\nUse `/track ign:{actual_ign1}` to start tracking."
        elif not is_tracked2:
            warning_msg = f"`{actual_ign2}` is not currently tracked. Only all-time stats are available for them.\nUse `/track ign:{actual_ign2}` to start tracking."

        if file:
            if warning_msg:
                message = await interaction.followup.send(content=warning_msg, file=file, view=view)
            else:
                message = await interaction.followup.send(view=view, file=file)
            view.message = message  # Store message reference for timeout handling
        else:
            if warning_msg:
                await interaction.followup.send(content=warning_msg, embed=embed, view=view)
            else:
                await interaction.followup.send(embed=embed, view=view)

        # Schedule cleanup for untracked users
        if not is_tracked1:
            # Register user in database for leaderboard accuracy (but don't actively track)
            register_user(actual_ign1)
        if not is_tracked2:
            # Register user in database for leaderboard accuracy (but don't actively track)
            register_user(actual_ign2)

    except subprocess.TimeoutExpired:
        await send_error_with_report(
            interaction,
            "The command took too long to complete. The Hypixel API might be slow.",
            "Command timed out (30s limit)",
            "/compare",
            f"Player 1: {ign1}, Player 2: {ign2}"
        )
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        await send_error_with_report(
            interaction,
            "An unexpected error occurred while comparing players.",
            f"{str(e)}\n\n{error_traceback}",
            "/compare",
            f"Player 1: {ign1}, Player 2: {ign2}"
        )


@bot.tree.command(name="killdistribution", description="View kill-type distribution as a pie chart")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def killdistribution(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    user_data = None
    try:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode != 0 and not (result.stderr and "429" in result.stderr):
            error_msg = result.stderr or result.stdout or "Unknown error"
            await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
            return

        # Optimistically update cache
        try:
            if result.stdout:
                data = json.loads(result.stdout)
                if "processed_stats" in data and "username" in data:
                    user_data = await STATS_CACHE.update_cache_entry(data["username"], data["processed_stats"])
                    ign = data["username"]
        except Exception as e:
            print(f"[WARNING] Failed to update cache from output: {e}")

        EXCEL_FILE = BOT_DIR / "stats.xlsx"
        if not EXCEL_FILE.exists():
            await interaction.followup.send("[ERROR] Excel file not found")
            return

        if not user_data:
            cache_data = await STATS_CACHE.get_data()
            key = ign.casefold()
            actual_ign = ign
            for name, data in cache_data.items():
                if name.casefold() == key:
                    user_data = data
                    actual_ign = name
                    break
        else:
            actual_ign = ign
        
        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return

        view = DistributionView(user_data, actual_ign, mode="kill")
        embed, file = view.generate_distribution("all-time")

        if file:
            message = await interaction.followup.send(file=file, view=view)
            view.message = message  # Store message reference for timeout handling
        else:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")


@bot.tree.command(name="deathdistribution", description="View death-type distribution as a pie chart")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def deathdistribution(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return

    user_data = None
    try:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode != 0 and not (result.stderr and "429" in result.stderr):
            error_msg = result.stderr or result.stdout or "Unknown error"
            await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
            return

        # Optimistically update cache
        try:
            if result.stdout:
                data = json.loads(result.stdout)
                if "processed_stats" in data and "username" in data:
                    user_data = await STATS_CACHE.update_cache_entry(data["username"], data["processed_stats"])
                    ign = data["username"]
        except Exception as e:
            print(f"[WARNING] Failed to update cache from output: {e}")

        EXCEL_FILE = BOT_DIR / "stats.xlsx"
        if not EXCEL_FILE.exists():
            await interaction.followup.send("[ERROR] Excel file not found")
            return

        if not user_data:
            cache_data = await STATS_CACHE.get_data()
            key = ign.casefold()
            actual_ign = ign
            for name, data in cache_data.items():
                if name.casefold() == key:
                    user_data = data
                    actual_ign = name
                    break
        else:
            actual_ign = ign
        
        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return

        view = DistributionView(user_data, actual_ign, mode="death")
        embed, file = view.generate_distribution("all-time")

        if file:
            message = await interaction.followup.send(file=file, view=view)
            view.message = message  # Store message reference for timeout handling
        else:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
    except subprocess.TimeoutExpired:
        await interaction.followup.send("[ERROR] Command timed out (30s limit)")
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

@bot.tree.command(name="ratios", description="Predict when you'll reach the next WLR and KDR milestones")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def ratios(interaction: discord.Interaction, ign: str = None):
    """Display milestone predictions for Win/Loss and Kill/Death ratios."""
    print(f"[DEBUG] /ratios triggered for IGN: {ign} by user: {interaction.user.name} in guild: {interaction.guild.name if interaction.guild else 'DM'}")
    
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign
    
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[DEBUG] Defer failed for {ign} in /ratios: {e}")
            return
    
    try:
        # Fetch fresh stats
        print(f"[DEBUG] Running api_get.py for IGN: {ign} (/ratios)")
        result = run_script("api_get.py", ["-ign", ign])
        print(f"[DEBUG] api_get.py returncode (/ratios): {result.returncode}")
        
        if result.returncode != 0:
            if result.stderr and "429" in result.stderr:
                print(f"[DEBUG] Rate limited for {ign} (/ratios), attempting to use existing data")
            else:
                error_msg = result.stderr or result.stdout or "Unknown error"
                await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
                return
        
        # Optimistically update cache if we have JSON output
        try:
            if result.stdout:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                    ign = json_data["username"]
        except Exception as e:
            print(f"[WARNING] Failed to update cache from output: {e}")
        
        # Get user data from cache
        cache_data = await STATS_CACHE.get_data()
        
        # Find user in cache case-insensitively
        key = ign.casefold()
        user_data = None
        actual_ign = ign
        for name, data in cache_data.items():
            if name.casefold() == key:
                user_data = data
                actual_ign = name
                break
        
        if not user_data:
            await interaction.followup.send(f"[ERROR] Player sheet '{ign}' not found")
            return
        
        # Check if user is tracked (UUID-aware)
        is_tracked = is_tracked_user(actual_ign)
        
        # Create view and generate image
        view = RatiosView(user_data, actual_ign)
        file = view.generate_ratios_image("all-time")
        
        if is_tracked:
            message = await interaction.followup.send(file=file, view=view)
            view.message = message  # Store message reference for timeout handling
        else:
            msg = f"`{actual_ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{actual_ign}` to start tracking and enable session/daily/monthly stats."
            await interaction.followup.send(content=msg, file=file)
            # Register user in database for leaderboard accuracy (but don't actively track)
            register_user(actual_ign)
    
    except subprocess.TimeoutExpired:
        await send_error_with_report(
            interaction,
            "The command took too long to complete.",
            "Command timed out (30s limit)",
            "/ratios",
            f"IGN: {ign}"
        )
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        await send_error_with_report(
            interaction,
            "An unexpected error occurred while calculating ratio predictions.",
            f"{str(e)}\n\n{error_traceback}",
            "/ratios",
            f"IGN: {ign}"
        )

@bot.tree.command(name="sheepwars", description="Get player stats with deltas")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def sheepwars(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign

    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    EXCEL_FILE = BOT_DIR / "stats.xlsx"
    if not EXCEL_FILE.exists():
        await interaction.followup.send("Stats file not found.")
        return

    # Try to find in cache first
    cache_data = await STATS_CACHE.get_data()
    user_data = None
    key = ign.casefold()
    for name, data in cache_data.items():
        if name.casefold() == key:
            user_data = data
            ign = name
            break

    # Check if user is tracked
    user_is_tracked = is_tracked_user(ign)
    
    # OPTIMIZATION: Tracked users ALWAYS use cached data (no API call)
    if user_is_tracked and user_data:
        print(f"[OPTIMIZE] User {ign} is tracked in /sheepwars, using cached data (no API call)")
        # Continue to display cached data below
    # For untracked users: only fetch if not in cache
    elif not user_data:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode == 0 and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                # Check if the API request was skipped due to an error
                if json_data and json_data.get("skipped"):
                    error_msg = json_data.get("error", "Unknown error")
                    reason = json_data.get("reason", "unknown")
                    
                    # RATE LIMITED - Use cached data instead of showing error
                    if reason == "rate_limited":
                        print(f"[DEBUG] Rate limited for {ign} in /sheepwars, using cached data")
                        # Refresh cache to get latest data (snapshots were written)
                        cache_data = await STATS_CACHE.get_data()
                        user_data = None
                        key = ign.casefold()
                        for name, data in cache_data.items():
                            if name.casefold() == key:
                                user_data = data
                                ign = name
                                break
                        # Continue to show cached data (don't return here)
                    else:
                        # Other errors - show error message
                        stdout_msg = result.stdout or ""
                        stderr_msg = result.stderr or ""
                        full_output = f"STDOUT:\n{stdout_msg}\n\nSTDERR:\n{stderr_msg}"
                        
                        if "403" in error_msg or "Forbidden" in error_msg:
                            await interaction.followup.send("Invalid API key. Chuck is slow and probably broke something.")
                            return
                        elif "404" in error_msg:
                            user_msg = f"Player `{ign}` not found on Hypixel.\n\n**Technical details:** {error_msg}"
                        elif reason == "api_error":
                            user_msg = f"API Error occurred while fetching data.\n\n**Technical details:** {error_msg}"
                        else:
                            user_msg = f"Unable to fetch data for `{ign}`.\n\n**Technical details:** {error_msg}"
                        
                        await send_error_with_report(
                            interaction,
                            user_msg,
                            full_output,
                            "/sheepwars",
                            f"IGN: {ign}"
                        )
                        return
                
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    user_data = await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                    ign = json_data["username"]
            except Exception as e:
                print(f"[WARNING] Failed to parse api_get output in sheepwars: {e}")
        elif result.returncode != 0:
            # If api_get failed, log it and potentially inform user
            error_msg = result.stderr or result.stdout or "Unknown error"
            print(f"[ERROR] api_get failed for {ign} in sheepwars: {error_msg}")

    if not user_data:
        await interaction.followup.send(
            f"`{ign}` has no Wool Games data. They have either never played the game (they are missing out) or on stat freeze. Alternatively, chuck managed to corrupt the database again."
        )
        return

    # Build all_data from cache
    all_data = {}
    stats = user_data.get("stats", {})
    
    # Map tab names to cache keys
    tab_map = {
        "all-time": "lifetime",
        "session": "session",
        "daily": "daily",
        "yesterday": "yesterday",
        "weekly": "weekly",
        "monthly": "monthly"
    }
    
    for tab, cache_key in tab_map.items():
        w = stats.get("wins", {}).get(cache_key, 0)
        l = stats.get("losses", {}).get(cache_key, 0)
        k = stats.get("kills", {}).get(cache_key, 0)
        d = stats.get("deaths", {}).get(cache_key, 0)
        p = stats.get("playtime", {}).get(cache_key, 0)
        
        all_data[tab] = {
            'wins': w, 'losses': l, 'kills': k, 'deaths': d, 'playtime': p,
            'wlr': w/l if l > 0 else w,
            'kdr': k/d if d > 0 else k
        }

    meta = user_data.get("meta", {})
    level = meta.get("level", 0)
    icon = meta.get("icon", "")
    ign_color = meta.get("ign_color")
    guild_tag = meta.get("guild_tag")
    guild_hex = meta.get("guild_hex")
    
    # Get real-time status
    # Parallelize status and skin fetching to reduce load time
    status_task = asyncio.to_thread(get_player_status, ign)
    skin_task = asyncio.to_thread(get_player_body, ign)
    (status_text, status_color), skin_image = await asyncio.gather(status_task, skin_task)
    
    view = StatsTabView(all_data, ign, int(level), icon, 
                        ign_color=ign_color, guild_tag=guild_tag, guild_hex=guild_hex,
                        status_text=status_text, status_color=status_color, skin_image=skin_image)
    
    # Check if tracked (UUID-aware)
    is_tracked = is_tracked_user(ign)
    
    file = view.generate_composite_image("all-time")
    if is_tracked:
        message = await interaction.followup.send(file=file, view=view)
        view.message = message  # Store message reference for timeout handling
    else:
        msg = f"`{ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{ign}` to start tracking and enable session/daily/monthly stats."
        await interaction.followup.send(content=msg, file=file)
        # Register user in database for leaderboard accuracy (but don't actively track)
        register_user(ign)

@bot.tree.command(name="ww", description="Get Wool Wars stats with class breakdown")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def ww(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign

    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    EXCEL_FILE = BOT_DIR / "stats.xlsx"
    if not EXCEL_FILE.exists():
        await interaction.followup.send("Stats file not found.")
        return

    # Try to find in cache first
    cache_data = await STATS_CACHE.get_data()
    user_data = None
    key = ign.casefold()
    for name, data in cache_data.items():
        if name.casefold() == key:
            user_data = data
            ign = name
            break

    # If not in cache, fetch fresh stats
    if not user_data:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode == 0 and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    user_data = await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                    ign = json_data["username"]
            except Exception as e:
                print(f"[WARNING] Failed to parse api_get output in ww: {e}")
        elif result.returncode != 0:
            print(f"[ERROR] api_get failed for {ign} in ww: {result.stderr}")

    if not user_data:
        await interaction.followup.send("Player not found in database or API error.")
        return

    # Build all_data from cache - WW stats
    all_data = {}
    stats = user_data.get("stats", {})
    
    # Map tab names to cache keys
    tab_map = {
        "all-time": "lifetime",
        "session": "session",
        "daily": "daily",
        "yesterday": "yesterday",
        "weekly": "weekly",
        "monthly": "monthly"
    }
    
    for tab, cache_key in tab_map.items():
        w = stats.get("ww_wins", {}).get(cache_key, 0)
        g = stats.get("ww_games_played", {}).get(cache_key, 0)
        k = stats.get("ww_kills", {}).get(cache_key, 0)
        d = stats.get("ww_deaths", {}).get(cache_key, 0)
        a = stats.get("ww_assists", {}).get(cache_key, 0)
        p = stats.get("playtime", {}).get(cache_key, 0)
        
        # Class-specific stats
        tab_data = {
            'wins': w, 'games': g, 'kills': k, 'deaths': d, 'assists': a, 'playtime': p
        }
        
        # Add class-specific stats
        for class_name in ['tank', 'assault', 'golem', 'swordsman', 'archer', 'engineer']:
            tab_data[f'{class_name}_kills'] = stats.get(f'ww_{class_name}_kills', {}).get(cache_key, 0)
            tab_data[f'{class_name}_deaths'] = stats.get(f'ww_{class_name}_deaths', {}).get(cache_key, 0)
            tab_data[f'{class_name}_assists'] = stats.get(f'ww_{class_name}_assists', {}).get(cache_key, 0)
        
        all_data[tab] = tab_data

    meta = user_data.get("meta", {})
    level = meta.get("level", 0)
    icon = meta.get("icon", "")
    ign_color = meta.get("ign_color")
    guild_tag = meta.get("guild_tag")
    guild_hex = meta.get("guild_hex")
    
    # Get real-time status
    status_task = asyncio.to_thread(get_player_status, ign)
    skin_task = asyncio.to_thread(get_player_body, ign)
    (status_text, status_color), skin_image = await asyncio.gather(status_task, skin_task)
    
    # Check if tracked (UUID-aware)
    is_tracked = is_tracked_user(ign)
    
    view = WWStatsView(all_data, ign, int(level), icon, 
                       ign_color=ign_color, guild_tag=guild_tag, guild_hex=guild_hex,
                       status_text=status_text, status_color=status_color, skin_image=skin_image,
                       show_period_buttons=is_tracked)
    
    file = view.generate_composite_image("all-time")
    if is_tracked:
        message = await interaction.followup.send(file=file, view=view)
        view.message = message  # Store message reference for timeout handling
    else:
        msg = f"`{ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{ign}` to start tracking and enable session/daily/monthly stats."
        message = await interaction.followup.send(content=msg, file=file, view=view)
        view.message = message  # Store message reference for timeout handling
        # Register user in database for leaderboard accuracy (but don't actively track)
        register_user(ign)

@bot.tree.command(name="ctw", description="Get Capture the Wool stats")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def ctw(interaction: discord.Interaction, ign: str = None):
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign

    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign

    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    EXCEL_FILE = BOT_DIR / "stats.xlsx"
    if not EXCEL_FILE.exists():
        await interaction.followup.send("Stats file not found.")
        return

    # Try to find in cache first
    cache_data = await STATS_CACHE.get_data()
    user_data = None
    key = ign.casefold()
    for name, data in cache_data.items():
        if name.casefold() == key:
            user_data = data
            ign = name
            break

    # If not in cache, fetch fresh stats
    if not user_data:
        result = run_script("api_get.py", ["-ign", ign])
        if result.returncode == 0 and result.stdout:
            try:
                json_data = None
                for line in reversed(result.stdout.splitlines()):
                    line = line.strip()
                    if line.startswith('{') and line.endswith('}'):
                        try:
                            json_data = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
                
                if json_data and "processed_stats" in json_data and "username" in json_data:
                    user_data = await STATS_CACHE.update_cache_entry(json_data["username"], json_data["processed_stats"])
                    ign = json_data["username"]
            except Exception as e:
                print(f"[WARNING] Failed to parse api_get output in ctw: {e}")
        elif result.returncode != 0:
            print(f"[ERROR] api_get failed for {ign} in ctw: {result.stderr}")

    if not user_data:
        await interaction.followup.send("Player not found in database or API error.")
        return

    # Build all_data from cache - CTW stats
    all_data = {}
    stats = user_data.get("stats", {})
    
    # Map tab names to cache keys
    tab_map = {
        "all-time": "lifetime",
        "session": "session",
        "daily": "daily",
        "yesterday": "yesterday",
        "weekly": "weekly",
        "monthly": "monthly"
    }
    
    for tab, cache_key in tab_map.items():
        w = stats.get("ctw_experienced_wins", {}).get(cache_key, 0)
        l = stats.get("ctw_experienced_losses", {}).get(cache_key, 0)
        k = stats.get("ctw_kills", {}).get(cache_key, 0)
        d = stats.get("ctw_deaths", {}).get(cache_key, 0)
        a = stats.get("ctw_assists", {}).get(cache_key, 0)
        p = stats.get("playtime", {}).get(cache_key, 0)
        
        wools_picked = stats.get("ctw_wools_stolen", {}).get(cache_key, 0)
        wools_captured = stats.get("ctw_wools_captured", {}).get(cache_key, 0)
        
        kills_on_wh = stats.get("ctw_kills_on_woolholder", {}).get(cache_key, 0)
        deaths_to_wh = stats.get("ctw_deaths_to_woolholder", {}).get(cache_key, 0)
        kills_as_wh = stats.get("ctw_kills_with_wool", {}).get(cache_key, 0)
        deaths_as_wh = stats.get("ctw_deaths_with_wool", {}).get(cache_key, 0)
        
        gold_earned = stats.get("ctw_gold_earned", {}).get(cache_key, 0)
        gold_spent = stats.get("ctw_gold_spent", {}).get(cache_key, 0)
        
        # Get draws directly from API (ctw_experienced_draws)
        draws = stats.get("ctw_experienced_draws", {}).get(cache_key, 0)
        
        all_data[tab] = {
            'wins': w, 'losses': l, 'draws': draws,
            'kills': k, 'deaths': d, 'assists': a,
            'wools_picked': wools_picked, 'wools_captured': wools_captured,
            'kills_on_wh': kills_on_wh, 'deaths_to_wh': deaths_to_wh,
            'kills_as_wh': kills_as_wh, 'deaths_as_wh': deaths_as_wh,
            'gold_earned': gold_earned, 'gold_spent': gold_spent,
            'playtime': p
        }

    meta = user_data.get("meta", {})
    level = meta.get("level", 0)
    icon = meta.get("icon", "")
    ign_color = meta.get("ign_color")
    guild_tag = meta.get("guild_tag")
    guild_hex = meta.get("guild_hex")
    
    # Get real-time status
    status_task = asyncio.to_thread(get_player_status, ign)
    skin_task = asyncio.to_thread(get_player_body, ign)
    (status_text, status_color), skin_image = await asyncio.gather(status_task, skin_task)
    
    view = CTWStatsView(all_data, ign, int(level), icon, 
                        ign_color=ign_color, guild_tag=guild_tag, guild_hex=guild_hex,
                        status_text=status_text, status_color=status_color, skin_image=skin_image)
    
    # Check if tracked (UUID-aware)
    is_tracked = is_tracked_user(ign)
    
    file = view.generate_composite_image("all-time")
    if is_tracked:
        message = await interaction.followup.send(file=file, view=view)
        view.message = message  # Store message reference for timeout handling
    else:
        msg = f"`{ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{ign}` to start tracking and enable session/daily/monthly stats."
        await interaction.followup.send(content=msg, file=file)
        # Register user in database for leaderboard accuracy (but don't actively track)
        register_user(ign)

# Standalone leaderboard commands
# Define category metrics globally for autocomplete
CATEGORY_METRICS = {
    "general": {
        "level": "Level",
        "experience": "Experience",
        "coins": "Coins",
        "playtime": "Playtime",
        "available_layers": "Available Layers",
        # Combined stats from Sheep Wars, Wool Wars, and CTW
        "total_kills": "Kills",
        "total_deaths": "Deaths",
        "total_kdr": "K/D Ratio",
        "total_wins": "Wins",
        "total_losses": "Losses",
        "total_wlr": "W/L Ratio",
        "total_games_played": "Games Played",
    },
    "sheepwars": {
        "kills": "Total Kills",
        "deaths": "Total Deaths",
        "kdr": "K/D Ratio",
        "wins": "Wins",
        "losses": "Losses",
        "wlr": "W/L Ratio",
        "damage_dealt": "Damage Dealt",
        "games_played": "Games Played",
        "sheep_thrown": "Sheep Thrown",
        "magic_wool_hit": "Magic Wool Hit",
        "kills_void": "Void Kills",
        "kills_explosive": "Explosive Kills",
        "kills_melee": "Melee Kills",
        "kills_bow": "Bow Kills",
        "deaths_void": "Void Deaths",
        "deaths_explosive": "Explosive Deaths",
        "deaths_melee": "Melee Deaths",
        "deaths_bow": "Bow Deaths",
        # Ratios
        "wl_ratio": "Win/Loss Ratio",
        "kd_ratio": "Kill/Death Ratio",
        "kills_per_game": "Kills per Game",
        "kills_per_win": "Kills per Win",
        "kills_per_hour": "Kills per Hour",
        "damage_per_game": "Damage per Game",
        "damage_per_sheep": "Damage per Sheep",
        "wools_per_game": "Wools per Game",
        "sheeps_per_game": "Sheeps per Game",
        "void_kd_ratio": "Void K/D Ratio",
        "explosive_kd_ratio": "Explosive K/D Ratio",
        "bow_kd_ratio": "Bow K/D Ratio",
        "melee_kd_ratio": "Melee K/D Ratio",
        "wins_per_hour": "Wins per Hour",
        "exp_per_hour": "EXP per Hour",
        "exp_per_game": "EXP per Game",
        "survival_rate": "Survival Rate",
        "carried_score": "Carried Score",
    },
    "ctw": {
        "ctw_kills": "Kills",
        "ctw_deaths": "Deaths",
        "ctw_assists": "Assists",
        "ctw_wools_captured": "Wools Captured",
        "ctw_wools_stolen": "Wools Stolen",
        "ctw_kills_on_woolholder": "Kills on Wool Holder",
        "ctw_kills_with_wool": "Kills with Wool",
        "ctw_deaths_to_woolholder": "Deaths to Wool Holder",
        "ctw_deaths_with_wool": "Deaths with Wool",
        "ctw_gold_earned": "Gold Earned",
        "ctw_gold_spent": "Gold Spent",
        "ctw_experienced_wins": "Experienced Wins",
        "ctw_experienced_losses": "Experienced Losses",
        "ctw_experienced_draws": "Draws",
        "ctw_participated_wins": "Participated Wins",
        "ctw_participated_losses": "Participated Losses",
        "ctw_games_played": "Games Played",
        "ctw_fastest_win": "Fastest Win",
        "ctw_fastest_wool_capture": "Fastest Wool Capture",
        "ctw_longest_game": "Longest Game",
        "ctw_most_kills_and_assists": "Most Kills & Assists",
        "ctw_most_gold_earned": "Most Gold Earned",
        # CTW Ratios
        "ctw_wl_ratio": "Win/Loss Ratio",
        "ctw_kd_ratio": "Kill/Death Ratio",
        "ctw_kills_per_game": "Kills per Game",
        "ctw_deaths_per_game": "Deaths per Game",
        "ctw_kd_on_woolholder": "K/D vs Wool Holder",
        "ctw_kd_as_woolholder": "K/D as Wool Holder",
        "ctw_woolholder_kills_per_game": "Wool Holder Kills per Game",
        "ctw_woolholder_kills_per_kill": "Wool Holder Kills per Kill",
        "ctw_wools_captured_per_game": "Wools Captured per Game",
        "ctw_wools_captured_per_death": "Wools Captured per Death",
        "ctw_gold_earned_per_game": "Gold Earned per Game",
        "ctw_gold_spent_per_game": "Gold Spent per Game",
        "ctw_wools_stolen_per_game": "Wools Stolen per Game",
    },
    "ww": {
        "ww_kills": "Kills",
        "ww_deaths": "Deaths",
        "ww_assists": "Assists",
        "ww_wins": "Wins",
        "ww_games_played": "Games Played",
        "ww_wool_placed": "Wool Placed",
        "ww_blocks_broken": "Blocks Broken",
        "ww_powerups_gotten": "Powerups Gotten",
        # Class-specific stats
        "ww_tank_kills": "Tank Kills",
        "ww_tank_deaths": "Tank Deaths",
        "ww_tank_assists": "Tank Assists",
        "ww_assault_kills": "Assault Kills",
        "ww_assault_deaths": "Assault Deaths",
        "ww_assault_assists": "Assault Assists",
        "ww_golem_kills": "Golem Kills",
        "ww_golem_deaths": "Golem Deaths",
        "ww_golem_assists": "Golem Assists",
        "ww_swordsman_kills": "Swordsman Kills",
        "ww_swordsman_deaths": "Swordsman Deaths",
        "ww_swordsman_assists": "Swordsman Assists",
        "ww_archer_kills": "Archer Kills",
        "ww_archer_deaths": "Archer Deaths",
        "ww_archer_assists": "Archer Assists",
        "ww_engineer_kills": "Engineer Kills",
        "ww_engineer_deaths": "Engineer Deaths",
        "ww_engineer_assists": "Engineer Assists",
        # WW Ratios
        "ww_wl_ratio": "Win/Loss Ratio",
        "ww_kd_ratio": "Kill/Death Ratio",
        "ww_kills_per_game": "Kills per Game",
        "ww_assists_per_game": "Assists per Game",
        "ww_kill_assist_ratio": "Kill/Assist Ratio",
        "ww_assists_per_death": "Assists per Death",
        # Class-specific ratios
        "ww_tank_kd_ratio": "Tank K/D Ratio",
        "ww_tank_assists_per_death": "Tank Assists per Death",
        "ww_tank_kill_assist_ratio": "Tank Kill/Assist Ratio",
        "ww_assault_kd_ratio": "Assault K/D Ratio",
        "ww_assault_assists_per_death": "Assault Assists per Death",
        "ww_assault_kill_assist_ratio": "Assault Kill/Assist Ratio",
        "ww_golem_kd_ratio": "Golem K/D Ratio",
        "ww_golem_assists_per_death": "Golem Assists per Death",
        "ww_golem_kill_assist_ratio": "Golem Kill/Assist Ratio",
        "ww_swordsman_kd_ratio": "Swordsman K/D Ratio",
        "ww_swordsman_assists_per_death": "Swordsman Assists per Death",
        "ww_swordsman_kill_assist_ratio": "Swordsman Kill/Assist Ratio",
        "ww_archer_kd_ratio": "Archer K/D Ratio",
        "ww_archer_assists_per_death": "Archer Assists per Death",
        "ww_archer_kill_assist_ratio": "Archer Kill/Assist Ratio",
        "ww_engineer_kd_ratio": "Engineer K/D Ratio",
        "ww_engineer_assists_per_death": "Engineer Assists per Death",
        "ww_engineer_kill_assist_ratio": "Engineer Kill/Assist Ratio",
    },
    "guild": {
        "GENERAL": "Total Guild Experience",
        "WOOL_GAMES": "Wool Games",
        "BATTLEGROUND": "Battleground",
        "VAMPIREZ": "VampireZ",
        "HOUSING": "Housing",
        "WALLS": "Walls",
        "SKYWARS": "SkyWars",
        "TNTGAMES": "TNT Games",
        "DUELS": "Duels",
        "UHC": "UHC",
        "ARCADE": "Arcade",
        "MURDER_MYSTERY": "Murder Mystery",
        "PIT": "Pit",
        "BUILD_BATTLE": "Build Battle",
        "MCGO": "Cops and Crims",
        "PAINTBALL": "Paintball",
        "SUPER_SMASH": "Super Smash",
        "WALLS3": "Mega Walls",
        "PROTOTYPE": "Prototype",
        "ARENA": "Arena",
        "SKYBLOCK": "SkyBlock",
        "SURVIVAL_GAMES": "Blitz SG",
        "QUAKECRAFT": "Quakecraft",
        "BEDWARS": "Bed Wars",
        "GINGERBREAD": "Gingerbread",
        "LEGACY": "Legacy",
        "SPEED_UHC": "Speed UHC",
        "REPLAY": "Replay",
        "SMP": "SMP",
    }
}

# Create leaderboard command group
leaderboard_group = discord.app_commands.Group(name="leaderboard", description="View player leaderboards")

async def general_metric_autocomplete(interaction: discord.Interaction, current: str) -> list[discord.app_commands.Choice[str]]:
    """Autocomplete for general stats metrics."""
    metrics = CATEGORY_METRICS.get("general", {})
    current_lower = current.lower()
    matches = [
        discord.app_commands.Choice(name=f"{display_name}", value=key)
        for key, display_name in metrics.items()
        if current_lower in key.lower() or current_lower in display_name.lower()
    ]
    return matches[:25]

async def sheepwars_metric_autocomplete(interaction: discord.Interaction, current: str) -> list[discord.app_commands.Choice[str]]:
    """Autocomplete for sheep wars stats metrics."""
    metrics = CATEGORY_METRICS.get("sheepwars", {})
    current_lower = current.lower()
    matches = [
        discord.app_commands.Choice(name=f"{display_name}", value=key)
        for key, display_name in metrics.items()
        if current_lower in key.lower() or current_lower in display_name.lower()
    ]
    return matches[:25]

async def ctw_metric_autocomplete(interaction: discord.Interaction, current: str) -> list[discord.app_commands.Choice[str]]:
    """Autocomplete for CTW stats metrics."""
    metrics = CATEGORY_METRICS.get("ctw", {})
    current_lower = current.lower()
    matches = [
        discord.app_commands.Choice(name=f"{display_name}", value=key)
        for key, display_name in metrics.items()
        if current_lower in key.lower() or current_lower in display_name.lower()
    ]
    return matches[:25]

async def ww_metric_autocomplete(interaction: discord.Interaction, current: str) -> list[discord.app_commands.Choice[str]]:
    """Autocomplete for Wool Wars stats metrics."""
    metrics = CATEGORY_METRICS.get("ww", {})
    current_lower = current.lower()
    matches = [
        discord.app_commands.Choice(name=f"{display_name}", value=key)
        for key, display_name in metrics.items()
        if current_lower in key.lower() or current_lower in display_name.lower()
    ]
    return matches[:25]

async def guild_metric_autocomplete(interaction: discord.Interaction, current: str) -> list[discord.app_commands.Choice[str]]:
    """Autocomplete for guild experience metrics."""
    metrics = CATEGORY_METRICS.get("guild", {})
    current_lower = current.lower()
    matches = [
        discord.app_commands.Choice(name=f"{display_name}", value=key)
        for key, display_name in metrics.items()
        if current_lower in key.lower() or current_lower in display_name.lower()
    ]
    return matches[:25]

@leaderboard_group.command(name="general", description="View general stats leaderboards")
@discord.app_commands.describe(metric="Choose a stat to rank players by")
@discord.app_commands.autocomplete(metric=general_metric_autocomplete)
async def leaderboard_general(interaction: discord.Interaction, metric: str):
    """General stats leaderboard."""
    await _handle_leaderboard(interaction, "general", metric)

@leaderboard_group.command(name="sheepwars", description="View Sheep Wars stats leaderboards")
@discord.app_commands.describe(metric="Choose a stat to rank players by")
@discord.app_commands.autocomplete(metric=sheepwars_metric_autocomplete)
async def leaderboard_sheepwars(interaction: discord.Interaction, metric: str):
    """Sheep Wars stats leaderboard."""
    await _handle_leaderboard(interaction, "sheepwars", metric)

@leaderboard_group.command(name="ctw", description="View Capture the Wool stats leaderboards")
@discord.app_commands.describe(metric="Choose a stat to rank players by")
@discord.app_commands.autocomplete(metric=ctw_metric_autocomplete)
async def leaderboard_ctw(interaction: discord.Interaction, metric: str):
    """Capture the Wool stats leaderboard."""
    await _handle_leaderboard(interaction, "ctw", metric)

@leaderboard_group.command(name="ww", description="View Wool Wars stats leaderboards")
@discord.app_commands.describe(metric="Choose a stat to rank players by")
@discord.app_commands.autocomplete(metric=ww_metric_autocomplete)
async def leaderboard_ww(interaction: discord.Interaction, metric: str):
    """Wool Wars stats leaderboard."""
    await _handle_leaderboard(interaction, "ww", metric)

@leaderboard_group.command(name="guild", description="View guild experience leaderboards")
@discord.app_commands.describe(stat="Choose a game type to rank guilds by")
@discord.app_commands.autocomplete(stat=guild_metric_autocomplete)
async def leaderboard_guild(interaction: discord.Interaction, stat: str):
    """Guild experience leaderboard."""
    await _handle_guild_leaderboard(interaction, stat)

# Add the leaderboard group to the bot
bot.tree.add_command(leaderboard_group)


# ==============================
# RANKINGS COMMANDS
# ==============================

def _calculate_user_rankings(username: str, category: str):
    """Calculate rankings for a specific user across all metrics in a category.
    
    OPTIMIZED: Loads all users once and calculates all rankings in memory,
    instead of loading all users separately for each metric.
    
    POTENTIAL FUTURE OPTIMIZATION: Could cache leaderboard calculations for 
    lifetime rankings for ~30-60 seconds since they change less frequently.
    
    Returns:
        dict: {
            period: {
                metric: (rank, total_players, value)
            }
        }
    """
    periods = ["lifetime", "session", "daily", "yesterday", "weekly", "monthly"]
    metrics = CATEGORY_METRICS.get(category, {})
    
    # Initialize result structure
    result = {period: {} for period in periods}
    user_is_tracked = is_tracked_user(username)
    
    # Define ratio metrics
    ratio_metrics = {
        "kdr", "wlr", "wl_ratio", "kd_ratio", "kills_per_game", "kills_per_win",
        "kills_per_hour", "damage_per_game", "damage_per_sheep", "wools_per_game",
        "sheeps_per_game", "void_kd_ratio", "explosive_kd_ratio", "bow_kd_ratio",
        "melee_kd_ratio", "wins_per_hour", "exp_per_hour", "exp_per_game",
        "survival_rate", "carried_score",
        # General combined stats ratios
        "total_kdr", "total_wlr",
        # CTW ratios
        "ctw_wl_ratio", "ctw_kd_ratio", "ctw_kills_per_game", "ctw_deaths_per_game",
        "ctw_kd_on_woolholder", "ctw_kd_as_woolholder", "ctw_woolholder_kills_per_game",
        "ctw_woolholder_kills_per_kill", "ctw_wools_captured_per_game", "ctw_wools_captured_per_death",
        "ctw_gold_earned_per_game", "ctw_gold_spent_per_game", "ctw_wools_stolen_per_game",
        # WW ratios
        "ww_wl_ratio", "ww_kd_ratio", "ww_kills_per_game", "ww_assists_per_game",
        "ww_kill_assist_ratio", "ww_assists_per_death",
        # WW class-specific ratios
        "ww_tank_kd_ratio", "ww_tank_assists_per_death", "ww_tank_kill_assist_ratio",
        "ww_assault_kd_ratio", "ww_assault_assists_per_death", "ww_assault_kill_assist_ratio",
        "ww_golem_kd_ratio", "ww_golem_assists_per_death", "ww_golem_kill_assist_ratio",
        "ww_swordsman_kd_ratio", "ww_swordsman_assists_per_death", "ww_swordsman_kill_assist_ratio",
        "ww_archer_kd_ratio", "ww_archer_assists_per_death", "ww_archer_kill_assist_ratio",
        "ww_engineer_kd_ratio", "ww_engineer_assists_per_death", "ww_engineer_kill_assist_ratio",
    }
    
    try:
        # OPTIMIZATION: Load all users' stats once instead of loading for each metric
        print(f"[RANKINGS] Loading all users for category '{category}'...")
        start_time = time.time()
        
        all_usernames = get_all_usernames()
        user_colors = load_user_colors()
        
        # Build a cache of all users' stats
        all_users_data = {}
        for uname in all_usernames:
            try:
                stats_dict = get_user_stats_with_deltas(uname)
                if not stats_dict:
                    continue
                
                user_meta = user_colors.get(uname.lower(), {})
                all_users_data[uname] = {
                    'stats': stats_dict,
                    'meta': user_meta,
                    'is_tracked': is_tracked_user(uname)
                }
            except Exception as e:
                print(f"[RANKINGS] Error loading user {uname}: {e}")
                continue
        
        print(f"[RANKINGS] Loaded {len(all_users_data)} users in {time.time() - start_time:.2f}s")
        
        # Get the target user's stats
        user_data = all_users_data.get(username)
        if not user_data:
            print(f"[RANKINGS] User '{username}' not found in database")
            return None
        
        user_stats = user_data['stats']
        
        # For each metric, calculate leaderboard and find user's position
        for metric_key in metrics.keys():
            try:
                is_ratio = metric_key in ratio_metrics
                
                # Calculate leaderboard for each period
                for period in periods:
                    # Skip non-lifetime periods if user is not tracked
                    if period != "lifetime" and not user_is_tracked:
                        continue
                    
                    # Build leaderboard for this metric and period
                    leaderboard = []
                    for uname, data in all_users_data.items():
                        u_is_tracked = data['is_tracked']
                        
                        # Skip untracked users for non-lifetime periods
                        if period != "lifetime" and not u_is_tracked:
                            continue
                        
                        u_stats = data['stats']
                        value = _calculate_metric_value(u_stats, metric_key, period, is_ratio)
                        
                        # Filter out invalid values for specific metrics
                        if metric_key in ["survival_rate", "carried_score"] and (value is None or value == 0):
                            continue
                        
                        if value is not None:
                            leaderboard.append((uname, value))
                    
                    # Sort: carried_score is ascending (lower is better), others descending
                    if metric_key == "carried_score":
                        leaderboard.sort(key=lambda x: x[1], reverse=False)
                    else:
                        leaderboard.sort(key=lambda x: x[1], reverse=True)
                    
                    # Find user's position
                    total_players = len(leaderboard)
                    for idx, (uname, value) in enumerate(leaderboard):
                        if uname.lower() == username.lower():
                            rank = idx + 1
                            user_value = _calculate_metric_value(user_stats, metric_key, period, is_ratio)
                            result[period][metric_key] = (rank, total_players, user_value)
                            break
                        
            except Exception as e:
                print(f"[RANKINGS] Error processing metric {metric_key}: {e}")
                continue
        
        print(f"[RANKINGS] Calculated all rankings in {time.time() - start_time:.2f}s total")
        return result
        
    except Exception as e:
        print(f"[RANKINGS] Fatal error: {e}")
        import traceback
        traceback.print_exc()
        return None


def _calculate_metric_value(stats_dict: dict, metric: str, period: str, is_ratio: bool):
    """Calculate the value for a specific metric for a user.
    
    Args:
        stats_dict: User's stats dictionary
        metric: The metric to calculate
        period: The time period (lifetime, session, etc.)
        is_ratio: Whether this is a ratio metric
    
    Returns:
        The calculated value or None
    """
    try:
        # Use the existing comprehensive ratio calculation function
        if is_ratio:
            return _calculate_ratio_value_from_excel(stats_dict, period, metric)
        
        # Handle combined general stats (Sheep Wars + Wool Wars + CTW)
        if metric == "total_kills":
            sw_kills = stats_dict.get("kills", {}).get(period, 0)
            ctw_kills = stats_dict.get("ctw_kills", {}).get(period, 0)
            ww_kills = stats_dict.get("ww_kills", {}).get(period, 0)
            return float(sw_kills + ctw_kills + ww_kills)
        elif metric == "total_deaths":
            sw_deaths = stats_dict.get("deaths", {}).get(period, 0)
            ctw_deaths = stats_dict.get("ctw_deaths", {}).get(period, 0)
            ww_deaths = stats_dict.get("ww_deaths", {}).get(period, 0)
            return float(sw_deaths + ctw_deaths + ww_deaths)
        elif metric == "total_wins":
            sw_wins = stats_dict.get("wins", {}).get(period, 0)
            ctw_wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            ww_wins = stats_dict.get("ww_wins", {}).get(period, 0)
            return float(sw_wins + ctw_wins + ww_wins)
        elif metric == "total_losses":
            sw_losses = stats_dict.get("losses", {}).get(period, 0)
            ctw_losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            ww_games = stats_dict.get("ww_games_played", {}).get(period, 0)
            ww_wins = stats_dict.get("ww_wins", {}).get(period, 0)
            ww_losses = ww_games - ww_wins
            return float(sw_losses + ctw_losses + ww_losses)
        elif metric == "total_games_played":
            sw_games = stats_dict.get("games_played", {}).get(period, 0)
            ctw_wins = stats_dict.get("ctw_experienced_wins", {}).get(period, 0)
            ctw_losses = stats_dict.get("ctw_experienced_losses", {}).get(period, 0)
            ctw_games = ctw_wins + ctw_losses
            ww_games = stats_dict.get("ww_games_played", {}).get(period, 0)
            return float(sw_games + ctw_games + ww_games)
        
        # Handle regular stats
        if metric in stats_dict:
            val = stats_dict.get(metric, {}).get(period, 0)
            # Make gold_spent positive
            if metric == "ctw_gold_spent":
                val = abs(val)
            return float(val)
        
        # If metric doesn't exist, return 0
        return 0.0
    except Exception as e:
        print(f"[RANKINGS] Error calculating {metric}: {e}")
        return 0.0


def _calculate_guild_rankings(guild_name: str):
    """Calculate rankings for a specific guild across all game types.
    
    Returns:
        dict: {
            period: {
                game_type: (rank, total_guilds, exp_value)
            }
        }
    """
    periods = ["lifetime", "session", "daily", "yesterday", "weekly", "monthly"]
    
    # Initialize result structure
    result = {period: {} for period in periods}
    guild_is_tracked = is_tracked_guild(guild_name)
    
    try:
        print(f"[GUILD RANKINGS] Loading all guilds...")
        start_time = time.time()
        
        # Get all guilds
        all_guild_names = get_all_guilds()
        
        # Load all guild data
        all_guilds_data = {}
        for gname in all_guild_names:
            try:
                exp_dict = get_guild_exp(gname)
                if not exp_dict:
                    continue
                
                # Get tracking status
                with get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('SELECT is_tracked FROM tracked_guilds WHERE name = ?', (gname,))
                    row = cursor.fetchone()
                    g_is_tracked = row['is_tracked'] == 1 if row else False
                
                all_guilds_data[gname] = {
                    'exp': exp_dict,
                    'is_tracked': g_is_tracked
                }
            except Exception as e:
                print(f"[GUILD RANKINGS] Error loading guild {gname}: {e}")
                continue
        
        print(f"[GUILD RANKINGS] Loaded {len(all_guilds_data)} guilds in {time.time() - start_time:.2f}s")
        
        # Get the target guild's data
        guild_data = all_guilds_data.get(guild_name)
        if not guild_data:
            print(f"[GUILD RANKINGS] Guild '{guild_name}' not found in database")
            return None
        
        guild_exp = guild_data['exp']
        
        # Get all game types from the guild's exp data (dynamically)
        game_types = list(guild_exp.keys())
        print(f"[GUILD RANKINGS] Found {len(game_types)} game types for guild {guild_name}: {game_types}")
        
        # For each game type, calculate leaderboard and find guild's position
        for game_type in game_types:
            try:
                # Calculate leaderboard for each period
                for period in periods:
                    # Skip non-lifetime periods if guild is not tracked
                    if period != "lifetime" and not guild_is_tracked:
                        continue
                    
                    # Get guild's value first
                    if period == "lifetime":
                        guild_value = guild_exp[game_type]['lifetime']
                    elif period == "session":
                        guild_value = guild_exp[game_type]['lifetime'] - guild_exp[game_type]['session']
                    elif period == "daily":
                        guild_value = guild_exp[game_type]['lifetime'] - guild_exp[game_type]['daily']
                    elif period == "yesterday":
                        guild_value = guild_exp[game_type]['daily'] - guild_exp[game_type]['yesterday']
                    elif period == "weekly":
                        guild_value = guild_exp[game_type]['lifetime'] - guild_exp[game_type]['weekly']
                    elif period == "monthly":
                        guild_value = guild_exp[game_type]['lifetime'] - guild_exp[game_type]['monthly']
                    else:
                        guild_value = 0
                    
                    # Build leaderboard for this game type and period
                    leaderboard = []
                    for gname, data in all_guilds_data.items():
                        g_is_tracked = data['is_tracked']
                        
                        # Skip untracked guilds for non-lifetime periods
                        if period != "lifetime" and not g_is_tracked:
                            continue
                        
                        g_exp = data['exp']
                        
                        # Skip if guild doesn't have this game type
                        if game_type not in g_exp:
                            continue
                        
                        # Get exp value for this period
                        if period == "lifetime":
                            value = g_exp[game_type]['lifetime']
                        elif period == "session":
                            value = g_exp[game_type]['lifetime'] - g_exp[game_type]['session']
                        elif period == "daily":
                            value = g_exp[game_type]['lifetime'] - g_exp[game_type]['daily']
                        elif period == "yesterday":
                            value = g_exp[game_type]['daily'] - g_exp[game_type]['yesterday']
                        elif period == "weekly":
                            value = g_exp[game_type]['lifetime'] - g_exp[game_type]['weekly']
                        elif period == "monthly":
                            value = g_exp[game_type]['lifetime'] - g_exp[game_type]['monthly']
                        else:
                            value = 0
                        
                        # Include all values, even 0
                        leaderboard.append((gname, value))
                    
                    # Sort descending
                    leaderboard.sort(key=lambda x: x[1], reverse=True)
                    
                    # Find guild's position
                    total_guilds = len(leaderboard)
                    rank = None
                    for idx, (gname, value) in enumerate(leaderboard):
                        if gname.lower() == guild_name.lower():
                            rank = idx + 1
                            break
                    
                    # Always add the ranking, even if guild has 0 exp (will be ranked last)
                    if rank is not None:
                        result[period][game_type] = (rank, total_guilds, guild_value)
                    
            except Exception as e:
                print(f"[GUILD RANKINGS] Error processing game type {game_type}: {e}")
                continue
        
        print(f"[GUILD RANKINGS] Calculated all rankings in {time.time() - start_time:.2f}s total")
        return result
        
    except Exception as e:
        print(f"[GUILD RANKINGS] Fatal error: {e}")
        import traceback
        traceback.print_exc()
        return None


class RankingsTabView(discord.ui.View):
    """View for displaying user rankings across all metrics with period tabs."""
    
    def __init__(self, username: str, category: str, rankings_data: dict, user_is_tracked: bool):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.username = username
        self.category = category
        self.rankings_data = rankings_data
        self.user_is_tracked = user_is_tracked
        self.current_period = "lifetime"
        self.page = 0
        self.page_size = 10
        self.message = None  # Store message reference for timeout handling
        
        # Map category to display name
        self.category_display = {
            "general": "Wool Games",
            "sheepwars": "Sheep Wars",
            "ctw": "Capture the Wool",
            "ww": "Wool Wars"
        }.get(category, category.title())
        
        # Add buttons conditionally
        self._add_buttons()
        
        # Update button styles
        self.update_button_styles()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
    def _add_buttons(self):
        """Add buttons and dropdowns."""
        # Add period selector dropdown (only if user is tracked)
        if self.user_is_tracked:
            self.period_select = RankingsPeriodSelect(self)
            self.add_item(self.period_select)
        
        # Add navigation buttons (always show)
        self.prev_button = discord.ui.Button(
            label="Prev Page",
            custom_id="page_prev",
            style=discord.ButtonStyle.secondary,
            row=1
        )
        self.prev_button.callback = self.prev_page
        self.add_item(self.prev_button)
        
        self.next_button = discord.ui.Button(
            label="Next Page",
            custom_id="page_next",
            style=discord.ButtonStyle.secondary,
            row=1
        )
        self.next_button.callback = self.next_page
        self.add_item(self.next_button)
    
    def update_button_styles(self):
        """Update button styles (not needed anymore since we use dropdown)."""
        pass
    
    def _paginate(self, rankings_list: list, page: int):
        """Paginate rankings list."""
        total_pages = max(1, (len(rankings_list) + self.page_size - 1) // self.page_size)
        clamped_page = max(0, min(page, total_pages - 1))
        start_index = clamped_page * self.page_size
        return rankings_list[start_index:start_index + self.page_size], total_pages, clamped_page
    
    def generate_rankings_image(self):
        """Generate image showing all rankings for current period."""
        period_data = self.rankings_data.get(self.current_period, {})
        
        if not period_data:
            embed = discord.Embed(
                title=f"🏆 All Rankings for {self.username}",
                description=f"No {self.current_period} data available for this user.",
                color=discord.Color.from_rgb(54, 57, 63)
            )
            if not self.user_is_tracked:
                embed.set_footer(text=f"`{self.username}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{self.username}` to start tracking and enable session/daily/monthly stats.")
            return embed, None, 1
        
        # Get metric labels
        metrics = CATEGORY_METRICS.get(self.category, {})
        
        # Build ranking data for image
        rankings_list = []
        # Define ratio metrics
        ratio_metrics = {
            "kdr", "wlr", "wl_ratio", "kd_ratio", "kills_per_game", "kills_per_win",
            "kills_per_hour", "damage_per_game", "damage_per_sheep", "wools_per_game",
            "sheeps_per_game", "void_kd_ratio", "explosive_kd_ratio", "bow_kd_ratio",
            "melee_kd_ratio", "wins_per_hour", "exp_per_hour", "exp_per_game",
            "survival_rate", "carried_score",
            # General combined stats ratios
            "total_kdr", "total_wlr",
            # CTW ratios
            "ctw_wl_ratio", "ctw_kd_ratio", "ctw_kills_per_game", "ctw_deaths_per_game",
            "ctw_kd_on_woolholder", "ctw_kd_as_woolholder", "ctw_woolholder_kills_per_game",
            "ctw_woolholder_kills_per_kill", "ctw_wools_captured_per_game", "ctw_wools_captured_per_death",
            "ctw_gold_earned_per_game", "ctw_gold_spent_per_game", "ctw_wools_stolen_per_game",
            # WW ratios
            "ww_wl_ratio", "ww_kd_ratio", "ww_kills_per_game", "ww_assists_per_game",
            "ww_kill_assist_ratio", "ww_assists_per_death",
            # WW class-specific ratios
            "ww_tank_kd_ratio", "ww_tank_assists_per_death", "ww_tank_kill_assist_ratio",
            "ww_assault_kd_ratio", "ww_assault_assists_per_death", "ww_assault_kill_assist_ratio",
            "ww_golem_kd_ratio", "ww_golem_assists_per_death", "ww_golem_kill_assist_ratio",
            "ww_swordsman_kd_ratio", "ww_swordsman_assists_per_death", "ww_swordsman_kill_assist_ratio",
            "ww_archer_kd_ratio", "ww_archer_assists_per_death", "ww_archer_kill_assist_ratio",
            "ww_engineer_kd_ratio", "ww_engineer_assists_per_death", "ww_engineer_kill_assist_ratio",
        }
        
        for metric_key, (rank, total, value) in sorted(period_data.items(), key=lambda x: x[1][0]):
            metric_label = metrics.get(metric_key, metric_key)
            
            # Format value based on metric type
            if "playtime" in metric_key.lower():
                formatted_value = format_playtime(int(value))
            elif metric_key in ratio_metrics:
                # Always show exactly 2 decimal places for ratios
                formatted_value = f"{float(value):.2f}"
            else:
                formatted_value = f"{int(value):,}"
            
            rankings_list.append((rank, metric_label, formatted_value))
        
        # Paginate the rankings
        paginated_list, total_pages, clamped_page = self._paginate(rankings_list, self.page)
        self.page = clamped_page
        
        # Generate image if Pillow is available
        if Image is not None:
            try:
                img_io = create_rankings_image(
                    self.username,
                    self.category_display,
                    self.current_period,
                    paginated_list,
                    page=clamped_page,
                    total_pages=total_pages
                )
                filename = f"rankings_{self.username}_{self.category}_{self.current_period}_p{clamped_page + 1}.png"
                file = discord.File(img_io, filename=filename)
                
                # Add footer message if user is not tracked
                footer_embed = None
                if not self.user_is_tracked:
                    footer_embed = discord.Embed(
                        description=f"`{self.username}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{self.username}` to start tracking and enable session/daily/monthly stats.",
                        color=discord.Color.from_rgb(54, 57, 63)
                    )
                
                return footer_embed, file, total_pages
            except Exception as e:
                print(f"[WARNING] Rankings image generation failed: {e}")
                # Fall back to embed if image generation fails
                pass
        
        # Fallback embed if Pillow not available or image generation failed
        paginated_list, total_pages, clamped_page = self._paginate(rankings_list, self.page)
        self.page = clamped_page
        
        ranking_lines = []
        for rank, metric_label, formatted_value in paginated_list:
            ranking_lines.append(f"#{rank} {metric_label}: {formatted_value}")
        
        title = f"{self.current_period.title()} {self.category_display} Rankings for {self.username}"
        description = f"```ansi\n" + "\n".join(ranking_lines) + "\n```"
        
        embed = discord.Embed(
            title=title,
            description=description,
            color=discord.Color.from_rgb(54, 57, 63)
        )
        
        if not self.user_is_tracked:
            embed.set_footer(text=f"`{self.username}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{self.username}` to start tracking and enable session/daily/monthly stats.")
        else:
            embed.set_footer(text=f"Rankings updated in real-time | Page {clamped_page + 1}/{total_pages}")
        
        return embed, None, total_pages
    
    async def _refresh_period(self, interaction: discord.Interaction, new_period: str):
        """Refresh the view with a new period."""
        self.current_period = new_period
        self.page = 0  # Reset to first page when changing period
        
        # Update dropdown selection
        if self.user_is_tracked and hasattr(self, 'period_select'):
            for option in self.period_select.options:
                option.default = option.value == new_period
        
        embed, file, _ = self.generate_rankings_image()
        
        if file:
            await interaction.response.edit_message(view=self, attachments=[file], embed=embed)
        else:
            await interaction.response.edit_message(embed=embed, view=self)
    
    async def _refresh_page(self, interaction: discord.Interaction, page_delta: int):
        """Refresh the view with a different page."""
        self.page += page_delta
        
        embed, file, _ = self.generate_rankings_image()
        
        if file:
            await interaction.response.edit_message(view=self, attachments=[file], embed=embed)
        else:
            await interaction.response.edit_message(embed=embed, view=self)
    
    async def prev_page(self, interaction: discord.Interaction):
        """Go to previous page."""
        await self._refresh_page(interaction, -1)
    
    async def next_page(self, interaction: discord.Interaction):
        """Go to next page."""
        await self._refresh_page(interaction, 1)


class RankingsPeriodSelect(discord.ui.Select):
    """Dropdown for selecting the time period in rankings view."""
    
    def __init__(self, view: RankingsTabView):
        options = [
            discord.SelectOption(label="Lifetime", value="lifetime", default=True),
            discord.SelectOption(label="Session", value="session"),
            discord.SelectOption(label="Daily", value="daily"),
            discord.SelectOption(label="Yesterday", value="yesterday"),
            discord.SelectOption(label="Weekly", value="weekly"),
            discord.SelectOption(label="Monthly", value="monthly"),
        ]
        super().__init__(
            placeholder="Select rankings period",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="rankings_period_select",
            row=0
        )
        self.view_ref = view
    
    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref._refresh_period(interaction, selected)


class GuildRankingsTabView(discord.ui.View):
    """View for displaying guild rankings across all game types with period tabs."""
    
    def __init__(self, guild_name: str, rankings_data: dict, guild_is_tracked: bool):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.guild_name = guild_name
        self.rankings_data = rankings_data
        self.guild_is_tracked = guild_is_tracked
        self.current_period = "lifetime"
        self.page = 0
        self.page_size = 10
        self.message = None  # Store message reference for timeout handling
        
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
        # Add buttons conditionally
        self._add_buttons()
    
    def _add_buttons(self):
        """Add buttons and dropdowns."""
        # Add period selector dropdown (only if guild is tracked)
        if self.guild_is_tracked:
            self.period_select = GuildRankingsPeriodSelect(self)
            self.add_item(self.period_select)
        
        # Add navigation buttons (always show)
        self.prev_button = discord.ui.Button(
            label="Prev Page",
            custom_id="guild_page_prev",
            style=discord.ButtonStyle.secondary,
            row=1
        )
        self.prev_button.callback = self.prev_page
        self.add_item(self.prev_button)
        
        self.next_button = discord.ui.Button(
            label="Next Page",
            custom_id="guild_page_next",
            style=discord.ButtonStyle.secondary,
            row=1
        )
        self.next_button.callback = self.next_page
        self.add_item(self.next_button)
    
    def _paginate(self, rankings_list: list, page: int):
        """Paginate rankings list."""
        total_pages = max(1, (len(rankings_list) + self.page_size - 1) // self.page_size)
        clamped_page = max(0, min(page, total_pages - 1))
        start_index = clamped_page * self.page_size
        return rankings_list[start_index:start_index + self.page_size], total_pages, clamped_page
    
    def generate_rankings_image(self):
        """Generate image showing all rankings for current period."""
        period_data = self.rankings_data.get(self.current_period, {})
        
        if not period_data:
            embed = discord.Embed(
                title=f"🏆 Guild Rankings for {self.guild_name}",
                description=f"No {self.current_period} data available for this guild.",
                color=discord.Color.from_rgb(54, 57, 63)
            )
            if not self.guild_is_tracked:
                embed.set_footer(text=f"`{self.guild_name}` is not currently tracked. Only all-time stats are available.\nUse `/trackguild guild_name:{self.guild_name}` to start tracking.")
            return embed, None, 1
        
        # Build ranking data for image
        rankings_list = []
        for game_type, (rank, total, value) in sorted(period_data.items(), key=lambda x: x[1][0]):
            # Format game type name: replace underscores with spaces and title case
            game_label = game_type.replace("_", " ").title()
            formatted_value = f"{int(value):,}"
            rankings_list.append((rank, game_label, formatted_value))
        
        # Paginate the rankings
        paginated_list, total_pages, clamped_page = self._paginate(rankings_list, self.page)
        self.page = clamped_page
        
        # Generate image if Pillow is available
        if Image is not None:
            try:
                img_io = create_rankings_image(
                    self.guild_name,
                    "Guild Experience",
                    self.current_period,
                    paginated_list,
                    page=clamped_page,
                    total_pages=total_pages
                )
                filename = f"guild_rankings_{self.guild_name}_{self.current_period}_p{clamped_page + 1}.png"
                file = discord.File(img_io, filename=filename)
                
                # Add footer message if guild is not tracked
                footer_embed = None
                if not self.guild_is_tracked:
                    footer_embed = discord.Embed(
                        description=f"`{self.guild_name}` is not currently tracked. Only all-time stats are available.\nUse `/trackguild guild_name:{self.guild_name}` to start tracking.",
                        color=discord.Color.from_rgb(54, 57, 63)
                    )
                
                return footer_embed, file, total_pages
            except Exception as e:
                print(f"[WARNING] Guild rankings image generation failed: {e}")
                # Fall back to embed if image generation fails
                pass
        
        # Fallback embed if Pillow not available or image generation failed
        paginated_list, total_pages, clamped_page = self._paginate(rankings_list, self.page)
        self.page = clamped_page
        
        ranking_lines = []
        for rank, game_label, formatted_value in paginated_list:
            ranking_lines.append(f"#{rank} {game_label}: {formatted_value}")
        
        title = f"{self.current_period.title()} Guild Rankings for {self.guild_name}"
        description = f"```ansi\n" + "\n".join(ranking_lines) + "\n```"
        
        embed = discord.Embed(
            title=title,
            description=description,
            color=discord.Color.from_rgb(54, 57, 63)
        )
        
        if not self.guild_is_tracked:
            embed.set_footer(text=f"`{self.guild_name}` is not currently tracked. Only all-time stats are available.\nUse `/trackguild guild_name:{self.guild_name}` to start tracking.")
        else:
            embed.set_footer(text=f"Rankings updated in real-time | Page {clamped_page + 1}/{total_pages}")
        
        return embed, None, total_pages
    
    async def _refresh_period(self, interaction: discord.Interaction, new_period: str):
        """Refresh the view with a new period."""
        self.current_period = new_period
        self.page = 0  # Reset to first page when changing period
        
        # Update dropdown selection
        if self.guild_is_tracked and hasattr(self, 'period_select'):
            for option in self.period_select.options:
                option.default = option.value == new_period
        
        embed, file, _ = self.generate_rankings_image()
        
        if file:
            await interaction.response.edit_message(view=self, attachments=[file], embed=embed)
        else:
            await interaction.response.edit_message(embed=embed, view=self)
    
    async def _refresh_page(self, interaction: discord.Interaction, page_delta: int):
        """Refresh the view with a different page."""
        self.page += page_delta
        
        embed, file, _ = self.generate_rankings_image()
        
        if file:
            await interaction.response.edit_message(view=self, attachments=[file], embed=embed)
        else:
            await interaction.response.edit_message(embed=embed, view=self)
    
    async def prev_page(self, interaction: discord.Interaction):
        """Go to previous page."""
        await self._refresh_page(interaction, -1)
    
    async def next_page(self, interaction: discord.Interaction):
        """Go to next page."""
        await self._refresh_page(interaction, 1)


class GuildRankingsPeriodSelect(discord.ui.Select):
    """Dropdown for selecting the time period in guild rankings view."""
    
    def __init__(self, view: GuildRankingsTabView):
        options = [
            discord.SelectOption(label="Lifetime", value="lifetime", default=True),
            discord.SelectOption(label="Session", value="session"),
            discord.SelectOption(label="Daily", value="daily"),
            discord.SelectOption(label="Yesterday", value="yesterday"),
            discord.SelectOption(label="Weekly", value="weekly"),
            discord.SelectOption(label="Monthly", value="monthly"),
        ]
        super().__init__(
            placeholder="Select rankings period",
            min_values=1,
            max_values=1,
            options=options,
            custom_id="guild_rankings_period_select",
            row=0
        )
        self.view_ref = view
    
    async def callback(self, interaction: discord.Interaction):
        selected = self.values[0]
        for opt in self.options:
            opt.default = opt.value == selected
        await self.view_ref._refresh_period(interaction, selected)


# Create rankings command group
rankings_group = discord.app_commands.Group(name="rankings", description="View all leaderboard positions for a player")

@rankings_group.command(name="general", description="View all general stat rankings for a player")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def rankings_general(interaction: discord.Interaction, ign: str = None):
    """General stats rankings."""
    await _handle_rankings(interaction, "general", ign)

@rankings_group.command(name="sheepwars", description="View all Sheep Wars stat rankings for a player")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def rankings_sheepwars(interaction: discord.Interaction, ign: str = None):
    """Sheep Wars stats rankings."""
    await _handle_rankings(interaction, "sheepwars", ign)

@rankings_group.command(name="ctw", description="View all Capture the Wool stat rankings for a player")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def rankings_ctw(interaction: discord.Interaction, ign: str = None):
    """Capture the Wool stats rankings."""
    await _handle_rankings(interaction, "ctw", ign)

@rankings_group.command(name="ww", description="View all Wool Wars stat rankings for a player")
@discord.app_commands.describe(ign="Minecraft IGN (optional if you set /default)")
async def rankings_ww(interaction: discord.Interaction, ign: str = None):
    """Wool Wars stats rankings."""
    await _handle_rankings(interaction, "ww", ign)

@rankings_group.command(name="guild", description="View all guild experience rankings")
@discord.app_commands.describe(guild_name="Guild name")
async def rankings_guild(interaction: discord.Interaction, guild_name: str):
    """Guild experience rankings."""
    await _handle_guild_rankings(interaction, guild_name)

# Add the rankings group to the bot
bot.tree.add_command(rankings_group)

async def _handle_rankings(interaction: discord.Interaction, category: str, ign: str = None):
    """Shared handler for all rankings commands."""
    # Resolve default IGN if not provided
    if ign is None or str(ign).strip() == "":
        default_ign = get_default_user(interaction.user.id)
        if not default_ign:
            await interaction.response.send_message("You don't have a default username set. Use /default to set one.", ephemeral=True)
            return
        ign = default_ign
    
    # Validate username early
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign
    
    # Defer response
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        # Check if user is tracked
        user_is_tracked = is_tracked_user(ign)
        
        # Check if user exists in database
        user_exists_in_db = user_exists(ign)
        
        # Track if we need to clean up (remove user from DB after calculation)
        should_cleanup = False
        
        # If user doesn't exist in database, fetch from API
        if not user_exists_in_db:
            result = await asyncio.to_thread(run_script, "api_get.py", ["-ign", ign])
            if result.returncode == 0 and result.stdout:
                try:
                    json_data = None
                    for line in reversed(result.stdout.splitlines()):
                        line = line.strip()
                        if line.startswith('{') and line.endswith('}'):
                            try:
                                json_data = json.loads(line)
                                break
                            except json.JSONDecodeError:
                                continue
                    
                    if json_data and "username" in json_data:
                        ign = json_data["username"]  # Use proper capitalization from API
                        print(f"[RANKINGS] Fetched stats for untracked user: {ign}")
                        # Mark for cleanup - we added them temporarily
                        should_cleanup = True
                    else:
                        await interaction.followup.send(f"❌ Could not find stats for user '{ign}' on the API.")
                        return
                except Exception as e:
                    print(f"[RANKINGS] Failed to parse api_get output: {e}")
                    await interaction.followup.send(f"❌ Could not find stats for user '{ign}'.")
                    return
            else:
                await interaction.followup.send(f"❌ Could not find stats for user '{ign}' on the API.")
                return
        
        # Calculate rankings for all metrics
        rankings_data = await asyncio.to_thread(_calculate_user_rankings, ign, category)
        
        if not rankings_data:
            await interaction.followup.send(f"❌ Could not find stats for user '{ign}'.")
            return
        
        # Check if user has any rankings
        has_data = any(len(period_data) > 0 for period_data in rankings_data.values())
        if not has_data:
            await interaction.followup.send(f"❌ No ranking data found for user '{ign}' in category '{category}'.")
            return
        
        # Create view and send
        view = RankingsTabView(ign, category, rankings_data, user_is_tracked)
        embed, file, _ = view.generate_rankings_image()
        
        if file:
            message = await interaction.followup.send(view=view, file=file, embed=embed)
            view.message = message  # Store message reference for timeout handling
        else:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
        
        # Clean up: Remove user from database if they were temporarily added
        if should_cleanup:
            try:
                await asyncio.to_thread(delete_user, ign)
                print(f"[RANKINGS] Cleaned up temporarily added user: {ign}")
            except Exception as cleanup_error:
                print(f"[RANKINGS] Failed to cleanup user {ign}: {cleanup_error}")
        
    except Exception as e:
        print(f"[RANKINGS] Error: {e}")
        import traceback
        traceback.print_exc()
        await interaction.followup.send(f"❌ [ERROR] {str(e)}")


async def _handle_guild_rankings(interaction: discord.Interaction, guild_name: str):
    """Handler for guild rankings command."""
    # Defer response
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except Exception as defer_error:
            print(f"[GUILD RANKINGS] Failed to defer: {defer_error}")
    
    try:
        # Check if guild exists in database
        guild_exists_in_db = is_registered_guild(guild_name)
        guild_is_tracked = is_tracked_guild(guild_name)
        should_cleanup = False
        
        # If guild doesn't exist in database, fetch from API
        if not guild_exists_in_db:
            # Import here to avoid circular imports
            from api_get import api_update_guild_database
            
            try:
                result = await asyncio.to_thread(api_update_guild_database, guild_name)
                if result:
                    print(f"[GUILD RANKINGS] Fetched stats for untracked guild: {guild_name}")
                    should_cleanup = True  # Clean up after displaying rankings
                else:
                    await interaction.followup.send(f"❌ Could not find guild '{guild_name}' on the API.")
                    return
            except Exception as e:
                print(f"[GUILD RANKINGS] Failed to fetch guild from API: {e}")
                await interaction.followup.send(f"❌ Could not find guild '{guild_name}' on the API.")
                return
        
        # Calculate rankings for all game types
        rankings_data = await asyncio.to_thread(_calculate_guild_rankings, guild_name)
        
        if not rankings_data:
            await interaction.followup.send(f"❌ Could not find stats for guild '{guild_name}'.")
            return
        
        # Check if guild has any rankings
        has_data = any(len(period_data) > 0 for period_data in rankings_data.values())
        if not has_data:
            await interaction.followup.send(f"❌ No ranking data found for guild '{guild_name}'.")
            return
        
        # Create view and send
        view = GuildRankingsTabView(guild_name, rankings_data, guild_is_tracked)
        embed, file, _ = view.generate_rankings_image()
        
        if file:
            message = await interaction.followup.send(view=view, file=file, embed=embed)
            view.message = message  # Store message reference for timeout handling
        else:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
        
        # Clean up: Remove guild from database if it was temporarily added
        if should_cleanup:
            try:
                with get_db_connection() as conn:
                    cursor = conn.cursor()
                    cursor.execute('DELETE FROM tracked_guilds WHERE name = ?', (guild_name,))
                    cursor.execute('DELETE FROM gexp WHERE name = ?', (guild_name,))
                    conn.commit()
                print(f"[GUILD RANKINGS] Cleaned up temporarily added guild: {guild_name}")
            except Exception as cleanup_error:
                print(f"[GUILD RANKINGS] Failed to cleanup guild {guild_name}: {cleanup_error}")
        
    except Exception as e:
        print(f"[GUILD RANKINGS] Error: {e}")
        import traceback
        traceback.print_exc()
        await interaction.followup.send(f"❌ [ERROR] {str(e)}")


async def _handle_leaderboard(interaction: discord.Interaction, category: str, metric: str):
    """Shared handler for all leaderboard commands."""
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        available_metrics = CATEGORY_METRICS.get(category, {})
        
        # Validate metric
        if metric not in available_metrics:
            metric_list = "\n".join([f"• {k}: {v}" for k, v in available_metrics.items()])
            await interaction.followup.send(
                f"❌ Invalid metric '{metric}' for category '{category}'.\n\n"
                f"**Available metrics:**\n{metric_list}"
            )
            return
        
        # Determine if this is a ratio metric
        ratio_metrics = {
            "kdr", "wlr", "wl_ratio", "kd_ratio", "kills_per_game", "kills_per_win",
            "kills_per_hour", "damage_per_game", "damage_per_sheep", "wools_per_game",
            "sheeps_per_game", "void_kd_ratio", "explosive_kd_ratio", "bow_kd_ratio",
            "melee_kd_ratio", "wins_per_hour", "exp_per_hour", "exp_per_game",
            "survival_rate", "carried_score",
            # General combined stats ratios
            "total_kdr", "total_wlr",
            # CTW ratios
            "ctw_wl_ratio", "ctw_kd_ratio", "ctw_kills_per_game", "ctw_deaths_per_game",
            "ctw_kd_on_woolholder", "ctw_kd_as_woolholder", "ctw_woolholder_kills_per_game",
            "ctw_woolholder_kills_per_kill", "ctw_wools_captured_per_game", "ctw_wools_captured_per_death",
            "ctw_gold_earned_per_game", "ctw_gold_spent_per_game", "ctw_wools_stolen_per_game",
            # WW ratios
            "ww_wl_ratio", "ww_kd_ratio", "ww_kills_per_game", "ww_assists_per_game",
            "ww_kill_assist_ratio", "ww_assists_per_death",
            # WW class-specific ratios
            "ww_tank_kd_ratio", "ww_tank_assists_per_death", "ww_tank_kill_assist_ratio",
            "ww_assault_kd_ratio", "ww_assault_assists_per_death", "ww_assault_kill_assist_ratio",
            "ww_golem_kd_ratio", "ww_golem_assists_per_death", "ww_golem_kill_assist_ratio",
            "ww_swordsman_kd_ratio", "ww_swordsman_assists_per_death", "ww_swordsman_kill_assist_ratio",
            "ww_archer_kd_ratio", "ww_archer_assists_per_death", "ww_archer_kill_assist_ratio",
            "ww_engineer_kd_ratio", "ww_engineer_assists_per_death", "ww_engineer_kill_assist_ratio",
        }
        
        is_ratio = metric in ratio_metrics
        
        # Load data from database
        if is_ratio:
            processed_data = await asyncio.to_thread(
                _load_ratio_leaderboard_data_from_excel, metric, category
            )
            view = RatioLeaderboardView(metric, processed_data, category)
        else:
            processed_data = await asyncio.to_thread(
                _load_leaderboard_data_from_excel, metric, category
            )
            view = LeaderboardView(metric, processed_data, category)
        
        embed, file, _ = await asyncio.to_thread(view.generate_leaderboard_image, "lifetime", 0)
        if file:
            message = await interaction.followup.send(view=view, file=file)
            view.message = message  # Store message reference for timeout handling
        else:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
    except Exception as e:
        await interaction.followup.send(f"❌ [ERROR] {str(e)}")


async def _handle_guild_leaderboard(interaction: discord.Interaction, game: str):
    """Handler for guild experience leaderboard commands."""
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    
    try:
        available_games = CATEGORY_METRICS.get("guild", {})
        
        # Validate game
        if game not in available_games:
            game_list = "\n".join([f"• {k}: {v}" for k, v in available_games.items()])
            await interaction.followup.send(
                f"❌ Invalid game type '{game}'.\n\n"
                f"**Available game types:**\n{game_list}"
            )
            return
        
        # Load guild data from database
        processed_data = await asyncio.to_thread(_load_guild_leaderboard_data, game)
        view = GuildLeaderboardView(game, processed_data)
        
        embed, file, _ = await asyncio.to_thread(view.generate_leaderboard_image, "lifetime", 0)
        if file:
            message = await interaction.followup.send(view=view, file=file)
            view.message = message  # Store message reference for timeout handling
        else:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
    except Exception as e:
        await interaction.followup.send(f"❌ [ERROR] {str(e)}")


def _load_guild_leaderboard_data(game: str) -> dict:
    """Load guild leaderboard data for a specific game type.
    
    Returns dict with period keys mapping to sorted lists of (guild_name, exp_value, tag, tag_color) tuples.
    Values are DELTAS (lifetime - snapshot) except for lifetime which is the actual lifetime value.
    Only tracked guilds (is_tracked=1) appear in delta-based periods; all guilds appear in lifetime.
    """
    with get_db_connection() as conn:
        cursor = conn.cursor()
        
        # Get all guilds' data for this game along with tag info and tracking status from tracked_guilds
        cursor.execute('''
            SELECT g.name, g.exp, g.lifetime, g.session, g.daily, g.yesterday, g.weekly, g.monthly,
                   t.guild_tag, t.guild_hex, COALESCE(t.is_tracked, 0) as is_tracked
            FROM gexp g
            LEFT JOIN tracked_guilds t ON g.name = t.name
            WHERE g.game = ?
            ORDER BY g.lifetime DESC
        ''', (game,))
        
        rows = cursor.fetchall()
        
        # Organize by period
        periods = ["lifetime", "session", "daily", "yesterday", "weekly", "monthly"]
        result = {period: [] for period in periods}
        
        for row in rows:
            guild_name = row['name']
            guild_tag = row['guild_tag']
            guild_tag_color = row['guild_hex']
            lifetime_value = row['lifetime']
            is_tracked = row['is_tracked'] == 1
            
            # Calculate deltas for each period (like player stats)
            period_values = {
                'lifetime': lifetime_value,  # Lifetime is absolute value
                'session': lifetime_value - row['session'],  # Delta since session start
                'daily': lifetime_value - row['daily'],  # Delta since daily reset
                'yesterday': row['daily'] - row['yesterday'],  # Yesterday's delta
                'weekly': lifetime_value - row['weekly'],  # Delta since weekly reset
                'monthly': lifetime_value - row['monthly']  # Delta since monthly reset
            }
            
            # Debug output for first guild
            if guild_name == rows[0]['name']:
                print(f"[DEBUG] Guild: {guild_name} ({game}) - Tracked: {is_tracked}")
                print(f"[DEBUG]   Lifetime: {lifetime_value}")
                print(f"[DEBUG]   Snapshots - S:{row['session']} D:{row['daily']} Y:{row['yesterday']} W:{row['weekly']} M:{row['monthly']}")
                print(f"[DEBUG]   Deltas - S:{period_values['session']} D:{period_values['daily']} Y:{period_values['yesterday']} W:{period_values['weekly']} M:{period_values['monthly']}")
            
            for period in periods:
                value = period_values[period]
                
                # For delta-based periods, only include tracked guilds
                # For lifetime, include all guilds
                if period == "lifetime":
                    if value > 0 or period == "lifetime":
                        result[period].append((guild_name, value, guild_tag, guild_tag_color))
                else:
                    # Delta-based periods: only tracked guilds
                    if is_tracked and value > 0:
                        result[period].append((guild_name, value, guild_tag, guild_tag_color))
        
        # Sort each period by value descending
        for period in periods:
            result[period].sort(key=lambda x: x[1], reverse=True)
        
        return result


class GuildLeaderboardPeriodSelect(discord.ui.Select):
    def __init__(self, parent_view):
        self.parent_view = parent_view
        options = [
            discord.SelectOption(label="Lifetime", value="lifetime", default=True),
            discord.SelectOption(label="Session", value="session"),
            discord.SelectOption(label="Daily", value="daily"),
            discord.SelectOption(label="Yesterday", value="yesterday"),
            discord.SelectOption(label="Weekly", value="weekly"),
            discord.SelectOption(label="Monthly", value="monthly"),
        ]
        super().__init__(placeholder="Select time period...", options=options, row=0)
    
    async def callback(self, interaction: discord.Interaction):
        period = self.values[0]
        self.parent_view.current_period = period
        
        # Update default selection
        for option in self.options:
            option.default = (option.value == period)
        
        embed, file, _ = self.parent_view.generate_leaderboard_image(period, self.parent_view.page)
        if file:
            await interaction.response.edit_message(attachments=[file], view=self.parent_view)
        else:
            await interaction.response.edit_message(embed=embed, view=self.parent_view)


class GuildLeaderboardView(discord.ui.View):
    def __init__(self, game: str, data_cache: dict):
        super().__init__(timeout=180)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.game = game
        self.data_cache = data_cache
        self.current_period = "lifetime"
        self.page = 0
        self.page_size = 10
        self.message = None  # Store message reference for timeout handling
        
        # Game display names
        self.game_labels = CATEGORY_METRICS.get("guild", {})
        self.game_display = self.game_labels.get(game, game)
        
        # Period selector dropdown
        self.period_select = GuildLeaderboardPeriodSelect(self)
        self.add_item(self.period_select)
        
        # Navigation buttons
        self.prev_button = discord.ui.Button(label="◀ Previous", style=discord.ButtonStyle.primary, row=1)
        self.prev_button.callback = self.prev_page_callback
        self.add_item(self.prev_button)
        
        self.next_button = discord.ui.Button(label="Next ▶", style=discord.ButtonStyle.primary, row=1)
        self.next_button.callback = self.next_page_callback
        self.add_item(self.next_button)
        
        # Search button
        self.search_button = discord.ui.Button(label="🔍 Search", style=discord.ButtonStyle.secondary, row=1)
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
        self.search_button.callback = self.search_callback
        self.add_item(self.search_button)
    
    def _get_leaderboard(self, period: str):
        return self.data_cache.get(period, [])
    
    def _paginate(self, leaderboard: list, page: int):
        total_pages = max(1, (len(leaderboard) + self.page_size - 1) // self.page_size)
        clamped_page = max(0, min(page, total_pages - 1))
        start_index = clamped_page * self.page_size
        return leaderboard[start_index:start_index + self.page_size], total_pages, clamped_page, start_index
    
    def generate_leaderboard_image(self, period: str, page: int):
        leaderboard = self._get_leaderboard(period)
        
        if not leaderboard:
            empty_embed = self.get_leaderboard_embed(period, page=0, total_pages=1, leaderboard=leaderboard)
            return empty_embed, None, 1
        
        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard, page)
        self.page = clamped_page
        
        # Format data for image generation
        # For guilds: (rank, guild_name, level=0, icon="", name_hex, guild_tag, guild_tag_hex, value, is_playtime=False)
        
        image_data = []
        for idx, entry in enumerate(sliced):
            rank = start_index + idx + 1
            
            # Unpack based on whether we have tag data
            if len(entry) == 4:
                guild_name, exp_value, guild_tag, guild_tag_color = entry
            elif len(entry) == 2:
                guild_name, exp_value = entry
                guild_tag, guild_tag_color = None, None
            else:
                # Fallback for unexpected format
                guild_name = str(entry[0])
                exp_value = entry[1] if len(entry) > 1 else 0
                guild_tag = None
                guild_tag_color = None
            
            # Debug output
            if idx == 0:  # Only print first entry to avoid spam
                print(f"[DEBUG] Guild leaderboard entry: name={guild_name}, tag={guild_tag}, color={guild_tag_color}, exp={exp_value}")
            
            # Convert tag color name to hex using Minecraft color mapping
            if guild_tag_color:
                tag_hex = MINECRAFT_NAME_TO_HEX.get(guild_tag_color.upper(), "#AAAAAA")
            else:
                tag_hex = "#AAAAAA"
            
            # No prestige for guilds, name shown in white, tag with its color
            image_data.append((rank, guild_name, 0, "", "#FFFFFF", guild_tag or "", tag_hex, exp_value, False))
        
        if Image is not None:
            try:
                title_with_period = f"{period.title()} Guild {self.game_display}"
                img_io = create_leaderboard_image(title_with_period, "Experience", image_data, page=clamped_page, total_pages=total_pages)
                filename = f"guild_leaderboard_{self.game}_{period}_p{clamped_page + 1}.png"
                return None, discord.File(img_io, filename=filename), total_pages
            except Exception as e:
                print(f"[WARNING] Guild leaderboard image generation failed: {e}")
                import traceback
                traceback.print_exc()
                return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages
        else:
            return self.get_leaderboard_embed(period, clamped_page, total_pages, leaderboard), None, total_pages
    
    def get_leaderboard_embed(self, period: str, page: int = 0, total_pages: int = 1, leaderboard: Optional[list] = None):
        leaderboard_data = self._get_leaderboard(period) if leaderboard is None else leaderboard
        
        if not leaderboard_data:
            embed = discord.Embed(
                title=f"{period.title()} Guild {self.game_display} Leaderboard",
                description="No data available",
                color=discord.Color.from_rgb(54, 57, 63)
            )
            return embed
        
        sliced, total_pages, clamped_page, start_index = self._paginate(leaderboard_data, page)
        
        # Build leaderboard text
        lines = []
        for idx, entry in enumerate(sliced):
            rank = start_index + idx + 1
            
            # Unpack based on whether we have tag data
            if len(entry) == 4:
                guild_name, exp_value, guild_tag, guild_tag_color = entry
            elif len(entry) == 2:
                guild_name, exp_value = entry
                guild_tag = None
            else:
                guild_name = str(entry[0])
                exp_value = entry[1] if len(entry) > 1 else 0
                guild_tag = None
            
            # Format rank emoji
            if rank == 1:
                rank_emoji = "🥇"
            elif rank == 2:
                rank_emoji = "🥈"
            elif rank == 3:
                rank_emoji = "🥉"
            else:
                rank_emoji = f"`#{rank:>2}`"
            
            # Format exp value
            exp_str = f"{exp_value:,.0f}"
            
            # Build display name with tag if available
            display_name = f"**{guild_name}**"
            if guild_tag:
                display_name += f" [{guild_tag}]"
            
            lines.append(f"{rank_emoji} {display_name} - {exp_str} exp")
        
        description = "\n".join(lines)
        
        embed = discord.Embed(
            title=f"{period.title()} Guild {self.game_display} Leaderboard",
            description=description,
            color=discord.Color.gold()
        )
        embed.set_footer(text=f"Page {clamped_page + 1}/{total_pages} • {len(leaderboard_data)} guilds")
        
        return embed
    
    async def prev_page_callback(self, interaction: discord.Interaction):
        self.page = max(0, self.page - 1)
        embed, file, _ = self.generate_leaderboard_image(self.current_period, self.page)
        if file:
            await interaction.response.edit_message(attachments=[file], view=self)
        else:
            await interaction.response.edit_message(embed=embed, view=self)
    
    async def next_page_callback(self, interaction: discord.Interaction):
        leaderboard = self._get_leaderboard(self.current_period)
        total_pages = max(1, (len(leaderboard) + self.page_size - 1) // self.page_size)
        self.page = min(total_pages - 1, self.page + 1)
        embed, file, _ = self.generate_leaderboard_image(self.current_period, self.page)
        if file:
            await interaction.response.edit_message(attachments=[file], view=self)
        else:
            await interaction.response.edit_message(embed=embed, view=self)
    
    async def search_callback(self, interaction: discord.Interaction):
        modal = GuildSearchModal(self)
        await interaction.response.send_modal(modal)


class GuildSearchModal(discord.ui.Modal, title="Search Guild Leaderboard"):
    def __init__(self, parent_view):
        super().__init__()
        self.parent_view = parent_view
        
        self.search_input = discord.ui.TextInput(
            label="Guild Name or Position",
            placeholder="Enter guild name or position number (e.g., '5' or 'GS 2077')",
            style=discord.TextStyle.short,
            required=True
        )
        self.add_item(self.search_input)
    
    async def on_submit(self, interaction: discord.Interaction):
        search_term = self.search_input.value.strip()
        leaderboard = self.parent_view._get_leaderboard(self.parent_view.current_period)
        
        # Try parsing as position number first
        try:
            position = int(search_term)
            if 1 <= position <= len(leaderboard):
                # Jump to the page containing this position
                page = (position - 1) // self.parent_view.page_size
                self.parent_view.page = page
                embed, file, _ = self.parent_view.generate_leaderboard_image(self.parent_view.current_period, page)
                if file:
                    await interaction.response.edit_message(attachments=[file], view=self.parent_view)
                else:
                    await interaction.response.edit_message(embed=embed, view=self.parent_view)
                return
            else:
                await interaction.response.send_message(
                    f"❌ Position {position} is out of range (1-{len(leaderboard)}).",
                    ephemeral=True
                )
                return
        except ValueError:
            # Not a number, search by guild name
            pass
        
        # Search for guild name
        search_lower = search_term.lower()
        for idx, entry in enumerate(leaderboard):
            # Handle both old format (2 elements) and new format (4 elements)
            guild_name = entry[0]
            if search_lower in guild_name.lower():
                # Found guild - jump to its page
                page = idx // self.parent_view.page_size
                self.parent_view.page = page
                embed, file, _ = self.parent_view.generate_leaderboard_image(self.parent_view.current_period, page)
                if file:
                    await interaction.response.edit_message(attachments=[file], view=self.parent_view)
                else:
                    await interaction.response.edit_message(embed=embed, view=self.parent_view)
                return
        
        # Guild not found
        await interaction.response.send_message(
            f"❌ Guild '{search_term}' not found in {self.parent_view.current_period} leaderboard.",
            ephemeral=True
        )


def calculate_carried_score_average(wins, losses, kills, deaths, games_played):
    """Calculate average carried score from all 4 formulas.
    
    This matches the calculation used in /aretheycarried command.
    Returns average of all 4 formulas (0-5 scale).
    """
    if games_played == 0 or deaths == 0:
        return 0
    
    win_rate = wins / games_played
    kd_ratio = kills / deaths
    wl_ratio = wins / losses if losses > 0 else wins
    kills_per_game = kills / games_played
    deaths_per_game = deaths / games_played
    
    # Formula 1: Win-Performance Disparity
    normalized_kd_f1 = min(1.0, kd_ratio / 5.0)
    disparity_f1 = win_rate - normalized_kd_f1
    baseline_f1 = max(0, 0.3 - (kd_ratio * 0.02))
    score_formula1 = max(0, min(5, (disparity_f1 * 9) + baseline_f1))
    
    # Formula 2: Multi-Factor Weighted
    kd_factor = max(0, (4.0 - kd_ratio) / 4.0)
    kills_factor = max(0, (2.0 - kills_per_game) / 2.0)
    deaths_factor = min(1, deaths_per_game / 1.5)
    disparity_factor = max(0, win_rate - (kd_ratio / 5))
    
    score_formula2 = (
        kd_factor * 1.3 +
        kills_factor * 1.0 +
        deaths_factor * 0.4 +
        disparity_factor * 2.0
    )
    score_formula2 = max(0, min(5, score_formula2 * 0.9))
    
    # Formula 3: Ratio-Based
    if losses > 0 and kd_ratio > 0:
        ratio_gap = wl_ratio / kd_ratio
        score_formula3 = max(0, min(5, (ratio_gap - 0.6) * 2.2))
    else:
        score_formula3 = 0
    
    # Formula 4: Impact Score
    impact_score = (
        (min(5, kd_ratio) / 5) * 0.65 +
        (min(2.5, kills_per_game) / 2.5) * 0.35
    )
    performance_gap = win_rate - impact_score
    baseline_f4 = 0.15
    score_formula4 = max(0, min(5, (performance_gap * 8) + baseline_f4))
    
    # Return average of all 4 formulas
    avg_score = (score_formula1 + score_formula2 + score_formula3 + score_formula4) / 4
    return round(avg_score, 2)





@bot.tree.command(name="prestiges", description="List all prestige prefixes with their colors")
async def prestiges(interaction: discord.Interaction):
    # Defer in case composing takes a moment
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException):
            return
    try:
        # Use image rendering if Pillow is available
        if Image is not None:
            try:
                combined = render_all_prestiges_combined()
                await interaction.followup.send(file=discord.File(combined, filename="Wool Games prestiges 0-4000.png"))
            except Exception:
                # If combining fails, fall back to sending individual images
                for base in sorted(PRESTIGE_RAW_PATTERNS.keys()):
                    end_display = base + 99
                    try:
                        imgio = render_prestige_range_image(base, end_display)
                        fname = f"prestige_{base}.png"
                        await interaction.followup.send(file=discord.File(imgio, filename=fname))
                    except Exception:
                        prestige_str = format_prestige_ansi(base, '')
                        await interaction.followup.send(prestige_str)
        else:
            # Pillow not installed; fallback to ANSI list
            lines = []
            for base in sorted(PRESTIGE_RAW_PATTERNS.keys()):
                prestige_str = format_prestige_ansi(base, '')
                lines.append(prestige_str)
            await _send_paged_ansi_followups(interaction, lines, block='ansi')
    except Exception as e:
        await interaction.followup.send(f"[ERROR] {str(e)}")

class CarriedView(discord.ui.View):
    def __init__(self, ign: str, stats_data: dict):
        super().__init__(timeout=840)  # 14 minutes (840 seconds) - remove buttons before Discord's 15-minute limit
        self.ign = ign
        self.stats_data = stats_data
        self.current_tab = "lifetime"
        self.message = None  # Store message reference for timeout handling
        self.update_buttons()
    
    async def on_timeout(self):
        """Remove buttons when the view times out."""
        if self.message:
            try:
                await self.message.edit(view=None)
            except Exception:
                pass  # Message might be deleted or inaccessible
    
    def update_buttons(self):
        for child in self.children:
            if isinstance(child, discord.ui.Button):
                child.style = discord.ButtonStyle.primary if child.custom_id == self.current_tab else discord.ButtonStyle.secondary
    
    def calculate_carried_scores(self, tab_name: str):
        """Calculate all 4 carried scores for a given time period."""
        # Extract stats for the specified period
        wins = self.stats_data.get('wins', {}).get(tab_name, 0)
        losses = self.stats_data.get('losses', {}).get(tab_name, 0)
        kills = self.stats_data.get('kills', {}).get(tab_name, 0)
        deaths = self.stats_data.get('deaths', {}).get(tab_name, 0)
        games_played = self.stats_data.get('games_played', {}).get(tab_name, 0)
        
        if games_played == 0 or deaths == 0:
            return None
        
        # Calculate derived stats
        win_rate = (wins / games_played * 100) if games_played > 0 else 0
        kd_ratio = kills / deaths if deaths > 0 else kills
        wl_ratio = wins / losses if losses > 0 else wins
        kills_per_game = kills / games_played if games_played > 0 else 0
        
        # Formula 1: Win-Performance Disparity
        normalized_kd_f1 = min(1.0, kd_ratio / 5.0)
        disparity_f1 = (win_rate/100) - normalized_kd_f1
        baseline_f1 = max(0, 0.3 - (kd_ratio * 0.02))
        score_formula1 = max(0, min(5, (disparity_f1 * 9) + baseline_f1))
        score_formula1 = round(score_formula1, 1)
        
        # Formula 2: Multi-Factor Weighted (Used for assessment)
        kd_factor = max(0, (4.0 - kd_ratio) / 4.0)
        kills_factor = max(0, (2.0 - kills_per_game) / 2.0)
        deaths_per_game = deaths / games_played
        deaths_factor = min(1, deaths_per_game / 1.5)
        disparity_factor = max(0, (win_rate/100) - (kd_ratio / 5))
        
        score_formula2 = (
            kd_factor * 1.3 +
            kills_factor * 1.0 +
            deaths_factor * 0.4 +
            disparity_factor * 2.0
        )
        score_formula2 = max(0, min(5, score_formula2 * 0.9))
        score_formula2 = round(score_formula2, 1)
        
        # Formula 3: Ratio-Based
        if losses > 0 and deaths > 0:
            ratio_gap = wl_ratio / kd_ratio
            score_formula3 = max(0, min(5, (ratio_gap - 0.6) * 2.2))
        else:
            score_formula3 = 0
        score_formula3 = round(score_formula3, 1)
        
        # Formula 4: Impact Score
        impact_score = (
            (min(5, kd_ratio) / 5) * 0.65 +
            (min(2.5, kills_per_game) / 2.5) * 0.35
        )
        performance_gap = (win_rate/100) - impact_score
        baseline_f4 = 0.15
        score_formula4 = max(0, min(5, (performance_gap * 8) + baseline_f4))
        score_formula4 = round(score_formula4, 1)
        
        # Calculate average of all 4 formulas
        avg_score = round((score_formula1 + score_formula2 + score_formula3 + score_formula4) / 4, 1)
        
        return {
            'formula1': score_formula1,
            'formula2': score_formula2,
            'formula3': score_formula3,
            'formula4': score_formula4,
            'assessment_score': avg_score,  # Use average for assessment
            'wins': int(wins),
            'losses': int(losses),
            'kills': int(kills),
            'deaths': int(deaths),
            'games': int(games_played),
            'win_rate': win_rate,
            'wl_ratio': wl_ratio,
            'kd_ratio': kd_ratio,
            'kills_per_game': kills_per_game
        }
    
    def generate_embed(self, tab_name: str):
        """Generate the embed for a specific time period."""
        scores = self.calculate_carried_scores(tab_name)
        
        if scores is None:
            embed = discord.Embed(
                title=f"Are they carried? - {self.ign}",
                description=f"Insufficient data for {tab_name} period.",
                color=0x808080
            )
            return embed
        
        # Determine assessment based on Formula 2 only
        assessment_score = scores['assessment_score']
        
        if assessment_score >= 4.5:
            assessment = "Definitely carried"
            color = 0xFF0000  # Red
        elif assessment_score >= 3.5:
            assessment = "Most likely carried"
            color = 0xFF6600  # Orange-red
        elif assessment_score >= 2.5:
            assessment = "Carried"
            color = 0xFF9900  # Orange
        elif assessment_score >= 1.5:
            assessment = "Slightly carried"
            color = 0xFFCC00  # Yellow-orange
        elif assessment_score >= 0.5:
            assessment = "Not carried"
            color = 0x99FF00  # Yellow-green
        else:
            assessment = "Definitely not carried"
            color = 0x00FF00  # Green
        
        # Create embed
        tab_display = "All-time" if tab_name == "lifetime" else tab_name.title()
        embed = discord.Embed(
            title=f"Are they carried? - {self.ign}",
            description=f"**Period:** {tab_display}",
            color=color
        )
        
        # Add calculated rate field (Average of all 4 formulas)
        embed.add_field(
            name="Calculated Rate",
            value=f"{assessment_score:.1f}",
            inline=False
        )
        
        # Add assessment field
        embed.add_field(
            name="Assessment",
            value=assessment,
            inline=False
        )
        
        # Add stats used
        stats_text = (
            f"**Games Played:** {scores['games']:,}\n"
            f"**Wins:** {scores['wins']:,} | **Losses:** {scores['losses']:,}\n"
            f"**Win Rate:** {scores['win_rate']:.1f}% | **W/L Ratio:** {scores['wl_ratio']:.2f}\n"
            f"**Kills:** {scores['kills']:,} | **Deaths:** {scores['deaths']:,}\n"
            f"**K/D Ratio:** {scores['kd_ratio']:.2f} | **Kills/Game:** {scores['kills_per_game']:.2f}"
        )
        embed.add_field(
            name="Stats Used",
            value=stats_text,
            inline=False
        )
        
        # Add all formula scores
        formula_text = (
            f"**Formula 1 (Win-Performance Disparity):** {scores['formula1']:.1f}\n"
            f"**Formula 2 (Multi-Factor Weighted):** {scores['formula2']:.1f}\n"
            f"**Formula 3 (Ratio-Based):** {scores['formula3']:.1f}\n"
            f"**Formula 4 (Impact Score):** {scores['formula4']:.1f}\n"
            f"**Average:** {scores['assessment_score']:.1f}"
        )
        embed.add_field(
            name="Formula Breakdown",
            value=formula_text,
            inline=False
        )
        
        # Add footer with timestamp
        embed.set_footer(text=f"Data from Hypixel API • Calculated at {time.strftime('%H:%M', time.localtime())}")
        
        return embed
    
    async def handle_tab_click(self, interaction: discord.Interaction, tab_name: str):
        self.current_tab = tab_name
        self.update_buttons()
        embed = self.generate_embed(tab_name)
        await interaction.response.edit_message(embed=embed, view=self)
    
    @discord.ui.button(label="All-time", custom_id="lifetime")
    async def all_time(self, interaction, button):
        await self.handle_tab_click(interaction, "lifetime")
    
    @discord.ui.button(label="Session", custom_id="session")
    async def session(self, interaction, button):
        await self.handle_tab_click(interaction, "session")
    
    @discord.ui.button(label="Daily", custom_id="daily")
    async def daily(self, interaction, button):
        await self.handle_tab_click(interaction, "daily")
    
    @discord.ui.button(label="Yesterday", custom_id="yesterday")
    async def yesterday(self, interaction, button):
        await self.handle_tab_click(interaction, "yesterday")
    
    @discord.ui.button(label="Weekly", custom_id="weekly")
    async def weekly(self, interaction, button):
        await self.handle_tab_click(interaction, "weekly")
    
    @discord.ui.button(label="Monthly", custom_id="monthly")
    async def monthly(self, interaction, button):
        await self.handle_tab_click(interaction, "monthly")


@bot.tree.command(name="aretheycarried", description="Calculate how 'carried' a player is based on their stats")
@discord.app_commands.describe(ign="Minecraft IGN (required)")
async def aretheycarried(interaction: discord.Interaction, ign: str):
    print(f"[DEBUG] /aretheycarried triggered for IGN: {ign} by user: {interaction.user.name}")
    
    # Validate username
    ok, proper_ign = validate_and_normalize_ign(ign)
    if not ok:
        await interaction.response.send_message(f"The username {ign} is invalid.", ephemeral=True)
        return
    ign = proper_ign
    
    if not interaction.response.is_done():
        try:
            await interaction.response.defer()
        except (discord.errors.NotFound, discord.errors.HTTPException) as e:
            print(f"[DEBUG] Defer failed for {ign} in /aretheycarried: {e}")
            return

    try:
        # Fetch fresh stats
        print(f"[DEBUG] Running api_get.py for IGN: {ign} (/aretheycarried)")
        result = run_script("api_get.py", ["-ign", ign])
        
        if result.returncode != 0:
            error_msg = result.stderr or result.stdout or "Unknown error"
            await interaction.followup.send(f"[ERROR] Failed to fetch stats:\n```{error_msg[:500]}```")
            return

        # Get stats from database with calculated deltas
        stats = get_user_stats_with_deltas(ign)
        
        if not stats:
            await interaction.followup.send(f"[ERROR] No stats found for player '{ign}'")
            return
        
        # Check if user is tracked (UUID-aware)
        is_tracked = is_tracked_user(ign)
        
        # Create view with tabs
        view = CarriedView(ign, stats)
        embed = view.generate_embed("lifetime")
        
        if is_tracked:
            message = await interaction.followup.send(embed=embed, view=view)
            view.message = message  # Store message reference for timeout handling
        else:
            msg = f"`{ign}` is not currently tracked. Only all-time stats are available.\nUse `/track ign:{ign}` to start tracking and enable session/daily/monthly stats."
            await interaction.followup.send(content=msg, embed=embed)
        
    except Exception as e:
        import traceback
        error_traceback = traceback.format_exc()
        await send_error_with_report(
            interaction,
            "An unexpected error occurred while calculating carried scores.",
            f"{str(e)}\n\n{error_traceback}",
            "/aretheycarried",
            f"IGN: {ign}"
        )


@bot.tree.command(name="stopbot", description="Gracefully shutdown the bot (admin only)")
async def stopbot(interaction: discord.Interaction):
    if not is_admin(interaction.user):
        await interaction.response.send_message("❌ This command is only available to bot administrators.", ephemeral=True)
        return
    
    await interaction.response.send_message("🛑 Shutting down bot gracefully... Please wait for all tasks to complete.", ephemeral=True)
    print(f"[SHUTDOWN] Bot shutdown initiated by {interaction.user.name} ({interaction.user.id})")
    print(f"[SHUTDOWN] Waiting for background tasks to complete...")
    
    # Cancel background tasks gracefully
    if hasattr(bot, 'background_tasks'):
        for task in bot.background_tasks:
            if not task.done():
                task.cancel()
        
        # Wait for tasks to complete their cleanup with timeout
        results = await asyncio.gather(*bot.background_tasks, return_exceptions=True)
        for i, result in enumerate(results):
            if isinstance(result, asyncio.CancelledError):
                print(f"[SHUTDOWN] Background task {i+1} cancelled successfully")
            elif isinstance(result, Exception):
                print(f"[SHUTDOWN] Background task {i+1} raised exception: {result}")
            else:
                print(f"[SHUTDOWN] Background task {i+1} completed normally")
    
    # Wait for any pending Discord operations
    print(f"[SHUTDOWN] Waiting for pending operations...")
    await asyncio.sleep(2)  # Give time for any pending messages/edits
    
    print(f"[SHUTDOWN] All cleanup complete, closing bot...")
    await bot.close()

# Run bot
if __name__ == "__main__":
    # Prevent multiple instances with a lock file
    import fcntl
    lock_file = BOT_DIR / "bot.lock"
    
    try:
        # Try to acquire exclusive lock
        lock_fp = open(lock_file, 'w')
        fcntl.flock(lock_fp.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
        
        # Write PID to lock file
        lock_fp.write(str(os.getpid()))
        lock_fp.flush()
        
        print(f"[STARTUP] Bot starting with PID {os.getpid()}")
        print(f"[STARTUP] Lock file created at {lock_file}")
        
        try:
            bot.run(DISCORD_TOKEN)
        finally:
            # Clean up lock file on exit
            try:
                fcntl.flock(lock_fp.fileno(), fcntl.LOCK_UN)
                lock_fp.close()
                if lock_file.exists():
                    lock_file.unlink()
                print("[SHUTDOWN] Lock file removed")
            except:
                pass
                
    except IOError:
        # Lock file exists - check if process is actually running
        print(f"[WARNING] Lock file exists at {lock_file}")
        stale_lock = False
        
        try:
            with open(lock_file, 'r') as f:
                existing_pid = f.read().strip()
                
            if existing_pid.isdigit():
                existing_pid = int(existing_pid)
                # Check if process with that PID exists
                try:
                    os.kill(existing_pid, 0)  # Signal 0 just checks if process exists
                    # Process exists - real duplicate instance
                    print(f"[ERROR] Another instance of the bot is already running!")
                    print(f"[ERROR] Existing process PID: {existing_pid}")
                    print(f"[ERROR] Use 'systemctl stop sheepbot' or 'kill {existing_pid}' to stop it")
                    sys.exit(1)
                except OSError:
                    # Process doesn't exist - stale lock file
                    stale_lock = True
                    print(f"[INFO] Found stale lock file from PID {existing_pid} (process not running)")
            else:
                stale_lock = True
                print("[INFO] Lock file contains invalid PID")
                
        except (FileNotFoundError, ValueError):
            stale_lock = True
            print("[INFO] Lock file is empty or unreadable")
        
        # If lock is stale, remove it and retry
        if stale_lock:
            print("[INFO] Removing stale lock file and starting bot...")
            try:
                lock_file.unlink()
            except:
                pass
            
            # Retry acquiring lock
            try:
                lock_fp = open(lock_file, 'w')
                fcntl.flock(lock_fp.fileno(), fcntl.LOCK_EX | fcntl.LOCK_NB)
                lock_fp.write(str(os.getpid()))
                lock_fp.flush()
                
                print(f"[STARTUP] Bot starting with PID {os.getpid()}")
                print(f"[STARTUP] Lock file created at {lock_file}")
                
                try:
                    bot.run(DISCORD_TOKEN)
                finally:
                    try:
                        fcntl.flock(lock_fp.fileno(), fcntl.LOCK_UN)
                        lock_fp.close()
                        if lock_file.exists():
                            lock_file.unlink()
                        print("[SHUTDOWN] Lock file removed")
                    except:
                        pass
            except IOError:
                print("[ERROR] Failed to acquire lock even after removing stale lock file")
                sys.exit(1) 